# -*- coding: utf-8 -*-
"""
Build the 3 PHD KoboToolbox XLSForms from the final source files.

Source files → Form mapping:
  Form 1  Registration         ← Master list.xlsx (Motherlist-BBFSWs + Parameters)
  Form 2  Patient Services     ← Patient Record Register + HTC Service register
                                  + Counselling Report + referral register
  Form 3  Activity & Ops       ← group health education + event database
                                  + material database + gbv corner database
                                  + Stock register

Rules:
  - PHD only — organisation is a hidden calculate field, no choice shown.
  - No fields that are not in the source files.
  - Bangla labels from the source files where available.
"""
import os
import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.core.management.base import BaseCommand
from programs.models import ServiceCenter

HERE   = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.normpath(os.path.join(HERE, '..', '..', '..', '..', 'koboforms'))
KOBO_BASE = 'https://kf.kobotoolbox.org'

_HFILL = PatternFill("solid", fgColor="003F72")
_HFONT = Font(color="FFFFFF", bold=True, size=10)


# ─── XLSForm helpers ──────────────────────────────────────────────────────────

SURVEY_HDR = [
    'type','name','label::English','label::Bangla',
    'hint','required','relevant','constraint','constraint_message',
    'default','appearance','calculation',
]
CHOICES_HDR = ['list_name','name','label::English','label::Bangla']
SETTINGS_HDR = ['form_title','form_id','version','default_language','style']


def _sr(qtype, name, en='', bn='', hint='', required='',
        relevant='', constraint='', cmsg='', default='', app='', calc=''):
    return [qtype, name, en, bn, hint, required, relevant,
            constraint, cmsg, default, app, calc]


def _ch(lst, name, en, bn=''):
    return [lst, name, en, bn]


def _wb(form_id, form_title, survey, choices):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in [
        ('survey',  SURVEY_HDR,   survey),
        ('choices', CHOICES_HDR,  choices),
        # 'theme-grid' (no 'pages') → single scrolling page like the MPDSR
        # form. 'pages' makes Enketo show one section at a time with
        # Next/Previous buttons — slow on phones and gets in the way of
        # the relevant logic since the page boundary breaks visibility.
        ('settings', SETTINGS_HDR, [[form_title, form_id, '20260606', 'English', 'theme-grid']]),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for ci in range(1, len(headers)+1):
            c = ws.cell(1, ci)
            c.font = _HFONT
            c.fill = _HFILL
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for ci in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(ci)].width = 28
        ws.freeze_panes = 'A2'
    return wb


# ─── Shared header (GPS + PHD hidden + centre + enumerator) ──────────────────

def _meta(center_required=True):
    req = 'yes' if center_required else ''
    return [
        _sr('begin_group','grp_meta','Submission info','তথ্য প্রেরণ'),
        _sr('calculate','organisation','','',calc="'PHD'"),
        _sr('geopoint','location',
            'GPS location (auto — captured when available)',
            'জিপিএস অবস্থান (পাওয়া গেলে স্বয়ংক্রিয়)', required=''),
        _sr('date','collection_date','Submission date (today)','প্রেরণের তারিখ (আজ)', required='yes'),
        # Dropdown of PHD's 9 wellness centres (PHD request, 2026-06-08).
        # The choice VALUE is the official Wellness Centre ID (R001..D009);
        # the webhook _get_center resolves it via code__iexact. Only PHD's
        # own centres are listed — nothing else leaks from the backend.
        _sr('select_one wellness_centre','centre_id',
            'Wellness Centre',
            'ওয়েলনেস সেন্টার', required=req,
            hint='Select your wellness centre.'),
        _sr('select_one medical_assistant','enumerator',
            'Your name (person filling this form)',
            'আপনার নাম (কে পূরণ করছেন)', required='yes',
            hint='Select your name from the list.'),
        _sr('text','enumerator_other',
            'If your name is not in the list, type it here',
            'আপনার নাম তালিকায় না থাকলে এখানে লিখুন',
            required='yes',
            relevant="${enumerator}='other'"),
        # The rest of the system reads enumerator_name (counsellor_name,
        # facilitator_name, the webhook name-stamp …). Resolve it to the human
        # name: the typed name for 'Other', else the chosen slug with its
        # underscores turned back into spaces ('Nilufar_Yesmin' → 'Nilufar
        # Yesmin'). translate() is already used elsewhere in this form.
        _sr('calculate','enumerator_name',
            calc="if(${enumerator}='other', ${enumerator_other}, "
                 "translate(${enumerator},'_',' '))"),
        _sr('end_group','grp_meta'),
    ]


def _centre_choices():
    """The 9 PHD wellness centres as a select_one 'wellness_centre' list.
    Single-sourced from seed_centers.PHD_BROTHELS so the dropdown can never
    drift from the seeded ServiceCenter rows. Choice value = Wellness Centre
    ID (R001..D009); label shows the name + ID for the field worker."""
    from .seed_centers import PHD_BROTHELS
    return [
        _ch('wellness_centre', c['code'],
            f"{c['name']} ({c['code']})",
            f"{c.get('name_bangla', c['name'])} ({c['code']})")
        for c in PHD_BROTHELS
    ]


# Medical Assistants (PHD field staff) — source "List Of Medical Assistants.xlsx"
# (2026-06). enumerator_name is a dropdown of these so (a) we know which MA
# submitted — they all share the one ciprb123 KoboCollect login — and (b) the
# names stay consistent. Tuple = (value slug, display name, centre/location).
PHD_MEDICAL_ASSISTANTS = [
    ('Nilufar_Yesmin',       'Nilufar Yesmin',       'Daulatdia'),
    ('AL_Mondna_Mim',        'AL Mondna Mim',        'Daulatdia'),
    ('Mst_Munzira_Khatun',   'Mst. Munzira Khatun',  'Faridpur'),
    ('Sharmin_Akter_Asa',    'Sharmin Akter Asa',    'Tangail'),
    ('Mansura_Khatun',       'Mansura Khatun',        'Jeshore'),
    ('Shipra_Roy',           'Shipra Roy',            'Baniashanta'),
    ('Mahaboba_Sharin_Moni', 'Mahaboba Sharin Moni',  'Jamalpur'),
    ('Sathi_Khatun',         'Sathi Khatun',          'Moymonsing'),
    ('Khukumoni_Adhikary',   'Khukumoni Adhikary',    'Pouthokhilae'),
    ('Moumita_Montho',       'Moumita Montho',        'Bagharhat'),
]


def _ma_choices():
    """select_one 'medical_assistant' — the PHD field staff who fill the forms.
    Label = name + location so each MA finds herself; value = a stable slug
    stored in enumerator_name. Plus an 'Other' escape so an unlisted submitter
    is never blocked (she types her name in enumerator_other)."""
    rows = [
        _ch('medical_assistant', slug, f'{name} ({loc})', f'{name} ({loc})')
        for slug, name, loc in PHD_MEDICAL_ASSISTANTS
    ]
    rows.append(_ch('medical_assistant', 'other',
                    'Other (not in the list)', 'অন্য (তালিকায় নেই)'))
    return rows


# ─── FORM 1: FSW Registration ─────────────────────────────────────────────────
# Source: Master list.xlsx  →  Motherlist-BBFSWs columns + Parameters sheet

def _form1_survey():
    rows = _meta()
    rows += [
        _sr('begin_group','grp_fsw','FSW Registration','যৌনকর্মী নিবন্ধন'),

        # Expected ID prefix for the chosen centre: the last digit of the
        # centre code is its serial → '1-' (R001) … '9-' (D009). Forces the
        # typed ID to belong to the selected wellness centre (1- Daulatdia,
        # 2- Maroawary, …). NOTE: use ODK 'substr' (0-indexed), NOT the XPath
        # 'substring' — JavaRosa/Enketo cannot evaluate 'substring'.
        _sr('calculate','_exp_prefix',
            calc="concat(substr(${centre_id},3),'-')"),

        _sr('note','_prefix_hint',
            'For this centre, the ID must start with: ${_exp_prefix} (e.g. ${_exp_prefix}0001)',
            'এই কেন্দ্রের জন্য আইডি শুরু হতে হবে: ${_exp_prefix} দিয়ে (যেমন ${_exp_prefix}0001)',
            relevant="${centre_id}!=''"),

        _sr('text','id_no',
            'ID No. (unique per FSW)',
            'আইডি নম্বর (অনন্য)',
            required='yes',
            # Two hard rules — the form will not advance unless BOTH hold:
            #   (1) the ID starts with the selected centre's serial prefix
            #       (_exp_prefix), so a Daulatdia worker can't file a 2-… ID;
            #   (2) the ID is not already in the Master List (phd_clients.csv).
            # Same trim+upper normalisation as _dup_name below, so '1-0001' /
            # ' 1-0001 ' all collide. The developer reassigns IDs from the
            # dashboard, not this field form, so they are exempt.
            constraint=("starts-with(normalize-space(.), ${_exp_prefix}) and "
                        "pulldata('phd_clients','name','id_no',"
                        "translate(normalize-space(.),"
                        "'abcdefghijklmnopqrstuvwxyz',"
                        "'ABCDEFGHIJKLMNOPQRSTUVWXYZ'))=''"),
            cmsg='⚠ This ID cannot be saved — see the red message above: it either '
                 'does not start with this centre number, or it is already registered. / '
                 'এই আইডি সংরক্ষণ করা যাবে না — উপরের লাল বার্তা দেখুন।',
            hint='Format: centre number + serial, e.g. 1-0001 (Daulatdia), '
                 '2-0001 (Jashore). Use the same ID in every Service Log.'),

        # Duplicate-ID warning. Looks up the typed ID in phd_clients.csv;
        # if she's already there, blocks accidental re-registration.
        # XPath 1.0 equivalent of trim+upper: translate(normalize-space(...)).
        # (Kobo/ODK rejected upper-case() — that's XPath 2.0.)
        _sr('calculate','_dup_name',
            calc=("pulldata('phd_clients','name','id_no',"
                  "translate(normalize-space(${id_no}),"
                  "'abcdefghijklmnopqrstuvwxyz',"
                  "'ABCDEFGHIJKLMNOPQRSTUVWXYZ'))")),
        _sr('note','_dup_warn',
            '⚠ This ID is already registered for ${_dup_name}. '
            'Do not re-register her — use her existing ID in the Service Log instead.',
            '⚠ এই আইডি ইতিমধ্যে ${_dup_name} এর জন্য নিবন্ধিত। '
            'পুনঃনিবন্ধন করবেন না — সেবা লগে তাঁর বিদ্যমান আইডি ব্যবহার করুন।',
            relevant="${id_no}!='' and ${_dup_name}!=''"),

        # Centre-mismatch warning — inline hint shown the moment the typed ID
        # does not start with the selected centre's prefix (the constraint
        # above also blocks it; this just tells her the expected prefix).
        _sr('note','_centre_warn',
            '⚠ This ID does not match the selected centre. It must start with ${_exp_prefix}',
            '⚠ এই আইডি নির্বাচিত কেন্দ্রের সাথে মেলে না। এটি ${_exp_prefix} দিয়ে শুরু হতে হবে।',
            relevant=("${id_no}!='' and ${centre_id}!='' and "
                      "not(starts-with(normalize-space(${id_no}), ${_exp_prefix}))")),

        _sr('text','name',
            'Name','নাম', required='yes'),

        _sr('text','mother_name',
            "Mother's name","মাতার নাম"),

        _sr('integer','birth_year',
            'Birth year','জন্ম সাল',
            hint='4-digit year, e.g. 1990 / ৪-সংখ্যার সাল, যেমন ১৯৯০',
            constraint='. >= 1940 and . <= 2010',
            cmsg='Must be a year between 1940 and 2010. / ১৯৪০–২০১০ এর মধ্যে একটি সাল হতে হবে।'),

        _sr('text','permanent_address',
            'Permanent address','স্থায়ী ঠিকানা', app='multiline'),

        _sr('select_one education','education',
            'Education','শিক্ষা'),

        _sr('select_one marital_status','marital_status',
            'Marital status','বৈবাহিক অবস্থা'),

        _sr('integer','years_in_profession',
            'How many years in this profession?',
            'কত বছর ধরে এই পেশায়?'),

        # Average clients — value + period (per day/week/month/year)
        # 4 separate sub-columns exactly as in masterlist r6 (day/week/month/year)
        _sr('begin_group','grp_avg_clients',
            'Average sex-work contacts (fill applicable period)',
            'সাধারণত গড়ে কতবার যৌনকাজ করেন (প্রযোজ্য ঘরটি পূরণ করুন)'),
        _sr('integer','avg_clients_per_day',
            'Per day','দিনে'),
        _sr('integer','avg_clients_per_week',
            'Per week','সপ্তাহে'),
        _sr('integer','avg_clients_per_month',
            'Per month','মাসে'),
        _sr('integer','avg_clients_per_year',
            'Per year','বছরে'),
        _sr('end_group','grp_avg_clients'),

        _sr('integer','children_under_18',
            'Number of children under 18',
            '১৮ বছরের নিচে সন্তান সংখ্যা'),

        _sr('select_one yes_no','tobacco_use',
            'Uses tobacco or drugs? (Yes/No)',
            'নেশা গ্রহণ করেন কি না?'),

        _sr('select_one yes_no','has_nid',
            'Has National ID? (Yes/No)',
            'জাতীয় পরিচয়পত্র আছে কি না?'),

        _sr('select_one yes_no','uses_fp',
            'Uses any family-planning method? (Yes/No)',
            'পরিবার পরিকল্পনার কোনো পদ্ধতি ব্যবহার করেন?'),

        _sr('text','fp_method',
            'Which FP method?','কোন পদ্ধতি?',
            relevant="${uses_fp}='yes'"),

        # RiCH quarterly tracking — 3 sub-columns from masterlist r5
        _sr('begin_group','grp_rich',
            'RiCH details (quarterly months)',
            'রিচ বিবরণ (কোয়ার্টারের মাস)'),
        _sr('text','rich_q_month_1',
            '1st month of quarter',
            'কোয়ার্টারের প্রথম মাস'),
        _sr('text','rich_q_month_2',
            '2nd month of quarter',
            'কোয়ার্টারের দ্বিতীয় মাস'),
        _sr('text','rich_q_month_3',
            '3rd month of quarter',
            'কোয়ার্টারের তৃতীয় মাস'),
        _sr('end_group','grp_rich'),

        _sr('text','remarks',
            'Remarks','মন্তব্য', app='multiline'),

        _sr('end_group','grp_fsw'),
    ]
    return rows


def _form1_choices():
    rows = []
    rows += [_ch('yes_no','yes','Yes','হ্যাঁ'), _ch('yes_no','no','No','না')]
    for v,en,bn in [
        ('1','Illiterate','নিরক্ষর'),
        ('2','Primary','প্রাথমিক'),
        ('3','Secondary','মাধ্যমিক'),
        ('4','Higher Secondary','উচ্চ মাধ্যমিক'),
        ('5','Graduate / Masters','স্নাতক/স্নাতকোত্তর'),
    ]:
        rows.append(_ch('education', v, en, bn))
    for v,en,bn in [
        ('1','Single — never married','অবিবাহিত'),
        ('2','Married','বিবাহিত'),
        ('3','Widowed','বিধবা'),
        ('4','Separated','আলাদা'),
        ('5','Divorced','তালাকপ্রাপ্ত'),
        ('6','Others','অন্যান্য'),
    ]:
        rows.append(_ch('marital_status', v, en, bn))
    rows += _centre_choices()
    rows += _ma_choices()
    return rows


# ─── FORM 2: Patient Services ─────────────────────────────────────────────────
# Sections: clinic | htc | counselling | referral
# Sources:
#   clinic      ← Patient Record Register_Rev_May 2026.xlsx
#   htc         ← HTC Service register.xlsx
#   counselling ← Counselling Report.docx  (monthly aggregate per counsellor)
#   referral    ← referral register.xlsx

def _form2_survey():
    rows = _meta()
    rows += [
        _sr('select_one service_type','service_type',
            'What service are you recording?',
            'কোন সেবা নথিভুক্ত করছেন?', required='yes'),
    ]

    # ── Clinic Visit (Patient Record Register) ────────────────────────────────
    REL_C = "${record_type}='clinic'"
    rows += [
        _sr('begin_group','sec_clinic',
            'Clinic Visit / Patient Record',
            'ক্লিনিক ভিজিট / রোগীর তথ্য',
            relevant=REL_C),

        _sr('date','clinic_date','Clinic visit date','ক্লিনিক ভিজিটের তারিখ', required='yes'),
        # client_id + age + sex live in the shared patient_id_group at the
        # top of the form — pulled from the Master List CSV via pulldata().
        # Don't repeat them here.

        # Screenings
        _sr('begin_group','grp_clinic_screen','Screenings','স্ক্রিনিং'),
        _sr('select_one yes_no','clinic_hiv_screen','HIV Screening','এইচআইভি স্ক্রিনিং'),
        _sr('select_one yes_no','clinic_syphilis_screen','Syphilis Screening','সিফিলিস স্ক্রিনিং'),
        _sr('select_one yes_no','clinic_hepb_screen','Hepatitis B Screening','হেপাটাইটিস বি স্ক্রিনিং'),
        _sr('select_one yes_no','clinic_hepc_screen','Hepatitis C Screening','হেপাটাইটিস সি স্ক্রিনিং'),
        _sr('select_one yes_no','clinic_tb_screen','TB Screening','যক্ষ্মা স্ক্রিনিং'),
        _sr('select_one yes_no','clinic_gbv_screen','GBV Screening','জিবিভি স্ক্রিনিং'),
        _sr('select_one yes_no','clinic_mh_screen','Mental Health Screening','মানসিক স্বাস্থ্য স্ক্রিনিং'),
        _sr('end_group','grp_clinic_screen'),

        _sr('text','clinic_chief_complaints','Chief Complaints','প্রধান অভিযোগ', app='multiline'),
        _sr('select_one visit_type','clinic_visit_type','STI Case Type','এসটিআই কেস ধরন'),

        # STI Diagnoses
        _sr('begin_group','grp_clinic_diag','STI Diagnosis','এসটিআই নির্ণয়'),
        _sr('select_one yes_no','clinic_diag_uds','UDS','ইউডিএস'),
        _sr('select_one yes_no','clinic_diag_vds','VDS','ভিডিএস'),
        _sr('select_one yes_no','clinic_diag_gu','GU','জিইউ'),
        _sr('select_one yes_no','clinic_diag_pid','PID','পিআইডি'),
        _sr('select_one yes_no','clinic_diag_ss','SS','এসএস'),
        _sr('select_one yes_no','clinic_diag_ib','IB','আইবি'),
        _sr('select_one yes_no','clinic_diag_anal_sti','Anal STIs','অ্যানাল এসটিআই'),
        _sr('select_one yes_no','clinic_diag_hiv','HIV','এইচআইভি'),
        _sr('select_one yes_no','clinic_diag_gh','GH (General Health)','জিএইচ (সাধারণ স্বাস্থ্য)'),
        _sr('end_group','grp_clinic_diag'),

        _sr('text','clinic_treatment','Treatment Provided (medicine name & quantity)',
            'প্রদত্ত চিকিৎসা (ওষুধের নাম ও পরিমাণ)', app='multiline'),
        _sr('select_one treatment_timing','clinic_treatment_timing',
            'Seeking treatment after onset of STI symptoms',
            'এসটিআই লক্ষণ শুরুর পর চিকিৎসা নিতে এসেছেন'),

        _sr('integer','clinic_condom_demo',
            '# of condom demonstrations','কনডম প্রদর্শনীর সংখ্যা'),

        # Follow up
        _sr('begin_group','grp_clinic_fu','Follow Up','ফলো-আপ'),
        _sr('date','clinic_fu_due','Follow-up due date','ফলো-আপের নির্ধারিত তারিখ'),
        _sr('date','clinic_fu_done','Follow-up done date','ফলো-আপ সম্পন্নের তারিখ'),
        _sr('select_one yes_no','clinic_adr','Adverse Drug Reaction (Y/N)','এডিআর (হ্যাঁ/না)'),
        _sr('end_group','grp_clinic_fu'),

        # Referrals (from source: STI-confirmatory, STI-non-responsive, STI-Partner,
        #            MCH/RH, GBV, MHPSS, MR, TB, General Health, EPI)
        _sr('begin_group','grp_clinic_ref','Referral Cases','রেফারেল কেস'),
        _sr('select_one yes_no','clinic_ref_sti_confirm','STI-confirmatory test','এসটিআই নিশ্চিত পরীক্ষা'),
        _sr('select_one yes_no','clinic_ref_sti_nonres','STI-non-responsive / complicated','এসটিআই অ-প্রতিক্রিয়াশীল'),
        _sr('select_one yes_no','clinic_ref_sti_partner','STI-Partner','এসটিআই সঙ্গী'),
        _sr('select_one yes_no','clinic_ref_mch','MCH / Reproductive Health','এমসিএইচ / প্রজনন স্বাস্থ্য'),
        _sr('select_one yes_no','clinic_ref_gbv','GBV referral','জিবিভি রেফারেল'),
        _sr('select_one yes_no','clinic_ref_mhpss','MHPSS referral','এমএইচপিএসএস রেফারেল'),
        _sr('select_one yes_no','clinic_ref_mr','MR','এমআর'),
        _sr('select_one yes_no','clinic_ref_tb','TB','যক্ষ্মা'),
        _sr('select_one yes_no','clinic_ref_gh','General Health','সাধারণ স্বাস্থ্য'),
        _sr('select_one yes_no','clinic_ref_epi','EPI','ইপিআই'),
        _sr('end_group','grp_clinic_ref'),

        _sr('select_one yes_no','clinic_contraception',
            'Use of current contraception (Y/N)',
            'বর্তমান গর্ভনিরোধক ব্যবহার (হ্যাঁ/না)'),
        _sr('text','clinic_contraception_method',
            'Which method?','কোন পদ্ধতি?',
            relevant="${clinic_contraception}='yes'"),

        _sr('end_group','sec_clinic'),
    ]

    # ── HTC Service Register ──────────────────────────────────────────────────
    REL_H = "${record_type}='htc'"
    rows += [
        _sr('begin_group','sec_htc',
            'HTC Service Register','এইচটিসি সেবা রেজিস্টার',
            relevant=REL_H),

        _sr('date','htc_date','Date (dd/mm/yy)','তারিখ', required='yes'),
        # client_id + age + sex are in the shared patient_id_group at
        # the top of the form. The partner-of-FSW case (P-prefix IDs) is
        # handled by the normalised lookup — no separate field here.
        _sr('select_one test_occasion','htc_test_type',
            'New test or Re-test',
            'নতুন পরীক্ষা নাকি পুনরায় পরীক্ষা'),
        _sr('select_one tested_at','htc_tested_at',
            'Tested at (DIC=1 / Outreach or Mobile Camp=2)',
            'কোথায় পরীক্ষা (DIC=1 / আউটরিচ বা মোবাইল ক্যাম্প=2)'),
        _sr('select_one tested_by','htc_tested_by',
            'Tested by (Paramedic/MA=1 / CO/FM/PV/CW=2)',
            'কে পরীক্ষা করলেন (প্যারামেডিক/এমএ=1 / CO/FM/PV/CW=2)'),
        _sr('select_one yes_no','htc_pretest_counsel',
            'Pre-test counselling (Y/N)','পরীক্ষা-পূর্ব কাউন্সেলিং (হ্যাঁ/না)'),

        # 3-test algorithm
        _sr('begin_group','grp_htc_tests','HIV Test Results','এইচআইভি পরীক্ষার ফলাফল'),
        _sr('select_one rnrinv','htc_test1',
            'Test 1: Determine (R / NR / INV)',
            'পরীক্ষা ১: ডিটারমাইন'),
        _sr('select_one rnrinv','htc_test2',
            'Test 2: Uni Gold (R / NR / INV)',
            'পরীক্ষা ২: ইউনি গোল্ড',
            relevant="${htc_test1}='R'"),
        _sr('select_one rnrinv','htc_test3',
            'Test 3: First Response (R / NR / INV)',
            'পরীক্ষা ৩: ফার্স্ট রেসপন্স',
            relevant="${htc_test1}='R'"),
        _sr('select_one final_result','htc_final_result',
            'Final Result (Positive / Negative / Indeterminate)',
            'চূড়ান্ত ফলাফল'),
        _sr('end_group','grp_htc_tests'),

        _sr('select_one yes_no','htc_eqa',
            'Send for EQA? (tick if yes)','ইকিউএ-তে পাঠাবেন?'),
        _sr('date','htc_dbs_date',
            'DBS prepared on (dd/mm/yy)','ডিবিএস তৈরির তারিখ',
            relevant="${htc_eqa}='yes'"),
        _sr('text','htc_eqa_result',
            'Result from EQA / Retesting','ইকিউএ / পুনঃপরীক্ষার ফলাফল',
            relevant="${htc_eqa}='yes'"),

        _sr('select_one yes_no','htc_posttest_counsel',
            'Post-test counselling (Y/N)','পরীক্ষা-পরবর্তী কাউন্সেলিং'),
        _sr('select_one yes_no','htc_client_received_result',
            'Client received result (Y/N)','ক্লায়েন্ট ফলাফল পেয়েছেন?'),
        _sr('text','htc_remarks','Remarks','মন্তব্য', app='multiline'),

        _sr('end_group','sec_htc'),
    ]

    # ── Counselling Report (monthly aggregate per counsellor) ─────────────────
    REL_CO = "${record_type}='counselling'"
    rows += [
        _sr('begin_group','sec_counsel',
            'Counselling Report (monthly)',
            'কাউন্সেলিং রিপোর্ট (মাসিক)',
            relevant=REL_CO),

        _sr('note','_counsel_aggregate_note',
            'ℹ This is a monthly summary — counts are stored as a report, not per-patient records.',
            'ℹ এটি মাসিক সারসংক্ষেপ — গণনা রিপোর্ট হিসেবে সংরক্ষিত হয়, রোগীভিত্তিক নয়।'),

        _sr('text','counsel_prepared_by',
            'Prepared by','প্রস্তুতকারী', required='yes'),
        _sr('date','counsel_date',
            'Date','তারিখ', required='yes'),
        _sr('text','counsel_month',
            'Name of the month','মাসের নাম', required='yes',
            hint='e.g. June 2026'),
        _sr('text','counsel_counsellor',
            'Counsellor (Medical Assistant / Midwife cum Counsellor)',
            'কাউন্সেলর (মেডিকেল অ্যাসিস্ট্যান্ট / মিডওয়াইফ কাম কাউন্সেলর)',
            required='yes'),

        _sr('integer','counsel_hiv_test',
            'HIV Test & Counselling (count)',
            'এইচআইভি পরীক্ষা ও কাউন্সেলিং (সংখ্যা)'),
        _sr('integer','counsel_sti',
            'STI Counselling (count)',
            'এসটিআই কাউন্সেলিং (সংখ্যা)'),
        _sr('integer','counsel_srhr',
            'SRHR Counselling (count)',
            'এসআরএইচআর কাউন্সেলিং (সংখ্যা)'),
        _sr('integer','counsel_gbv',
            'GBV Counselling (count)',
            'জিবিভি কাউন্সেলিং (সংখ্যা)'),
        _sr('integer','counsel_art',
            'ART Counselling (count)',
            'এআরটি কাউন্সেলিং (সংখ্যা)'),
        _sr('integer','counsel_mh',
            'Mental Health Counselling (count)',
            'মানসিক স্বাস্থ্য কাউন্সেলিং (সংখ্যা)'),
        _sr('calculate','counsel_total',
            'Total number of individual counselling (auto)',
            'মোট ব্যক্তিগত কাউন্সেলিং (স্বয়ংক্রিয়)',
            calc='${counsel_hiv_test} + ${counsel_sti} + ${counsel_srhr} + ${counsel_gbv} + ${counsel_art} + ${counsel_mh}'),
        _sr('integer','counsel_group_mh',
            'Total Mental Health Group Counselling (count)',
            'মোট মানসিক স্বাস্থ্য গ্রুপ কাউন্সেলিং (সংখ্যা)'),
        _sr('text','counsel_note','Note','মন্তব্য', app='multiline'),

        _sr('end_group','sec_counsel'),
    ]

    # ── Referral Register ─────────────────────────────────────────────────────
    REL_R = "${record_type}='referral'"
    rows += [
        _sr('begin_group','sec_referral',
            'Referral Register','রেফারেল রেজিস্টার',
            relevant=REL_R),

        _sr('text','ref_month_year',
            'Month and Year','মাস ও বছর',
            required='yes', hint='e.g. June 2026'),
        _sr('date','ref_date','Date','তারিখ', required='yes'),
        # client_id lives in the shared patient_id_group at the top of the form.
        _sr('text','ref_referred_for',
            'Referred for','কী কারণে রেফার'),
        _sr('text','ref_referred_to',
            'Referred to','কোথায় রেফার'),
        _sr('date','ref_date_received',
            'Date of receiving service',
            'সেবা প্রাপ্তির তারিখ'),
        _sr('date','ref_followup_date',
            'Date of follow-up (if applicable)',
            'ফলো-আপের তারিখ (প্রযোজ্য হলে)'),
        _sr('text','ref_remarks',
            'Remarks (for TB: write the result)',
            'মন্তব্য (যক্ষ্মার ক্ষেত্রে ফলাফল লিখুন)',
            app='multiline'),

        _sr('end_group','sec_referral'),
    ]

    return rows


def _form2_choices():
    rows = []
    rows += [_ch('yes_no','yes','Yes','হ্যাঁ'), _ch('yes_no','no','No','না')]
    for v,en,bn in [
        ('clinic','Clinic Visit / Patient Record','ক্লিনিক ভিজিট / রোগীর তথ্য'),
        ('htc','HTC Service (HIV test)','এইচটিসি সেবা (এইচআইভি পরীক্ষা)'),
        ('counselling','Counselling Report (monthly)','কাউন্সেলিং রিপোর্ট (মাসিক)'),
        ('referral','Referral','রেফারেল'),
    ]:
        rows.append(_ch('service_type', v, en, bn))
    for v,en,bn in [
        ('new','New','নতুন'),
        ('follow_up','Follow-up','ফলো-আপ'),
        ('recurrent','Recurrent (within last 6 months)','পুনরাবৃত্তি (শেষ ৬ মাসে)'),
    ]:
        rows.append(_ch('visit_type', v, en, bn))
    for v,en,bn in [
        ('within_7','Within 7 days','৭ দিনের মধ্যে'),
        ('more_7','More than 7 days','৭ দিনের বেশি'),
    ]:
        rows.append(_ch('treatment_timing', v, en, bn))
    for v,en,bn in [
        ('M','Male','পুরুষ'), ('F','Female','মহিলা'),
    ]:
        rows.append(_ch('sex', v, en, bn))
    for v,en,bn in [
        ('new','New test','নতুন পরীক্ষা'),
        ('retest','Re-test (2nd/3rd time same year)','পুনরায় পরীক্ষা'),
    ]:
        rows.append(_ch('test_occasion', v, en, bn))
    for v,en,bn in [
        ('1','DIC','ডিআইসি'),
        ('2','Outreach / Mobile Camp','আউটরিচ / মোবাইল ক্যাম্প'),
    ]:
        rows.append(_ch('tested_at', v, en, bn))
    for v,en,bn in [
        ('1','Paramedic / MA','প্যারামেডিক / এমএ'),
        ('2','CO / FM / PV / CW','CO / FM / PV / CW'),
    ]:
        rows.append(_ch('tested_by', v, en, bn))
    for v,en in [('R','R (Reactive)'),('NR','NR (Non-Reactive)'),('INV','INV (Invalid)')]:
        rows.append(_ch('rnrinv', v, en, ''))
    for v,en,bn in [
        ('positive','Positive','পজিটিভ'),
        ('negative','Negative','নেগেটিভ'),
        ('indeterminate','Indeterminate','অনির্ধারিত'),
    ]:
        rows.append(_ch('final_result', v, en, bn))
    rows += _centre_choices()
    return rows


# ─── FORM 3: Activity & Operations ────────────────────────────────────────────
# Sections: group_edu | event | material | gbv_corner | stock
# Sources:
#   group_edu  ← group health education.docx (monthly by topic)
#   event      ← event database.docx
#   material   ← material database.docx
#   gbv_corner ← gbv corner establishment database.docx
#   stock      ← Stock register.xlsx

# Group education topics by audience — from both source files:
#   group health education.docx   (monthly summary, 8 topics)
#   3. Register Group Education   (per-session register, 4 tables by audience)
#
# Table 1 & 2 (Adult FSW / Old Aged FSW): 5 topic columns
# Table 3 (General): 2 topic columns
# Table 4 (Clients of FSW): 3 topic columns
#
# Topics shown per audience type via relevant conditions.
FSW_TOPICS = [
    # slug, English, Bangla
    ('personal_hygiene', 'Personal hygiene / cleanliness',
     'ব্যক্তিগত পরিষ্কার পরিচ্ছন্নতা'),
    ('unsafe_sex',       'Unsafe sex behaviour / unexpected pregnancy',
     'অনিরাপদ যৌন আচরণ ও অপ্রত্যাশিত গর্ভধারণ'),
    ('gbv',              'What is GBV / types of GBV / prevention',
     'জিবিভি কী, জিবিভি-র ধরনসমূহ; জিবিভি প্রতিরোধ'),
    ('hiv_sti',          'HIV, AIDS and STI awareness',
     'এইচআইভি, এইডস ও এসটিআই সম্পর্কে আলোচনা'),
    ('cancer_screen',    'Cervical and breast cancer screening for sex workers',
     'জরায়ু ও স্তন ক্যান্সার স্ক্রিনিং যৌনকর্মীদের জন্য'),
]
GENERAL_TOPICS = [
    ('social_safety',    'Social safety-net',
     'সোশ্যাল সেফটি-নেট'),
    ('small_business',   'Small / micro business / income generation',
     'ক্ষুদ্র ব্যবসায় আয়ের উদ্যোগ গ্রহণ সম্পর্কে'),
]
CLIENT_TOPICS = [
    ('safe_sex_client',  'Safe sex — how to protect yourself and partner',
     'নিরাপদ যৌন আচরণ কীভাবে নিজে ও সঙ্গীকে সুরক্ষিত রাখবেন'),
    ('hiv_sti_client',   'HIV, AIDS and STI awareness',
     'এইচআইভি, এইডস ও এসটিআই সম্পর্কে আলোচনা'),
    ('safe_condom_drugs','Safe condom use / extra drug use',
     'নিরাপদ চ্যাম্প সেক্স বা অতিরিক্ত মাদক সেবন করে যৌনকাজ'),
]


def _form3_survey():
    rows = _meta(center_required=False)
    rows += [
        _sr('select_one activity_type','activity_type',
            'What are you recording?',
            'কী নথিভুক্ত করছেন?', required='yes'),
    ]

    # ── Group Health Education — per-session register ─────────────────────────
    # Source: 3. Register Group Education_rev_May_2026.docx (4 tables by audience)
    #       + group health education.docx (monthly summary same topics)
    # Structure: one Kobo submission = one group session
    # Audience selector → relevant topics shown per audience type
    REL_G = "${record_type}='group_edu'"
    IS_FSW    = "(${gedu_audience}='adult_fsw' or ${gedu_audience}='old_fsw')"
    IS_GEN    = "${gedu_audience}='general'"
    IS_CLIENT = "${gedu_audience}='client'"
    rows += [
        _sr('begin_group','sec_group_edu',
            'Group Health Education (per session)',
            'দলগত স্বাস্থ্য শিক্ষা (প্রতি সেশন)',
            relevant=REL_G),
        _sr('date','gedu_date',
            'Date of session','সেশনের তারিখ', required='yes'),
        _sr('text','gedu_venue',
            'Venue (house number / location)',
            'স্থান (বাড়ির নম্বর / অবস্থান)', required='yes'),
        _sr('select_one gedu_audience','gedu_audience',
            'Target audience','লক্ষ্য দল', required='yes'),
        _sr('integer','gedu_participant_count',
            'Number of participants','অংশগ্রহণকারীর সংখ্যা'),
    ]
    # Topics for Adult FSW / Old Aged FSW (Tables 1 & 2)
    for slug, en, bn in FSW_TOPICS:
        rows.append(_sr('select_one yes_no', f'gedu_{slug}', en, bn,
                        relevant=IS_FSW))
    # Topics for General (Table 3)
    for slug, en, bn in GENERAL_TOPICS:
        rows.append(_sr('select_one yes_no', f'gedu_{slug}', en, bn,
                        relevant=IS_GEN))
    # Topics for Clients of FSW (Table 4)
    for slug, en, bn in CLIENT_TOPICS:
        rows.append(_sr('select_one yes_no', f'gedu_{slug}', en, bn,
                        relevant=IS_CLIENT))
    rows.append(_sr('end_group','sec_group_edu'))

    # ── Event Database ────────────────────────────────────────────────────────
    REL_E = "${record_type}='event'"
    rows += [
        _sr('begin_group','sec_event',
            'Event Database','ইভেন্ট ডেটাবেস',
            relevant=REL_E),
        _sr('select_one event_subtype','event_subtype',
            'Type of event','ইভেন্টের ধরন', required='yes'),
        # Participant category — shown only for training/orientation, where the
        # category routes the indicator (SL10 HM, SL11 GOB, SL12 MW, SL13 PE).
        # Without it SL11/SL12/SL13 sum over an empty set and read 0 forever.
        _sr('select_one event_ptype','event_participant_type',
            'Who were the participants?','অংশগ্রহণকারী কারা ছিলেন?',
            relevant="${event_subtype}='training' or ${event_subtype}='orientation'",
            required='yes'),
        _sr('text','event_title',
            'Title of the event','ইভেন্টের শিরোনাম', required='yes'),
        _sr('date','event_date',
            'Date of the event','ইভেন্টের তারিখ', required='yes'),
        _sr('text','event_place',
            'Place of the event','ইভেন্টের স্থান'),
        _sr('integer','event_participants',
            'Total participants','মোট অংশগ্রহণকারী'),
        _sr('text','event_notes',
            'Notes','মন্তব্য', app='multiline'),
        _sr('end_group','sec_event'),
    ]

    # ── Material Database ─────────────────────────────────────────────────────
    REL_M = "${record_type}='material'"
    rows += [
        _sr('begin_group','sec_material',
            'Material Database','উপকরণ ডেটাবেস',
            relevant=REL_M),
        _sr('text','mat_name',
            'Name of the material','উপকরণের নাম', required='yes'),
        _sr('date','mat_date',
            'Date of installation','স্থাপনের তারিখ', required='yes'),
        _sr('text','mat_place',
            'Place / Centre / Brothel of installation',
            'স্থাপনের স্থান / কেন্দ্র / ব্রথেল'),
        _sr('integer','mat_quantity',
            'Total material installed','মোট স্থাপিত উপকরণ'),
        _sr('text','mat_notes',
            'Notes','মন্তব্য', app='multiline'),
        _sr('end_group','sec_material'),
    ]

    # ── GBV Corner Establishment Database ────────────────────────────────────
    REL_GBV = "${record_type}='gbv_corner'"
    rows += [
        _sr('begin_group','sec_gbv_corner',
            'GBV Corner Establishment',
            'জিবিভি কর্নার স্থাপন',
            relevant=REL_GBV),
        _sr('text','gbv_place',
            'Place of establishment','স্থাপনের স্থান', required='yes'),
        _sr('date','gbv_date',
            'Date of establishment','স্থাপনের তারিখ', required='yes'),
        _sr('integer','gbv_furniture',
            'Furniture (add numbers)','আসবাবপত্র (সংখ্যা)'),
        _sr('integer','gbv_equipment',
            'Essential equipment / commodities (add numbers)',
            'প্রয়োজনীয় সরঞ্জাম / পণ্যসামগ্রী (সংখ্যা)'),
        _sr('select_one yes_no','gbv_functional',
            'Fully functional? (Yes/No)',
            'সম্পূর্ণ কার্যকর? (হ্যাঁ/না)'),
        _sr('end_group','sec_gbv_corner'),
    ]

    # ── Stock Register ────────────────────────────────────────────────────────
    REL_S = "${record_type}='stock'"
    rows += [
        _sr('begin_group','sec_stock',
            'Stock Register','স্টক রেজিস্টার',
            relevant=REL_S),
        _sr('text','stock_item',
            'Item description','পণ্যের বিবরণ', required='yes'),
        _sr('date','stock_date',
            'Date','তারিখ', required='yes'),
        _sr('text','stock_received_issued_to',
            'Received from* / Issued to**',
            'প্রাপ্তির উৎস* / বিতরণের গন্তব্য**'),
        _sr('text','stock_challan',
            'Delivery challan / issue voucher no.',
            'ডেলিভারি চালান / ইস্যু ভাউচার নম্বর'),
        _sr('text','stock_batch',
            'Batch no.','ব্যাচ নম্বর'),
        _sr('date','stock_expiry',
            'Expiry date','মেয়াদ উত্তীর্ণের তারিখ'),
        _sr('integer','stock_opening',
            'Opening balance','প্রারম্ভিক মজুদ'),
        _sr('integer','stock_received',
            'Quantity received','গৃহীত পরিমাণ'),
        _sr('integer','stock_issued',
            'Quantity issued','বিতরণকৃত পরিমাণ'),
        _sr('integer','stock_expired_lost',
            'Quantity expired / loss / adjustment / QA sample',
            'মেয়াদোত্তীর্ণ / ক্ষতি / সমন্বয় / QA নমুনা'),
        _sr('calculate','stock_closing',
            'Closing stock balance (auto)',
            'সমাপনী মজুদ (স্বয়ংক্রিয়)',
            calc='${stock_opening} + ${stock_received} - ${stock_issued} - ${stock_expired_lost}'),
        _sr('text','stock_comments',
            'Comments','মন্তব্য', app='multiline'),
        _sr('end_group','sec_stock'),
    ]

    return rows


def _form3_choices():
    rows = []
    rows += [_ch('yes_no','yes','Yes','হ্যাঁ'), _ch('yes_no','no','No','না')]
    for v,en,bn in [
        ('adult_fsw', 'Adult Sex Workers',     'প্রাপ্তবয়স্ক যৌনকর্মী'),
        ('old_fsw',   'Old Aged Sex Workers',  'বয়স্ক যৌনকর্মী'),
        ('general',   'General (all)',          'সাধারণ (সকলের জন্য)'),
        ('client',    'Clients of Sex Workers','যৌনকর্মীদের ক্লায়েন্ট'),
    ]:
        rows.append(_ch('gedu_audience', v, en, bn))
    for v,en,bn in [
        ('training',      'Training',              'প্রশিক্ষণ'),
        ('orientation',   'Orientation/Workshop',  'ওরিয়েন্টেশন/কর্মশালা'),
        ('camp',          'Mobile Health Camp',    'মোবাইল স্বাস্থ্য ক্যাম্প'),
        ('coord_meeting', 'Coordination Meeting',  'সমন্বয় সভা'),
    ]:
        rows.append(_ch('event_subtype', v, en, bn))
    for v,en,bn in [
        ('hm',    'Health managers (DGFP/DGHS/DGNM focal points)', 'স্বাস্থ্য ব্যবস্থাপক (ডিজিএফপি/ডিজিএইচএস/ডিজিএনএম ফোকাল)'),
        ('gob',   'District/Upazila GOB health staff',             'জেলা/উপজেলা সরকারি স্বাস্থ্যকর্মী'),
        ('mw',    'Medical Assistants / Midwives / Counsellors',   'মেডিকেল অ্যাসিস্ট্যান্ট/মিডওয়াইফ/কাউন্সেলর'),
        ('pe',    'Peer educators / community leaders',            'পিয়ার এডুকেটর/কমিউনিটি লিডার'),
        ('mixed', 'Mixed / other',                                 'মিশ্র/অন্যান্য'),
    ]:
        rows.append(_ch('event_ptype', v, en, bn))
    for v,en,bn in [
        ('group_edu','Group Health Education (per session)','দলগত স্বাস্থ্য শিক্ষা (প্রতি সেশন)'),
        ('event','Event','ইভেন্ট'),
        ('material','Material installation','উপকরণ স্থাপন'),
        ('gbv_corner','GBV Corner establishment','জিবিভি কর্নার স্থাপন'),
        ('stock','Stock entry','স্টক এন্ট্রি'),
    ]:
        rows.append(_ch('activity_type', v, en, bn))
    rows += _centre_choices()
    return rows


# ─── MERGED FORM 2 — Service Log ──────────────────────────────────────────────
# Single Kobo form with a "What are you recording today?" selector at the top.
# Only the section matching the picked record_type shows; the other 8 stay
# hidden. Different roles fill different sections — medical assistant fills
# Clinic/HTC, counsellor fills Counselling, peer educator fills Group Education,
# storekeeper fills Stock, programme officer fills Event/GBV-corner/Material.
#
# Replaces the previous Forms 2 & 3. Field name conventions and webhook
# dispatch use the section prefix in the field name (clinic_*, htc_*, ref_*,
# gedu_*, event_*, mat_*, gbv_*, stock_*, counsel_*).

def _form_service_log_survey():
    """Build the merged Service Log survey:
       meta → record_type selector → patient_id_group (pulldata)
       → 9 gated sections."""
    rows = _meta(center_required=False)
    rows += [
        _sr('select_one record_type','record_type',
            'What are you recording today?',
            'আজ কী নথিভুক্ত করছেন?', required='yes',
            hint='Pick one. Only the matching section appears below. / একটি নির্বাচন করুন — শুধু সংশ্লিষ্ট অংশ দেখাবে।'),
    ]
    rows += _patient_id_group()
    # Reuse all 9 section bodies. _form2_survey() and _form3_survey() each
    # prepend their own meta + selector; strip those off so we don't repeat.
    f2 = _form2_survey()
    f3 = _form3_survey()
    rows += _strip_meta_and_selector(f2)
    rows += _strip_meta_and_selector(f3)
    return rows


def _patient_id_group():
    """One ID field shared across all patient-level record types
    (clinic / HTC / referral). pulldata() reads phd_clients.csv (uploaded
    via 'manage.py export_phd_clients --upload') and shows her name,
    mother, age, address read-only — so the enumerator confirms they
    have the right woman BEFORE recording any clinical data.

    Counselling is a monthly aggregate report (no per-client ID), so
    this group is hidden for that record type.

    Activity-level types (group_edu, event, material, gbv_corner,
    stock) also hide this group — they aren't patient-specific.

    ID format: '{centre number}-{4-digit serial}', e.g. 1-0001 (Daulatdia).
    Free-text + trim/upper normalisation so ' 1-0001 ' still matches the
    canonical '1-0001' in the CSV.
    """
    REL = ("${record_type}='clinic' or ${record_type}='htc' "
           "or ${record_type}='referral'")
    # XPath 1.0 trim+upper — Kobo's engine does not support upper-case().
    NORM = ("translate(normalize-space(${{client_id}}),"
            "'abcdefghijklmnopqrstuvwxyz',"
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ')")
    PULL = "pulldata('phd_clients','{col}','id_no'," + NORM + ")"
    NOT_FOUND = "${client_id}!='' and ${_pull_name}=''"
    FOUND     = "${_pull_name}!=''"
    return [
        _sr('begin_group','patient_id_group',
            'FSW Identity (auto-fill from Master List)',
            'যৌনকর্মীর পরিচয় (মাদারলিস্ট থেকে স্বয়ংক্রিয়)',
            relevant=REL),

        _sr('text','client_id',
            'FSW ID No.','যৌনকর্মীর আইডি নম্বর',
            required='yes',
            hint='Type her registered ID, e.g. 1-0001 (centre number + serial).'),

        # pulldata() lookups — all keyed on the upper-cased client_id.
        _sr('calculate','_pull_name',     calc=PULL.format(col='name')),
        _sr('calculate','_pull_mother',   calc=PULL.format(col='mother_name')),
        _sr('calculate','_pull_birth',    calc=PULL.format(col='birth_year')),
        _sr('calculate','_pull_age',      calc=PULL.format(col='age')),
        _sr('calculate','_pull_address',  calc=PULL.format(col='address')),
        _sr('calculate','_pull_status',   calc=PULL.format(col='status')),

        # Read-only confirmation block — appears when an ID matches.
        _sr('note','_show_name',
            'Name: ${_pull_name}','নাম: ${_pull_name}',
            relevant=FOUND),
        _sr('note','_show_mother',
            "Mother's name: ${_pull_mother}",
            'মাতার নাম: ${_pull_mother}',
            relevant=FOUND + " and ${_pull_mother}!=''"),
        _sr('note','_show_age',
            'Age: ${_pull_age}   (Born: ${_pull_birth})',
            'বয়স: ${_pull_age}   (জন্ম: ${_pull_birth})',
            relevant=FOUND + " and ${_pull_age}!=''"),
        _sr('note','_show_address',
            'Address: ${_pull_address}',
            'ঠিকানা: ${_pull_address}',
            relevant=FOUND + " and ${_pull_address}!=''"),
        _sr('note','_show_status',
            'Status: ${_pull_status}',
            'অবস্থা: ${_pull_status}',
            relevant=FOUND + " and ${_pull_status}!=''"),

        # Warning when ID typed but not found in CSV.
        _sr('note','_id_not_found',
            '⚠ This ID is NOT registered in the Master List. '
            'Register her in PHD 1 (FSW Registration) first, then come back.',
            '⚠ এই আইডি মাদারলিস্টে নিবন্ধিত নয়। '
            'প্রথমে PHD 1 (যৌনকর্মী নিবন্ধন) ফর্মে নিবন্ধন করুন।',
            relevant=NOT_FOUND),

        _sr('end_group','patient_id_group'),
    ]


def _strip_meta_and_selector(rows):
    """Drop the leading grp_meta block and the per-form select_one selector
    so the section bodies can be appended after our combined selector."""
    out = list(rows)
    # 1. Drop everything from start through end_group grp_meta
    for i, r in enumerate(out):
        if r[0] == 'end_group' and r[1] == 'grp_meta':
            out = out[i+1:]
            break
    # 2. Drop the leading select_one (the old per-form selector that's been
    #    rewritten by replace_all to use ${record_type})
    if out and out[0][0].startswith('select_one'):
        out = out[1:]
    return out


def _form_service_log_choices():
    """Merge Form 2 + Form 3 choices into one sheet. The dropped per-form
    selectors (service_type, activity_type) are replaced by the unified
    record_type list with all 9 options."""
    rows = []
    # yes_no — used everywhere
    rows += [_ch('yes_no','yes','Yes','হ্যাঁ'), _ch('yes_no','no','No','না')]
    # Medical Assistant dropdown (same enumerator list as Form 1).
    rows += _ma_choices()

    # ── record_type — the top-level "What are you recording today?" ──
    for v, en, bn in [
        ('clinic',      'Clinic visit (Patient Record)',
            'ক্লিনিক ভিজিট (রোগীর রেকর্ড)'),
        ('htc',         'HIV / STI test (HTC)',
            'এইচআইভি / এসটিআই পরীক্ষা (HTC)'),
        ('counselling', 'Counselling Report (monthly)',
            'কাউন্সেলিং রিপোর্ট (মাসিক)'),
        ('referral',    'Referral',
            'রেফারেল'),
        ('group_edu',   'Group Education session',
            'দলগত স্বাস্থ্য শিক্ষা সেশন'),
        ('event',       'Event (training / orientation / camp / coord. meeting)',
            'ইভেন্ট (প্রশিক্ষণ / ওরিয়েন্টেশন / ক্যাম্প / সমন্বয় সভা)'),
        ('material',    'IEC material installed',
            'আইইসি উপকরণ স্থাপন'),
        ('gbv_corner',  'GBV corner established',
            'জিবিভি কর্নার স্থাপন'),
        ('stock',       'Stock entry',
            'স্টক এন্ট্রি'),
    ]:
        rows.append(_ch('record_type', v, en, bn))

    # ── reuse all other choice lists from F2 + F3 (visit_type, treatment_timing,
    #    sex, test_occasion, tested_at, tested_by, rnrinv, final_result,
    #    gedu_audience, event_subtype) ──
    # Deduplicate yes_no (already added above).
    seen = {('yes_no','yes'), ('yes_no','no')}
    for src_choices in (_form2_choices(), _form3_choices()):
        for row in src_choices:
            key = (row[0], row[1])
            # Skip the dead per-form selectors (we use record_type instead)
            if row[0] in ('service_type', 'activity_type'):
                continue
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
    return rows


# ─── Kobo upload helpers (mirror build_ciprb_forms) ───────────────────────────

def _kobo_token():
    return (getattr(settings, 'KOBO_API_TOKEN', '')
            or os.environ.get('KOBO_TOKEN', '')).strip()


def _import_xlsform(xlsx_path, asset_uid, token, stdout):
    """Replace an existing asset's survey by importing the xlsx into it."""
    headers = {'Authorization': f'Token {token}'}
    api = f'{KOBO_BASE}/api/v2'
    with open(xlsx_path, 'rb') as fh:
        files = {'file': (os.path.basename(xlsx_path), fh,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'destination': f'{api}/assets/{asset_uid}/', 'library': 'false'}
        r = requests.post(f'{api}/imports/', headers=headers,
                          files=files, data=data, timeout=120)
    if r.status_code not in (200, 201):
        stdout.write(f'    import FAILED ({r.status_code}): {r.text[:200]}')
        return False
    stdout.write('    imported')
    return True


def _deploy(asset_uid, token, stdout):
    """Redeploy the asset's latest version so Enketo serves the new form."""
    headers = {'Authorization': f'Token {token}'}
    api = f'{KOBO_BASE}/api/v2'
    v = requests.get(f'{api}/assets/{asset_uid}/versions/?limit=1',
                     headers=headers, timeout=30)
    try:
        vhash = v.json()['results'][0]['uid']
    except Exception:
        stdout.write('    no version yet — skipping deploy')
        return False
    r = requests.patch(
        f'{api}/assets/{asset_uid}/deployment/',
        headers=headers, json={'version_id': vhash, 'active': True}, timeout=60)
    if r.status_code in (200, 201):
        stdout.write('    redeployed')
        return True
    r2 = requests.post(
        f'{api}/assets/{asset_uid}/deployment/',
        headers=headers, json={'version_id': vhash, 'active': True}, timeout=60)
    if r2.status_code in (200, 201):
        stdout.write('    deployed (POST)')
        return True
    stdout.write(f'    deploy FAILED ({r.status_code}/{r2.status_code}): {r2.text[:160]}')
    return False


# ─── Command ──────────────────────────────────────────────────────────────────

FORMS = [
    {
        'file': 'PHD-1_Registration.xlsx',
        'id':   'phd_registration_v1',
        'title':'PHD 1 — FSW Registration',
        'survey': _form1_survey,
        'choices': _form1_choices,
    },
    {
        'file': 'PHD-2_Service_Log.xlsx',
        'id':   'phd_service_log_v1',
        'title':'PHD 2 — Service Log',
        'survey': _form_service_log_survey,
        'choices': _form_service_log_choices,
    },
]


class Command(BaseCommand):
    help = 'Build the 2 PHD XLSForms (Registration + Service Log) from final source files.'

    def add_arguments(self, parser):
        parser.add_argument('--output-dir', default=OUTDIR)
        parser.add_argument('--upload', action='store_true',
            help='Import each xlsx into its existing Kobo asset and redeploy.')

    def handle(self, *args, **options):
        out = options['output_dir']
        os.makedirs(out, exist_ok=True)
        token = _kobo_token() if options['upload'] else ''
        if options['upload'] and not token:
            self.stdout.write(self.style.ERROR('KOBO_TOKEN not set — cannot --upload.'))
            return

        for f in FORMS:
            survey  = f['survey']()
            choices = f['choices']()
            wb = _wb(f['id'], f['title'], survey, choices)
            path = os.path.join(out, f['file'])
            wb.save(path)
            self.stdout.write(self.style.SUCCESS(
                f"  OK  {f['file']:35s}  {len(survey)} rows  id: {f['id']}"))

            if options['upload']:
                self.stdout.write('     uploading…')
                api = f'{KOBO_BASE}/api/v2'
                headers = {'Authorization': f'Token {token}'}
                # Resolve the asset by its id_string (form_id in settings).
                q = requests.get(
                    f'{api}/assets/?q=settings__id_string:{f["id"]}',
                    headers=headers, timeout=30).json()
                asset_uid = None
                for a in q.get('results', []):
                    if a.get('settings', {}).get('id_string') == f['id']:
                        asset_uid = a.get('uid')
                        break
                if not asset_uid:
                    # Kobo often auto-generates its own id_string on first
                    # upload, so the id_string query returns nothing. Fall back
                    # to a FRESH asset listing matched by title (the em-dash in
                    # the title is preserved on both sides).
                    allq = requests.get(f'{api}/assets/?limit=300',
                                        headers=headers, timeout=30).json()
                    for a in allq.get('results', []):
                        if (a.get('name') or '') == f['title']:
                            asset_uid = a.get('uid')
                            break
                if not asset_uid:
                    self.stdout.write(self.style.ERROR(
                        f'    no Kobo asset found for {f["id"]} — skipping.'))
                    continue
                if _import_xlsform(path, asset_uid, token, self.stdout):
                    _deploy(asset_uid, token, self.stdout)

        self.stdout.write(f'\nWritten to {os.path.abspath(out)}/')
