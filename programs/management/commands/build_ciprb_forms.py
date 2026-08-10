# -*- coding: utf-8 -*-
"""
Build the 9 CIPRB KoboToolbox XLSForms — Phase 2 deliverable.

  1. CIPRB Fistula Question Bank          ← Fistula Question.xlsx
  2. MPDSR Form 01 Community Maternal     ← MPDSR Form 01 (Community Mother).pdf
  3. MPDSR Form 02 Community Neonatal     ← MPDSR Form 02 (Community Neonate).pdf
  4. MPDSR Form 04 Facility Maternal      ← MPDSR Form 04 (Facility Maternal).pdf
  5. MPDSR Form 05 Facility Neonatal      ← MPDSR Form 05 (Facility Neonates).pdf
  6. Social Autopsy (Maternal Death)      ← Social Autopsy.pdf (Bangla)
  7. Death Notification Slip 01           ← notification slip01.pdf
  8. Death Notification Slip 02           ← notification slip02.pdf
  9. Maternal Near Miss audit             ← Tool_Near Miss.xlsx (WHO MNM)

Conventions (mirror build_phd_forms.py):
  - All forms CIPRB-tagged. organisation hidden calculate = 'CIPRB'.
  - Bilingual: English + Bangla. Govt MoHFW phrasing where the source
    PDF uses it; otherwise CIPRB-confirmed labels.
  - Single-page (theme-grid, no pages) — Enketo's pages mode hides the
    relevant-logic conditions, slows phones, breaks the staged workflow.
  - Case-ID = system UUID + visible CIPRB serial number per form.
  - 18 CIPRB districts from the Near Miss tool — the canonical
    working footprint across every form.

Run:
    python manage.py build_ciprb_forms                # writes xlsx
    python manage.py build_ciprb_forms --upload       # also uploads to Kobo
"""
import os
import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.core.management.base import BaseCommand

HERE   = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.normpath(os.path.join(HERE, '..', '..', '..', '..', 'koboforms_ciprb'))

KOBO_BASE = 'https://kf.kobotoolbox.org'

_HFILL = PatternFill("solid", fgColor="003F72")
_HFONT = Font(color="FFFFFF", bold=True, size=10)


# ─── XLSForm helpers ──────────────────────────────────────────────────────────

SURVEY_HDR = [
    'type', 'name', 'label::English', 'label::Bangla',
    'hint', 'required', 'relevant', 'constraint', 'constraint_message',
    'default', 'appearance', 'calculation', 'choice_filter',
]
CHOICES_HDR  = ['list_name', 'name', 'label::English', 'label::Bangla']
SETTINGS_HDR = ['form_title', 'form_id', 'version', 'default_language', 'style']


def _sr(qtype, name, en='', bn='', hint='', required='',
        relevant='', constraint='', cmsg='', default='', app='', calc='', cf=''):
    return [qtype, name, en, bn, hint, required, relevant,
            constraint, cmsg, default, app, calc, cf]


def _ch(lst, name, en, bn=''):
    return [lst, name, en, bn]


def _wb(form_id, form_title, survey, choices):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in [
        ('survey',   SURVEY_HDR,  survey),
        ('choices',  CHOICES_HDR, choices),
        ('settings', SETTINGS_HDR,
         [[form_title, form_id, '20260620', 'English', 'theme-grid']]),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for ci in range(1, len(headers) + 1):
            c = ws.cell(1, ci)
            c.font = _HFONT
            c.fill = _HFILL
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 30
        ws.freeze_panes = 'A2'
    return wb


# ─── 18 canonical CIPRB districts ────────────────────────────────────────────
CIPRB_DISTRICTS = [
    'Sunamganj', 'Sherpur', 'Bhola', 'Kurigram', 'Gaibandha',
    'Khagrachari', 'Noakhali', 'Patuakhali', 'Sirajganj', 'Barguna',
    'Jamalpur', 'Bagerhat', 'Habiganj', 'Moulavibazar', 'Sylhet',
    'Bandarban', 'Chandpur', 'Rangpur', 'Dhaka',
]
DISTRICT_BANGLA = {
    'Sunamganj': 'সুনামগঞ্জ', 'Sherpur': 'শেরপুর', 'Bhola': 'ভোলা',
    'Kurigram': 'কুড়িগ্রাম', 'Gaibandha': 'গাইবান্ধা',
    'Khagrachari': 'খাগড়াছড়ি', 'Noakhali': 'নোয়াখালী',
    'Patuakhali': 'পটুয়াখালী', 'Sirajganj': 'সিরাজগঞ্জ',
    'Barguna': 'বরগুনা', 'Jamalpur': 'জামালপুর', 'Bagerhat': 'বাগেরহাট',
    'Habiganj': 'হবিগঞ্জ', 'Moulavibazar': 'মৌলভীবাজার',
    'Sylhet': 'সিলেট', 'Bandarban': 'বান্দরবান', 'Chandpur': 'চাঁদপুর',
    'Rangpur': 'রংপুর', 'Dhaka': 'ঢাকা',
}

DISTRICT_CHOICES = [
    _ch('district', d.lower().replace(' ', '_'), d, DISTRICT_BANGLA[d])
    for d in CIPRB_DISTRICTS
] + [_ch('district', 'other', 'Other', 'অন্যান্য')]

YES_NO = [
    _ch('yes_no', 'yes',     'Yes',     'হ্যাঁ'),
    _ch('yes_no', 'no',      'No',      'না'),
    _ch('yes_no', 'unknown', 'Unknown', 'অজানা'),
]


# ─── Shared submission header (CIPRB org + GPS + district + enumerator) ─────
def _meta(form_id_visible='', form_id_visible_bn=''):
    """Common opening section every CIPRB form starts with.

    `form_id_visible` is the human-visible serial number CIPRB uses on
    paper (e.g. মাতৃমৃত্যুর বাৎসরিক ক্রমিক নং on MPDSR Form 1). It is
    captured as text so the field worker can copy what the paper carries;
    Django persists a UUID separately."""
    bn_label = form_id_visible_bn or form_id_visible
    return [
        _sr('begin_group', 'grp_meta', 'Submission info', 'তথ্য প্রেরণ'),
        _sr('calculate', 'organisation', '', '', calc="'CIPRB'"),
        _sr('geopoint', 'location',
            'GPS location (required — step outside if no signal)',
            'জিপিএস অবস্থান (প্রয়োজনীয়)', required='yes'),
        _sr('date', 'collection_date', 'Date', 'তারিখ', required='yes'),
        _sr('select_one district', 'district',
            'District', 'জেলা', required='yes'),
        _sr('text', 'upazila',  'Upazila',  'উপজেলা', required='yes'),
        _sr('text', 'union',    'Union / Pourashava', 'ইউনিয়ন / পৌরসভা'),
        _sr('text', 'ward',     'Ward',     'ওয়ার্ড'),
        _sr('text', 'village',  'Village / Mahalla', 'গ্রাম / মহল্লা'),
        _sr('text', 'enumerator_name',
            'Your name (person filling this form)',
            'আপনার নাম (কে পূরণ করছেন)', required='yes'),
        _sr('text', 'enumerator_designation',
            'Designation', 'পদবী'),
        _sr('text', 'enumerator_institution',
            'Institution', 'প্রতিষ্ঠান'),
        _sr('text', 'enumerator_mobile',
            'Mobile number', 'মোবাইল নম্বর',
            constraint='regex(., "^[0-9+ -]{6,20}$")',
            cmsg='Enter a valid phone number.'),
        *([_sr('text', 'case_serial', form_id_visible, bn_label,
               hint='The handwritten serial number from the paper form, if any.')]
          if form_id_visible else []),
        _sr('end_group', 'grp_meta'),
    ]


def _office_use_block(serial_en, serial_bn):
    """The 'For office use' (অফিসের ব্যবহারের জন্য) box printed at the top of the
    MPDSR paper forms: form submission date, annual serial number, and the
    receiver's name/signature. Office-filled, not by the field worker."""
    return [
        _sr('begin_group', 'grp_office', 'For office use',
            'অফিসের ব্যবহারের জন্য'),
        _sr('date', 'office_submission_date', 'Date of form submission',
            'ফর্ম জমাদানের তারিখ'),
        _sr('text', 'case_serial', serial_en, serial_bn,
            hint='Annual serial number from the paper form.'),
        _sr('text', 'office_receiver', 'Name & signature of form receiver',
            'ফর্ম গ্রহণকারীর নাম ও স্বাক্ষর'),
        _sr('end_group', 'grp_office'),
    ]


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 1 — CIPRB Fistula Question Bank                                  ║
# ║   Source: Fistula Question.xlsx                                          ║
# ║   5 staged sections: Suspected → Diagnosed → Referred → Repaired →       ║
# ║                      Rehabilitated & Reintegrated                        ║
# ╚══════════════════════════════════════════════════════════════════════════╝

# District slug → numeric code for the fistula patient-ID prefix. The spec
# names 10 base districts (codes 1–10); the remaining CIPRB districts get
# 11–18 in slug order so every option yields a deterministic prefix. The
# slug keys MUST match the `district` choice values (lower-cased, spaces→'_').
FISTULA_DISTRICT_CODE = {
    'sunamganj': 1, 'bhola': 2, 'noakhali': 3, 'gaibandha': 4,
    'kurigram': 5, 'sirajganj': 6, 'sherpur': 7, 'patuakhali': 8,
    'khagrachari': 9, 'dhaka': 10,
    # extras (no spec code) — kept deterministic so their IDs still validate:
    'barguna': 11, 'jamalpur': 12, 'bagerhat': 13, 'habiganj': 14,
    'moulavibazar': 15, 'sylhet': 16, 'bandarban': 17, 'rangpur': 18,
    'chandpur': 19,
}


def _fistula_dist_code_calc():
    """Build the nested if() that maps the selected district slug → its
    numeric code string. Enketo is XPath 1.0, so we use an explicit if()
    chain rather than a lookup — and we never use substr() on the code,
    which avoids the '10-' vs '1-' starts-with collision (the regex
    constraint anchors the full prefix instead)."""
    expr = "''"
    for slug, code in reversed(list(FISTULA_DISTRICT_CODE.items())):
        expr = f"if(${{district}}='{slug}','{code}',{expr})"
    return expr


# District slug → UPPERCASE LETTER code for the MPDSR Action-ID prefix
# (Patuakhali=PA, Dhaka=DH). Distinct from FISTULA_DISTRICT_CODE (numeric — those
# stay for fistula patient IDs). Action IDs read PA-001, DH-001.
#
# DERIVED from the canonical mpdsr.models.DISTRICT_ACTION_CODE so the prefix the
# form tells a worker to type can never diverge from what the server allocates /
# validates (audit FIX 2026-06: the two maps had silently drifted, 13/19 codes
# disagreed). The model dict is keyed by proper-case district name; the form
# `district` choice values are lower-cased, spaces→'_', so we re-key here.
from mpdsr.models import DISTRICT_ACTION_CODE as _DISTRICT_ACTION_CODE
MPDSR_DISTRICT_CODE = {
    name.lower().replace(' ', '_'): code
    for name, code in _DISTRICT_ACTION_CODE.items()
}


def _mpdsr_dist_code_calc():
    """Nested if() mapping ${district} slug → its UPPERCASE letter code (PA, DH).
    XPath 1.0 (Enketo) — explicit if() chain, not a lookup."""
    expr = "''"
    for slug, code in reversed(list(MPDSR_DISTRICT_CODE.items())):
        expr = f"if(${{district}}='{slug}','{code}',{expr})"
    return expr


_FISTULA_STAGES = [
    ('suspected',     'Suspected (community identification)',
     'সন্দেহজনক (সম্প্রদায়ে শনাক্ত)'),
    ('diagnosed',     'Diagnosed (at Fistula Corner)',
     'নির্ণীত (ফিস্টুলা কর্নারে)'),
    ('referred',      'Referred for Surgical Management',
     'অস্ত্রোপচারের জন্য প্রেরিত'),
    ('repaired',      'Surgically Repaired',
     'অস্ত্রোপচার সম্পন্ন'),
    ('rehabilitated', 'Rehabilitated & Reintegrated',
     'পুনর্বাসিত ও পুনরায় সমাজে অন্তর্ভুক্ত'),
]


def _fistula_survey():
    rows = _meta('Annual fistula serial number', 'ফিস্টুলা বাৎসরিক ক্রমিক নং')

    # ── Stage selector — drives every relevant() below.
    rows += [
        _sr('select_one stage', 'stage',
            'Are you registering a NEW patient, or updating an EXISTING one?',
            'নতুন রোগী নিবন্ধন করছেন, না কি আগের রোগীর তথ্য হালনাগাদ করছেন?',
            required='yes',
            hint='Pick "Suspected" to REGISTER a new woman. Pick any later stage '
                 '(Diagnosed / Referred / Repaired / Rehabilitated) to UPDATE a woman '
                 'already registered — you will choose her from a list.'),
    ]

    # ── Patient identity (captured ONCE, at the Suspected stage only).
    #    Later stages identify the woman via the grp_lookup dropdown below.
    rows += [
        _sr('begin_group', 'grp_patient', 'Patient identity', 'রোগীর পরিচয়',
            relevant="${stage}='suspected'"),

        # The Patient ID is no longer typed. The 2026-08-08 Bhola incident
        # (2-0028 registered twice, 12 minutes apart) proved the pulldata
        # duplicate check only sees the CSV copy cached on the device, so the
        # server now ISSUES the ID on submission — same design as the Bandhu
        # Mother List. This note manages expectations at registration.
        _sr('note', '_auto_id_note',
            'The Patient ID is assigned AUTOMATICALLY after you submit — you '
            'do not type it. She will appear in the update list (with her ID) '
            'once the form is re-downloaded.',
            'রোগীর আইডি জমা দেওয়ার পর স্বয়ংক্রিয়ভাবে তৈরি হবে — আপনাকে লিখতে হবে না। '
            'ফর্মটি নতুন করে ডাউনলোড করলে তালিকায় রোগীকে (আইডিসহ) পাওয়া যাবে।',
            relevant="${stage}='suspected'"),

        _sr('text', 'name', 'Name of woman', 'মহিলার নাম', required='yes'),
        _sr('integer', 'age', 'Age (years)', 'বয়স (বছর)',
            required='yes', constraint='. >= 8 and . <= 80'),
        _sr('select_one education', 'education',
            'Education', 'শিক্ষা'),
        _sr('text', 'husband', "Husband's name", 'স্বামীর নাম'),
        _sr('text', 'husband_profession',
            "Husband's profession", 'স্বামীর পেশা'),
        _sr('text', 'profession_patient',
            'Profession of patient', 'রোগীর পেশা'),
        _sr('text', 'current_condition',
            'Present condition of patient',
            'রোগীর বর্তমান অবস্থা'),
        _sr('text', 'contact_number',
            'Contact number', 'যোগাযোগের নম্বর'),
        _sr('select_one marital_status', 'marital_status',
            'Marital status', 'বৈবাহিক অবস্থা'),
        _sr('integer', 'age_at_marriage',
            'Age at marriage', 'বিবাহের সময় বয়স'),
        _sr('integer', 'age_at_first_delivery',
            'Age at first delivery', 'প্রথম প্রসবের সময় বয়স'),
        _sr('integer', 'number_of_children',
            'Number of children', 'সন্তান সংখ্যা'),
        _sr('end_group', 'grp_patient'),
    ]

    # ── Obstetric history (captured ONCE, at the Suspected stage only).
    rows += [
        _sr('begin_group', 'grp_obs_history',
            'Obstetric history', 'প্রসূতি ইতিহাস',
            relevant="${stage}='suspected'"),
        _sr('text', 'delivery_complication',
            'Delivery complication', 'প্রসব জটিলতা'),
        _sr('text', 'last_delivery_labour_duration',
            'Duration of labour during last delivery',
            'শেষ প্রসবের সময় শ্রমকালের সময়কাল'),
        _sr('select_one mode_of_delivery', 'mode_of_last_delivery',
            'Mode of last delivery', 'শেষ প্রসবের পদ্ধতি'),
        _sr('select_one place_of_delivery', 'place_of_last_delivery',
            'Place of last delivery', 'শেষ প্রসবের স্থান'),
        _sr('select_one delivery_conductor', 'conducted_last_delivery',
            'Who conducted the last delivery',
            'শেষ প্রসব কে পরিচালনা করেন'),
        _sr('select_one delivery_outcome', 'delivery_outcome',
            'Delivery outcome', 'প্রসবের ফলাফল'),
        _sr('select_multiple reasons_no_facility',
            'reasons_no_institutional_delivery',
            'Reasons for not availing institutional delivery',
            'প্রাতিষ্ঠানিক প্রসব না নেওয়ার কারণ'),
        _sr('text', 'time_duration_fistula_occurrence',
            'Time duration of fistula occurrence after delivery',
            'প্রসবের পর ফিস্টুলা সংঘটনের সময়কাল'),
        _sr('text', 'duration_suffering',
            'Duration of suffering from fistula (years / months)',
            'ফিস্টুলায় ভোগার সময়কাল (বছর / মাস)'),
        _sr('end_group', 'grp_obs_history'),
    ]

    # ── Patient lookup (stages 2–5). The woman was already registered at the
    #    Suspected stage; here the field worker picks her from the CSV-backed
    #    dropdown (fistula_clients.csv) instead of re-typing identity. The
    #    selected value is her id_no; pulldata() shows her details read-only so
    #    the worker confirms the right woman before recording the new stage.
    LATER = ("${stage}='diagnosed' or ${stage}='referred' or "
             "${stage}='repaired' or ${stage}='rehabilitated'")
    # Normalise the dropdown value the same way the CSV / handler key is stored.
    # Braces around the field name are doubled so str.format() (below) leaves
    # the ${...} XPath reference intact and only substitutes {col}.
    NORM_SEL = ("translate(normalize-space(${{patient_code_sel}}),"
                "'abcdefghijklmnopqrstuvwxyz',"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ')")
    PULL = "pulldata('fistula_clients','{col}','id_no'," + NORM_SEL + ")"
    FOUND = "${_pull_name}!=''"
    rows += [
        _sr('begin_group', 'grp_lookup',
            'Select the registered patient', 'নিবন্ধিত রোগী নির্বাচন করুন',
            relevant=LATER),
        # Dropdown sourced from the attached CSV — the stored value is id_no,
        # the label shows "54 | Rahima | Nurabad | 1-0001" (paper serial first:
        # that is the number the worker knows from her own register).
        _sr('select_one_from_file fistula_clients.csv', 'patient_code_sel',
            'Registered patient (ID — name)',
            'নিবন্ধিত রোগী (আইডি — নাম)',
            required='yes',
            relevant=LATER,
            app='autocomplete'),
        _sr('calculate', '_pull_name',    calc=PULL.format(col='patient_name')),
        _sr('calculate', '_pull_age',     calc=PULL.format(col='age')),
        _sr('calculate', '_pull_husband', calc=PULL.format(col='husband')),
        _sr('calculate', '_pull_village', calc=PULL.format(col='village')),
        _sr('calculate', '_pull_susp',    calc=PULL.format(col='suspected_date')),
        _sr('note', '_show_name',
            'Name: ${_pull_name}', 'নাম: ${_pull_name}', relevant=FOUND),
        _sr('note', '_show_age',
            'Age: ${_pull_age} · Husband: ${_pull_husband}',
            'বয়স: ${_pull_age} · স্বামী: ${_pull_husband}',
            relevant=FOUND + " and (${_pull_age}!='' or ${_pull_husband}!='')"),
        _sr('note', '_show_addr',
            'Village: ${_pull_village} · Suspected: ${_pull_susp}',
            'গ্রাম: ${_pull_village} · সন্দেহ: ${_pull_susp}',
            relevant=FOUND + " and (${_pull_village}!='' or ${_pull_susp}!='')"),
        _sr('note', '_lookup_missing',
            '⚠ No registered patient found for this ID. Register her at the '
            'Suspected stage first, then record this stage.',
            '⚠ এই আইডির জন্য কোনো নিবন্ধিত রোগী পাওয়া যায়নি। প্রথমে সন্দেহজনক '
            'ধাপে নিবন্ধন করুন।',
            relevant="${patient_code_sel}!='' and ${_pull_name}=''"),
        _sr('end_group', 'grp_lookup'),
    ]

    # ── The handler's single ID key. Registration sends it EMPTY (the server
    #    issues the ID); later stages carry the dropdown pick. The field is kept
    #    (rather than dropped) so old and new form versions share one payload
    #    shape and the id write-back has a stable column to land in.
    rows += [
        _sr('calculate', 'patient_code_final', calc=(
            "if(${stage}='suspected', '', ${patient_code_sel})")),
    ]

    # ── STAGE 1 · Suspected.
    rows += [
        _sr('begin_group', 'grp_suspected',
            'Stage 1 · Suspected',
            'ধাপ ১ · সন্দেহজনক',
            relevant="${stage}='suspected'"),
        _sr('date', 'suspected_date',
            'Date of suspected', 'সন্দেহের তারিখ'),
        _sr('text', 'source_information',
            "Source of patient's information",
            'রোগীর তথ্যের উৎস'),
        _sr('end_group', 'grp_suspected'),
    ]

    # ── STAGE 2 · Diagnosed.
    rows += [
        _sr('begin_group', 'grp_diagnosed',
            'Stage 2 · Diagnosed (Fistula Corner)',
            'ধাপ ২ · নির্ণীত (ফিস্টুলা কর্নার)',
            relevant="${stage}='diagnosed'"),
        _sr('date', 'diagnosed_date',
            'Date of diagnosis', 'নির্ণয়ের তারিখ', required='yes'),
        _sr('text', 'diagnosed_place',
            'Place of diagnosis', 'নির্ণয়ের স্থান'),
        _sr('text', 'diagnosed_by',
            'Diagnosed by', 'নির্ণয়কারী'),
        # Anatomical fistula type (VVF / RVF / …) — recorded at the Fistula
        # Corner when the diagnosis is made, mirroring the paper register's
        # "ফিস্টুলার ধরন" column. Moved here from the surgery stage so a case
        # carries its type as soon as it is diagnosed (most cases are
        # diagnosed long before they are operated on).
        _sr('select_one genital_fistula_type', 'genital_fistula_type',
            'Type of genital fistula (VVF / RVF / …)',
            'যৌনাঙ্গের ফিস্টুলার ধরন (VVF / RVF / …)', required='yes'),
        _sr('text', 'genital_fistula_type_other',
            'Other — please specify the type',
            'অন্যান্য হলে ধরন উল্লেখ করুন',
            relevant="${genital_fistula_type}='other'", required='yes'),
        _sr('end_group', 'grp_diagnosed'),
    ]

    # ── STAGE 3 · Referred for Surgical Management.
    rows += [
        _sr('begin_group', 'grp_referred',
            'Stage 3 · Referred for Surgical Management',
            'ধাপ ৩ · অস্ত্রোপচারের জন্য প্রেরিত',
            relevant="${stage}='referred'"),
        _sr('date', 'refer_date',
            'Referral date', 'প্রেরণের তারিখ', required='yes'),
        _sr('text', 'refer_place',
            'Refer place (facility)', 'প্রেরণের স্থান'),
        _sr('text', 'referred_by_person',
            'Person who referred the woman to the facility',
            'প্রেরণকারী ব্যক্তি'),
        _sr('select_one refer_outcome', 'refer_outcome',
            'Refer outcome', 'প্রেরণের ফলাফল'),
        _sr('end_group', 'grp_referred'),
    ]

    # ── STAGE 4 · Surgically Repaired.
    rows += [
        _sr('begin_group', 'grp_repaired',
            'Stage 4 · Surgically Repaired',
            'ধাপ ৪ · অস্ত্রোপচার সম্পন্ন',
            relevant="${stage}='repaired'"),
        _sr('date', 'operation_date',
            'Date of operation', 'অস্ত্রোপচারের তারিখ', required='yes'),
        _sr('text', 'operation_place',
            'Place of operation', 'অস্ত্রোপচারের স্থান'),
        _sr('integer', 'hospital_stay_days',
            'Duration of hospital stay (days)',
            'হাসপাতালে অবস্থানের সময়কাল (দিন)'),
        _sr('integer', 'times_of_operations',
            'Times of operations', 'অস্ত্রোপচারের সংখ্যা'),
        # The 4 fistula types CIPRB asked for on the dashboard donut.
        _sr('select_one fistula_type', 'fistula_type_v2',
            'Type of fistula (cause)',
            'ফিস্টুলার ধরন (কারণ)'),
        _sr('select_one iatrogenic_cause', 'iatrogenic_cause',
            'Cause of iatrogenic fistula',
            'চিকিৎসাজনিত ফিস্টুলার কারণ',
            relevant="${fistula_type_v2}='iatrogenic'"),
        # NOTE: the anatomical type (genital_fistula_type — VVF/RVF/…) now
        # lives on the Diagnosed stage (it is classified at the Fistula
        # Corner, not in theatre). The surgery stage keeps the cause
        # classification + operative detail only.
        _sr('select_one operation_route', 'operation_route',
            'Route of operation',
            'অস্ত্রোপচারের পথ'),
        _sr('select_one surgery_outcome_v2', 'surgery_outcome_v2',
            'Outcome of surgery',
            'অস্ত্রোপচারের ফলাফল', required='yes'),
        _sr('end_group', 'grp_repaired'),
    ]

    # ── STAGE 5 · Rehabilitated & Reintegrated.
    rows += [
        _sr('begin_group', 'grp_rehab',
            'Stage 5 · Rehabilitated & Reintegrated',
            'ধাপ ৫ · পুনর্বাসিত ও পুনরায় সমাজে অন্তর্ভুক্ত',
            relevant="${stage}='rehabilitated'"),
        _sr('select_one yes_no', 'rehabilitation_received',
            'Receive rehabilitation support? (Y/N)',
            'পুনর্বাসন সহায়তা পেয়েছেন? (হ্যাঁ/না)'),
        _sr('date', 'rehabilitation_date',
            'Date of receiving support',
            'সহায়তা প্রাপ্তির তারিখ',
            relevant="${rehabilitation_received}='yes'"),
        _sr('select_one rehab_place', 'rehab_place',
            'Place of receiving support',
            'সহায়তা প্রাপ্তির স্থান',
            relevant="${rehabilitation_received}='yes'"),
        _sr('select_multiple rehab_support_types', 'rehab_support_types',
            'Types of rehabilitation support received',
            'পুনর্বাসন সহায়তার ধরন',
            relevant="${rehabilitation_received}='yes'"),
        _sr('text', 'rehab_notes',
            'Notes (optional)', 'মন্তব্য (ঐচ্ছিক)',
            relevant="${stage}='rehabilitated'"),
        _sr('end_group', 'grp_rehab'),
    ]
    return rows


def _fistula_choices():
    # District labels carry the fistula code (e.g. "Sunamganj (1)", "Dhaka (10)")
    # so the field worker sees the code that prefixes the patient ID right in the
    # dropdown. The choice VALUE stays the bare slug, so _dist_code / the ID
    # constraint are unaffected.
    fistula_districts = [
        _ch('district', d.lower().replace(' ', '_'),
            '%s (%s)' % (d, FISTULA_DISTRICT_CODE[d.lower().replace(' ', '_')]),
            '%s (%s)' % (DISTRICT_BANGLA[d], FISTULA_DISTRICT_CODE[d.lower().replace(' ', '_')]))
        for d in CIPRB_DISTRICTS
    ]
    ch = list(fistula_districts) + list(YES_NO)
    ch += [_ch('stage', s, en, bn) for s, en, bn in _FISTULA_STAGES]

    ch += [
        _ch('education', 'no_education',       'No education',
            'কোন শিক্ষা নাই'),
        _ch('education', 'primary_incomplete', 'Primary incomplete',
            'প্রাথমিক অসম্পূর্ণ'),
        _ch('education', 'primary',            'Primary',
            'প্রাথমিক'),
        _ch('education', 'secondary',          'Secondary',
            'মাধ্যমিক'),
        _ch('education', 'higher_secondary',   'Higher secondary',
            'উচ্চ মাধ্যমিক'),
        _ch('education', 'graduate',           'Graduate / Masters',
            'স্নাতক / স্নাতকোত্তর'),
    ]
    ch += [
        _ch('marital_status', 'married',   'Married',  'বিবাহিত'),
        _ch('marital_status', 'separated', 'Separated','পৃথক'),
        _ch('marital_status', 'divorced',  'Divorced', 'তালাকপ্রাপ্ত'),
        _ch('marital_status', 'widowed',   'Widowed',  'বিধবা'),
        _ch('marital_status', 'other',     'Other',    'অন্যান্য'),
    ]
    ch += [
        _ch('mode_of_delivery', 'nvd',
            'NVD (Normal Vaginal Delivery)',
            'এনভিডি (স্বাভাবিক প্রসব)'),
        _ch('mode_of_delivery', 'csection',
            'C-section', 'সিজারিয়ান'),
        _ch('mode_of_delivery', 'assisted_vaginal',
            'Assisted vaginal delivery',
            'সহায়ক স্বাভাবিক প্রসব'),
    ]
    ch += [
        _ch('place_of_delivery', 'gov_facility',
            'Government facility', 'সরকারি প্রতিষ্ঠান'),
        _ch('place_of_delivery', 'private_facility',
            'Private facility', 'বেসরকারি প্রতিষ্ঠান'),
        _ch('place_of_delivery', 'home', 'Home', 'বাড়ি'),
    ]
    ch += [
        _ch('delivery_conductor', 'relatives', 'Relatives',
            'আত্মীয়'),
        _ch('delivery_conductor', 'tba',       'TBA (traditional birth attendant)',
            'টিবিএ'),
        _ch('delivery_conductor', 'nurse',     'Nurse', 'নার্স'),
        _ch('delivery_conductor', 'midwife',   'Midwife', 'মিডওয়াইফ'),
        _ch('delivery_conductor', 'doctor',    'Doctor', 'ডাক্তার'),
    ]
    ch += [
        _ch('delivery_outcome', 'livebirth',  'Livebirth', 'জীবিত জন্ম'),
        _ch('delivery_outcome', 'stillbirth', 'Stillbirth','মৃত জন্ম'),
    ]
    ch += [
        _ch('reasons_no_facility', 'traditional', 'Traditional belief',
            'ঐতিহ্যবাহী বিশ্বাস'),
        _ch('reasons_no_facility', 'transport',   'Transport problem',
            'পরিবহন সমস্যা'),
        _ch('reasons_no_facility', 'financial',   'Financial problem',
            'আর্থিক সমস্যা'),
        _ch('reasons_no_facility', 'no_idea',     'No idea about hospital',
            'হাসপাতাল সম্পর্কে ধারণা নাই'),
        _ch('reasons_no_facility', 'no_faith',    'No faith in hospital service',
            'হাসপাতাল সেবায় আস্থা নাই'),
        _ch('reasons_no_facility', 'other',       'Other', 'অন্যান্য'),
    ]
    ch += [
        _ch('refer_outcome', 'reached',    'Reached the facility',
            'প্রতিষ্ঠানে পৌঁছেছেন'),
        _ch('refer_outcome', 'not_reached', 'Did not reach',
            'পৌঁছাননি'),
        _ch('refer_outcome', 'pending',    'Pending follow-up',
            'অপেক্ষমান'),
        _ch('refer_outcome', 'refused',    'Refused', 'অস্বীকার'),
    ]
    # The 4 fistula types CIPRB asked for (drop Other; add Congenital + Traumatic).
    ch += [
        _ch('fistula_type', 'obstetric',  'Obstetric',  'প্রসূতিজনিত'),
        _ch('fistula_type', 'iatrogenic', 'Iatrogenic', 'চিকিৎসাজনিত'),
        _ch('fistula_type', 'congenital', 'Congenital', 'জন্মগত'),
        _ch('fistula_type', 'traumatic',  'Traumatic',  'আঘাতজনিত'),
    ]
    ch += [
        _ch('iatrogenic_cause', 'hysterectomy', 'Hysterectomy',
            'জরায়ু অপসারণ'),
        _ch('iatrogenic_cause', 'csection',     'C-section',
            'সিজারিয়ান'),
        _ch('iatrogenic_cause', 'laparoscopy',  'Laparoscopy',
            'ল্যাপারোস্কপি'),
    ]
    ch += [
        _ch('genital_fistula_type', 'vvf', 'Vesico-vaginal fistula (VVF)',
            'ভেসিকো-ভ্যাজাইনাল ফিস্টুলা (VVF)'),
        _ch('genital_fistula_type', 'rvf', 'Recto vaginal fistula (RVF)',
            'রেক্টো-ভ্যাজাইনাল ফিস্টুলা (RVF)'),
        _ch('genital_fistula_type', 'ureterovaginal',
            'Uretero-vaginal', 'ইউরেটেরো-ভ্যাজাইনাল'),
        _ch('genital_fistula_type', 'urethrovaginal',
            'Urethro-vaginal', 'ইউরেথ্রো-ভ্যাজাইনাল'),
        _ch('genital_fistula_type', 'vesicouterine',
            'Vesico-Uterine', 'ভেসিকো-ইউটেরাইন'),
        _ch('genital_fistula_type', 'vesicocervical',
            'Vesico-Cervical', 'ভেসিকো-সারভিকাল'),
        _ch('genital_fistula_type', 'other', 'Other', 'অন্যান্য'),
    ]
    ch += [
        _ch('operation_route', 'vaginal',           'Vaginal', 'যৌনাঙ্গপথ'),
        _ch('operation_route', 'abdominal',         'Abdominal', 'উদর'),
        _ch('operation_route', 'abdomino_perineal', 'Abdomino-perineal',
            'উদর-পেরিনিয়াল'),
        _ch('operation_route', 'laparoscopy',       'Laparoscopy', 'ল্যাপারোস্কপি'),
    ]
    ch += [
        _ch('surgery_outcome_v2', 'success_dry',
            'Successfully repaired with a dry',
            'সফলভাবে নিরাময়, শুকনো'),
        _ch('surgery_outcome_v2', 'success_not_dry',
            'Successfully repaired but not dry',
            'সফলভাবে নিরাময় কিন্তু শুকনো নয়'),
        _ch('surgery_outcome_v2', 'failed',
            'Failed', 'ব্যর্থ'),
    ]
    ch += [
        _ch('rehab_place', 'individual',        'Individual support',
            'ব্যক্তিগত সহায়তা'),
        _ch('rehab_place', 'ngo',               'NGO', 'এনজিও'),
        _ch('rehab_place', 'union_council',     'Union Council',
            'ইউনিয়ন পরিষদ'),
        _ch('rehab_place', 'womens_affairs',    'Women Affairs Office',
            'মহিলা বিষয়ক অধিদপ্তর'),
        _ch('rehab_place', 'uno',               'UNO Office',
            'উপজেলা নির্বাহী কর্মকর্তার কার্যালয়'),
        _ch('rehab_place', 'social_welfare',    'Social Welfare Department',
            'সমাজসেবা অধিদপ্তর'),
        _ch('rehab_place', 'other',             'Other', 'অন্যান্য'),
    ]
    ch += [
        _ch('rehab_support_types', 'cash',         'Cash', 'নগদ অর্থ'),
        _ch('rehab_support_types', 'livestock',    'Livestock', 'পশুসম্পদ'),
        _ch('rehab_support_types', 'training',     'Training', 'প্রশিক্ষণ'),
        _ch('rehab_support_types', 'tree_plant',   'Tree plant', 'বৃক্ষরোপণ'),
        _ch('rehab_support_types', 'sewing',       'Sewing machine',
            'সেলাই মেশিন'),
        _ch('rehab_support_types', 'vgf_card',     'VGF Card',
            'ভিজিএফ কার্ড'),
        _ch('rehab_support_types', 'disability',   'Disability card',
            'প্রতিবন্ধী কার্ড'),
        _ch('rehab_support_types', 'psychosocial', 'Psychosocial support',
            'মনস্তাত্ত্বিক সহায়তা'),
        _ch('rehab_support_types', 'reintegration', 'Reintegration support',
            'পুনঃসংহতি সহায়তা'),
    ]
    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 2 — MPDSR Form 01 · Community Maternal Death Review              ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _shared_consent_block():
    """Standard Bangla consent paragraph that opens every MPDSR review."""
    return [
        _sr('note', 'consent_note',
            'CONSENT: How are you? My name is __ from the DGHS/DGFP. '
            'I would like to talk to you about the death and ask some '
            'questions as part of the Maternal & Perinatal Death '
            'Surveillance and Response (MPDSR) work jointly conducted '
            'by DGHS and DGFP. Your answers will be kept confidential. '
            'The interview takes 30–45 minutes. You may refuse or stop '
            'at any time. Your name will not be used in any report.',
            'সম্মতি: কেমন আছেন? আমার নাম __ আমি স্বাস্থ্য/পরিবার পরিকল্পনা '
            'অধিদপ্তরের একজন। মৃত্যু সম্পর্কে কিছু প্রশ্ন করব। সাক্ষাৎকার '
            '৩০–৪৫ মিনিট লাগবে। সকল তথ্য গোপন রাখা হবে; আপনি যেকোনো সময়ে '
            'সাক্ষাৎকার বন্ধ করতে পারেন। আপনার নাম প্রতিবেদনে উল্লেখ হবে না।'),
        _sr('select_one yes_no', 'consent_given',
            'Consent given by respondent?',
            'উত্তরদাতা সম্মতি দিয়েছেন?', required='yes'),
        _sr('date', 'interview_date',
            'Date of interview', 'সাক্ষাৎকারের তারিখ'),
    ]


def _respondent_block():
    """Respondent table from the paper: main + 2 associate respondents,
    each with name / relationship / present-at-death, plus respondent mobile."""
    return [
        _sr('begin_group', 'grp_respondent',
            "Respondent's information", 'উত্তরদাতার তথ্য'),
        _sr('text', 'respondent_mobile',
            'Respondent mobile number', 'উত্তরদাতার মোবাইল নম্বর',
            constraint='regex(., "^[0-9+ -]{6,20}$") or .=""',
            cmsg='Enter a valid phone number.'),
        _sr('text', 'respondent_main_name',
            'Main respondent — name', 'মুখ্য উত্তরদাতা — নাম'),
        _sr('select_one relationship', 'respondent_main_rel',
            'Main respondent — relationship with deceased',
            'মুখ্য উত্তরদাতা — মৃতের সাথে সম্পর্ক'),
        _sr('select_one yes_no', 'respondent_main_present',
            'Main respondent — present at time of death?',
            'মুখ্য উত্তরদাতা — মৃত্যুর সময় উপস্থিত ছিলেন?'),
        _sr('text', 'respondent_alt1_name',
            'Associate respondent 1 — name', 'সহযোগী উত্তরদাতা ১ — নাম'),
        _sr('select_one relationship', 'respondent_alt1_rel',
            'Associate 1 — relationship', 'সহযোগী ১ — সম্পর্ক'),
        _sr('select_one yes_no', 'respondent_alt1_present',
            'Associate 1 — present at time of death?',
            'সহযোগী ১ — উপস্থিত ছিলেন?'),
        _sr('text', 'respondent_alt2_name',
            'Associate respondent 2 — name', 'সহযোগী উত্তরদাতা ২ — নাম'),
        _sr('select_one relationship', 'respondent_alt2_rel',
            'Associate 2 — relationship', 'সহযোগী ২ — সম্পর্ক'),
        _sr('select_one yes_no', 'respondent_alt2_present',
            'Associate 2 — present at time of death?',
            'সহযোগী ২ — উপস্থিত ছিলেন?'),
        _sr('end_group', 'grp_respondent'),
    ]


def _community_maternal_survey():
    rows = _meta()
    rows += _office_use_block('Annual maternal death serial number',
                              'মাতৃমৃত্যুর বাৎসরিক ক্রমিক নং')
    rows += _shared_consent_block()
    rows += _respondent_block()

    # ── General information & mother identity (paper page 3) ──────────────
    rows += [
        _sr('begin_group', 'grp_identity',
            'General information & mother identity',
            'সাধারণ তথ্য ও মায়ের পরিচিতি',
            relevant="${consent_given}='yes'"),
        _sr('text', 'clinic_name',
            'Community Clinic name', 'কমিউনিটি ক্লিনিকের নাম'),
        _sr('text', 'clinic_code',
            'Community Clinic code (must be filled)',
            'কমিউনিটি ক্লিনিকের কোড (অবশ্যই পূরণ করতে হবে)',
            required='yes',
            hint='Numeric code printed on the paper form.'),
        _sr('text', 'deceased_name', "Mother's name", 'মায়ের নাম',
            required='yes'),
        _sr('text', 'mother_dhis2_code',
            "Mother's online registration (DHIS-2) code (must be filled)",
            'মায়ের অনলাইন রেজিস্ট্রেশন (ডিএইচআইএস-২) কোড '
            '(অবশ্যই পূরণ করতে হবে)',
            required='yes',
            hint='DHIS-2 registration number from the paper form.'),
        _sr('integer', 'deceased_age', "Mother's age (years)",
            'মায়ের বয়স (বছর)', required='yes',
            constraint='. > 9 and . < 60', cmsg='10–59'),
        _sr('select_one education_level', 'mother_education',
            "Mother's education", 'মায়ের শিক্ষাগত যোগ্যতা'),
        _sr('select_one ses', 'household_ses',
            'Household socio-economic status', 'খানার আর্থ-সামাজিক অবস্থা'),
        _sr('text', 'deceased_husband', "Husband's name", 'স্বামীর নাম'),
        _sr('integer', 'husband_age', "Husband's age (years)",
            'স্বামীর বয়স (বছর)', constraint='. > 0 and . < 110'),
        _sr('end_group', 'grp_identity'),
    ]

    # ── Section 1: Primary information (Q1–Q6) ────────────────────────────
    rows += [
        _sr('begin_group', 'grp_section1',
            'Section 1: Primary information', 'সেকশন ১: প্রাথমিক তথ্য',
            relevant="${consent_given}='yes'"),
        _sr('date', 'death_date', '1. Date of death', '১. মৃত্যুর তারিখ',
            required='yes'),
        _sr('time', 'death_time', '1a. Time of death (24-hour)',
            '১. মৃত্যুর সময় (২৪ ঘন্টা)'),
        _sr('select_one death_period', 'death_period',
            '2. When did the death occur?', '২. কোন সময় মৃত্যু ঘটেছিল?',
            required='yes'),
        _sr('select_one death_place', 'death_place',
            '3. Where did the death occur?', '৩. কোথায় মৃত্যু ঘটেছিল?',
            required='yes'),
        _sr('text', 'death_place_other', 'Other place (specify)',
            'অন্যান্য স্থান (উল্লেখ করুন)', relevant="${death_place}='other'"),
        _sr('integer', 'gestation_month',
            '4. Month of pregnancy at death', '৪. গর্ভাবস্থার কোন মাসে মৃত্যু',
            constraint='. >= 0 and . <= 10'),
        _sr('integer', 'gestation_week',
            '4a. Week of pregnancy at death', '৪. গর্ভাবস্থার কোন সপ্তাহে মৃত্যু',
            constraint='. >= 0 and . <= 45'),
        _sr('integer', 'children_born',
            '5. How many children has the mother delivered?',
            '৫. মা কতটি সন্তানের জন্ম দিয়েছেন?',
            constraint='. >= 0 and . <= 25'),
        _sr('integer', 'abortion_count',
            '6. How many abortions / miscarriages has the mother had?',
            '৬. মায়ের কতবার গর্ভপাত হয়েছে?', constraint='. >= 0 and . <= 25'),
        _sr('select_one yes_no', 'abortion_unknown',
            '6a. Is the number of abortions not known?',
            '৬. গর্ভপাতের সংখ্যা জানা নেই?'),
        _sr('end_group', 'grp_section1'),
    ]

    # ── Section 2: Maternal complications (Q7–Q9) ─────────────────────────
    comps = [
        ('comp_high_bp',          '1. High blood pressure',              '১. উচ্চ রক্তচাপ'),
        ('comp_diabetes',         '2. Diabetes',                         '২. ডায়াবেটিস'),
        ('comp_abortion',         '3. Abortion',                         '৩. গর্ভপাত'),
        ('comp_haemorrhage',      '4. Haemorrhage',                      '৪. রক্তক্ষরণ'),
        ('comp_high_fever',       '5. High fever',                       '৫. প্রচন্ড জ্বর'),
        ('comp_oedema',           '6. Water in face, legs and hands',    '৬. মুখ, পা ও হাতে পানি আসা'),
        ('comp_convulsion',       '7. Convulsion / eclampsia / fainting','৭. খিচুনি/একলামসিয়া/বেহুঁশ হওয়া'),
        ('comp_jaundice',         '8. Jaundice',                         '৮. জন্ডিস'),
        ('comp_anaemia',          '9. Anaemia',                          '৯. রক্তস্বল্পতা'),
        ('comp_blurred_vision',   '10. Blurred vision in the eyes',      '১০. চোখে ঝাপসা দেখা'),
        ('comp_prolonged_labour', '11. Labour pain for more than 12 hours','১১. ১২ ঘন্টার বেশী প্রসববেদনা'),
        ('comp_reduced_movement', '12. Reduced foetal movement or no movement for a long time','১২. ভ্রূণের নড়াচড়া কম হওয়া অথবা দীর্ঘ সময় নড়াচড়া না হওয়া'),
        ('comp_uterine_rupture',  '13. Tearing of the uterus',           '১৩. জরায়ু ছিঁড়ে যাওয়া'),
        ('comp_malpresentation',  '14. A part other than the head coming out','১৪. মাথা ছাড়া অন্যকোন অঙ্গ বের হওয়া'),
        ('comp_retained_placenta','15. Retained placenta',               '১৫. গর্ভফুল থেকে যাওয়া'),
        ('comp_foul_discharge',   '16. Foul-smelling discharge',         '১৬. দুর্গন্ধ যুক্ত স্রাব'),
        ('comp_abdominal_pain',   '17. Abnormal (severe) pain in lower abdomen','১৭. তলপেটে অস্বাভাবিক ব্যাথা (তীব্র)'),
        ('comp_other',            '18. Other, specify',                  '১৮. অন্যান্য, উল্লেখ করুন'),
    ]
    rows += [
        _sr('begin_group', 'grp_section2',
            'Section 2: Maternal complications', 'সেকশন ২: মাতৃত্বকালীন জটিলতা',
            relevant="${consent_given}='yes'"),
        _sr('select_one yes_no', 'prepreg_disease',
            '7. Did the mother suffer any disease before pregnancy?',
            '৭. গর্ভধারণের পূর্বে মা কোন রোগে আক্রান্ত ছিলেন কি?'),
        _sr('select_multiple prepreg_disease_type', 'prepreg_disease_types',
            '7a. If yes, which disease(s)?', '৭. হ্যাঁ হলে কোন রোগ(সমূহ)?',
            relevant="${prepreg_disease}='yes'"),
        _sr('text', 'prepreg_disease_other', 'Other disease (specify)',
            'অন্যান্য রোগ (উল্লেখ করুন)',
            relevant="selected(${prepreg_disease_types}, 'other')"),
        _sr('select_one last_pregnancy_outcome', 'last_pregnancy_outcome',
            '8. Outcome of the last pregnancy / delivery',
            '৮. শেষ গর্ভ / প্রসবের ফলাফল?'),
        _sr('note', 'q9_note',
            '9. Which complications occurred during the current pregnancy? '
            'For each, tick the phase(s) — antepartum, intrapartum, postpartum, '
            "or don't know.",
            '৯. বর্তমান গর্ভকালীন সময়ে কি কি জটিলতা ছিল? প্রতিটির জন্য '
            'পর্যায় চিহ্নিত করুন।'),
        _sr('begin_group', 'grp_q9',
            '9. Complications (current pregnancy)',
            '৯. জটিলতা (বর্তমান গর্ভকাল)', app='field-list'),
        _sr('select_multiple complication_phase', 'q9_header',
            'Complication ▸ phase', 'জটিলতা ▸ পর্যায়', app='label'),
    ]
    rows += [
        _sr('select_multiple complication_phase', name, en, bn,
            app='list-nolabel')
        for name, en, bn in comps
    ]
    rows += [
        _sr('text', 'comp_other_specify', 'Other complication (specify)',
            'অন্যান্য জটিলতা (উল্লেখ করুন)',
            relevant="selected(${comp_other}, 'antepartum') or "
                     "selected(${comp_other}, 'intrapartum') or "
                     "selected(${comp_other}, 'postpartum') or "
                     "selected(${comp_other}, 'dont_know')"),
        _sr('end_group', 'grp_q9'),
        _sr('end_group', 'grp_section2'),
    ]

    # ── Section 3: Antenatal care (Q10–Q13) ───────────────────────────────
    anc_seen = "${anc_count}!='none' and ${anc_count}!='unknown'"
    rows += [
        _sr('begin_group', 'grp_section3',
            'Section 3: Antenatal care (ANC)', 'সেকশন ৩: প্রসব পূর্বসেবা',
            relevant="${consent_given}='yes'"),
        _sr('select_one anc_count', 'anc_count',
            '10. How many times did the mother receive ANC?',
            '১০. কতবার প্রসবপূর্ব সেবা গ্রহণ করেছেন?'),
        _sr('select_multiple facility_place', 'anc_place',
            '11. Where was ANC received?', '১১. প্রসবপূর্ব সেবা কোথা থেকে নেওয়া হয়েছে?',
            relevant=anc_seen),
        _sr('text', 'anc_place_other', 'Other place (specify)',
            'অন্যান্য স্থান (উল্লেখ)', relevant="selected(${anc_place}, 'other')"),
        _sr('select_multiple provider_cadre', 'anc_provider',
            '12. Who provided ANC?', '১২. প্রসবপূর্ব সেবা কে প্রদান করেছেন?',
            relevant=anc_seen),
        _sr('text', 'anc_provider_other', 'Other provider (specify)',
            'অন্যান্য (উল্লেখ)', relevant="selected(${anc_provider}, 'other')"),
        _sr('select_multiple birth_plan', 'birth_plan',
            '13. What birth-plan preparations had been made?',
            '১৩. প্রসব পরিকল্পনায় কী কী ছিল?'),
        _sr('end_group', 'grp_section3'),
    ]

    # ── Section 4: Delivery / intrapartum (Q14–Q17) ───────────────────────
    delivered = ("${death_period}='delivery' or "
                 "${death_period}='postpartum_42d'")
    rows += [
        _sr('begin_group', 'grp_section4',
            'Section 4: Delivery information', 'সেকশন ৪: প্রসবকালীন তথ্য',
            relevant="${consent_given}='yes'"),
        _sr('note', 's4_note',
            'If the mother died before delivery, skip Q14–16 and go to Q17.',
            'মা প্রসবের পূর্বে মারা গেলে ১৪–১৬ বাদ দিয়ে ১৭ নং প্রশ্নে যান।'),
        _sr('select_one facility_place', 'delivery_place',
            '14. Where was the delivery conducted?',
            '১৪. কোথায় প্রসব করা হয়েছিল?', relevant=delivered),
        _sr('text', 'delivery_place_other', 'Other place (specify)',
            'অন্যান্য স্থান (উল্লেখ)', relevant="${delivery_place}='other'"),
        _sr('select_one provider_cadre', 'delivery_conductor',
            '15. Who conducted the delivery?',
            '১৫. কার দ্বারা প্রসব সংঘটিত হয়েছিল?', relevant=delivered),
        _sr('text', 'delivery_conductor_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ)', relevant="${delivery_conductor}='other'"),
        _sr('select_one delivery_mode', 'delivery_mode',
            '16. Mode of delivery', '১৬. কোন পদ্ধতিতে প্রসব হয়েছিল?',
            relevant=delivered),
        _sr('select_one delivery_outcome', 'delivery_outcome',
            '17. Outcome of the current pregnancy', '১৭. বর্তমান গর্ভের ফলাফল'),
        _sr('end_group', 'grp_section4'),
    ]

    # ── Section 5: Care received before death (Q18–Q21) ───────────────────
    rows += [
        _sr('begin_group', 'grp_section5',
            'Section 5: Care received before death',
            'সেকশন ৫: মৃত্যুর পূর্বে গ্রহণকৃত চিকিৎসা',
            relevant="${consent_given}='yes'"),
        _sr('select_one yes_no', 'treatment_received',
            '18. Did the mother receive any treatment before death?',
            '১৮. মৃত্যুর পূর্বে মা কোন চিকিৎসা গ্রহণ করেছিলেন কি?'),
        _sr('select_multiple facility_place', 'treatment_place',
            '19. If yes, where was treatment received?',
            '১৯. হ্যাঁ হলে চিকিৎসা গ্রহণের স্থান',
            relevant="${treatment_received}='yes'"),
        _sr('text', 'treatment_place_other', 'Other place (specify)',
            'অন্যান্য স্থান (উল্লেখ)',
            relevant="selected(${treatment_place}, 'other')"),
        _sr('select_multiple provider_cadre', 'treatment_provider',
            '20. Who provided the treatment?', '২০. কে চিকিৎসাসেবা দিয়েছিলেন?',
            relevant="${treatment_received}='yes'"),
        _sr('text', 'treatment_provider_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ)',
            relevant="selected(${treatment_provider}, 'other')"),
        _sr('select_multiple no_treatment_reason', 'no_treatment_reasons',
            '21. If no treatment was received, why not?',
            '২১. চিকিৎসাসেবা গ্রহণ না করে থাকলে তার কারণ কি ছিল?',
            relevant="${treatment_received}='no'"),
        _sr('text', 'no_treatment_other', 'Other reason (specify)',
            'অন্যান্য কারণ (উল্লেখ)',
            relevant="selected(${no_treatment_reasons}, 'other')"),
        _sr('end_group', 'grp_section5'),
    ]

    # ── Section 6: Postpartum period (Q22–Q26) — only if died postpartum ──
    pnc_seen = "${pnc_count} > 0 and ${pnc_count} != 99"
    rows += [
        _sr('begin_group', 'grp_section6',
            'Section 6: Postpartum period', 'সেকশন ৬: প্রসবোত্তরকাল',
            # Paper: Section 6 is N/A only for pregnancy or delivery deaths,
            # so it must also show for post-abortion deaths.
            relevant="${consent_given}='yes' and "
                     "${death_period}!='pregnancy' and "
                     "${death_period}!='delivery'"),
        _sr('note', 's6_note',
            'Section 6 applies only when the mother died in the postpartum '
            'period.',
            'মা প্রসবোত্তর সময়ে মারা গেলে এই সেকশন প্রযোজ্য।'),
        _sr('integer', 'death_after_delivery_days',
            '22. Days after delivery until death',
            '২২. প্রসবের কত দিন পর মৃত্যু', constraint='. >= 0 and . <= 42'),
        _sr('integer', 'death_after_delivery_hours', '22a. Hours', 'ঘন্টা',
            constraint='. >= 0 and . <= 23'),
        _sr('integer', 'death_after_delivery_minutes', '22b. Minutes', 'মিনিট',
            constraint='. >= 0 and . <= 59'),
        _sr('integer', 'pnc_count',
            '23. How many PNC visits did the mother receive?',
            '২৩. কতটি প্রসবোত্তর (পিএনসি) সেবা গ্রহণ করেছিলেন?',
            constraint='. >= 0 and . <= 99', hint='Enter 99 if not known.'),
        _sr('integer', 'pnc_first_days',
            '24. Days after delivery to first PNC',
            '২৪. প্রথম পিএনসি প্রসবের কত দিন পর', relevant=pnc_seen),
        _sr('integer', 'pnc_first_hours', '24a. Hours', 'ঘন্টা',
            relevant=pnc_seen),
        _sr('select_multiple facility_place', 'pnc_place',
            '25. Where was PNC received?', '২৫. প্রসবোত্তর সেবা গ্রহণের স্থান',
            relevant=pnc_seen),
        _sr('text', 'pnc_place_other', 'Other place (specify)',
            'অন্যান্য স্থান (উল্লেখ)', relevant="selected(${pnc_place}, 'other')"),
        _sr('select_multiple provider_cadre', 'pnc_provider',
            '26. Who provided PNC?', '২৬. প্রসবোত্তর সেবাদানকারী?',
            relevant=pnc_seen),
        _sr('text', 'pnc_provider_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ)', relevant="selected(${pnc_provider}, 'other')"),
        _sr('end_group', 'grp_section6'),
    ]

    # ── Narrative & opinion of cause (Q27–Q29) ────────────────────────────
    rows += [
        _sr('begin_group', 'grp_narrative', 'Narrative & cause',
            'বিবরণ ও কারণ', relevant="${consent_given}='yes'"),
        _sr('text', 'narrative_before_death',
            '27. Describe what happened just before the death (events, '
            'complications, how treatment started, social factors).',
            '২৭. মৃত্যুর ঠিক পূর্বে যা ঘটেছিল তা লিখুন।', app='multiline'),
        _sr('text', 'cause_opinion',
            '28. In your opinion, what caused this maternal death?',
            '২৮. মাতৃমৃত্যুর কি কারণ হতে পারে বলে আপনি মনে করেন?',
            app='multiline'),
        _sr('text', 'certificate_cause',
            '29. If a death certificate was issued, the cause per the '
            'certificate',
            '২৯. মৃত্যু সার্টিফিকেট প্রদান করা হলে সার্টিফিকেট অনুযায়ী কারণ',
            app='multiline'),
        _sr('end_group', 'grp_narrative'),
    ]

    # ── Consultant / doctor ICD-10 cause coding (mandatory by consultant) ─
    rows += [
        _sr('begin_group', 'grp_icd',
            'To be filled by the Consultant / Doctor',
            'কনসালটেন্ট/ডাক্তার কর্তৃক পূরণের জন্য '
            '(অবশ্যই কনসালটেন্ট/ডাক্তার দ্বারা পূরণ করতে হবে)'),
        _sr('note', 'icd_ref_note',
            'Identify / mark the cause as per the ICD 10 table.',
            'ICD 10 ছক অনুযায়ী কারণ নিরূপণ/চিহ্নিত করুন।'),
        _sr('select_one icd_cause', 'icd_cause',
            'Cause', 'কারণ'),
        _sr('text', 'icd_code', 'ICD code', 'ICD code',
            hint='e.g. O72'),
        _sr('text', 'icd_disease_name', 'Name of disease', 'রোগের নাম'),
        _sr('text', 'icd_diagnoser_name', 'Name of person coding',
            'নিরূপণকারীর নাম'),
        _sr('text', 'icd_diagnoser_designation', 'Designation', 'পদবী'),
        _sr('text', 'icd_diagnoser_institution', 'Institution', 'প্রতিষ্ঠান'),
        _sr('date', 'icd_date', 'Date', 'তারিখ'),
        _sr('end_group', 'grp_icd'),
    ]
    return rows


def _community_maternal_choices():
    ch = list(DISTRICT_CHOICES) + list(YES_NO)

    # Relationship list (per MPDSR Form 01 paper).
    rel = [
        ('husband',      'Husband',                'স্বামী'),
        ('father_inlaw', 'Father-in-law',          'শ্বশুর'),
        ('mother',       'Mother',                 'মা'),
        ('aunt_uncle',   'Aunt / uncle',           'খালা / খালু / চাচী / চাচা / মামী / মামা / ফুপু / ফুপা'),
        ('neighbour',    'Neighbour',              'প্রতিবেশী'),
        ('mother_inlaw', 'Mother-in-law',          'শাশুড়ী'),
        ('sister_inlaw', 'Sister-in-law',          'শ্যালিকা'),
        ('father',       'Father',                 'বাবা'),
        ('sibling',      'Elder brother / sister', 'বড় বোন / ভাই'),
        ('other',        'Other (specify)',        'অন্যান্য'),
    ]
    ch += [_ch('relationship', k, en, bn) for k, en, bn in rel]

    # Q2 — period of death (4).
    ch += [
        _ch('death_period', 'pregnancy', 'During pregnancy', 'গর্ভকালীন সময়'),
        _ch('death_period', 'post_abortion_28w',
            'After abortion (within 28 weeks)',
            'গর্ভপাতের পরে (২৮ সপ্তাহের মধ্যে)'),
        _ch('death_period', 'delivery', 'During delivery', 'প্রসবকালীন সময়'),
        _ch('death_period', 'postpartum_42d',
            'Postpartum (within 42 days)', 'প্রসবোত্তর সময় (৪২ দিনের মধ্যে)'),
    ]

    # Q3 — place of death (10).
    ch += [
        _ch('death_place', 'home', 'Home', 'বাড়িতে'),
        _ch('death_place', 'union_hfwc',
            'Union Health & Family Welfare Centre',
            'ইউনিয়ন স্বাস্থ্য ও পরিবার কল্যাণ কেন্দ্র'),
        _ch('death_place', 'district_hospital',
            'District / Sadar hospital', 'জেলা অথবা সদর হাসপাতাল'),
        _ch('death_place', 'medical_college',
            'Medical College hospital', 'মেডিকেল কলেজ হাসপাতাল'),
        _ch('death_place', 'private_clinic',
            'Private clinic / hospital', 'প্রাইভেট ক্লিনিক / হাসপাতাল'),
        _ch('death_place', 'in_transit', 'On the way (in transit)', 'পথে'),
        _ch('death_place', 'upazila_hc',
            'Upazila Health Complex', 'উপজেলা স্বাস্থ্য কমপ্লেক্স'),
        _ch('death_place', 'maternal_centre',
            'Maternal & Child Welfare Centre', 'মাতৃ মঙ্গল কেন্দ্র'),
        _ch('death_place', 'ngo_clinic', 'NGO clinic', 'এনজিও ক্লিনিক'),
        _ch('death_place', 'other', 'Other (specify)', 'অন্যান্য'),
    ]

    # Mother's education (5).
    ch += [
        _ch('education_level', 'none', 'None', 'নাই'),
        _ch('education_level', 'class5', 'Up to class 5', '৫ম শ্রেণী পর্যন্ত'),
        _ch('education_level', 'class10', 'Up to class 10', '১০ম শ্রেণী পর্যন্ত'),
        _ch('education_level', 'class12', 'Up to class 12', 'দ্বাদশ শ্রেণী পর্যন্ত'),
        _ch('education_level', 'graduate', 'Graduate / degree', 'স্নাতক / ডিগ্রি'),
    ]
    # Household socio-economic status (4).
    ch += [
        _ch('ses', 'extreme_poor', 'Extreme poor', 'অতিগরীব'),
        _ch('ses', 'poor', 'Poor', 'গরীব'),
        _ch('ses', 'middle', 'Middle class', 'মধ্যবিত্ত'),
        _ch('ses', 'rich', 'Rich', 'ধনী'),
    ]

    # Q7 — pre-pregnancy disease (9).
    ch += [
        _ch('prepreg_disease_type', 'high_bp', 'High blood pressure', 'উচ্চ রক্তচাপ'),
        _ch('prepreg_disease_type', 'diabetes', 'Diabetes', 'ডায়াবেটিস'),
        _ch('prepreg_disease_type', 'heart_disease', 'Heart disease', 'হৃদরোগ'),
        _ch('prepreg_disease_type', 'convulsion', 'Convulsion', 'খিচুনি'),
        _ch('prepreg_disease_type', 'anaemia', 'Anaemia', 'রক্তস্বল্পতা'),
        _ch('prepreg_disease_type', 'tb', 'Tuberculosis (TB)', 'যক্ষা'),
        _ch('prepreg_disease_type', 'asthma', 'Asthma', 'হাঁপানি'),
        _ch('prepreg_disease_type', 'jaundice', 'Jaundice', 'জন্ডিস'),
        _ch('prepreg_disease_type', 'other', 'Other (specify)', 'অন্যান্য'),
    ]
    # Q8 — outcome of last pregnancy (6).
    ch += [
        _ch('last_pregnancy_outcome', 'abortion', 'Abortion', 'গর্ভপাত'),
        _ch('last_pregnancy_outcome', 'stillbirth', 'Stillbirth', 'মৃতজন্ম'),
        _ch('last_pregnancy_outcome', 'livebirth', 'Live birth', 'জীবিতজন্ম'),
        _ch('last_pregnancy_outcome', 'preterm', 'Preterm delivery',
            'নির্দিষ্ট সময়ের পূর্বে প্রসব'),
        _ch('last_pregnancy_outcome', 'csection', 'Caesarean section', 'সিজারিয়ান'),
        _ch('last_pregnancy_outcome', 'other_operative',
            'Other operative delivery', 'অন্যান্য অপারেশন পদ্ধতিতে প্রসব'),
    ]
    # Q9 — complication phases / matrix columns (4).
    ch += [
        _ch('complication_phase', 'antepartum', 'Antepartum', 'প্রসবপূর্ব'),
        _ch('complication_phase', 'intrapartum', 'Intrapartum', 'প্রসবকালীন'),
        _ch('complication_phase', 'postpartum', 'Postpartum', 'প্রসবোত্তর'),
        _ch('complication_phase', 'dont_know', "Don't know", 'জানা নেই'),
    ]

    # Q10 — ANC count (also used as the count select).
    ch += [
        _ch('anc_count', 'none', 'None (00)', 'নাই'),
        _ch('anc_count', '1', '1', '১'),
        _ch('anc_count', '2', '2', '২'),
        _ch('anc_count', '3', '3', '৩'),
        _ch('anc_count', '4_plus', '4 or more', '৪ বা তার বেশি'),
        _ch('anc_count', 'unknown', 'Unknown (99)', 'জানা নেই'),
    ]

    # Shared facility list — Q11 / Q14 / Q19 / Q25 (11).
    fac = [
        ('home', 'Home', 'বাড়ি'),
        ('community_clinic', 'Community Clinic', 'কমিউনিটি ক্লিনিক'),
        ('union_hfwc', 'Union Health & Family Welfare Centre',
         'ইউনিয়ন স্বাস্থ্য ও পরিবার কল্যাণ কেন্দ্র'),
        ('upazila_hc', 'Upazila Health Complex', 'উপজেলা স্বাস্থ্য কমপ্লেক্স'),
        ('maternal_centre', 'Maternal & Child Welfare Centre', 'মাতৃ মঙ্গল কেন্দ্র'),
        ('district_hospital', 'District / Sadar hospital', 'জেলা অথবা সদর হাসপাতাল'),
        ('medical_college', 'Medical College hospital', 'মেডিকেল কলেজ হাসপাতাল'),
        ('private_clinic', 'Private clinic / hospital', 'প্রাইভেট ক্লিনিক / হাসপাতাল'),
        ('ngo_clinic', 'NGO clinic', 'এনজিও ক্লিনিক'),
        ('provider_home', "Provider's chamber / home",
         'চেম্বার / স্বাস্থ্য সেবা দানকারীর বাড়ী'),
        ('other', 'Other (specify)', 'অন্যান্য'),
    ]
    ch += [_ch('facility_place', k, en, bn) for k, en, bn in fac]

    # Shared provider-cadre list — Q12 / Q15 / Q20 / Q26 (12).
    prov = [
        ('doctor_mbbs', 'Doctor (MBBS)', 'ডাক্তার (MBBS)'),
        ('nurse', 'Nurse', 'নার্স'),
        ('fwv', 'Family Welfare Visitor (FWV)', 'পরিবার পরিকল্পনা ভিজিটর (FWV)'),
        ('csba', 'CSBA', 'সিএসবিএ (CSBA)'),
        ('ma', 'Medical Assistant (MA)', 'মেডিকেল অ্যাসিস্ট্যান্ট (MA)'),
        ('ha', 'Health Assistant (HA)', 'স্বাস্থ্য সহকারী (HA)'),
        ('fwa', 'Family Welfare Assistant (FWA)', 'পরিবার পরিকল্পনা সহকারী (FWA)'),
        ('dai', 'Dai (TBA)', 'দাই'),
        ('palli_chikitsok', 'Palli chikitsok (village doctor)', 'পল্লী চিকিৎসক'),
        ('ngo_worker', 'NGO worker', 'এনজিও কর্মী'),
        ('midwife', 'Midwife', 'মিড ওয়াইফ (Midwife)'),
        ('other', 'Other (specify)', 'অন্যান্য'),
    ]
    ch += [_ch('provider_cadre', k, en, bn) for k, en, bn in prov]

    # Q13 — birth-plan elements (8).
    bp = [
        ('place_planned', 'Place of delivery planned',
         'কোথায় প্রসব করবেন তা পরিকল্পনায় ছিল'),
        ('attendant_planned', 'Who will conduct delivery planned',
         'কার দ্বারা প্রসব তা পরিকল্পনায় ছিল'),
        ('transport_planned', 'How to reach facility planned',
         'কিভাবে কেন্দ্রে যাবেন তা পরিকল্পনায় ছিল'),
        ('referral_planned', 'Where to go on complication decided',
         'জটিলতায় কোথায় যাবেন তা ঠিক ছিল'),
        ('money_saved', 'Money saved for care / transport',
         'চিকিৎসা / যাতায়াতের টাকা জমানো ছিল'),
        ('caregiver_arranged', 'Caregiver for mother / home arranged',
         'মায়ের পরিচর্যাকারী ঠিক করা ছিল'),
        ('blood_donor', 'Blood donor identified / group tested',
         'রক্তদাতা নির্ধারণ / গ্রুপ পরীক্ষা ছিল'),
        ('newborn_items', 'Newborn-care items ready',
         'নবজাতক পরিচর্যার সরঞ্জাম প্রস্তুত ছিল'),
    ]
    ch += [_ch('birth_plan', k, en, bn) for k, en, bn in bp]

    # Q16 — mode of delivery (4) — verbatim from the paper.
    ch += [
        _ch('delivery_mode', 'vaginal_spontaneous', 'Vaginal-spontaneous',
            'স্পন্টিনিওয়াস'),
        _ch('delivery_mode', 'instrumental_vaginal', 'Instrumental vaginal',
            'ভ্যাকুয়াম/ফরসেপ'),
        _ch('delivery_mode', 'csection', 'Caesarean Section', 'সিজারিয়ান'),
        _ch('delivery_mode', 'destructive', 'Destructive operations',
            'ডেস্ট্রাক্টিভ অপারেশন'),
    ]
    # Q17 — delivery outcome (4).
    ch += [
        _ch('delivery_outcome', 'livebirth', 'Live birth', 'জীবিতজন্ম'),
        _ch('delivery_outcome', 'stillbirth', 'Stillbirth', 'মৃতজন্ম'),
        _ch('delivery_outcome', 'abortion', 'Abortion', 'গর্ভপাত'),
        _ch('delivery_outcome', 'not_delivered', 'Not delivered', 'প্রসব হয়নি'),
    ]

    # Q21 — reasons for not receiving treatment (16) — Bangla verbatim.
    reasons = [
        ('unnecessary', 'Thought unnecessary', 'অপ্রয়োজনীয় মনে করেছিল'),
        ('costly', 'Was costly', 'ব্যয়বহুল ছিল'),
        ('too_far', 'Distance was far', 'দূরত্ব বেশি ছিল'),
        ('no_escort', 'No companion to take', 'নিয়ে যাওয়ার সঙ্গী না ছিল না'),
        ('family_refusal', 'Family disagreement', 'পারিবারিক অসম্মতি ছিল'),
        ('no_way_known', 'Did not know the way to go', 'যাওয়ার উপায় জানা ছিল না'),
        ('dont_know_where', 'Did not know where to go', 'কোথায় যেতে হবে না জানা'),
        ('disaster_night', 'Natural disaster / bad weather / night',
         'প্রাকৃতিক দুর্যোগ/খারাপ আবহাওয়া/রাত্রি'),
        ('didnt_understand', 'Did not realise treatment was needed',
         'চিকিৎসার প্রয়োজন বুঝতে পারেনি'),
        ('no_money', 'Lack of money', 'অর্থের অভাব ছিল'),
        ('no_transport', 'Lack of transport', 'যানবাহনের অভাব ছিল'),
        ('poor_service', 'Service quality', 'সেবার মান অনুমত'),
        ('good_care_home', 'Got good care at home', 'বাড়ীতে ভাল সেবা পাওয়ায়'),
        ('no_time', 'Lack of time', 'সময়ের অভাব ছিল'),
        ('fear_care', 'Fear of receiving care', 'সেবাগ্রহণ ভয় পাওয়া'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    ch += [_ch('no_treatment_reason', k, en, bn) for k, en, bn in reasons]

    # Consultant ICD-10 cause table — transcribed VERBATIM from the paper's
    # English-only table; spelling kept exactly as printed (typos included).
    icd = [
        ('pph', 'PPH (O72)'),
        ('aph', 'APH (O46)'),
        ('puerperal_sepsis', 'Puerperal Sepsis (O85)'),
        ('ectopic', 'Ectopic Pregnancy (O00)'),
        ('eclampsia', 'Eclampsia (O15)'),
        ('haemorrhage_early', 'Haemorrhage in Early Pregnancy (O20)'),
        ('sequel', 'Death from sequel of direct obstetric cause (O97)'),
        ('failed_abortion', 'Failed Attempt abortion (O07)'),
        ('obstructed_labour',
         'Obstructed Labour due to Malposition and Malpresentation of foetus (O64)'),
        ('anaesthesia_ld',
         'Complication of Anaesthesia durind Labour & Delivery (O74)'),
        ('placenta_previa', 'Placenta Previa (O44)'),
        ('abruptio', 'Abruptio placentae (O45)'),
        ('medical_abortion', 'Medical abortion (O04)'),
        ('repture_uterus', 'Repture Uterus (O71)'),
        ('anaesthesia_preg', 'Complications of anaesthesia during pregnancy (O29)'),
        ('obstetric_emblism', 'Obstetric Emblism (O88)'),
        ('malnutrition', 'Malnutrition in pregnancy (O25)'),
    ]
    ch += [_ch('icd_cause', k, en, en) for k, en in icd]

    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 3 — MPDSR Form 02 · Community Neonatal Death Review              ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _community_neonatal_survey():
    rows = _meta()
    rows += _office_use_block('Annual neonatal death serial number',
                              'নবজাতক মৃত্যুর বাৎসরিক ক্রমিক নং')
    rows += _shared_consent_block()
    rows += _respondent_block()

    consented = "${consent_given}='yes'"

    # ── General information (paper page 3: সাধারণ তথ্য) ────────────────────
    rows += [
        _sr('begin_group', 'grp_identity',
            'General information', 'সাধারণ তথ্য',
            relevant=consented),
        _sr('text', 'clinic_name',
            'Community Clinic name', 'কমিউনিটি ক্লিনিকের নামঃ'),
        _sr('text', 'clinic_code',
            'Community Clinic code (must be filled)',
            'কমিউনিটি ক্লিনিকের কোডঃ (অবশ্যই পূরণ করতে হবে)',
            required='yes',
            hint='Numeric code printed on the paper form.'),
        _sr('select_one district', 'address_district',
            'District (address)', 'জেলা (ঠিকানা)'),
        _sr('text', 'address_upazila', 'Upazila (address)', 'উপজেলা (ঠিকানা)'),
        _sr('text', 'mother_name', "Mother's name", 'মায়ের নামঃ',
            required='yes'),
        _sr('text', 'mother_dhis2_code',
            "Mother's online registration (DHIS-2) coding",
            'মায়ের অনলাইন রেজিস্ট্রেশন (ডিএইচআইএস-২) কোডিংঃ',
            hint='DHIS-2 registration number from the paper form, if available.'),
        _sr('integer', 'mother_age', "Mother's age (years)",
            'মায়ের বয়সঃ (বৎসর)', constraint='. > 9 and . < 60', cmsg='10–59'),
        _sr('select_one education_level', 'mother_education',
            "Mother's education", 'মায়ের শিক্ষাগত যোগ্যতা ঃ'),
        _sr('select_one ses', 'household_ses',
            'Household socio-economic status', 'খানার আর্থ-সামাজিক অবস্থাঃ'),
        _sr('text', 'father_name', "Father's name", 'পিতার নামঃ'),
        _sr('integer', 'father_age', "Father's age (years)",
            'পিতার বয়সঃ (বৎসর)', constraint='. > 0 and . < 110'),
        _sr('text', 'child_name', "Child's name", 'শিশুর নামঃ'),
        # Paper page 3: child DHIS-2 box carries NO "must be filled" mark
        # (unlike clinic code + mother DHIS-2), so it is NOT required.
        _sr('text', 'child_dhis2_code',
            "Child's online registration number (DHIS-2)",
            'শিশুর অনলাইন রেজিস্ট্রেশন নম্বর (ডিএইচআইএস-২) :',
            hint='DHIS-2 registration number from the paper form, if any.'),
        _sr('end_group', 'grp_identity'),
    ]

    # ── Section 1: Neonate information (Q1–Q5) ────────────────────────────
    rows += [
        _sr('begin_group', 'grp_section1',
            "Section 1: Neonate's information", 'সেকশন ১: নবজাতকের তথ্য',
            relevant=consented),
        _sr('date', 'birth_date', '1. Date of birth', '১। জন্ম তারিখঃ'),
        _sr('time', 'birth_time', '1. Time of birth (24-hour)',
            '১। সময় (২৪ঘন্টা)ঃ'),
        _sr('date', 'death_date', '2. Date of death', '২। মৃত্যুর তারিখঃ',
            required='yes'),
        _sr('time', 'death_time', '2. Time of death (24-hour)',
            '২। সময় (২৪ঘন্টা)ঃ'),
        # Paper Q2: "নবজাতকের বয়স ___ দিন · ২৪ ঘন্টার কম হলে <১ লিখুন".
        # Field is integer, so "<1" cannot be typed — capture 0 for any
        # death within the first 24 hours; the hint explains this.
        _sr('integer', 'age_at_death_days',
            "2. Neonate's age (days) — write <1 if under 24 hours",
            '২। নবজাতকের বয়স (দিন)  ২৪ ঘন্টার কম হলে <১ লিখুন',
            constraint='. >= 0 and . <= 28',
            hint='Whole days. Enter 0 if the neonate died within 24 hours '
                 '(paper "<1").'),
        _sr('select_one death_place', 'death_place',
            '3. Where did the death occur? (tick the correct box)',
            '৩। কোথায় মৃত্যু ঘটেছিল? (সঠিক বক্সে ঠিক দিন)',
            required='yes',
            hint='Tick the correct box.'),
        _sr('text', 'death_place_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)', relevant="${death_place}='other'"),
        _sr('integer', 'sick_duration_days',
            '4. How long was the neonate sick before death?',
            '৪। নবজাতক মৃত্যুর পূর্বে কত সময় অসুস্থ ছিল ?',
            constraint='. >= 0 and . <= 28'),
        # Paper Q5: হ্যাঁ / না / জানি না — yes_no (Yes/No/Unknown) maps the
        # "জানি না" tick onto Unknown.
        _sr('select_one yes_no', 'death_by_injury',
            '5. Did the neonate die from any physical injury?',
            '৫। নবজাতক কি কোন শারীরিক আঘাতে মৃত্যুবরন করেছে?'),
        _sr('text', 'death_injury_type',
            '5. If yes, what type of injury (specify)',
            '৫। হ্যাঁ হলে কি ধরনের আঘাত, উল্লেখ করুন',
            relevant="${death_by_injury}='yes'"),
        _sr('end_group', 'grp_section1'),
    ]

    # ── Section 2: Mother's information (Q6–Q16) ──────────────────────────
    anc_seen = "${anc_count}!='none' and ${anc_count}!='unknown'"
    rows += [
        _sr('begin_group', 'grp_section2',
            "Section 2: Mother's information", 'সেকশন ২: মায়ের তথ্য',
            relevant=consented),
        # Paper Q6 is ONE question with a মাস / সপ্তাহ box; kept as two
        # integer fields (month + week) but both carry the paper's Q6 text.
        _sr('integer', 'gestation_month',
            '6. At how many months or weeks of gestation was the delivery? '
            '(months)',
            '৬। গর্ভের কত মাস বা সপ্তাহে প্রসব হয়েছে? (মাস)',
            constraint='. >= 0 and . <= 10'),
        _sr('integer', 'gestation_week',
            '6. At how many months or weeks of gestation was the delivery? '
            '(weeks)',
            '৬। গর্ভের কত মাস বা সপ্তাহে প্রসব হয়েছে? (সপ্তাহ)',
            constraint='. >= 0 and . <= 45'),
        _sr('select_multiple pregnancy_complication', 'pregnancy_complications',
            '7. Were there any complications related to delivery during the '
            'pregnancy period? (more than one answer possible)',
            '৭। গর্ভকালীন সময়কালে প্রসব সংক্রান্ত কোন জটিলতা হয়েছিল কিনা? '
            '(একাধিক উত্তর হতে পারে )'),
        _sr('text', 'pregnancy_complication_other',
            'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="selected(${pregnancy_complications}, 'other')"),
        _sr('select_one anc_count', 'anc_count',
            '8. How many times did you receive antenatal (ANC) care?',
            '৮। কতবার প্রসবপূর্ব সেবা (এএনসি) গ্রহণ করেছেন?',
            hint='If no ANC or not known, go to Q11.'),
        _sr('select_multiple facility_place', 'anc_place',
            '9. From which place was antenatal care received? '
            '(more than one answer may be ticked)',
            '৯। প্রসবপূর্ব সেবা কোন স্থান হতে গ্রহণ করা হয়েছে? '
            '(একাধিক উত্তর ✓ টিক চিহ্ন দেয়া যাবে)',
            relevant=anc_seen),
        _sr('text', 'anc_place_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="selected(${anc_place}, 'other')"),
        _sr('select_multiple provider_cadre', 'anc_provider',
            '10. Who provided the antenatal care? '
            '(more than one answer ✓ tick)',
            '১০। প্রসবপূর্ব সেবা কে প্রদান করেছেন? (একাধিক উত্তর ✓ টিক চিহ্ন দিন)',
            relevant=anc_seen),
        _sr('text', 'anc_provider_other', 'Other (specify)',
            'অন্যান্য উল্লেখ করুন',
            relevant="selected(${anc_provider}, 'other')"),
        _sr('integer', 'parity',
            '11. How many times has the mother given birth? Write the number.',
            '১১। মা কত বার সন্তান প্রসব করেছেন? সংখ্যা লিখুন',
            constraint='. >= 0 and . <= 25'),
        _sr('integer', 'abortion_count',
            '12. How many times has the mother had an abortion? '
            'Write the number.',
            '১২। মায়ের কতবার গর্ভপাত হয়েছে? সংখ্যা লিখুন',
            constraint='. >= 0 and . <= 25'),
        _sr('select_one yes_no', 'abortion_unknown',
            '12. Tick if not known.',
            '১২। জানা না থাকলে টিক দিন'),
        _sr('select_one facility_place', 'birth_place',
            '13. Place of birth of the neonate?', '১৩। নবজাতক প্রসবের স্থান?',
            required='yes'),
        _sr('text', 'birth_place_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="${birth_place}='other'"),
        _sr('select_one provider_cadre', 'delivery_conductor',
            '14. By whom was the delivery conducted?',
            '১৪। কার দ্বারা প্রসব সংগঠিত হয়েছিল?'),
        _sr('text', 'delivery_conductor_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="${delivery_conductor}='other'"),
        _sr('select_one delivery_mode', 'delivery_mode',
            '15. By which method was the delivery done?',
            '১৫। কোন পদ্ধতিতে প্রসব হয়েছিল?'),
        _sr('select_multiple delivery_complication', 'delivery_complications',
            '16. Did the mother show any complications during delivery? '
            'What type of complications occurred? (more than one answer ✓ tick)',
            '১৬। ডেলিভারির সময় মায়ের কোন জটিলতা দেখা দিয়েছিল? কোন ধরনের '
            'জটিলতা হয়েছিল? (একাধিক উত্তর ✓ টিক চিহ্ন দিন)'),
        _sr('end_group', 'grp_section2'),
    ]

    # ── Section 3: Post-delivery information (Q17–Q27) ────────────────────
    rows += [
        _sr('begin_group', 'grp_section3',
            'Section 3: Post-delivery information', 'সেকশন ৩ : প্রসবোত্তর তথ্য',
            relevant=consented),
        _sr('select_one yes_no', 'twin_birth',
            '17. Did the mother deliver twins?',
            '১৭। মা যমজ শিশু প্রসব  করেছেন কি না ?'),
        _sr('select_one birth_weight_band', 'birth_weight_band',
            "18. What was the neonate's weight at birth?",
            '১৮। জন্ম কালীন সময় শিশুর ওজন কতো ছিল?'),
        _sr('select_multiple congenital_defect', 'congenital_defects',
            '19. Were there any congenital defects?',
            '১৯ । কোন জন্মগত ক্রটি ছিল কিনা?'),
        _sr('text', 'congenital_defect_other', 'Other (specify)',
            'অন্যান্য ((উল্লেখ করুন)',
            relevant="selected(${congenital_defects}, 'other')"),
        _sr('select_one cried_breathed', 'cried_breathed',
            '20. After birth, did the neonate cry / breathe?',
            '২০ । শিশু জন্মের পরে কান্না করেছিলো/শ্বাস নিয়েছিল কিনা?',
            hint='If the answer is "Yes / breathed normally" or "Don\'t know", '
                 'go to Q27.'),
        _sr('select_multiple resuscitation_action', 'resuscitation_actions',
            '21. If the answer was weak cry or breathed after a long time, '
            'what measures were taken?',
            '২১ । যদি উত্তরে দুর্বল কান্না  বা অনেক সময় পরে শ্বাস নিয়েছিল  হয় তবে '
            'কি ব্যবস্থা নেয়া হয়েছিলো',
            relevant="${cried_breathed}='weak_cry' or "
                     "${cried_breathed}='delayed_breath'"),
        _sr('text', 'resuscitation_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="selected(${resuscitation_actions}, 'other')"),
        # Q22 — neonatal danger signs (multi-tick, 13 options).
        _sr('select_multiple danger_sign', 'danger_signs',
            '22. What danger signs did the neonate show? '
            '(more than one tick possible)',
            '২২. নবজাতককে কোন বিপদজনক চিহ্ন দেখা দিয়েছিল? '
            '(একের অধিক টিক চিহ্ন দিতে পারেন)'),
        _sr('text', 'danger_sign_other', 'Other (specify)',
            'অন্যান্য (',
            relevant="selected(${danger_signs}, 'other')"),
        _sr('select_one yes_no', 'treatment_received',
            '23. Did the child receive any treatment before death?',
            '২৩ । শিশু মৃত্যুর পূর্বে কোন চিকিৎসা গ্রহন করেছেন কিনা?',
            hint='If yes, go to Q24. If no, go to Q25.'),
        _sr('select_multiple facility_place', 'treatment_place',
            '24. Place where the post-delivery care was received? '
            '(more than one answer tick)',
            '২৪ । প্রসবোত্তর সেবা গ্রহণের স্থান? (একাধিক উত্তরটিক চিহ্ন দিন)',
            relevant="${treatment_received}='yes'"),
        _sr('text', 'treatment_place_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="selected(${treatment_place}, 'other')"),
        _sr('select_multiple no_treatment_reason', 'no_treatment_reasons',
            '25. If treatment care was not received, what was the reason? '
            '(more than one answer)',
            '২৫ । চিকিৎসাসেবা গ্রহণ না করলে তার কারণ? (একাধিক উত্তর)',
            relevant="${treatment_received}='no'"),
        _sr('text', 'no_treatment_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="selected(${no_treatment_reasons}, 'other')"),
        _sr('text', 'cause_opinion',
            '26. In your opinion, write the probable cause of this death?',
            '২৬ । আপনার মতে এই মৃত্যুর সম্ভাব্য কারণ লিখুন?',
            app='multiline'),
        _sr('text', 'certificate_cause',
            '27. If a death certificate of this neonate exists, what was the '
            'cause of death according to the certificate?',
            '২৭ । যদি এই নবজাতকের মৃত্যুর সার্টিফিকেট থাকে তাহলে সার্টিফিকেট '
            'অনুযায়ী মৃত্যুর কারণ কি ছিল?',
            app='multiline'),
        _sr('end_group', 'grp_section3'),
    ]

    # ── Consultant / doctor ICD-10 cause coding (must be filled by doctor) ─
    # Paper page 8: "কনসালটেন্ট/ডাক্তার কর্তৃক পূরণের জন্য
    # (অবশ্যই কনসালটেন্ট/ডাক্তার দ্বারা পূরণ করতে হবে)".
    rows += [
        _sr('begin_group', 'grp_icd',
            'For consultant / doctor (must be filled by consultant / doctor)',
            'কনসালটেন্ট/ডাক্তার কর্তৃক পূরণের জন্য (অবশ্যই কনসালটেন্ট/'
            'ডাক্তার দ্বারা পূরণ করতে হবে)'),
        _sr('note', 'icd_ref_note',
            'Identify / mark the cause per the ICD 10 table — '
            'Birth Asphyxia P 21 · Low Birth Weight P 07 · Meningitis G 00 · '
            'Congenital Anomalies Q 00 · Pneumonia J 10-20 · Septicaemia P 36 · '
            'Birth Trauma P 10-15.',
            'ICD 10 ছক অনুযায়ী কারণ নিরুপণ/চিহ্নিত করুন'),
        _sr('select_one icd_cause', 'icd_cause',
            'CAUSE', 'রোগের নাম'),
        _sr('text', 'icd_code', 'ICD code', 'ICD code',
            hint='e.g. P 21'),
        _sr('text', 'icd_disease_name', 'Name of disease', 'রোগের নাম'),
        _sr('text', 'icd_diagnoser_name', 'Name of identifier',
            'নিরুপণকারীর নাম'),
        _sr('text', 'icd_diagnoser_designation', 'Designation', 'পদবী'),
        _sr('text', 'icd_diagnoser_institution', 'Institution', 'প্রতিষ্ঠান'),
        _sr('date', 'icd_date', 'Date', 'তারিখ'),
        _sr('end_group', 'grp_icd'),
    ]
    return rows


def _community_neonatal_choices():
    ch = list(DISTRICT_CHOICES) + list(YES_NO)

    # Relationship list (per MPDSR Form 02 respondent table, page 2:
    # "সম্পর্ক: (উদাহরণ স্বরূপ)"). Values/order unchanged; labels verbatim.
    rel = [
        ('mother', 'Mother', 'মা'),
        ('grandparent', 'Grandfather / grandmother (paternal / maternal)',
         'দাদা / দাদী / নানা / নানী'),
        ('sibling', 'Sister / brother', 'বোন/ভাই'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন'),
        ('father', 'Father', 'বাবা'),
        ('aunt_uncle',
         'Aunt / uncle (khala / khalu, chachi / chacha, mami / mama, fupu / fupa)',
         'খালা/খালু, চাচী/চাচা, মামী/মামা, ফুপু/ফুপা'),
        ('neighbour', 'Neighbour', 'প্রতিবেশী'),
    ]
    ch += [_ch('relationship', k, en, bn) for k, en, bn in rel]

    # Mother's education (5) — page 3.
    ch += [
        _ch('education_level', 'none', 'None', 'নাই'),
        _ch('education_level', 'class5', 'Up to class 5', '৫ম শ্রেণী পর্যন্ত'),
        _ch('education_level', 'class10', 'Up to class 10', '১০ম শ্রেণী পর্যন্ত'),
        _ch('education_level', 'class12', 'Up to class 12', 'দ্বাদশ শ্রেণী পর্যন্ত'),
        _ch('education_level', 'graduate', 'Graduate / degree', 'স্নাতক/ডিগ্রি'),
    ]
    # Household socio-economic status (4) — page 3.
    ch += [
        _ch('ses', 'extreme_poor', 'Extreme poor', 'অতিগরীব'),
        _ch('ses', 'poor', 'Poor', 'গরীব'),
        _ch('ses', 'middle', 'Middle class', 'মধ্যবিত্ত'),
        _ch('ses', 'rich', 'Rich', 'ধনী'),
    ]

    # Q3 — place of death (10) — page 3. (No "Community Clinic" on paper.)
    ch += [
        _ch('death_place', 'home', 'Home', 'বাড়ীতে'),
        _ch('death_place', 'union_hfwc',
            'Union Health & Family Welfare Centre',
            'ইউনিয়ন স্বাস্থ্য ও পরিবার কল্যাণ কেন্দ্র'),
        _ch('death_place', 'district_hospital',
            'District / Sadar hospital', 'জেলা অথবা সদর হাসপাতালে'),
        _ch('death_place', 'medical_college',
            'Medical College hospital', 'মেডিকেল কলেজ হাসপাতাল'),
        _ch('death_place', 'private_clinic',
            'Private clinic / hospital', 'প্রাইভেট ক্লিনিক/হাসপাতাল'),
        _ch('death_place', 'in_transit', 'On the way', 'পথে'),
        _ch('death_place', 'upazila_hc',
            'Upazila Health Complex', 'উপজেলা স্বাস্থ্য কমপ্লেক্সে'),
        _ch('death_place', 'maternal_centre',
            'Maternal & Child Welfare Centre', 'মাতৃ মঙ্গল কেন্দ্রে'),
        _ch('death_place', 'ngo_clinic', 'NGO clinic', 'এনজিও ক্লিনিক'),
        _ch('death_place', 'other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন'),
    ]

    # Q7 — pregnancy / antenatal complications (multi, 12) — page 4.
    preg = [
        ('excess_bleeding', 'Excessive bleeding', 'অতিরিক্ত রক্তক্ষরণ'),
        ('high_bp', 'High blood pressure', 'উচ্চ রক্তচাপ'),
        ('convulsion', 'Convulsion', 'খিঁচুনি'),
        ('high_fever', 'High fever', 'প্রচন্ড জ্বর'),
        ('prolonged_labour', 'Prolonged labour', 'বিলম্বিত প্রসব'),
        ('abnormal_behaviour', 'Abnormal behaviour', 'অসংলগ্ন আচরণ'),
        ('obstructed_labour', 'Obstructed labour', 'বাধাপ্রাপ্ত প্রসব'),
        ('twin_pregnancy', 'Twin pregnancy', 'যমজ গর্ভধারন'),
        ('injury', 'Injury', 'আঘাত'),
        ('no_illness', 'No, the mother had no disease',
         'না মায়ের কোন রোগ ছিল না'),
        ('dont_know', "I don't know", 'আমি জানি না'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন'),
    ]
    ch += [_ch('pregnancy_complication', k, en, bn) for k, en, bn in preg]

    # Q8 — ANC count (6) — page 4.
    ch += [
        _ch('anc_count', '1', '1 time', '১ বার'),
        _ch('anc_count', '2', '2 times', '২ বার'),
        _ch('anc_count', '3', '3 times', '৩ বার'),
        _ch('anc_count', '4_plus', '4 times and more', '৪ বার এবং তার বেশী'),
        _ch('anc_count', 'none', 'Received no ANC care', 'কোন এএনসি সেবা পায়নি'),
        _ch('anc_count', 'unknown', "I don't know", 'আমি জানি না'),
    ]

    # Shared facility list — Q9 / Q13 / Q24 (11) — pages 4-7.
    fac = [
        ('home', 'Home', 'বাড়ী'),
        ('community_clinic', 'Community Clinic', 'কমিউনিটি ক্লিনিক'),
        ('union_hfwc', 'Union Health & Family Welfare Centre',
         'ইউনিয়ন স্বাস্থ্য ও পরিবার কল্যাণ কেন্দ্র'),
        ('upazila_hc', 'Upazila Health Complex', 'উপজেলা স্বাস্থ্য কমপ্লেক্স'),
        ('maternal_centre', 'Maternal & Child Welfare Centre', 'মাতৃ মঙ্গল কেন্দ্র'),
        ('district_hospital', 'District / Sadar hospital', 'জেলা অথবা সদর হাসপাতাল'),
        ('medical_college', 'Medical College hospital', 'মেডিকেল কলেজ হাসপাতাল'),
        ('private_clinic', 'Private clinic / hospital', 'প্রাইভেট ক্লিনিক/হাসপাতাল'),
        ('ngo_clinic', 'NGO clinic', 'এনজিও ক্লিনিক'),
        ('provider_home', "Provider's chamber / home",
         'চেম্বার/স্বাস্থ্য সেবাদানকারীর বাড়ী'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন'),
    ]
    ch += [_ch('facility_place', k, en, bn) for k, en, bn in fac]

    # Shared provider-cadre list — Q10 / Q14 (12) — pages 4-5.
    prov = [
        ('doctor_mbbs', 'Doctor (MBBS)', 'ডাক্তার (MBBS)'),
        ('nurse', 'Nurse', 'নার্স'),
        ('fwv', 'Family Welfare Visitor (FWV)', 'পরিবার পরিকল্পনা ভিজিটর (FWV)'),
        ('csba', 'CSBA', 'সিএসবিএ (CSBA)'),
        ('ma', 'Medical Assistant (MA)', 'মেডিকেল অ্যাসিস্ট্যান্ট (MA)'),
        ('ha', 'Health Assistant (HA)', 'স্বাস্থ্য সহকারী (HA)'),
        ('fwa', 'Family Welfare Assistant (FWA)', 'পরিবার পরিকল্পনা সহকারী(FWA)'),
        ('dai', 'Dai', 'দাই'),
        ('palli_chikitsok', 'Palli chikitsok', 'পল্লী চিকিৎসক'),
        ('ngo_worker', 'NGO worker', 'এনজিও কর্মী'),
        ('midwife', 'Mid wife', 'মিড ওয়াইফ (Mid wife)'),
        ('other', 'Other (specify)', 'অন্যান্য উল্লেখ করুন'),
    ]
    ch += [_ch('provider_cadre', k, en, bn) for k, en, bn in prov]

    # Q15 — mode of delivery (3) — page 5. Paper prints Bangla + English
    # together per option; English copied verbatim from the paper.
    ch += [
        _ch('delivery_mode', 'vaginal_spontaneous', 'Vaginal-spontaneous',
            'স্পন্টিনিওয়াস'),
        _ch('delivery_mode', 'csection', 'Caesarean Section', 'সিজারিয়ান'),
        _ch('delivery_mode', 'instrumental_vaginal',
            'Instrumental vaginal (vacuum/forceps)',
            'সহায়ক প্রসব  (ভ্যাকুয়াম/ফোরসেপ)'),
    ]

    # Q16 — delivery / intrapartum complications (multi, 16) — page 5.
    # Each option on the paper carries Bangla + bracketed English; the
    # English here is copied VERBATIM from the paper (typos kept).
    delcomp = [
        ('no_complication', 'No complication', 'মায়ের কোন জটিলতা ছিল না'),
        ('dont_know', "Don't know", 'এ বিষয়ে আমি কিছু জানি না'),
        ('high_bp', 'High Blood Pressure/Hypertension', 'উচ্চ রক্তচাপ'),
        ('convulsion', 'Convulsion/eclampsia', 'খিচুনি/একলামসিয়া'),
        ('aph', 'Bleeding during pregnancy/ Ante Partum Hemorrhage',
         'প্রসবেরপূর্বে /প্রসব কালীন সময় রক্তক্ষরণ'),
        ('obstructed_labour', 'Obstructed labour', 'বাধাগ্রস্থ প্রসব'),
        ('prolonged_labour', 'Labour pain more than 12 hrs/Prolonged labour ',
         '১২ ঘন্টার বেশী প্রসববেদনা/বিলম্বিত প্রসব'),
        ('premature_labour', 'Premature labour/labour pain before 37 weeks ',
         '৩৭ সপ্তাহের পূর্বে প্রসব বেদনা'),
        ('prom', 'Leakage of fluid before labour/ Ruptured membrane/PROM',
         'প্রসবের পূর্বে পানি ভাঙ্গা'),
        ('fetal_distress',
         'Less fetal movement/ fetal distress/ meconium stained liquor',
         'ভ্রূণের নড়াচড়া কম হওয়া অথবা দীর্ঘ সময় নড়াচড়া না হওয়া'),
        ('malpresentation',
         'Abnormal position of the fetus/ mal-presentation ',
         'ভ্রূণের অস্বাভাবিক অবস্থান'),
        ('ruptured_uterus', 'Ruptured uterus', 'জরায়ু ছিড়ে যাওয়া'),
        ('retained_placenta',
         'Placenta not delivered 30 minutes after delivery of the '
         'baby/Retained placenta',
         'প্রসবের ৩০ মিনিটের মধ্যে গর্ভ ফুল না পড়া'),
        ('pph', 'Bleeding after/ Post-Partum Hemorrhage',
         'প্রসব পরবর্তী রক্তক্ষরণ'),
        ('high_fever', 'High fever', 'প্রচন্ড জ্বর'),
        ('foul_discharge', 'Fowl smelling vaginal discharge', 'দুর্গন্ধযুক্ত স্রাব'),
    ]
    ch += [_ch('delivery_complication', k, en, bn) for k, en, bn in delcomp]

    # Q18 — birth-weight band (3) — page 6.
    ch += [
        _ch('birth_weight_band', 'low',
            'Lower than normal (< 2.5 kg)', 'স্বাভাবিকের কম (<২.৫ কেজি)'),
        _ch('birth_weight_band', 'normal',
            'Normal (2.5 kg-4.5 kg)', 'স্বাভাবিক (২.৫ কেজি-৪.৫ কেজি)'),
        _ch('birth_weight_band', 'high',
            'Higher than normal (> 4.5 kg)', 'স্বাভাবিকের বেশী (> ৪.৫ কেজি)'),
    ]

    # Q19 — congenital defect site (multi, 11) — page 6.
    defects = [
        ('head', 'Head', 'মাথা'),
        ('cleft_palate', 'Cleft palate', 'তালু কাটা'),
        ('lip', 'Lip', 'ঠোঁট'),
        ('elbow', 'Elbow', 'কনুই'),
        ('shoulder', 'Shoulder', 'কাঁধ'),
        ('anus', 'Anus', 'পায়ু পথ'),
        ('leg', 'Leg', 'পা'),
        ('genitalia', 'Genitalia', 'যোনি পথ'),
        ('dont_know', "I don't know about this", 'আমি এ বিষয়টি জানি না'),
        ('none_normal', 'No, the neonate was normal', 'না, নবজাতক স্বাভাবিক ছিল'),
        ('other', 'Other (specify)', 'অন্যান্য'),
    ]
    ch += [_ch('congenital_defect', k, en, bn) for k, en, bn in defects]

    # Q20 — cried / breathed after birth (4) — page 6.
    ch += [
        _ch('cried_breathed', 'normal_breath',
            'Yes / breathed normally', 'হ্যাঁ / স্বাভাবিক ভাবে শ্বাস  নিয়েছিল'),
        _ch('cried_breathed', 'weak_cry', 'Weak cry', 'দুর্বল কান্না'),
        _ch('cried_breathed', 'delayed_breath',
            'Breathed after a long time', 'অনেক সময় পরে শ্বাস নিয়েছিল'),
        _ch('cried_breathed', 'dont_know', "Don't know", 'জানি না'),
    ]

    # Q21 — resuscitation actions taken (multi, 9) — page 6. Each option
    # on the paper carries Bangla + bracketed English; English copied
    # VERBATIM from the paper (typos like "atthevertebral", "ambo" kept).
    resus = [
        ('dried_wrapped',
         'Dried the neonate and wrapped with dry clean cloth',
         'নবজাতকে শুকনো এবং শুকনো পরিক্ষার কাপড় দিয়ে আবৃত করা হয়েছে'),
        ('spinal_stimulation',
         'Stimulation was given atthevertebral column or back',
         'নবজাতকের পিঠে বা শিরদাড় চাপড় করা হয়েছে'),
        ('mouth_to_mouth', 'Mouth to mouth breathing was given',
         'মুখের মাধ্যমে শ্বাস দেয়া'),
        ('ambu_bag', 'Neonatal resuscitation by ambo bag',
         'আম্বো ব্যাগের মাধ্যমে শ্বাস দেয়া হয়েছিলো কিনা'),
        ('slap_back',
         'Slap forcefully on the back/chest/ hip holding both leg up '
         'and head below',
         'পিঠে অথবা বুকে হাতের সাহায্যে চাপড়দেয়া হয়েছে'),
        ('warm_water', 'Flash warm water in mouth',
         'মুখে গরম পানি ছিটানো হয়েছে'),
        ('transferred', 'Transfer in health care centre',
         'স্বাস্থ্য কেন্দ্রে পাঠানো হয়েছিলো'),
        ('nothing_done', 'Nothing was done significantly',
         'কিছুই করা হয় নাই'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন'),
    ]
    ch += [_ch('resuscitation_action', k, en, bn) for k, en, bn in resus]

    # Q22 — neonatal danger signs (multi, 13: 11 clinical + other + don't
    # know) — pages 6-7. Each option carries Bangla + bracketed English;
    # English copied VERBATIM (typos like "colouation" kept).
    danger = [
        ('convulsions', 'Convulsions', 'খিঁচুনি'),
        ('no_feed', 'No feed or reluctant to feed',
         'নবজাতক বুকের দুধ খেতে অক্ষম বা দুর্বল'),
        ('fast_breathing', 'Fast breathing', 'দ্রুত শ্বাসপ্রশ্বাস'),
        ('chest_indrawing', 'Chest indrawing with Fast breathing',
         'দ্রুত শ্বাসপ্রশ্বাসের সাথে বুক দেবে যাওয়া'),
        ('hypothermia', 'Hypothermia of hand and feet',
         'হাত ও পা ঠান্ডা হয়ে যাওয়া'),
        ('no_movement', 'Lack of movement or less movement',
         'নড়াচড়া নেই'),
        ('fever', 'Fever', 'জ্বর'),
        ('jaundice', 'Yellow colouation of eyes/ Skin',
         'চোখ ও গায়ের চামড়া হলুদ বর্ণের হওয়া'),
        ('umbilicus_infection',
         'Reddish umbilicus/ Pus secretion from umbilicus',
         'নাভি লাল বর্ণ ধারন বা নাভি থেকে পুজ বের হওয়া'),
        ('diarrhoea', 'Diarrhoea', 'পাতলা পায়খানা (ডায়রিয়া)'),
        ('skin_pustules', 'Pus contained vesicle in the skin',
         'চামরায় পুজ সহ ফোড়া'),
        ('other', 'Other (specify)', 'অন্যান্য ('),
        ('dont_know', "I don't know", 'আমার জানা নেই'),
    ]
    ch += [_ch('danger_sign', k, en, bn) for k, en, bn in danger]

    # Q25 — reasons for not receiving treatment (multi, 16) — page 7.
    reasons = [
        ('unnecessary', 'Thought unnecessary', 'অপ্রয়োজনীয় মনে করা'),
        ('didnt_understand', 'Could not understand the need for treatment',
         'চিকিৎসার প্রয়োজন বুঝতে না পারা'),
        ('costly', 'Costly', 'ব্যয়বহুল'),
        ('no_money', 'Lack of money', 'অর্থের অভাব'),
        ('too_far', 'Distance', 'দূরত্ব'),
        ('no_transport', 'Lack of transport', 'যানবাহনের অভাব'),
        ('no_escort', 'Did not get a companion to take', 'নিয়ে যাওয়ার সঙ্গী না পাওয়া'),
        ('poor_service', 'Service quality poor', 'সেবার মান অনুন্নত'),
        ('family_refusal', 'Family disagreement', 'পারিবারিক অসম্মতি'),
        ('good_care_home', 'Got good care at home',
         'বাড়ীতে ভাল সেবা পাওয়ায়'),
        ('no_way_known', 'Did not know the way to go', 'যাওয়ার উপায় না জানা'),
        ('no_time', 'Lack of time', 'সময়ের অভাব'),
        ('dont_know_where', 'Did not know where to go',
         'কোথায় যেতে হবে না জানা'),
        ('fear_care', 'Fear of receiving care', 'সেবাগ্রহণে ভয় পাওয়া'),
        ('disaster_night', 'Natural disaster/bad weather/night',
         'প্রাকৃতিক দুর্যোগ/খারাপ আবহাওয়া/রাত্রি'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন'),
    ]
    ch += [_ch('no_treatment_reason', k, en, bn) for k, en, bn in reasons]

    # Consultant ICD 10 causes (7 from the page-8 table + Other) — page 8.
    # The paper's ICD 10 table is ENGLISH-ONLY (CAUSE | ICD CODE), so the
    # English is transcribed VERBATIM including the space inside each code
    # ("P 21", "P 07", "J 10-20"...). Bangla mirrors the English (gold-
    # standard Form 01 discipline), since the table prints no Bangla.
    icd = [
        ('birth_asphyxia_p21', 'Birth Asphyxia (P 21)', 'Birth Asphyxia (P 21)'),
        ('low_birth_weight_p07', 'Low Birth Weight (P 07)',
         'Low Birth Weight (P 07)'),
        ('meningitis_g00', 'Meningitis (G 00)', 'Meningitis (G 00)'),
        ('congenital_anomalies_q00', 'Congenital Anomalies (Q 00)',
         'Congenital Anomalies (Q 00)'),
        ('pneumonia_j10_20', 'Pneumonia (J 10-20)', 'Pneumonia (J 10-20)'),
        ('septicaemia_p36', 'Septicaemia (P 36)', 'Septicaemia (P 36)'),
        ('birth_trauma_p10_15', 'Birth Trauma (P 10-15)',
         'Birth Trauma (P 10-15)'),
        ('other', 'Other (specify code)', 'অন্যান্য'),
    ]
    ch += [_ch('icd_cause', k, en, bn) for k, en, bn in icd]

    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 4 — MPDSR Form 04 · Facility Maternal Death Review               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _facility_maternal_survey():
    rows = _meta()
    rows += _office_use_block('Annual facility maternal death serial number',
                              'মাতৃমৃত্যুর বাৎসরিক ক্রমিক নং')

    # General information & facility / mother identity (paper page 2)
    rows += [
        _sr('begin_group', 'grp_identity', 'General information', 'সাধারণ তথ্য'),
        _sr('text', 'facility_name', 'Name of facility',
            'স্বাস্থ্য কেন্দ্রের নামঃ', required='yes'),
        _sr('text', 'facility_code',
            'Facility code (must be filled — 8 digits)',
            'স্বাস্থ্য কেন্দ্রের কোড নংঃ',
            required='yes',
            hint='8-digit code printed in the boxes on the paper form.',
            constraint='regex(normalize-space(.), "^[0-9]{8}$")',
            cmsg='Enter exactly 8 digits. / ঠিক ৮ অঙ্ক লিখুন।'),
        _sr('text', 'deceased_name', "Mother's name", 'মায়ের নামঃ',
            required='yes'),
        _sr('integer', 'deceased_age', "Mother's age (years)",
            'মায়ের বয়সঃ (বৎসর)', required='yes',
            constraint='. > 9 and . < 60', cmsg='10–59'),
        _sr('text', 'mother_hosp_reg_no',
            "Mother's Hospital Registration No. (must be filled)",
            'মায়ের হাসপাতাল রেজিস্ট্রেশন নং',
            required='yes',
            hint='Hospital registration number from the boxes on the paper form.'),
        _sr('note', 'mother_addr_note',
            "Mother's address", 'মায়ের ঠিকানাঃ'),
        _sr('select_one district', 'mother_district',
            'District', 'জেলাঃ', required='yes'),
        _sr('text', 'mother_upazila', 'Upazila', 'উপজেলাঃ'),
        _sr('text', 'mother_union', 'Union / Pourashava',
            'ইউনিয়ন/পৌরসভাঃ'),
        _sr('text', 'mother_ward', 'Ward', 'ওয়ার্ডঃ'),
        _sr('text', 'mother_village', 'Village', 'গ্রামঃ'),
        _sr('text', 'deceased_husband', "Husband's name", 'স্বামীর নামঃ'),
        _sr('text', 'family_phone', 'Phone number', 'ফোন নংঃ',
            constraint='regex(., "^[0-9+ -]{6,20}$") or .=""',
            cmsg='Enter a valid phone number.'),
        _sr('end_group', 'grp_identity'),
    ]

    # Clinical timeline — ER/OPD arrival, admission, death (page 2)
    rows += [
        _sr('begin_group', 'grp_timeline', 'Clinical timeline', 'সময়রেখা'),
        _sr('date', 'er_arrival_date',
            'Date of arrival in Emergency department / OPD',
            'জরুরি বিভাগ/ বহিঃবিভাগে রোগী আসার তারিখ'),
        _sr('time', 'er_arrival_time',
            'Time of arrival in Emergency department / OPD',
            'জরুরি বিভাগ/ বহিঃবিভাগে রোগী আসার সময় (২৪ঘন্টা)'),
        _sr('date', 'admission_date', 'Date of admission in inpatient',
            'অন্তঃবিভাগে ভর্তির তারিখ'),
        _sr('time', 'admission_time', 'Time of admission in inpatient',
            'অন্তঃবিভাগে ভর্তির সময় (২৪ঘন্টা)'),
        _sr('date', 'date_of_death', 'Date of death', 'মৃত্যুর তারিখ',
            required='yes'),
        _sr('time', 'time_of_death', 'Time of death',
            'মৃত্যুর সময় (২৪ঘন্টা)'),
        _sr('end_group', 'grp_timeline'),
    ]

    # Q2 condition at admission, Q3 pregnancy status in ER (page 2)
    rows += [
        _sr('begin_group', 'grp_admission', 'Admission status', 'ভর্তির অবস্থা'),
        _sr('select_one admission_condition', 'admission_condition',
            "2. Mother's condition at admission",
            '২। ভর্তির সময় মায়ের অবস্থা কেমন ছিলঃ'),
        _sr('select_one er_pregnancy_status', 'er_pregnancy_status',
            "3. Mother's pregnancy status in OPD / ER",
            '৩। বহিঃবিভাগ বা জরুরি বিভাগে মায়ের গর্ভাবস্থা কেমন ছিলঃ'),
        _sr('end_group', 'grp_admission'),
    ]

    # Q4 diagnosis at admission (multi-tick long menu, page 3)
    rows += [
        _sr('begin_group', 'grp_q4', '4. Diagnosis at admission',
            '৪। ভর্তির সময় নির্ণয়কৃত রোগ'),
        _sr('select_multiple admission_diagnosis', 'admission_diagnosis',
            '4. Diagnosis at admission',
            '৪। ভর্তির সময় নির্ণয়কৃত রোগ'),
        _sr('select_one retained_placenta_haem', 'admission_retained_placenta_haem',
            "4a. Retained placenta: with or without haemorrhage",
            'Retained placenta: a.with haemorrhage / b. without haemorrhage',
            relevant="selected(${admission_diagnosis}, 'retained_placenta')"),
        _sr('text', 'admission_diagnosis_other', 'Others (Specify):',
            'Others (Specify):',
            relevant="selected(${admission_diagnosis}, 'other')"),
        _sr('end_group', 'grp_q4'),
    ]

    # Q5 referred in? + source (page 3)
    rows += [
        _sr('begin_group', 'grp_q5', '5. Referred in', '৫। রেফার হয়ে এসেছিল'),
        _sr('select_one yes_no', 'referred_in',
            '5. Was the admitted mother referred in?',
            '৫। ভর্তিকৃত মা কি রেফার হয়ে এসেছিল?'),
        _sr('select_one referral_source', 'referral_source',
            '5a. If yes, from where was she referred?',
            'যদি হ্যাঁ হয় তাহলে কোথা হতে এসেছিল-',
            relevant="${referred_in}='yes'"),
        _sr('text', 'referral_source_other', 'অন্যান্য',
            'অন্যান্য',
            relevant="${referred_in}='yes' and ${referral_source}='other'"),
        _sr('end_group', 'grp_q5'),
    ]

    # Q6 first doctor/consultant observation date+time (page 3)
    rows += [
        _sr('begin_group', 'grp_q6',
            '6. First observation by doctor / consultant',
            '৬। ডাক্তার / কনসালট্যান্ট প্রথম পর্যবেক্ষণ'),
        _sr('select_one yes_no', 'first_obs_recorded',
            '6. Is the date/time of first observation recorded?',
            '৬। কখন ডাক্তার বা কনসালট্যান্ট প্রথম মাকে পর্যবেক্ষণ করেছেনঃ',
            hint="Choose 'No' if the form notes তথ্য সংগ্রহ করা নেই।"),
        _sr('date', 'first_obs_date', '6a. Date of first observation',
            'তারিখঃ',
            relevant="${first_obs_recorded}='yes'"),
        _sr('time', 'first_obs_time', '6b. Time of first observation',
            'সময় (২৪ঘন্টা)ঃ',
            relevant="${first_obs_recorded}='yes'"),
        _sr('end_group', 'grp_q6'),
    ]

    # Q7 inpatient diagnosis (multi-tick long menu, page 3)
    rows += [
        _sr('begin_group', 'grp_q7', '7. Inpatient diagnosis',
            '৭। অন্তঃবিভাগে রোগ/সমস্যা নির্ণয়;'),
        _sr('select_multiple inpatient_diagnosis', 'inpatient_diagnosis',
            '7. Disease / problem diagnosed in the inpatient ward',
            '৭। অন্তঃবিভাগে রোগ/সমস্যা নির্ণয়;'),
        _sr('select_one retained_placenta_haem', 'inpatient_retained_placenta_haem',
            "7a. Retained placenta: with or without haemorrhage",
            'Retained placenta: With haemorrhage / Without haemorrhage',
            relevant="selected(${inpatient_diagnosis}, 'retained_placenta')"),
        _sr('text', 'inpatient_diagnosis_other', 'Others (Specify):',
            'Others (Specify):',
            relevant="selected(${inpatient_diagnosis}, 'other')"),
        _sr('end_group', 'grp_q7'),
    ]

    # Q8 time management started after admission (page 3)
    rows += [
        _sr('begin_group', 'grp_q8', '8. Management started',
            '৮। ব্যবস্থাপনা শুরু'),
        _sr('select_one yes_no', 'management_recorded',
            '8. Is the date/time management started recorded?',
            '৮। অন্তঃবিভাগে ভর্তির পর কখন ব্যবস্থাপনা শুরু হয়েছেঃ',
            hint="Choose 'No' if the form notes তথ্য সংগ্রহ করা নেই।"),
        _sr('date', 'management_start_date', '8a. Date management started',
            'তারিখঃ',
            relevant="${management_recorded}='yes'"),
        _sr('time', 'management_start_time',
            '8b. Time management started',
            'সময় (২৪ঘন্টা)ঃ',
            relevant="${management_recorded}='yes'"),
        _sr('end_group', 'grp_q8'),
    ]

    # Q9 mode of delivery, Q10 outcome, Q11 birth weight, Q12 anomaly (page 4)
    rows += [
        _sr('begin_group', 'grp_delivery', 'Delivery & newborn',
            'প্রসব ও নবজাতক'),
        _sr('select_one delivery_mode', 'delivery_mode',
            '9. Mode of delivery', '৯। প্রসবের পদ্ধতিঃ'),
        _sr('select_one delivery_outcome', 'delivery_outcome',
            '10. Outcome of the current pregnancy',
            '১০। বর্তমান গর্ভবতীর ফলাফলঃ'),
        _sr('text', 'delivery_outcome_other', 'অন্যান্য উল্লেখ করুন',
            'অন্যান্য উল্লেখ করুন', relevant="${delivery_outcome}='other'"),
        _sr('integer', 'birth_weight_grams', "11. Baby's birth weight (grams)",
            '১১। জন্মের পর শিশুর ওজন (গ্রাম)',
            constraint='. >= 0 and . <= 8000',
            relevant="${delivery_outcome}!='abortion'"),
        _sr('select_one yes_no', 'baby_abnormality',
            '12. Was there any abnormality in the baby after birth?',
            '১২। জন্মের পর শিশুর কোন অস্বাভাবিকতা ছিল কিনাঃ',
            relevant="${delivery_outcome}!='abortion'"),
        _sr('end_group', 'grp_delivery'),
    ]

    # Q13 place of death within the facility (page 4)
    rows += [
        _sr('begin_group', 'grp_q13', '13. Place of death in facility',
            '১৩। মায়ের মৃত্যুর স্থান'),
        _sr('select_one death_place_facility', 'death_place_facility',
            "13. Place of the mother's death",
            '১৩। মায়ের মৃত্যুর স্থানঃ'),
        _sr('text', 'death_place_facility_other', 'অন্যান্য',
            'অন্যান্য',
            relevant="${death_place_facility}='other'"),
        _sr('end_group', 'grp_q13'),
    ]

    # Q14 most probable cause of death — multi-tick x ICD-10 (page 4)
    rows += [
        _sr('begin_group', 'grp_q14',
            '14. Most probable cause(s) of death (ICD-10)',
            '১৪। সবচেয়ে সম্ভাব্য মৃত্যুর কারণ'),
        _sr('note', 'q14_note',
            '14. Most probable cause of death — you may tick more than one '
            '(WHO Cause of death form, with separate ICD10 list).',
            '১৪। সবচেয়ে সম্ভাব্য মৃত্যুর কারণঃ (একাধিক উত্তর ✓ টিক চিহ্ন দিতে পারবেন) '
            '(WHO Cause of death form, with separate ICD10 list)'),
        _sr('select_multiple icd_cause', 'cause_of_death',
            'CAUSE', 'CAUSE',
            required='yes'),
        _sr('text', 'cause_of_death_other', 'Other cause (specify + ICD code)',
            'Other cause (specify + ICD code)',
            relevant="selected(${cause_of_death}, 'other')"),
        _sr('end_group', 'grp_q14'),
    ]

    # Q15 brief death narrative (page 4)
    rows += [
        _sr('begin_group', 'grp_q15', '15. Brief narrative of the death',
            '১৫। মৃত্যুর সংক্ষিপ্ত বিবরণ'),
        _sr('text', 'death_narrative',
            '15. Comments (describe the death in brief)',
            '১৫। মন্তব্যঃ (সংক্ষেপে মৃত্যুর বিবরণ উল্লেখ করুন)', app='multiline'),
        _sr('end_group', 'grp_q15'),
    ]

    # Q16 form-filler identity (page 4; signature N/A digitally)
    rows += [
        _sr('begin_group', 'grp_q16', '16. Form filled by', '১৬। ফর্ম পূরণকারী'),
        _sr('text', 'filler_name', '16. Name of person filling the form',
            '১৬। ফর্ম পূরণকারীর নামঃ', required='yes'),
        _sr('text', 'filler_designation', 'Designation', 'পদবীঃ'),
        _sr('date', 'filler_date', 'Date of data collection',
            'তথ্য সংগ্রহের তারিখঃ'),
        _sr('end_group', 'grp_q16'),
    ]
    return rows


def _facility_maternal_choices():
    ch = list(DISTRICT_CHOICES) + list(YES_NO)

    # Q2 — condition of the mother at admission. Paper Bangla verbatim;
    # 'Stable'/'Unstable' are printed in English alongside the Bangla.
    ch += [
        _ch('admission_condition', 'conscious', 'Conscious', 'সচেতন'),
        _ch('admission_condition', 'stable', 'Stable', 'স্থিতিশীল (Stable)'),
        _ch('admission_condition', 'unstable', 'Unstable',
            'অস্থিতিশীল (Unstable)'),
        _ch('admission_condition', 'unconscious', 'Unconscious', 'অচেতন'),
        _ch('admission_condition', 'not_recorded', 'Not recorded',
            'তথ্য সংগ্রহ করা হয়নি'),
    ]

    # Q3 — pregnancy status in OPD / ER (5). Paper Bangla verbatim.
    ch += [
        _ch('er_pregnancy_status', 'pregnant', 'Pregnant (antenatal)',
            'গর্ভকালীন'),
        _ch('er_pregnancy_status', 'in_labour', 'In labour',
            'প্রসব বেদনা ছিল'),
        _ch('er_pregnancy_status', 'after_delivery', 'Came after delivery',
            'প্রসবের পর এসেছিল'),
        _ch('er_pregnancy_status', 'other_problem', 'Had other problem',
            'অন্যান্য সমস্যা ছিল'),
        _ch('er_pregnancy_status', 'not_recorded', 'Not recorded',
            'তথ্য সংগ্রহ করা হয়নি'),
    ]

    # Q4 — diagnosis at admission. Paper prints these options in ENGLISH only,
    # so both label columns carry the EXACT printed English (gold-standard
    # 'en, en' discipline — no invented Bangla). Order = paper top-to-bottom.
    diag_adm = [
        ('ectopic', 'Ectopic pregnancy'),
        ('spontaneous_abortion', 'Spontaneous abortion'),
        ('induced_abortion', 'Induced abortion'),
        ('placenta_previa', 'Placenta previa'),
        ('molar_pregnancy', 'Molar pregnancy'),
        ('pph', 'Postpartum haemorrhage'),
        ('prolonged_labour', 'Prolonged labour'),
        ('rupture_uterus', 'Rupture uterus'),
        ('missed_abortion', 'Missed abortion'),
        ('medical_abortion', 'Medical abortion'),
        ('threatened_abortion', 'Threatened abortion'),
        ('abruptio_placenta', 'Abruptio placenta'),
        ('intrapartum_haemorrhage', 'Intra partum haemorrhage'),
        ('retained_placenta', 'Retained placenta:'),
        ('obstructed_labour', 'Obstructed labour'),
        ('not_recorded', 'Not recorded'),
        ('other', 'Others (Specify):'),
    ]
    ch += [_ch('admission_diagnosis', k, en, en) for k, en in diag_adm]

    # Q7 — inpatient diagnosis. SAME keys as Q4, but the paper prints SOME
    # words differently (Inducedabortion / Threatenedabortion run together;
    # 'Post partum haemorrhage' with a space; 'Rupture uterus:' with a colon).
    # Transcribed exactly as the paper prints them — typos kept.
    diag_inp = [
        ('ectopic', 'Ectopic pregnancy'),
        ('spontaneous_abortion', 'Spontaneous abortion'),
        ('induced_abortion', 'Inducedabortion'),
        ('placenta_previa', 'Placenta previa'),
        ('molar_pregnancy', 'Molar pregnancy'),
        ('pph', 'Post partum haemorrhage'),
        ('prolonged_labour', 'Prolonged labour'),
        ('rupture_uterus', 'Rupture uterus:'),
        ('missed_abortion', 'Missed abortion'),
        ('medical_abortion', 'Medical abortion'),
        ('threatened_abortion', 'Threatenedabortion'),
        ('abruptio_placenta', 'Abruptio placenta'),
        ('intrapartum_haemorrhage', 'Intra partum haemorrhage'),
        ('retained_placenta', 'Retained placenta:'),
        ('obstructed_labour', 'Obstructed labour'),
        ('not_recorded', 'Not recorded'),
        ('other', 'Others (Specify):'),
    ]
    ch += [_ch('inpatient_diagnosis', k, en, en) for k, en in diag_inp]

    # Q4a — retained-placenta sub-options. Q4 prints them with a./b. prefixes;
    # Q7 prints the same two without prefixes. Both English-only on the paper.
    # This single shared list feeds both 4a and 7a — Q4's printed wording kept.
    ch += [
        _ch('retained_placenta_haem', 'with_haem', 'a.with haemorrhage',
            'a.with haemorrhage'),
        _ch('retained_placenta_haem', 'without_haem', 'b. without haemorrhage',
            'b. without haemorrhage'),
    ]

    # Q5 — referral source. Paper Bangla verbatim (no English on the paper,
    # so the Bangla text fills both columns).
    ch += [
        _ch('referral_source', 'govt_facility', 'সরকারী স্বাস্থ্য কেন্দ্র',
            'সরকারী স্বাস্থ্য কেন্দ্র'),
        _ch('referral_source', 'private_facility',
            'ব্যক্তিগত স্বাস্থ্য কেন্দ্র', 'ব্যক্তিগত স্বাস্থ্য কেন্দ্র'),
        _ch('referral_source', 'home', 'বাড়ী থেকে', 'বাড়ী থেকে'),
        _ch('referral_source', 'other', 'অন্যান্য', 'অন্যান্য'),
        _ch('referral_source', 'not_recorded', 'তথ্য সংগ্রহ করা নেই',
            'তথ্য সংগ্রহ করা নেই'),
    ]

    # Q9 — mode of delivery (4). Paper Bangla verbatim.
    ch += [
        _ch('delivery_mode', 'normal', 'স্বাভাবিক প্রসব', 'স্বাভাবিক প্রসব'),
        _ch('delivery_mode', 'csection', 'সিজারিয়ান সেকশন',
            'সিজারিয়ান সেকশন'),
        _ch('delivery_mode', 'instrumental',
            'যন্ত্রপাতির মাধ্যমে', 'যন্ত্রপাতির মাধ্যমে'),
        _ch('delivery_mode', 'not_delivered', 'প্রসব হয়নি', 'প্রসব হয়নি'),
    ]

    # Q10 — outcome of current pregnancy (5). Paper Bangla verbatim.
    ch += [
        _ch('delivery_outcome', 'livebirth', 'জীবিত জন্ম', 'জীবিত জন্ম'),
        _ch('delivery_outcome', 'stillbirth', 'মৃত জন্ম', 'মৃত জন্ম'),
        _ch('delivery_outcome', 'low_birth_weight', 'কম জন্ম ওজন',
            'কম জন্ম ওজন'),
        _ch('delivery_outcome', 'abortion', 'গর্ভপাত', 'গর্ভপাত'),
        _ch('delivery_outcome', 'other', 'অন্যান্য উল্লেখ করুন',
            'অন্যান্য উল্লেখ করুন'),
    ]

    # Q13 — place of death within the facility (8). Paper Bangla verbatim,
    # in the paper's printed order.
    ch += [
        _ch('death_place_facility', 'er', 'জরুরি বিভাগ', 'জরুরি বিভাগ'),
        _ch('death_place_facility', 'trolley', 'ট্রলিতে', 'ট্রলিতে'),
        _ch('death_place_facility', 'ward', 'ওয়ার্ড', 'ওয়ার্ড'),
        _ch('death_place_facility', 'labour_room', 'লেবার রুম', 'লেবার রুম'),
        _ch('death_place_facility', 'operation_theatre', 'অপারেশন রুম',
            'অপারেশন রুম'),
        _ch('death_place_facility', 'icu', 'আইসিইউ', 'আইসিইউ'),
        _ch('death_place_facility', 'other', 'অন্যান্য', 'অন্যান্য'),
        _ch('death_place_facility', 'after_operation', 'অপারেশনের পর',
            'অপারেশনের পর'),
    ]

    # Q14 — ICD-10 cause-of-death table. English-only on the paper; codes
    # printed as "O 72" (letter O, space, digits). Transcribed VERBATIM in
    # the paper's table order (down each column, left then right). Both label
    # columns carry the exact printed English (gold-standard 'en, en').
    icd = [
        ('pph_o72', 'PPH (O 72)'),
        ('aph_o46', 'APH (O 46)'),
        ('puerperal_sepsis_o85', 'Puerperal Sepsis (O 85)'),
        ('ectopic_o00', 'Ectopic Pregnancy (O 00)'),
        ('eclampsia_o15', 'Eclampsia (O 15)'),
        ('early_haemorrhage_o20', 'Haemorrhage in Early Pregnancy (O 20)'),
        ('sequel_o97', 'Death from sequel of direct obstetric cause (O 97)'),
        ('anaesthesia_ld_o74',
         'Complication of Anaesthesia during Labour & Delivery (O 74)'),
        ('obstructed_labour_o64',
         'Obstructed Labour due to Malposition and Malpresentation of '
         'foetus (O 64)'),
        ('failed_abortion_o07', 'Failed Attempt abortion (O 07)'),
        ('placenta_previa_o44', 'Placenta Previa (O 44)'),
        ('abruptio_o45', 'Abruptio placentae (O 45)'),
        ('medical_abortion_o04', 'Medical abortion (O 04)'),
        ('rupture_uterus_o71', 'Rupture Uterus (O 71)'),
        ('anaesthesia_preg_o29',
         'Complications of anaesthesia during pregnancy (O 29)'),
        ('embolism_o88', 'Obstetric Embolism (O 88)'),
        ('malnutrition_o25', 'Malnutrition in pregnancy (O 25)'),
        ('other', 'Other (specify code)'),
    ]
    ch += [_ch('icd_cause', k, en, en) for k, en in icd]

    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 5 — MPDSR Form 05 · Facility Neonatal Death Review               ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _facility_neonatal_survey():
    rows = _meta()
    rows += _office_use_block('Annual neonatal death serial number',
                              'নবজাতক মৃত্যুর বাৎসরিক ক্রমিক নং')

    # ── Facility identity (paper page 2) — facility name + 9-box mandatory code ─
    rows += [
        _sr('begin_group', 'grp_facility',
            'Facility information', 'প্রতিষ্ঠানের তথ্য'),
        _sr('text', 'facility_name',
            'Name of facility', 'ফ্যাসিলিটির নাম:', required='yes'),
        # Paper prints an 8-box code grid → constraint allows 8 digits or blank.
        _sr('text', 'facility_code',
            'Facility code',
            'ফ্যাসিলিটি কোড',
            hint='The 8-digit facility code printed on the paper form.',
            constraint='regex(normalize-space(.), "^[0-9]{8}$") or .=""',
            cmsg='Enter the 8-digit facility code (digits only). / '
                 '৮ অঙ্কের ফ্যাসিলিটি কোড লিখুন।'),
        _sr('end_group', 'grp_facility'),
    ]

    # ── Mother & neonate identity (paper page 2) ──────────────────────────────
    rows += [
        _sr('begin_group', 'grp_identity',
            'Mother & neonate identity', 'মা ও নবজাতকের পরিচিতি'),
        _sr('text', 'mother_name', "Mother's name", 'মায়ের নাম:',
            required='yes'),
        _sr('integer', 'mother_age', "Mother's age (years)",
            'মায়ের বয়স: (বৎসর)', constraint='. > 9 and . < 60', cmsg='10–59'),
        _sr('text', 'neonate_name', "Neonate's name", 'নবজাতকের নাম:'),
        _sr('note', 'age_death_note',
            "Age of neonate at time of death (day/hour/minute)",
            'মৃত্যুর সময় নবজাতকের বয়স(দিন/ঘন্টা/মিনিট)'),
        _sr('integer', 'age_death_days', 'Day',
            'দিন', constraint='. >= 0 and . <= 28'),
        _sr('integer', 'age_death_hours', 'Hour',
            'ঘন্টা', constraint='. >= 0 and . <= 23'),
        _sr('integer', 'age_death_minutes', 'Minute',
            'মিনিট', constraint='. >= 0 and . <= 59'),
        # Paper prints a ~16-box hospital registration grid → permissive
        # numeric constraint (8–18 digits) so hand-entry box counts still pass.
        _sr('text', 'hospital_reg_no',
            'Mother / neonate hospital registration number',
            'মা/ নবজাতকের হাসপাতালের রেজিস্ট্রেশনের নম্বর',
            required='yes',
            hint='The ~16-digit hospital registration number from the paper form.',
            constraint='regex(normalize-space(.), "^[0-9]{8,18}$")',
            cmsg='Enter the numeric hospital registration number. / '
                 'হাসপাতালের রেজিস্ট্রেশন নম্বর (অঙ্কে) লিখুন।'),
        _sr('text', 'hospital_ward_no',
            'Hospital ward number', 'হাসপাতালের ওয়াড নম্বর'),
        _sr('text', 'hospital_bed_no',
            'Hospital bed number', 'হাসপাতালের বেড নম্বর'),
        _sr('end_group', 'grp_identity'),
    ]

    # ── Mother's address (paper page 2 — separate from enumerator geo) ────────
    rows += [
        _sr('begin_group', 'grp_mother_addr',
            "Mother's address", 'মায়ের ঠিকানা:'),
        _sr('select_one district', 'mother_district',
            'District', 'জেলাঃ'),
        _sr('text', 'mother_upazila', 'Upazila', 'উপজেলাঃ'),
        _sr('text', 'mother_union', 'Union / Pourashava', 'ইউনিয়ন/পৌরসভাঃ'),
        _sr('text', 'mother_ward', 'Ward', 'ওয়াডঃ'),
        _sr('text', 'mother_village', 'Village', 'গ্রামঃ'),
        _sr('text', 'father_name', "Father's name", 'পিতার নাম:'),
        _sr('text', 'family_phone', 'Phone number', 'ফোন নংঃ',
            constraint='regex(., "^[0-9+ -]{6,20}$") or .=""',
            cmsg='Enter a valid phone number.'),
        _sr('end_group', 'grp_mother_addr'),
    ]

    # ── Q1 · Place of birth (11 options) ─────────────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q1',
            'Q1. Place of birth', '১. নবজাতকের জন্মের স্থান'),
        _sr('select_one birth_place', 'birth_place',
            '1. Place of birth of the neonate?',
            '১। নবজাতকের জন্মের স্থান?', required='yes'),
        _sr('text', 'birth_place_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)', relevant="${birth_place}='other'"),
        _sr('end_group', 'grp_q1'),
    ]

    # ── Q2 · Type of neonatal death (inborn → skip to Q5) ────────────────────
    rows += [
        _sr('begin_group', 'grp_q2',
            'Q2. Type of neonatal death', '২. নবজাতকের মৃত্যুর ধরন'),
        _sr('select_one death_type', 'death_type',
            '2. What type of neonatal death was it? '
            '(If applicable, tick.)',
            '২। নবজাতকের কোন ধরনের মৃত্যু হয়েছিল? '
            '(প্রযোজ্য হলে টিক চিহ্ন দিন):', required='yes'),
        _sr('end_group', 'grp_q2'),
    ]

    # Outborn neonates (admitted from outside) — Q3, Q4 apply only to them.
    outborn = "${death_type}='outborn'"

    # ── Q3 · ER / emergency arrival date & time ──────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q3',
            'Q3. ER / Emergency arrival', '৩। বহি: বিভাগ/ জরুরী বিভাগে আসা',
            relevant=outborn),
        _sr('select_one recorded_status', 'er_arrival_recorded',
            '3. Was the ER / emergency arrival recorded?',
            '৩। বহি: বিভাগ/ জরুরী বিভাগে আসার তথ্য রেকর্ড করা হয়েছিল?'),
        _sr('date', 'er_arrival_date',
            '3. Date brought to ER / emergency department',
            '৩। বহি: বিভাগ/ জরুরী বিভাগে নিয়ে আসার তারিখ:',
            relevant="${er_arrival_recorded}='recorded'"),
        _sr('time', 'er_arrival_time', 'Time',
            'সময়:', relevant="${er_arrival_recorded}='recorded'"),
        _sr('end_group', 'grp_q3'),
    ]

    # ── Q4 · Inpatient admission date & time ─────────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q4',
            'Q4. Inpatient admission', '৪। অন্তর্বিভাগে ভর্তি',
            relevant=outborn),
        _sr('select_one recorded_status', 'admission_recorded',
            '4. Was the inpatient admission recorded?',
            '৪। অন্তর্বিভাগে ভর্তির তথ্য রেকর্ড করা হয়েছিল?'),
        _sr('date', 'admission_date',
            '4. Date of inpatient admission',
            '৪। অন্তর্বিভাগে ভর্তির তারিখ:',
            relevant="${admission_recorded}='recorded'"),
        _sr('time', 'admission_time', 'Time',
            'সময়:', relevant="${admission_recorded}='recorded'"),
        _sr('end_group', 'grp_q4'),
    ]

    # ── Q5 · Date & time of death (applies to all) ───────────────────────────
    rows += [
        _sr('begin_group', 'grp_q5',
            'Q5. Date of death', '৫। মৃত্যুর তারিখ'),
        _sr('select_one recorded_status', 'death_datetime_recorded',
            '5. Was the date / time of death recorded?',
            '৫। মৃত্যুর তারিখ / সময় রেকর্ড করা হয়েছিল?'),
        _sr('date', 'death_date',
            '5. Date of death', '৫। মৃত্যুর তারিখ',
            relevant="${death_datetime_recorded}='recorded'"),
        _sr('time', 'death_time', 'Time',
            'সময়:', relevant="${death_datetime_recorded}='recorded'"),
        _sr('end_group', 'grp_q5'),
    ]

    # ── Q6 · Condition at admission ──────────────────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q6',
            'Q6. Condition at admission', '৬। ভর্তির সময় নবজাতকের অবস্থা'),
        _sr('select_one admission_condition', 'admission_condition',
            '6. Condition of the neonate at admission',
            '৬। ভর্তির সময় নবজাতকের অবস্থা'),
        _sr('end_group', 'grp_q6'),
    ]

    # ── Q7 · Diagnosis at admission / after birth (multi-tick) ───────────────
    rows += [
        _sr('begin_group', 'grp_q7',
            'Q7. Diagnosis at admission / after birth',
            '৭. ভর্তির বা জন্মের পর রোগ নির্ণয়'),
        _sr('select_multiple admission_diagnosis', 'admission_diagnosis',
            '7. What disease was diagnosed at the admission of the neonate or '
            'after birth in the hospital (✓ tick)',
            '৭। নবজাতকের ভর্তির বা হাসপাতালে জন্মের পর কি রোগ নির্ণয় করা '
            'হয়েছিল (✓ টিক চিহ্ন দিন)'),
        _sr('text', 'admission_diagnosis_other', 'Others (specify)',
            'অন্যান্য Others (উল্লেখ করুন)',
            relevant="selected(${admission_diagnosis}, 'other')"),
        _sr('end_group', 'grp_q7'),
    ]

    # ── Q8 / Q9 · Referred patient? + referral source ────────────────────────
    rows += [
        _sr('begin_group', 'grp_q8',
            'Q8–Q9. Referral', '৮–৯. রেফারেল'),
        _sr('select_one yes_no', 'referred_in',
            '8. Was the neonate a referred patient? (✓ tick)',
            '৮। নবজাতক কি রেফারকৃত রোগী ছিল? (✓ টিক চিহ্ন দিন)'),
        _sr('select_one referral_source', 'referral_source',
            '9. If yes, from where was it referred? (✓ tick)',
            '৯। যদি হ্যাঁ হয় তাহলে কোথায় থেকে রেফার হয়েছিল? (✓ টিক চিহ্ন দিন)',
            relevant="${referred_in}='yes'"),
        _sr('text', 'referral_source_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="${referred_in}='yes' and ${referral_source}='other'"),
        _sr('end_group', 'grp_q8'),
    ]

    # ── Q10 · First doctor / consultant observation date & time ──────────────
    rows += [
        _sr('begin_group', 'grp_q10',
            'Q10. First observation by doctor / consultant',
            '১০. ডাক্তার / কনসালটেন্ট কর্তৃক প্রথম পর্যবেক্ষণ'),
        _sr('select_one doctor_observed', 'doctor_observed',
            '10. When did a doctor / consultant first observe the neonate?',
            '১০। কখন ডাক্তার/কনসালটেন্ট প্রথম নবজাতককে পর্যবেক্ষণ করেছেন?'),
        _sr('date', 'doctor_observe_date',
            '10. Date', 'তারিখ:',
            relevant="${doctor_observed}='observed'"),
        _sr('time', 'doctor_observe_time', 'Time (24h)',
            'সময় (২৪ঘন্টা):',
            relevant="${doctor_observed}='observed'"),
        _sr('end_group', 'grp_q10'),
    ]

    # ── Q11 · Neonatal danger signs (13-option multi) ────────────────────────
    rows += [
        _sr('begin_group', 'grp_q11',
            'Q11. Danger signs', '১১. বিপদ চিহ্ন'),
        _sr('select_multiple danger_signs', 'danger_signs',
            '11. Did the neonate have any danger sign? '
            '(tick the correct box)',
            '১১। নবজাতকের কি কোন বিপদ চিহ্ন ছিল? (সঠিক বক্সে ঠিক দিন)'),
        _sr('text', 'danger_signs_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="selected(${danger_signs}, 'other')"),
        _sr('end_group', 'grp_q11'),
    ]

    # ── Q12 · Specialist inpatient diagnosis (with ICD codes, multi) ─────────
    rows += [
        _sr('begin_group', 'grp_q12',
            'Q12. Specialist inpatient diagnosis',
            '১২. বিশেষজ্ঞ কর্তৃক রোগ নির্ণয়'),
        _sr('select_multiple specialist_diagnosis', 'specialist_diagnosis',
            '12. What disease was diagnosed for the neonate by the specialist '
            'in the inpatient department of the hospital?',
            '১২। হাসপাতালে অন্তর্বিভাগে বিশেষজ্ঞ কর্তৃক নবজাতকের কি রোগ '
            'নির্ণয় করা হয়েছিল?'),
        _sr('text', 'specialist_diagnosis_other', 'Others (specify)',
            'অন্যান্য Others (উল্লেখ করুন)',
            relevant="selected(${specialist_diagnosis}, 'other')"),
        _sr('end_group', 'grp_q12'),
    ]

    # ── Q13 · Time to treatment after admission ──────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q13',
            'Q13. Time to treatment after admission',
            '১৩। ভর্তির কতক্ষন পর চিকিৎসা শুরু'),
        _sr('date', 'treatment_start_date',
            '13. How long after admission did hospital inpatient-department '
            'treatment start — Date',
            '১৩। ভর্তির কতক্ষন পর হাসপাতাল অন্তর্বিভাগে চিকিৎসা শুরু হয় তারিখ:'),
        _sr('time', 'treatment_start_time', 'Time (24h)',
            'সময় (২৪ঘন্টা):'),
        _sr('end_group', 'grp_q13'),
    ]

    # ── Q14 · Birth weight ───────────────────────────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q14',
            'Q14. Birth weight', '১৪। জন্মের পর শিশুর ওজন'),
        _sr('integer', 'birth_weight_grams',
            '14. Weight of the child after birth (grams)',
            '১৪। জন্মের পর শিশুর ওজন (গ্রাম)',
            constraint='. >= 200 and . <= 7000', cmsg='200–7000 grams'),
        _sr('end_group', 'grp_q14'),
    ]

    # ── Q15 / Q16 · Congenital anomaly? + anomaly site (9-option multi) ──────
    rows += [
        _sr('begin_group', 'grp_q15',
            'Q15–Q16. Congenital anomaly', '১৫–১৬। জন্মগত ত্রুটি'),
        _sr('select_one yes_no', 'congenital_anomaly',
            '15. Did the neonate have any congenital anomaly?',
            '১৫। নবজাতকের কি কোন জন্মগত ত্রুটি ছিল?'),
        _sr('select_multiple anomaly_site', 'anomaly_site',
            '16. If yes, where was the congenital anomaly?',
            '১৬। যদি হ্যাঁ হয় তাহলে কোথায় জন্মগত ত্রুটি ছিল',
            relevant="${congenital_anomaly}='yes'"),
        _sr('text', 'anomaly_site_other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)',
            relevant="${congenital_anomaly}='yes' and "
                     "selected(${anomaly_site}, 'other')"),
        _sr('end_group', 'grp_q15'),
    ]

    # ── Q17 · Place of death within the facility ─────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q17',
            'Q17. Place of death in facility',
            '১৭। হাসপাতালের কোন বিভাগে মৃত্যু'),
        _sr('select_one place_of_death_facility', 'place_of_death_facility',
            '17. In which department or place of the hospital did the neonate '
            'die?',
            '১৭। হাসপাতালের কোন বিভাগে বা স্থানে নবজাতকের মৃত্যু হয়েছে?'),
        _sr('text', 'place_of_death_other', 'Other place (specify)',
            'অন্য কোন স্থানে উল্লেখ করুন',
            relevant="${place_of_death_facility}='other'"),
        _sr('end_group', 'grp_q17'),
    ]

    # ── Q18 · Most probable cause(s) of death — WHO cause-of-death checklist ──
    # The paper Q18 is a flat WHO "tick all that apply" list of 6 ICD-10-coded
    # causes (একাধিক উত্তর দিতে পারবেন). It is NOT a multi-column matrix, so it
    # is faithfully a single select_multiple — no field-list / label / shared
    # one-column "select" pseudo-matrix (that idiom is only for true N×M grids
    # like Form 01 Q9).
    rows += [
        _sr('begin_group', 'grp_q18',
            'Q18. Most probable cause of death (WHO Cause of death form, '
            'with separate ICD list)',
            '১৮। সবচেয়ে সম্ভাব্য মৃত্যুর কারণ '
            '(WHO Cause of death form, with separate ICD list)'),
        _sr('select_multiple cod_cause', 'cod_cause',
            '18. Most probable cause of death: '
            '(you may tick more than one answer ✓) '
            '(WHO Cause of death form, with separate ICD list)',
            '১৮। সবচেয়ে সম্ভাব্য মৃত্যুর কারণ: '
            '(একাধিক উত্তর ✓ টিক চিহ্ন দিতে পারবেন) '
            '(WHO Cause of death form, with separate ICD list)'),
        _sr('end_group', 'grp_q18'),
    ]

    # ── Q19 · Probable cause name + ICD-10 code (free text) ──────────────────
    rows += [
        _sr('begin_group', 'grp_q19',
            'Q19. Write the probable cause of death',
            '১৯। মৃত্যুর সম্ভাব্য কারণ লিখুন'),
        _sr('text', 'cause_name', 'Name of disease',
            'রোগের নামঃ'),
        _sr('text', 'icd10_code', 'ICD 10 code no.',
            'আইসিডি ১০ কোড নং', hint='e.g. P-21'),
        _sr('end_group', 'grp_q19'),
    ]

    # ── Q20 · Death narrative ────────────────────────────────────────────────
    rows += [
        _sr('begin_group', 'grp_q20',
            'Q20. Brief narrative of the death',
            '২০। রোগীর মৃত্যুর ঘটনার বিবরণ'),
        _sr('text', 'death_narrative',
            "20. Write briefly the description of the patient's death event.",
            '২০। রোগীর মৃত্যুর ঘটনার বিবরণ লিখুন সংক্ষেপে', app='multiline'),
        _sr('end_group', 'grp_q20'),
    ]

    # ── Reviewer block (paper footer: name / designation / date / signature) ─
    rows += [
        _sr('begin_group', 'grp_reviewer',
            'Death reviewer', 'মৃত্যু পর্যালোচনাকারী'),
        _sr('text', 'reviewer_name', 'Name of death reviewer',
            'মৃত্যু পর্যালোচনাকারীর নামঃ'),
        _sr('text', 'reviewer_designation', 'Designation', 'পদবিঃ'),
        _sr('date', 'data_collection_date', 'Date of data collection',
            'তথ্য সংগ্রহের তারিখ'),
        _sr('end_group', 'grp_reviewer'),
    ]

    return rows


def _facility_neonatal_choices():
    ch = list(DISTRICT_CHOICES) + list(YES_NO)

    # Recorded / Not recorded — mirrors the single "রেকর্ড করা হয়নি" tick box
    # on Q3 / Q4 / Q5 (binary: it was recorded, or it was not). A three-way
    # yes/no/not-recorded list is avoided because "No" and "Not recorded" are
    # the same answer here.
    ch += [
        _ch('recorded_status', 'recorded', 'Recorded', 'রেকর্ড করা হয়েছে'),
        _ch('recorded_status', 'not_recorded', 'Not recorded',
            'রেকর্ড করা হয়নি'),
    ]

    # Q1 — place of birth (11 options, matches the maternal facility list).
    birth = [
        ('home', 'Home', 'বাড়ী'),
        ('community_clinic', 'Community Clinic', 'কমিউনিটি ক্লিনিক'),
        ('union_hfwc', 'Union Health & Family Welfare Centre',
         'ইউনিয়ন স্বাস্থ্য ও পরিবার কল্যাণ কেন্দ্র'),
        ('upazila_hc', 'Upazila Health Complex', 'উপজেলা স্বাস্থ্য কমপ্লেক্স'),
        ('maternal_centre', 'Maternal Centre',
         'মাতৃমঙ্গল কেন্দ্র'),
        ('district_hospital', 'District or Sadar hospital',
         'জেলা অথবা সদর হাসপাতাল'),
        ('medical_college', 'Medical College hospital',
         'মেডিকেল কলেজ হাসপাতাল'),
        ('private_clinic', 'Private clinic/hospital',
         'প্রাইভেট ক্লিনিক/হাসপাতাল'),
        ('ngo_clinic', 'NGO clinic', 'এনজিওক্লিনিক'),
        ('provider_home', "Chamber/health provider's home",
         'চেম্বার/স্বাস্থ্য সেবাদানকারীর বাড়ী'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    ch += [_ch('birth_place', k, en, bn) for k, en, bn in birth]

    # Q2 — type of neonatal death (inborn vs outborn).
    ch += [
        _ch('death_type', 'inborn',
            'The neonate was born in this hospital and died '
            '(if ticked here, go to question No. 5)',
            'নবজাতক এই হাসপাতালে জন্মগ্রহণ করে এবং মারা যায় '
            '(এখানে টিক দিলে ৫ নং প্রশ্নে চলে যান)'),
        _ch('death_type', 'outborn',
            'The neonate was born outside the hospital and died after '
            'admission to the hospital',
            'নবজাতক হাসপাতালের বাইরে জন্মগ্রহণ এবং হাসপাতালে ভতির পর '
            'মৃত্যুবরণ করে'),
    ]

    # Q6 — condition at admission (4: Stable / Unstable / Unconscious /
    # Not recorded — paper order is Stable, Unconscious, Unstable, Not recorded).
    ch += [
        _ch('admission_condition', 'stable', 'Stable', 'স্থিতিশীল (Stable)'),
        _ch('admission_condition', 'unconscious', 'Unconscious',
            'অচেতন (Unconscious)'),
        _ch('admission_condition', 'unstable', 'Unstable',
            'অস্থিতিশীল (Unstable)'),
        _ch('admission_condition', 'not_recorded', 'Not recorded',
            'রেকর্ড করা হয়নি'),
    ]

    # Q7 — diagnosis at admission / after birth (multi, 9 options incl.
    # Not recorded + Other; no ICD codes on the paper for Q7).
    adx = [
        ('birth_asphyxia', 'Birth Asphyxia', 'শ্বাস কষ্ট (Birth Asphyxia)'),
        ('septicemia', 'Septicemia', 'সেপ্টিসেমিয়া (Septicemia)'),
        ('low_birth_weight', 'Low birth weight',
         'কম ওজন জন্ম (Low birth weight)'),
        ('severe_pneumonia', 'Severe Pneumonia',
         'মারাত্নক নিউমোনিয়া (Severe Pneumonia)'),
        ('meningitis', 'Meningitis', 'মেনিনজাইটিস (Meningitis)'),
        ('birth_trauma', 'Birth trauma', 'জন্মকালীন আঘাত (Birth trauma)'),
        ('congenital', 'Congenital anomalies',
         'জন্মগত ত্রুটি (Congenital anomalies)'),
        ('not_recorded', 'Not recorded', 'রেকর্ড করা হয়নি'),
        ('other', 'Others (specify)', 'অন্যান্য Others (উল্লেখ করুন)'),
    ]
    ch += [_ch('admission_diagnosis', k, en, bn) for k, en, bn in adx]

    # Q9 — referral source (5: matches the paper exactly).
    ch += [
        _ch('referral_source', 'govt', 'Government institution',
            'সরকারি প্রতিষ্ঠান'),
        _ch('referral_source', 'private', 'Private institution',
            'বেসরকারি প্রতিষ্ঠান'),
        _ch('referral_source', 'home', 'Home', 'বাড়ি'),
        _ch('referral_source', 'unknown', 'Not known', 'জানা নেই'),
        _ch('referral_source', 'other', 'Other (specify)',
            'অন্যান্য (উল্লেখ করুন)'),
    ]

    # Q10 — whether/when doctor first observed (3: observed / did-not-come /
    # not-recorded — the paper carries two tick boxes beside the date/time).
    ch += [
        _ch('doctor_observed', 'observed', 'Observed (record date & time)',
            'পর্যবেক্ষণ করা হয়েছে (তারিখ ও সময় উল্লেখ করুন)'),
        _ch('doctor_observed', 'not_come', 'Did not come',
            'অথবা আসে নাই'),
        _ch('doctor_observed', 'not_recorded', 'Not recorded',
            'রেকর্ড করা হয়নি'),
    ]

    # Q11 — neonatal danger signs (13 options).
    danger = [
        ('body_arching', 'Body arching / stiffening', 'শরীর বেঁকে যাওয়া'),
        ('not_feeding', 'Not feeding / refuses to feed',
         'খেতে না পারা বা খেতে চায় না'),
        ('fast_breathing', 'Fast breathing', 'দ্রুত শ্বাস'),
        ('fast_breathing_chest_indraw',
         'Fast breathing with chest indrawing',
         'দ্রুত শ্বাস ও বুক দেবে যাওয়া'),
        ('cold_extremities', 'Cold hands and feet',
         'হাত পা ঠান্ডা হয়ে যাওয়া'),
        ('fever', 'Fever', 'জ্বর'),
        ('no_movement', 'No or reduced movement',
         'নড়াচড়া না করা বা কম নড়াচড়া করা'),
        ('jaundice', 'Jaundice', 'জন্ডিস'),
        ('umbilical_infection', 'Red / discharging umbilicus',
         'নাভিতে লালচে দাগ বা পুজ পরা'),
        ('skin_pustules', 'Skin pustules / boils',
         'চামড়ার মধ্যে ফোড়া ও পুজ'),
        ('diarrhoea', 'Diarrhoea', 'ডায়রিয়া'),
        ('dont_know', "Don't know", 'জানি না'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    ch += [_ch('danger_signs', k, en, bn) for k, en, bn in danger]

    # Q12 — specialist inpatient diagnosis (8 options). ICD codes transcribed
    # VERBATIM from the paper exactly as printed, including its typos:
    # "মিনিনজাইটিস" (sic) and the "Meningitis) GO" / "Q 00" codes; Severe
    # Pneumonia carries no code on the paper.
    sdx = [
        ('birth_asphyxia', 'Birth Asphyxia P-21',
         'শ্বাস কষ্ট (Birth Asphyxia) P-21'),
        ('septicemia', 'Septicemia P-36',
         'সেপ্টিসেমিয়া (Septicemia) P-36'),
        ('low_birth_weight', 'Low birth weight P-07',
         'কম ওজন জন্ম (Low birth weight) P-07'),
        ('severe_pneumonia', 'Severe Pneumonia',
         'মারাত্নক নিউমোনিয়া (Severe Pneumonia)'),
        ('meningitis', 'Meningitis GO',
         'মিনিনজাইটিস (Meningitis) GO'),
        ('birth_trauma', 'Birth trauma P-10-15',
         'জন্মকালীন আঘাত (Birth trauma) P-10-15'),
        ('congenital', 'Congenital anomalies Q 00',
         'জন্মগত ত্রুটি (Congenital anomalies) Q 00'),
        ('other', 'Others (specify)', 'অন্যান্য Others (উল্লেখ করুন)'),
    ]
    ch += [_ch('specialist_diagnosis', k, en, bn) for k, en, bn in sdx]

    # Q16 — congenital-anomaly site (9 options).
    site = [
        ('head', 'Head', 'মাথা'),
        ('mouth_palate', 'Mouth / palate', 'মুখের তালু'),
        ('lip', 'Lip', 'ঠোঁট'),
        ('hands', 'Hands', 'হাতে'),
        ('chest', 'Chest', 'বুকে'),
        ('anus', 'Anus / back passage', 'পায়ু পথে'),
        ('foot_sole', 'Sole of foot', 'পায়ের পাতাতে'),
        ('vaginal', 'Vaginal passage', 'যোনী পথে'),
        ('other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    ch += [_ch('anomaly_site', k, en, bn) for k, en, bn in site]

    # Q17 — place of death within the facility (7 options).
    pod = [
        ('emergency_room', 'Emergency room', 'ইমারজেন্সি রুমে'),
        ('in_transit', 'During transfer', 'স্থানান্তরের সময়'),
        ('ward', 'Ward', 'ওয়ার্ডে'),
        ('labour_room', 'Labour room', 'লেবার রুমে'),
        ('operation_theatre', 'Operation room', 'অপারেশান রুমে'),
        ('scanu', 'Neonatal ward ( SCANU)', 'নবজাতক ওয়ার্ডে ( SCANU)'),
        ('other', 'In some other place (specify)',
         'অন্য কোন স্থানে উল্লেখ করুন'),
    ]
    ch += [_ch('place_of_death_facility', k, en, bn) for k, en, bn in pod]

    # Q18 — WHO Cause-of-death table (6 ICD-coded causes, multi-select).
    # Transcribed VERBATIM from the paper's English-only CAUSE / ICD CODE
    # table; codes kept EXACTLY as printed (Q OO and GOO are the paper's
    # letter-O renderings, not zeros; spacing P 07 / P 36 / P 10-15 as printed).
    # English fills both label slots, mirroring the gold-standard Form 01
    # `icd_cause` list — the paper's CAUSE column carries no Bangla.
    # Paper table order: Birth Asphyxia P-21 · Birth trauma P 10-15 ·
    #   Low birth weight P 07 · Congenital anomalies Q OO ·
    #   Meningitis GOO · Septicemia P 36.
    cod = [
        ('birth_asphyxia', 'Birth Asphyxia (P-21)'),
        ('birth_trauma', 'Birth trauma (P 10-15)'),
        ('low_birth_weight', 'Low birth weight (P 07)'),
        ('congenital', 'Congenital anomalies (Q OO)'),
        ('meningitis', 'Meningitis (GOO)'),
        ('septicemia', 'Septicemia (P 36)'),
    ]
    ch += [_ch('cod_cause', k, en, en) for k, en in cod]

    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 6 — Social Autopsy (Maternal Death)                              ║
# ║   Source: সামাজিক মৃত্যু পর্যালোচনা ফর্ম (Social Autopsy).pdf            ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _social_autopsy_survey():
    # Verbatim digitisation of the official paper "সামাজিক মৃত্যু পর্যালোচনা
    # রিপোর্টিং ফর্ম" (Social Autopsy reporting form). It is a committee
    # MEETING report — not a death-review with structured cause/place — keyed to
    # the death-notification slip number. Every visible question + its Bangla
    # wording matches the paper (incl. its spellings 'কল্যান', 'বর্ননা',
    # 'উপনিত'); only the hidden organisation calc and the operational GPS point
    # are added (location tagging is mandatory system-wide). No _meta() header:
    # the paper carries its own meeting header and puts the data-collector block
    # at the end.
    rows = [
        _sr('note', '_sa_gov', 'Government of the People’s Republic of Bangladesh',
            'গণপ্রজাতন্ত্রী বাংলাদেশ সরকার'),
        _sr('note', '_sa_ministry', 'Ministry of Health and Family Welfare',
            'স্বাস্থ্য ও পরিবার কল্যান মন্ত্রণালয়'),
        _sr('note', '_sa_title1',
            'Maternal death, neonatal death and stillbirth review',
            'মাতৃমৃত্যু, নবজাতকের মৃত্যু ও মৃতজন্ম পর্যালোচনা'),
        _sr('note', '_sa_title2', 'Social Autopsy reporting form',
            'সামাজিক মৃত্যু পর্যালোচনা রিপোর্টিং ফর্ম'),
        _sr('calculate', 'organisation', '', '', calc="'CIPRB'"),
        _sr('geopoint', 'location',
            'GPS location (required — step outside if no signal)',
            'জিপিএস অবস্থান (প্রয়োজনীয়)', required='yes'),
        # মাতৃমৃত্যু = ১, নবজাতকের মৃত্যু = ২, মৃতজন্ম = ৩
        _sr('select_one sa_death_type', 'sa_death_type',
            'Maternal death = 1, Neonatal death = 2, Stillbirth = 3',
            'মাতৃমৃত্যু = ১, নবজাতকের মৃত্যু = ২, মৃতজন্ম = ৩', required='yes'),
        _sr('date', 'meeting_date', 'Date of meeting', 'সভার তারিখ', required='yes'),
        _sr('time', 'meeting_time', 'Time of meeting (24 hours)',
            'সভার সময় [ ২৪ ঘণ্টায়]'),
        _sr('text', 'meeting_place', 'Place of meeting', 'সভার স্থান'),
        _sr('select_one district', 'district', 'District', 'জেলা', required='yes'),
        _sr('text', 'upazila', 'Upazila', 'উপজেলা'),
        _sr('text', 'union', 'Union', 'ইউনিয়ন'),
        _sr('text', 'ward', 'Ward', 'ওয়ার্ড'),
        _sr('text', 'village', 'Village name', 'গ্রামের নাম'),
        _sr('text', 'slip_number', 'Death notification slip number',
            'মৃত্যু অবহিতকরণ স্লিপ নম্বর'),
        _sr('text', 'deceased_name', 'Name of the deceased / mother',
            'মৃতের/মায়ের নাম', required='yes'),
        # মৃতের/মায়ের বয়স : বছর | মাস | দিন
        _sr('integer', 'age_years', 'Age of the deceased / mother — years',
            'মৃতের/মায়ের বয়স — বছর'),
        _sr('integer', 'age_months', 'Months', 'মাস'),
        _sr('integer', 'age_days', 'Days', 'দিন'),
        _sr('select_one sa_sex', 'sa_sex',
            'Sex (applicable for stillbirth / dead newborn): Boy = 1, Girl = 2',
            'লিঙ্গ : (মৃতজন্ম/ মৃত নবজাতকের ক্ষেত্রে প্রযোজ্য) ছেলে = ১, মেয়ে = ২'),
        _sr('text', 'death_narrative',
            'How the death occurred, in the words of the participants present '
            '(brief description)',
            'সভায় উপস্থিত অংশগ্রহণকারীদের ভাষ্যমতে কিভাবে মৃত্যুটি সংঘটিত হয়েছিল '
            '(সংক্ষিপ্ত বর্ননা)', app='multiline'),
    ]
    # কি করলে মৃত্যুটি রোধ করা যেত… (ক্রমানুসারে লিখুন) — ১..৪
    rows.append(_sr('note', '_sa_prevent_hdr',
        'What the participants think could have prevented the death (write in order)',
        'কি করলে মৃত্যুটি রোধ করা যেত বলে অংশগ্রহণকারীরা মনে করছেন (ক্রমানুসারে লিখুন)'))
    for i, bn in enumerate(['১', '২', '৩', '৪'], start=1):
        rows.append(_sr('text', 'prevention_%d' % i, '%d.' % i, '%s।' % bn))
    # উক্ত আলোচনা সভায় অংশগ্রহণকারীরা কি সিদ্ধান্তে উপনিত হলো (ক্রমানুসারে লিখুন) — ১..৪
    rows.append(_sr('note', '_sa_decision_hdr',
        'What decisions the meeting participants reached (write in order)',
        'উক্ত আলোচনা সভায় অংশগ্রহণকারীরা কি সিদ্ধান্তে উপনিত হলো (ক্রমানুসারে লিখুন)'))
    for i, bn in enumerate(['১', '২', '৩', '৪'], start=1):
        rows.append(_sr('text', 'decision_%d' % i, '%d.' % i, '%s।' % bn))
    rows += [
        _sr('integer', 'members_male',
            'Members present — male (persons)', 'উপস্থিত সদস্য/ সদস্যা : পুরুষ (জন)'),
        _sr('integer', 'members_female', 'Female (persons)', 'মহিলা (জন)'),
        _sr('integer', 'pregnant_women',
            'Pregnant women, if present (persons)',
            'গর্ভবতী মহিলা (যদি উপস্থিত থাকেন) (জন)'),
        _sr('text', 'collector_name', 'Name of the data collector',
            'তথ্য গ্রহণকারীর নাম', required='yes'),
        _sr('text', 'collector_designation', 'Designation', 'পদবী'),
    ]
    return rows


def _social_autopsy_choices():
    ch = list(DISTRICT_CHOICES)
    ch += [
        _ch('sa_death_type', '1', 'Maternal death', 'মাতৃমৃত্যু'),
        _ch('sa_death_type', '2', 'Neonatal death', 'নবজাতকের মৃত্যু'),
        _ch('sa_death_type', '3', 'Stillbirth', 'মৃতজন্ম'),
    ]
    ch += [
        _ch('sa_sex', '1', 'Boy', 'ছেলে'),
        _ch('sa_sex', '2', 'Girl', 'মেয়ে'),
    ]
    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 7 + 8 — Death Notification Slips 01 & 02                         ║
# ║   Small CHW-completed slips that NOTIFY a death and trigger the         ║
# ║   review committee. Same shape, two slip variants.                      ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _notification_slip_survey(slip_num: int):
    # Verbatim digitisation of the official "মাতৃমৃত্যু, নবজাতকের মৃত্যু ও
    # মৃতজন্ম অবহিতকরণ স্লিপ". Slip 01 = community version (কমিউনিটিতে প্রযোজ্য);
    # Slip 02 = hospital version (হাসপাতাল প্রযোজ্য). They share an identity block
    # and differ at the head (registration) and the clinical middle. No _meta():
    # the paper slip carries its own header and a data-collector block at the end.
    is_hosp = (slip_num == 2)
    banner_en = 'Applicable in hospital' if is_hosp else 'Applicable in the community'
    banner_bn = 'হাসপাতাল প্রযোজ্য' if is_hosp else 'কমিউনিটিতে প্রযোজ্য'
    rows = [
        _sr('note', '_ns_gov', 'Government of the People’s Republic of Bangladesh',
            'গণপ্রজাতন্ত্রী বাংলাদেশ সরকার'),
        _sr('note', '_ns_ministry', 'Ministry of Health and Family Welfare',
            'স্বাস্থ্য ও পরিবার কল্যাণ মন্ত্রণালয়'),
        _sr('note', '_ns_title',
            'Maternal death, neonatal death and stillbirth notification slip',
            'মাতৃমৃত্যু, নবজাতকের মৃত্যু ও মৃতজন্ম অবহিতকরণ স্লিপ'),
        _sr('note', '_ns_banner', banner_en, banner_bn),
        _sr('calculate', 'organisation', '', '', calc="'CIPRB'"),
        _sr('geopoint', 'location',
            'GPS location (required — step outside if no signal)',
            'জিপিএস অবস্থান (প্রয়োজনীয়)', required='yes'),
        _sr('text', 'case_serial', 'Serial no.', 'ক্রমিক নং'),
        _sr('date', 'slip_date', 'Date', 'তারিখ', required='yes'),
    ]
    if is_hosp:
        rows += [
            _sr('text', 'hospital_reg_no',
                'Hospital registration no. of mother / newborn / stillbirth',
                'মা/নবজাতক/মৃতজন্মের হাসপাতালের রেজিস্ট্রেশন নং'),
            _sr('text', 'ward_no', 'Ward no.', 'ওয়ার্ড নং'),
            _sr('text', 'bed_no', 'Bed no.', 'বিছানা নং'),
        ]
    else:
        rows += [
            _sr('text', 'mother_reg_no', "Mother's registration no.",
                'মায়ের রেজিস্ট্রেশন নং'),
            _sr('text', 'dhis2_newborn_reg',
                'Online (DHIS-2) newborn registration no.',
                'অনলাইনে (ডিএইচআইএস-২)-তে নবজাতকের রেজিস্ট্রেশন নং'),
        ]
    rows += [
        _sr('select_one death_kind', 'death_kind', 'Type of death',
            'মৃত্যুর ধরন', required='yes'),
        _sr('select_one ns_sex', 'sex', 'Sex', 'লিঙ্গ'),
        _sr('text', 'mother_name', "Mother's name", 'মায়ের নাম', required='yes'),
        _sr('integer', 'mother_age', "Mother's age (years)", 'মায়ের বয়স (বছর)'),
        _sr('text', 'father_husband_name', "Father's / husband's name",
            'পিতা/স্বামীর নাম'),
        _sr('text', 'para', 'Para', 'পাড়া'),
        _sr('text', 'village', 'Village', 'গ্রাম'),
        _sr('text', 'union', 'Union', 'ইউনিয়ন'),
        _sr('text', 'upazila', 'Upazila', 'উপজেলা'),
        _sr('select_one district', 'district', 'District', 'জেলা', required='yes'),
        _sr('text', 'family_mobile', "Deceased family's mobile no. (if any)",
            'মৃতের পরিবারের মোবাইল নং (যদি থাকে)'),
    ]
    if is_hosp:
        rows += [
            _sr('text', 'hospital_name', 'Name of hospital / health centre',
                'হাসপাতাল/স্বাস্থ্য কেন্দ্রের নাম'),
            _sr('date', 'admission_date',
                'Date of admission to hospital / health centre',
                'হাসপাতালে/স্বাস্থ্য কেন্দ্রে ভর্তির তারিখ'),
            _sr('time', 'admission_time', 'Time of admission (24 hours)',
                'ভর্তির সময় (২৪ ঘন্টায়)'),
            _sr('text', 'diagnosis_admission', 'Diagnosis at admission',
                'ভর্তির সময় রোগ নির্ণয়'),
            _sr('date', 'death_date', 'Date of death / stillbirth',
                'মৃত্যু/মৃত-জন্মের তারিখ', required='yes'),
            _sr('time', 'death_time', 'Time of death (24 hours)',
                'মৃত্যুর সময় (২৪ ঘন্টায়)'),
            _sr('text', 'cause_of_death', 'Cause of death', 'মৃত্যুর কারণ'),
        ]
    else:
        rows += [
            _sr('date', 'death_date', 'Date of death / stillbirth',
                'মৃত্যু/মৃত-জন্মের তারিখ', required='yes'),
            _sr('time', 'death_time', 'Time of death (24 hours)',
                'মৃত্যুর সময় (২৪ ঘন্টায়)'),
            _sr('date', 'delivery_date', 'Date of delivery', 'প্রসবের তারিখ'),
            _sr('time', 'delivery_time', 'Time of delivery (24 hours)',
                'প্রসবের সময় (২৪ ঘন্টায়)'),
            _sr('select_one ns_place_death', 'place_of_death', 'Place of death',
                'মৃত্যু স্থান'),
            _sr('text', 'place_of_death_other', 'Other place of death (specify)',
                'অন্যান্য মৃত্যু স্থান (উল্লেখ করুন)',
                relevant="${place_of_death}='other'"),
            _sr('select_one ns_place_delivery', 'place_of_delivery',
                'Place of delivery', 'প্রসবের স্থান'),
            _sr('text', 'place_of_delivery_other',
                'Other place of delivery (specify)',
                'অন্যান্য প্রসবের স্থান (উল্লেখ করুন)',
                relevant="${place_of_delivery}='other'"),
            _sr('select_one ns_attendant', 'delivery_attendant',
                'Who conducted the delivery?', 'কে প্রসব করিয়েছেন'),
            _sr('text', 'delivery_attendant_other', 'Other (specify)',
                'অন্যান্য (উল্লেখ করুন)',
                relevant="${delivery_attendant}='other'"),
        ]
    rows += [
        _sr('text', 'collector_name', 'Name of the data collector',
            'তথ্য গ্রহণকারীর নাম', required='yes'),
        _sr('text', 'collector_designation', 'Designation', 'পদবী'),
        _sr('text', 'collector_mobile', 'Mobile no.', 'মোবাইল নং'),
    ]
    if not is_hosp:
        rows += [
            _sr('text', 'cc_name', 'Name of community clinic',
                'কমিউনিটি ক্লিনিকের নাম'),
            _sr('text', 'cc_code', 'Community clinic code',
                'কমিউনিটি ক্লিনিকের কোড'),
        ]
    return rows


def _notification_slip_choices():
    ch = list(DISTRICT_CHOICES)
    # মৃত্যুর ধরন (order + spacing as printed: মাতৃ মৃত্যু / মৃত জন্ম / নবজাতকের মৃত্যু)
    ch += [
        _ch('death_kind', 'maternal',   'Maternal death',  'মাতৃ মৃত্যু'),
        _ch('death_kind', 'stillbirth', 'Stillbirth',      'মৃত জন্ম'),
        _ch('death_kind', 'neonatal',   'Neonatal death',  'নবজাতকের মৃত্যু'),
    ]
    ch += [
        _ch('ns_sex', 'boy',  'Boy',  'ছেলে'),
        _ch('ns_sex', 'girl', 'Girl', 'মেয়ে'),
    ]
    # মৃত্যু স্থান (slip 01)
    ch += [
        _ch('ns_place_death', 'home',          'At home', 'বাড়িতে'),
        _ch('ns_place_death', 'on_the_way',    'On the way', 'পথে'),
        _ch('ns_place_death', 'govt_facility', 'Govt health centre / hospital',
            'সরকারী স্বাস্থ্য কেন্দ্রে/হাসপাতালে'),
        _ch('ns_place_death', 'private_ngo',   'Private or NGO clinic / hospital',
            'বেসরকারী বা এনজিও ক্লিনিকে/হাসপাতালে'),
        _ch('ns_place_death', 'other',         'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    # প্রসবের স্থান (slip 01)
    ch += [
        _ch('ns_place_delivery', 'home',             'Home', 'বাড়ি'),
        _ch('ns_place_delivery', 'community_clinic', 'Community clinic', 'কমিউনিটি ক্লিনিক'),
        _ch('ns_place_delivery', 'uhfwc',
            'Union Health & Family Welfare Centre',
            'ইউনিয়ন স্বাস্থ্য ও পরিবার কল্যাণ কেন্দ্র'),
        _ch('ns_place_delivery', 'uhc', 'Upazila Health Complex',
            'উপজেলা স্বাস্থ্য কমপ্লেক্স'),
        _ch('ns_place_delivery', 'mcwc', 'Mother & Child Welfare Centre',
            'মা ও শিশু কল্যাণ কেন্দ্র'),
        _ch('ns_place_delivery', 'district_hospital', 'District hospital',
            'জেলা হাসপাতালে'),
        _ch('ns_place_delivery', 'medical_college', 'Medical college hospital',
            'মেডিকেল কলেজ হাসপাতালে'),
        _ch('ns_place_delivery', 'private_ngo', 'Private NGO clinic / hospital',
            'বেসরকারী এনজিও ক্লিনিক/হাসপাতাল'),
        _ch('ns_place_delivery', 'other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    # কে প্রসব করিয়েছেন (slip 01) — মিডওয়াইফ is printed twice on the paper; both kept.
    ch += [
        _ch('ns_attendant', 'doctor',        'Doctor (MBBS)', 'ডাক্তার (MBBS)'),
        _ch('ns_attendant', 'nurse',         'Nurse', 'নার্স'),
        _ch('ns_attendant', 'midwife',       'Midwife', 'মিডওয়াইফ'),
        _ch('ns_attendant', 'fwv',           'Family Welfare Visitor (FWV)',
            'পরিবার কল্যাণ পরিদর্শিকা (FWV)'),
        _ch('ns_attendant', 'csba',          'CSBA', 'সিএসবিএ (CSBA)'),
        _ch('ns_attendant', 'sacmo',         'Ma/SACMO', 'স্যাকমো (Ma/SACMO)'),
        _ch('ns_attendant', 'ha_fwa',        'HA/FWA',
            'স্বাস্থ্য সহকারী/পরিবার কল্যাণ সহকারী (HA/FWA)'),
        _ch('ns_attendant', 'village_doctor','Village doctor', 'পল্লী চিকিৎসক'),
        _ch('ns_attendant', 'dai',           'Dai (TBA)', 'দাই'),
        _ch('ns_attendant', 'midwife_2',     'Midwife', 'মিডওয়াইফ'),
        _ch('ns_attendant', 'ngo_worker',    'NGO worker', 'এনজিও কর্মী'),
        _ch('ns_attendant', 'other',         'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║   FORM 9 — Maternal Near Miss audit (WHO MNM)                           ║
# ║   Source: Tool_Near Miss.xlsx (18 district sheets, identical structure) ║
# ╚══════════════════════════════════════════════════════════════════════════╝

def _near_miss_survey():
    """Full WHO Maternal Near-Miss audit form, digitised verbatim from
    Tool_Near Miss.xlsx (column A — the 18 district sheets are identical).
    Ten sections: S1 screening (0-4 temporal coding), S2 general info,
    S3 maternal/perinatal, S4 process indicators, S5 underlying causes,
    S6 contributory conditions, S7-10 audit narrative. English labels are
    verbatim from the source; Bangla is a translation aid (source is EN only).

    Field names that the ingest handler / MaternalNearMissCase model map are
    preserved (woman_name, woman_age, gestational_weeks, facility_name,
    event_date, the 17 screening flags, mode_of_delivery, cause_of_near_miss,
    contributory_conditions, audit_summary). Everything else is carried in
    raw_payload."""
    rows = _meta('Annual near-miss case serial number',
                 'নিকট-মৃত্যুর বাৎসরিক ক্রমিক নং')

    # ── Section 1: Screening questions (WHO MNM Q1-3). Each criterion is
    # coded 0-4 per the WHO temporal scheme (verbatim from source A4-A5).
    rows += [
        _sr('begin_group', 'grp_sec1', 'Section 1 · Screening Questions',
            'প্রথম পর্যায় · স্ক্রিনিং প্রশ্ন'),
        _sr('note', 'sec1_coding',
            'In the questions 1-3, please specify:  0 = The condition was not '
            'present OR intervention not needed during the hospital stay;  '
            '1 = The condition was present at arrival;  2 = The condition was '
            'not present at arrival but developed within 12 hours of hospital '
            'admission;  3 = The condition was not present on admission but '
            'developed after 12 hours of hospital admission;  4 = Information '
            'not available / unknown or not applicable',
            'প্রশ্ন ১-৩ এ উত্তর দিন: ০=ছিল না/দরকার হয়নি; ১=ভর্তির সময় ছিল; '
            '২=ভর্তির ১২ ঘণ্টার মধ্যে দেখা দিয়েছে; ৩=ভর্তির ১২ ঘণ্টা পরে দেখা '
            'দিয়েছে; ৪=তথ্য নেই / অজানা / প্রযোজ্য নয়'),
        _sr('note', 'q1_head', '1. Severe maternal complications',
            '১. গুরুতর মাতৃস্বাস্থ্য জটিলতা'),
    ]
    for code, en, bn in [
        ('sev_pph',      'Severe Postpartum Hemorrhage', 'গুরুতর প্রসব-পরবর্তী রক্তক্ষরণ'),
        ('sev_preec',    'Severe Pre-eclampsia', 'গুরুতর প্রি-একলাম্পসিয়া'),
        ('eclampsia',    'Eclampsia', 'একলাম্পসিয়া'),
        ('sepsis',       'Sepsis or severe systemic infection', 'সেপসিস বা গুরুতর সংক্রমণ'),
        ('rupt_uterus',  'Ruptured uterus', 'জরায়ু ফেটে যাওয়া'),
        ('sev_abortion', 'Severe complication of abortion', 'গর্ভপাতের গুরুতর জটিলতা'),
    ]:
        rows.append(_sr('select_one mnm_code', code, en, bn))

    rows.append(_sr('note', 'q2_head', '2. Critical interventions',
                    '২. ক্রিটিক্যাল হস্তক্ষেপ'))
    for code, en, bn in [
        ('crit_blood',   'Use of blood products (includes any blood transfusion)',
         'রক্তজাত দ্রব্যের ব্যবহার (যেকোনো রক্ত পরিসঞ্চালন সহ)'),
        ('crit_radiol',  'Interventional radiology', 'ইন্টারভেনশনাল রেডিওলজি'),
        ('crit_laparot', 'Laparotomy (includes hysterectomy, excludes caesarean section)',
         'ল্যাপারোটমি (হিস্টেরেক্টমি সহ, সিজার বাদে)'),
        ('crit_icu',     'Admission to intensive care unit', 'নিবিড় পরিচর্যা কেন্দ্রে ভর্তি'),
    ]:
        rows.append(_sr('select_one mnm_code', code, en, bn))

    rows.append(_sr('note', 'q3_head',
        '3. Life-threatening conditions / Organ-dysfunction-based',
        '৩. জীবন-হুমকিস্বরূপ অবস্থা / অঙ্গ-ব্যর্থতা'))
    for code, en, bn in [
        ('life_cardio',  'Cardiovascular dysfunction [Shock, use of continuous vasoactive drugs, cardiac arrest, cardio pulmonary resuscitation, severe hypoperfusion (lactate >5 mmol/L or >45mg/dL), severe acidosis pH<7.1]', 'হৃদ-সংবহন ব্যর্থতা'),
        ('life_resp',    'Respiratory dysfunction [Acute cyanosis, gasping, severe tachypnea (respiratory rate >40 breaths per minute), Severe bradypnea (respiratory rate <6 breaths per minute), Severe hypoxemia (PAO2/FiO2 < 200, O2 saturation <90% for ≥60 min), intubation and ventilation not related to anesthesia]', 'শ্বাসতন্ত্রের ব্যর্থতা'),
        ('life_renal',   'Renal dysfunction: [Oliguria non-responsive to fluids or diuretics Dialysis for acute renal failure, severe acute azotemia (creatinine ≥300 μmol/mL or ≥3.5 mg/dL]', 'বৃক্কীয় ব্যর্থতা'),
        ('life_coag',    'Coagulation/ hematological dysfunction [Failure to form clots, massive transfusion of blood or red cells (≥5 units), severe acute thrombocytopenia (<50 000 platelets/mL)]', 'রক্ত জমাট / রক্তসংক্রান্ত ব্যর্থতা'),
        ('life_hepatic', 'Hepatic dysfunction [Jaundice in the presence of pre-eclampsia, Bilirubin >100 μmol/L or >6.0 mg/dL]', 'যকৃৎ ব্যর্থতা'),
        ('life_neuro',   'Neurological Dysfunction [Any unconsciousness (lasting ≥12 hours)/coma (including metabolic coma), stroke, Uncontrollable Fits/status epilepticus, total paralysis]', 'স্নায়বিক ব্যর্থতা'),
        ('life_uterine', 'Uterine dysfunction/Hysterectomy [Hemorrhage or infection leading to hysterectomy]', 'জরায়ু ব্যর্থতা / হিস্টেরেক্টমি'),
    ]:
        rows.append(_sr('select_one mnm_code', code, en, bn))

    rows.append(_sr('note', 'sec1_eligibility',
        'If the answer is "1" or "2" or "3" to any of the questions in '
        'Section 1 (Questions 1-3), go to the following sections (Sections '
        '2-10). If the answer is "0" or "4" to all of the questions in '
        'Section 1 (Questions 1-3), the woman is not eligible for this review '
        '— no need to answer the following sections. In case of doubt on '
        'questions 1 to 3, consult with the attending physician.',
        'প্রথম পর্যায়ের যেকোনো প্রশ্নে উত্তর "১/২/৩" হলে পরবর্তী অংশগুলো '
        '(২-১০) পূরণ করুন। সব উত্তর "০" বা "৪" হলে রোগী এই রিভিউয়ের জন্য '
        'যোগ্য নন — পরবর্তী অংশ পূরণের দরকার নেই। সন্দেহ থাকলে দায়িত্বরত '
        'চিকিৎসকের সাথে পরামর্শ করুন।'))
    rows.append(_sr('end_group', 'grp_sec1'))

    # ── Section 2: General information (Q4-14).
    rows += [
        _sr('begin_group', 'grp_sec2g', 'Section 2 · General information',
            'দ্বিতীয় পর্যায় · সাধারণ তথ্য'),
        _sr('text', 'facility_name', '4. Name of health facility',
            '৪. স্বাস্থ্য কেন্দ্রের নাম', required='yes'),
        _sr('text', 'hospital_reg_no', '5. Hospital registration number',
            '৫. হাসপাতাল রেজিস্ট্রেশন নম্বর'),
        _sr('text', 'woman_name', '6. Name of the patient',
            '৬. রোগীর নাম', required='yes'),
        _sr('integer', 'woman_age', '7. Age of the patient (Year)',
            '৭. রোগীর বয়স (বছর)', required='yes', constraint='. >= 10 and . <= 60'),
        _sr('text', 'patient_address', '8. Address', '৮. ঠিকানা'),
        _sr('date', 'event_date', '9. Date of hospital admission (dd/mm/yy)',
            '৯. হাসপাতালে ভর্তির তারিখ', required='yes'),
        _sr('text', 'admission_time',
            '10. Time of hospital admission (hr/min/am/pm)', '১০. ভর্তির সময়'),
        _sr('date', 'delivery_date',
            '11. Date of delivery or uterine evacuation (dd/mm/yy)',
            '১১. প্রসব/জরায়ু খালি করার তারিখ'),
        _sr('text', 'delivery_time',
            '12. Approximate time of delivery or uterine evacuation (hr/min/am/pm)',
            '১২. প্রসব/জরায়ু খালি করার আনুমানিক সময়'),
        _sr('date', 'discharge_date', '13. Date of hospital discharge (dd/mm/yy)',
            '১৩. হাসপাতাল থেকে ছাড়ার তারিখ'),
        _sr('date', 'review_date', '14. Date of Near-miss review (dd/mm/yy)',
            '১৪. নিকট-মৃত্যু রিভিউয়ের তারিখ'),
        _sr('end_group', 'grp_sec2g'),
    ]

    # ── Section 3: Maternal and Perinatal Information (Q15-17).
    # Q15 is digitised exactly as the Excel sheet structures it: a header row
    # plus 9 checkbox items, each coded 0=No/1=Yes — NOT a single-select.
    rows += [
        _sr('begin_group', 'grp_sec3m',
            'Section 3 · Maternal and Perinatal Information',
            'তৃতীয় পর্যায় · মাতৃ ও প্রসবকালীন তথ্য'),
        _sr('note', 'q15_head',
            '15. Final mode of delivery / end of pregnancy: (Tick appropriate '
            'response) (0=No; 1=Yes)',
            '১৫. প্রসব/গর্ভাবস্থার চূড়ান্ত পরিণতি (প্রযোজ্য উত্তরে টিক দিন) (০=না; ১=হ্যাঁ)'),
    ]
    for code, en, bn in [
        ('mode_vaginal',             '1. Vaginal delivery', '১. স্বাভাবিক প্রসব'),
        ('mode_csection',            '2. Caesarean section', '২. সিজারিয়ান'),
        ('mode_complete_abortion',   '3. Complete abortion', '৩. সম্পূর্ণ গর্ভপাত'),
        ('mode_curettage',           '4. Curettage / vacuum aspiration', '৪. কিউরেটাজ / ভ্যাকুয়াম অ্যাসপিরেশন'),
        ('mode_medical_evacuation',  '5. Medical methods for uterine evacuation', '৫. ওষুধে জরায়ু খালি করা'),
        ('mode_lap_ectopic',         '6. Laparotomy for ectopic pregnancy', '৬. একটোপিক প্রেগন্যান্সিতে ল্যাপারোটমি'),
        ('mode_lap_rupture',         '7. Laparotomy for ruptured uterus', '৭. জরায়ু ফাটায় ল্যাপারোটমি'),
        ('mode_discharged_pregnant', '8. Woman discharged while still pregnant', '৮. গর্ভাবস্থায় ছাড়পত্র'),
        ('mode_unknown_other',       '9. Unknown / other', '৯. অজানা / অন্যান্য'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))
    rows += [
        _sr('integer', 'gestational_weeks',
            '16.1 Best estimate of gestational age at delivery or abortion '
            '(completed weeks) — not applicable if Q15 = 8 or 9',
            '১৬.১ প্রসব/গর্ভপাতের সময় গর্ভকাল (সম্পূর্ণ সপ্তাহ)',
            constraint='. >= 0 and . <= 45'),
        _sr('integer', 'gestational_weeks_discharge',
            '16.2 Best estimate of gestational age at hospital discharge '
            '(completed weeks) — applicable if Q15 = 8',
            '১৬.২ ছাড়ার সময় গর্ভকাল (সম্পূর্ণ সপ্তাহ)',
            relevant="selected(${mode_discharged_pregnant},'yes')",
            constraint='. >= 0 and . <= 45'),
        _sr('note', 'q17_head',
            '17. Regarding the vital status of the infant, please specify',
            '১৭. শিশুর জীবিত অবস্থা'),
        _sr('select_one infant_vital', 'infant_status_birth',
            '17.1 At birth', '১৭.১ জন্মের সময়'),
        _sr('select_one infant_vital', 'infant_status_discharge',
            '17.2 At hospital discharge or on the 7th day of life, if still in the hospital',
            '১৭.২ ছাড়ার সময় বা জীবনের ৭ম দিনে (হাসপাতালে থাকলে)'),
        _sr('end_group', 'grp_sec3m'),
    ]

    # ── Section 4: Process Indicators (Q18-19).
    rows += [
        _sr('begin_group', 'grp_sec4', 'Section 4 · Process Indicators',
            'চতুর্থ পর্যায় · প্রক্রিয়া নির্দেশক'),
        _sr('note', 'q18_head',
            '18. About condition at arrival in the facility and the referral '
            'process, specify (0=No; 1=Yes)',
            '১৮. কেন্দ্রে পৌঁছানোর সময়ের অবস্থা ও রেফারেল প্রক্রিয়া'),
    ]
    for code, en, bn in [
        ('arr_delivery_before', 'Delivery or abortion occurred before arrival at hospital',
         'হাসপাতালে আসার আগে প্রসব/গর্ভপাত হয়েছে'),
        ('arr_delivery_3h', 'Delivery occurred within 3 hours of arrival in the hospital',
         'আসার ৩ ঘণ্টার মধ্যে প্রসব হয়েছে'),
        ('arr_lap_3h', 'Laparotomy done within 3 hours of hospital arrival',
         'আসার ৩ ঘণ্টার মধ্যে ল্যাপারোটমি হয়েছে'),
        ('arr_lap_before', 'Laparotomy done before arrival at this hospital',
         'এই হাসপাতালে আসার আগে ল্যাপারোটমি হয়েছে'),
        ('arr_ref_from', 'Woman referred from other hospital',
         'অন্য হাসপাতাল থেকে রেফার করা হয়েছে'),
        ('arr_ref_to', 'Woman referred to any higher-level hospital',
         'উচ্চতর হাসপাতালে রেফার করা হয়েছে'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))

    rows.append(_sr('note', 'q19_head',
        '19. About use of interventions [if applicable]: please specify '
        'whether the woman received any of the following (0=No; 1=Yes)',
        '১৯. হস্তক্ষেপের ব্যবহার (প্রযোজ্য হলে)'))
    rows.append(_sr('note', 'q19_1_head',
        '19.1 Uterotonic for prevention of PPH after delivery',
        '১৯.১ প্রসব-পরবর্তী রক্তক্ষরণ প্রতিরোধে ইউটেরোটনিক'))
    rows.append(_sr('select_one yes_no', 'utero_prev_oxytocin', 'Oxytocin', 'অক্সিটোসিন'))
    rows.append(_sr('select_one yes_no', 'utero_prev_other', 'Other uterotonic, specify',
                    'অন্যান্য ইউটেরোটনিক (উল্লেখ করুন)'))
    rows.append(_sr('text', 'utero_prev_other_specify',
                    'Other uterotonic — specify', 'অন্যান্য ইউটেরোটনিক — উল্লেখ করুন',
                    relevant="selected(${utero_prev_other},'yes')"))
    rows.append(_sr('note', 'q19_2_head', '19.2 For treatment of PPH',
                    '১৯.২ রক্তক্ষরণের চিকিৎসায়'))
    for code, en, bn in [
        ('pph_oxytocin',         'Oxytocin', 'অক্সিটোসিন'),
        ('pph_ergometrine',      'Ergometrine', 'এরগোমেট্রিন'),
        ('pph_misoprostol',      'Misoprostol', 'মিসোপ্রোস্টল'),
        ('pph_other_utero',      'Other uterotonics', 'অন্যান্য ইউটেরোটনিক'),
        ('pph_tranexamic',       'Tranexamic acid', 'ট্রানেক্সামিক অ্যাসিড'),
        ('pph_carbetocin',       'Carbetocin', 'কার্বেটোসিন'),
        ('pph_repair_tear',      'Repair of tear', 'ছিঁড়ে যাওয়া স্থান মেরামত'),
        ('pph_remove_products',  'Removal of retained products', 'অবশিষ্ট অংশ অপসারণ'),
        ('pph_tamponade',        'Balloon or condom tamponade', 'বেলুন/কন্ডম ট্যাম্পোনেড'),
        ('pph_artery_ligation',  'Artery ligation (uterine/hypogastric)', 'ধমনী বন্ধন (জরায়ু/হাইপোগ্যাস্ট্রিক)'),
        ('pph_hysterectomy',     'Hysterectomy', 'হিস্টেরেক্টমি'),
        ('pph_abdominal_packing','Abdominal packing', 'অ্যাবডমিনাল প্যাকিং'),
        ('pph_blood_transfusion','Blood transfusion', 'রক্ত পরিসঞ্চালন'),
        ('pph_manual_placenta',  'Manual removal of placenta', 'হাতে গর্ভফুল অপসারণ'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))
    rows.append(_sr('note', 'q19_3_head', '19.3 For treatment of Eclampsia',
                    '১৯.৩ একলাম্পসিয়ার চিকিৎসায়'))
    rows.append(_sr('select_one yes_no', 'ecl_anticonvulsant',
                    '19.3.1 Anticonvulsant', '১৯.৩.১ অ্যান্টিকনভালসেন্ট'))
    rows.append(_sr('select_one yes_no', 'ecl_antihypertensive',
                    '19.3.2 Antihypertensive (Oral)', '১৯.৩.২ অ্যান্টিহাইপারটেনসিভ (মুখে)'))
    rows.append(_sr('note', 'q19_4_head', '19.4 Antibiotics', '১৯.৪ অ্যান্টিবায়োটিক'))
    for code, en, bn in [
        ('abx_proph_csection',  'Prophylactic antibiotic for C-section',
         'সিজারে প্রতিরোধমূলক অ্যান্টিবায়োটিক'),
        ('abx_proph_laparotomy','Prophylactic antibiotic for Laparotomy',
         'ল্যাপারোটমিতে প্রতিরোধমূলক অ্যান্টিবায়োটিক'),
        ('abx_therapeutic',     'Parenteral, therapeutic antibiotic for severe systemic infections or sepsis',
         'গুরুতর সংক্রমণ/সেপসিসে প্যারেন্টেরাল থেরাপিউটিক অ্যান্টিবায়োটিক'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))
    rows.append(_sr('note', 'q19_5_head',
        '19.5 For fetal lung maturation in case of preterm birth',
        '১৯.৫ প্রিটার্ম জন্মে ভ্রূণের ফুসফুস পরিপক্বতায়'))
    rows.append(_sr('select_one yes_no', 'flm_corticosteroids',
        'Corticosteroids (Dexamethasone or Betamethason)',
        'কর্টিকোস্টেরয়েড (ডেক্সামেথাসন বা বিটামেথাসন)'))
    rows.append(_sr('note', 'q19_6_head', '19.6 For managing Rupture uterus',
                    '১৯.৬ জরায়ু ফাটা ব্যবস্থাপনায়'))
    for code, en, bn in [
        ('rupt_lap_repair', 'Laparotomy followed by repair', 'ল্যাপারোটমি ও মেরামত'),
        ('rupt_lap_hyster', 'Laparotomy followed by Hysterectomy', 'ল্যাপারোটমি ও হিস্টেরেক্টমি'),
        ('rupt_blood',      'Blood transfusion', 'রক্ত পরিসঞ্চালন'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))
    rows.append(_sr('note', 'q19_7_head', '19.7 For Severe complication of abortion',
                    '১৯.৭ গর্ভপাতের গুরুতর জটিলতায়'))
    for code, en, bn in [
        ('abrt_shock',     'Shock management', 'শক ব্যবস্থাপনা'),
        ('abrt_blood',     'Blood transfusion', 'রক্ত পরিসঞ্চালন'),
        ('abrt_antibiotic','Therapeutic antibiotic in septic abortion',
         'সেপটিক গর্ভপাতে থেরাপিউটিক অ্যান্টিবায়োটিক'),
        ('abrt_dc',        'D&C', 'ডি অ্যান্ড সি'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))
    rows.append(_sr('end_group', 'grp_sec4'))

    # ── Section 5: Underlying Causes of Near-Miss (Q20).
    rows += [
        _sr('begin_group', 'grp_sec5', 'Section 5 · Underlying Causes of Near-Miss',
            'পঞ্চম পর্যায় · নিকট-মৃত্যুর অন্তর্নিহিত কারণ'),
        _sr('note', 'q20_head', '20. Please specify (0=No; 1=Yes)', '২০. উল্লেখ করুন'),
    ]
    for code, en, bn in [
        ('un_abortive',        'Pregnancy with abortive outcome', 'গর্ভপাতমূলক পরিণতির গর্ভাবস্থা'),
        ('un_ectopic_molar',   'Ectopic /Molar Pregnancy', 'একটোপিক / মোলার গর্ভাবস্থা'),
        ('un_haemorrhage',     'Obstetric hemorrhage', 'প্রসূতি রক্তক্ষরণ'),
        ('un_hypertensive',    'Hypertensive disorders', 'উচ্চ রক্তচাপজনিত ব্যাধি'),
        ('un_infection',       'Pregnancy-related infection', 'গর্ভকালীন সংক্রমণ'),
        ('un_rupture',         'Ruptured uterus', 'জরায়ু ফেটে যাওয়া'),
        ('un_other_obstetric', 'Other obstetric diseases or complication', 'অন্যান্য প্রসূতি রোগ বা জটিলতা'),
        ('un_medical',         'Medical/surgical/mental disease or complication', 'মেডিকেল/সার্জিকাল/মানসিক রোগ বা জটিলতা'),
        ('un_unexpected',      'Unexpected complications of management', 'ব্যবস্থাপনার অপ্রত্যাশিত জটিলতা'),
        ('un_coincidental',    'Coincidental conditions', 'আকস্মিক অবস্থা'),
        ('un_unknown',         'Unknown', 'অজানা'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))
    rows.append(_sr('end_group', 'grp_sec5'))

    # ── Section 6: Contributory / Associated Conditions (Q21).
    rows += [
        _sr('begin_group', 'grp_sec6',
            'Section 6 · Contributory / Associated Conditions of Near-Miss',
            'ষষ্ঠ পর্যায় · অবদানকারী / সম্পর্কিত অবস্থা'),
        _sr('note', 'q21_head', '21. Please specify', '২১. উল্লেখ করুন'),
    ]
    for code, en, bn in [
        ('cc_anemia',      'Anemia', 'রক্তস্বল্পতা'),
        ('cc_hiv',         'HIV infection', 'এইচআইভি সংক্রমণ'),
        ('cc_prev_cs',     'Previous cesarean section', 'পূর্ববর্তী সিজার'),
        ('cc_obstructed',  'Prolonged/obstructed labor', 'দীর্ঘ/বাধাগ্রস্ত প্রসব'),
        ('cc_heart',       'Heart disease', 'হৃদরোগ'),
        ('cc_diabetes',    'Diabetes mellitus', 'ডায়াবেটিস মেলাইটাস'),
        ('cc_respiratory', 'Respiratory dysfunction;Asthma,TB', 'শ্বাসতন্ত্রের ব্যাধি; হাঁপানি, যক্ষ্মা'),
        ('cc_other',       'Other condition, specify', 'অন্যান্য অবস্থা (উল্লেখ করুন)'),
    ]:
        rows.append(_sr('select_one yes_no', code, en, bn))
    rows.append(_sr('text', 'cc_other_specify', 'Other condition — specify',
                    'অন্যান্য অবস্থা — উল্লেখ করুন',
                    relevant="selected(${cc_other},'yes')"))
    rows.append(_sr('end_group', 'grp_sec6'))

    # ── Sections 7-10: Audit narrative.
    rows += [
        _sr('begin_group', 'grp_sec7_10', 'Sections 7-10 · Audit narrative',
            'সপ্তম–দশম পর্যায় · অডিট বিবরণ'),
        _sr('text', 'interview_summary',
            '7. Summary of interview with mother/attendant [socio-economic / '
            'health-system factors that have the potential for developing this '
            'case and management events]',
            '৭. মা/সঙ্গীর সাক্ষাৎকারের সারসংক্ষেপ', app='multiline'),
        _sr('text', 'audit_summary',
            '8. Case summary (brief overview on clinical events)',
            '৮. কেস সারসংক্ষেপ (ক্লিনিক্যাল ঘটনার সংক্ষিপ্ত বিবরণ)', app='multiline'),
        _sr('note', 'sec9_head',
            '9. Short summary of the event strengths and weaknesses [short '
            'summary of the events surrounding the near-miss e.g. missed '
            'opportunities and standard care, infrastructural issues, '
            'availability of drugs, instrument, equipment or consumables etc.]',
            '৯. ঘটনার শক্তি ও দুর্বলতার সংক্ষিপ্ত সারসংক্ষেপ'),
        _sr('text', 'audit_strength',
            '9.1 Strength (explain Medical or Non-Medical factors behind '
            'successful management of this emergency case)',
            '৯.১ শক্তি', app='multiline'),
        _sr('text', 'audit_needs_improvement',
            '9.2 Needs improvement (explain Medical or Non-Medical factors that '
            'hinder the successful management of the patient with this case)',
            '৯.২ উন্নতি প্রয়োজন', app='multiline'),
        _sr('text', 'audit_recommendations',
            '10. Major Recommendations (short summary of the recommendations to '
            'address audit findings)',
            '১০. প্রধান সুপারিশ', app='multiline'),
        _sr('text', 'audit_team',
            'Name and signature of Audit team members',
            'অডিট দলের সদস্যদের নাম ও স্বাক্ষর', app='multiline'),
        _sr('end_group', 'grp_sec7_10'),
    ]
    return rows


def _near_miss_choices():
    ch = list(DISTRICT_CHOICES) + list(YES_NO)
    # WHO MNM screening temporal code (Section 1, verbatim 0-4 scheme).
    ch += [
        _ch('mnm_code', '0', '0 — Not present / intervention not needed', '০ — ছিল না / দরকার হয়নি'),
        _ch('mnm_code', '1', '1 — Present at arrival', '১ — ভর্তির সময় ছিল'),
        _ch('mnm_code', '2', '2 — Developed within 12 hours of admission', '২ — ভর্তির ১২ ঘণ্টার মধ্যে'),
        _ch('mnm_code', '3', '3 — Developed after 12 hours of admission', '৩ — ভর্তির ১২ ঘণ্টা পরে'),
        _ch('mnm_code', '4', '4 — Information not available / unknown / NA', '৪ — তথ্য নেই / অজানা'),
    ]
    # Q17 — infant vital status (0=Alive; 1=Dead).
    ch += [
        _ch('infant_vital', 'alive', 'Alive', 'জীবিত'),
        _ch('infant_vital', 'dead',  'Dead',  'মৃত'),
    ]
    return ch


# ─── Form catalogue ──────────────────────────────────────────────────────────

# ─── FORM 10 — MPDSR Action Plan (review-meeting action tracker) ─────────────
# Source: January to June Action Plan_2026_Kurigram.docx — digitised verbatim.
#   Table 1  MPDSR System Strengthening — sub-category column + Activities /
#            Responsible / Timeline / Indicator / Milestone / Considerations.
#   Table 2  Common modifiable factors, (Community verbal autopsy) and
#            (Facility death review) — Actions are taken / Responsible /
#            Timeline / Milestone / Considerations (no Indicator).
#   Each action is one repeat entry. The master carries no status column and no
#   meeting-metadata header, so neither is added.

def _response_plan_survey():
    # Verbatim digitisation of the official master "MPDSR Action Plan 2026"
    # (January to June Action Plan_2026, Kurigram). The two master tables become
    # three repeat sections so a review can log any number of actions, each as
    # its own entry:
    #   1. MPDSR System Strengthening — sub-category column (Community/Facility
    #      death review, assignment of causes, response-plan development &
    #      implementation, M&E) + Activities / Responsible / Timeline /
    #      Indicator / Milestone / Considerations.
    #   2-3. Common modifiable factors (Community verbal autopsy) and
    #      (Facility death review) — Actions are taken / Responsible / Timeline /
    #      Milestone / Considerations (no Indicator column).
    # The master has no status column or meeting-metadata header; a per-action
    # status is added by request, but every column wording is kept exactly as
    # the master.
    rows = _meta()
    rows.append(_sr('note', '_rp_title', 'MPDSR Action Plan 2026',
                    'এমপিডিএসআর অ্যাকশন প্ল্যান ২০২৬'))
    rows.append(_sr('select_one ap_mode', 'ap_mode',
                    'What do you want to do?', 'আপনি কী করতে চান?', required='yes'))

    # ═══ MODE A — register ONE action (mirrors the fistula suspected stage:
    #     the worker TYPES the action_id, the form shows the district-code
    #     prefix + blocks duplicates; one action per submission, no repeat). ═══
    NORM_AID = ("translate(normalize-space(${action_id}),"
                "'abcdefghijklmnopqrstuvwxyz',"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ')")
    rows.append(_sr('begin_group', 'grp_new_plan', 'Register a new action',
                    'নতুন পদক্ষেপ নিবন্ধন', relevant="${ap_mode}='new_plan'"))
    # District-code prefix — UPPERCASE LETTER codes (Patuakhali=PA, Dhaka=DH),
    # distinct from the fistula numeric codes. Action IDs now read PA-001, DH-001.
    rows.append(_sr('calculate', '_act_dist_code', calc=_mpdsr_dist_code_calc()))
    rows.append(_sr('note', '_act_code_show',
        'Your district code is ${_act_dist_code}. Type the Action ID as '
        '${_act_dist_code}-001, ${_act_dist_code}-002, … (3 digits after the dash).',
        'আপনার জেলা কোড ${_act_dist_code}। পদক্ষেপের আইডি এভাবে লিখুন: '
        '${_act_dist_code}-001, ${_act_dist_code}-002, … (ড্যাশের পরে ৩ অঙ্ক)।'))
    rows.append(_sr('select_one rp_section', 'rp_section',
                    'Which table / section is this action from?',
                    'এই পদক্ষেপটি কোন টেবিল / বিভাগের?', required='yes'))
    # The unique action ID — typed at registration (the fistula patient_code shape):
    #   (1) regex anchors the FULL prefix (^<code>-<digits>$);
    #   (2) pulldata(...)='' blocks re-using an action ID that already exists.
    rows.append(_sr('text', 'action_id',
        'Action ID (district letter code + 3-digit number, e.g. PA-001)',
        'পদক্ষেপ আইডি (জেলার অক্ষর কোড + ৩ অঙ্ক, যেমন PA-001)',
        required='yes',
        # regex uppercases the typed value first, so 'pa-001' is accepted and
        # stored canonical as PA-001 (the handler's _norm_id upper-cases too).
        constraint=("regex(translate(normalize-space(.),"
                    "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'), "
                    "concat('^', ${_act_dist_code}, '-[0-9]{3}$')) and "
                    "pulldata('mpdsr_actions','activity','action_id'," + NORM_AID + ")=''"),
        cmsg='⚠ Invalid or duplicate ID. It must be <district-letter-code>-<3 digits> '
             '(e.g. PA-001, Dhaka = DH-001) and not already used. / '
             'ভুল বা ডুপ্লিকেট আইডি — জেলার অক্ষর কোড + ৩ অঙ্ক হতে হবে এবং আগে ব্যবহৃত হওয়া যাবে না।',
        hint='Format: your district letter code + a 3-digit serial. Patuakhali = PA-001, Dhaka = DH-001.'))
    # Duplicate-ID soft warning — shows the activity already using this ID.
    rows.append(_sr('calculate', '_act_dup',
        calc=("pulldata('mpdsr_actions','activity','action_id'," + NORM_AID + ")")))
    rows.append(_sr('note', '_act_dup_warn',
        '⚠ This ID is already used for: ${_act_dup}. '
        'Use a new number, or switch to "Update an existing action".',
        '⚠ এই আইডি ইতিমধ্যে ব্যবহৃত: ${_act_dup}। '
        'নতুন নম্বর দিন, অথবা "বিদ্যমান একটি পদক্ষেপ হালনাগাদ" বেছে নিন।',
        relevant="${action_id}!='' and ${_act_dup}!=''"))
    # The action's details (flat — one action per submission).
    rows.append(_sr('select_one rp_subcat', 'act_subcat',
        'Sub-category (System Strengthening only)',
        'উপ-বিভাগ (শুধু সিস্টেম শক্তিশালীকরণ)',
        relevant="${rp_section}='system_strengthening'"))
    # The modifiable-factor row of master Table 2. Both factor sections share one
    # list: the master ships those rows BLANK (each district fills them from its
    # own verbal-autopsy / death-review findings), so the codes below are the
    # factors districts actually write, plus 'other' + free text so a district is
    # never forced into a wrong bucket. Asked before the action so the action is
    # always attached to the factor it answers.
    rows.append(_sr('select_one rp_factor', 'act_factor',
        'Which common modifiable factor is this action for?',
        'এই পদক্ষেপটি কোন সাধারণ পরিবর্তনযোগ্য কারণের জন্য?',
        required='yes',
        relevant="${rp_section}='community_va' or ${rp_section}='facility_dr'"))
    rows.append(_sr('text', 'act_factor_other',
        'Write the modifiable factor', 'পরিবর্তনযোগ্য কারণটি লিখুন',
        required='yes', app='multiline',
        relevant="${act_factor}='other' and "
                 "(${rp_section}='community_va' or ${rp_section}='facility_dr')"))
    rows.append(_sr('text', 'act_activity', 'Activity / action to be taken',
                    'কার্যক্রম / গৃহীত পদক্ষেপ', required='yes', app='multiline'))
    rows.append(_sr('text', 'act_responsible', 'Responsible',
                    'দায়িত্বপ্রাপ্ত (ব্যক্তি / দপ্তর)'))
    rows.append(_sr('date', 'act_timeline', 'Timeline', 'সময়সীমা'))
    rows.append(_sr('text', 'act_indicator', 'Indicator', 'নির্দেশক',
                    relevant="${rp_section}='system_strengthening'"))
    rows.append(_sr('text', 'act_milestone', 'Milestone', 'মাইলফলক'))
    rows.append(_sr('text', 'act_considerations', 'Considerations',
                    'বিবেচ্য বিষয়', app='multiline'))
    rows.append(_sr('select_one rp_status', 'act_status',
                    'Status of this action', 'এই পদক্ষেপের অবস্থা'))
    rows.append(_sr('end_group', 'grp_new_plan'))

    # ═══ MODE B — update an existing action (pick by id, auto-fill, advance) ═══
    # mpdsr_actions.csv (form media, kept current by export_mpdsr_actions) lists
    # every open action: name=action_id, label='D-01 — <activity> (<district>)'.
    NORM_SEL = ("translate(normalize-space(${ap_action_sel}),"
                "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')")
    rows.append(_sr('begin_group', 'grp_update', 'Update an action',
                    'একটি পদক্ষেপ হালনাগাদ', relevant="${ap_mode}='update_action'"))
    rows.append(_sr('select_one_from_file mpdsr_actions.csv', 'ap_action_sel',
                    'Select the action by its ID (e.g. D-01)',
                    'আইডি দিয়ে পদক্ষেপটি বেছে নিন (যেমন D-01)',
                    app='autocomplete', required='yes',
                    cf='district_slug=${district}'))
    for col, nm in [('activity', '_ap_act'), ('responsible', '_ap_resp'),
                    ('timeline', '_ap_time'), ('district', '_ap_dist'),
                    ('status', '_ap_curstat')]:
        rows.append(_sr('calculate', nm, '', '',
                        calc="pulldata('mpdsr_actions','%s','action_id',%s)" % (col, NORM_SEL)))
    rows.append(_sr('note', '_ap_show',
        'Action ${ap_action_sel}: ${_ap_act}\n'
        'Responsible: ${_ap_resp} · Timeline: ${_ap_time} (${_ap_dist})\n'
        'Current status: ${_ap_curstat}',
        'পদক্ষেপ ${ap_action_sel}: ${_ap_act}\n'
        'দায়িত্বপ্রাপ্ত: ${_ap_resp} · সময়সীমা: ${_ap_time} (${_ap_dist})',
        relevant="${_ap_act}!=''"))
    rows.append(_sr('note', '_ap_nomatch',
        'No action found for that ID — check the list, or it may not have synced yet.',
        'এই আইডিতে কোনো পদক্ষেপ পাওয়া যায়নি — তালিকা দেখুন।',
        relevant="${ap_action_sel}!='' and ${_ap_act}=''"))
    rows.append(_sr('select_one rp_status', 'ap_new_status', 'Updated status',
                    'হালনাগাদ অবস্থা', required='yes'))
    rows.append(_sr('select_one ap_completion', 'ap_new_completion',
                    'Completion %', 'সম্পন্নের হার (%)', required='yes'))
    rows.append(_sr('date', 'ap_completion_date',
                    'Date completed (if 100%)', 'সম্পন্নের তারিখ (যদি ১০০%)'))
    rows.append(_sr('text', 'ap_remarks', 'Remarks / progress note',
                    'মন্তব্য / অগ্রগতির নোট', app='multiline'))
    rows.append(_sr('end_group', 'grp_update'))
    return rows


def _response_plan_choices():
    # rp_subcat = the System-Strengthening sub-category column from the master
    # (Table 1, first column). The master carries no review-level or status list.
    ch = list(DISTRICT_CHOICES)
    for k, en, bn in [
        ('community_death_review', 'Community Death Review', 'কমিউনিটি ডেথ রিভিউ'),
        ('facility_death_review', 'Facility Death Review', 'ফ্যাসিলিটি ডেথ রিভিউ'),
        ('assignment_causes', 'Assignment causes of deaths', 'মৃত্যুর কারণ নির্ধারণ'),
        ('response_plan_dev', 'Response plan development', 'রেসপন্স প্ল্যান উন্নয়ন'),
        ('implementation_response', 'Implementation of response', 'রেসপন্স বাস্তবায়ন'),
        ('monitoring_evaluation', 'Monitoring and evaluation', 'পর্যবেক্ষণ ও মূল্যায়ন'),
    ]:
        ch.append(_ch('rp_subcat', k, en, bn))
    # Per-action implementation status (kept by request — not in the master).
    for k, en, bn in [
        ('implemented', 'Implemented', 'বাস্তবায়িত'),
        ('in_progress', 'In progress', 'চলমান'),
        ('pending',     'Pending / not started', 'অপেক্ষমাণ / শুরু হয়নি'),
        ('delayed',     'Delayed', 'বিলম্বিত'),
        ('dropped',     'Dropped', 'বাতিল'),
    ]:
        ch.append(_ch('rp_status', k, en, bn))
    # Which master table a single registered action belongs to (replaces the
    # three repeat groups; values match mpdsr.models.ActionSection).
    for k, en, bn in [
        ('system_strengthening', 'MPDSR System Strengthening',
         'এমপিডিএসআর সিস্টেম শক্তিশালীকরণ'),
        ('community_va', 'Common modifiable factors (Community verbal autopsy)',
         'সাধারণ পরিবর্তনযোগ্য কারণ (কমিউনিটি ভার্বাল অটোপসি)'),
        ('facility_dr', 'Common modifiable factors (Facility death review)',
         'সাধারণ পরিবর্তনযোগ্য কারণ (ফ্যাসিলিটি ডেথ রিভিউ)'),
    ]:
        ch.append(_ch('rp_section', k, en, bn))
    # rp_factor = the first column of master Table 2 (Common modifiable
    # factors), shown for BOTH the community-verbal-autopsy and the
    # facility-death-review sections. Wording follows the district action plans
    # as written; 'other' keeps the list open because the master template leaves
    # these rows blank for districts to fill.
    for k, en, bn in [
        ('quality_of_care',
         'Lack of quality maternal and newborn care service at facilities',
         'সুবিধাকেন্দ্রে মানসম্মত মাতৃ ও নবজাতক সেবার ঘাটতি'),
        ('pph_management', 'Management of Postpartum Haemorrhage (PPH)',
         'প্রসব-পরবর্তী রক্তক্ষরণ (পিপিএইচ) ব্যবস্থাপনা'),
        ('multiparity_htn_preeclampsia',
         'Multiparity, prolonged labour, HTN and pre-eclampsia',
         'বহুপ্রসব, দীর্ঘায়িত প্রসববেদনা, উচ্চ রক্তচাপ ও প্রি-এক্লাম্পসিয়া'),
        ('referral_linkages', 'Inadequate referral linkages',
         'রেফারেল সংযোগের অপর্যাপ্ততা'),
        ('delayed_anc', 'Delayed ANC initiation',
         'বিলম্বে এএনসি শুরু'),
        ('home_delivery_tba', 'Home delivery by TBA',
         'দাই (টিবিএ) দ্বারা বাড়িতে প্রসব'),
        ('death_reporting', 'Increase proper death reporting',
         'সঠিক মৃত্যু রিপোর্টিং বাড়ানো'),
        ('other', 'Other (write it below)', 'অন্যান্য (নিচে লিখুন)'),
    ]:
        ch.append(_ch('rp_factor', k, en, bn))
    for k, en, bn in [
        ('new_plan', 'Start a new action plan', 'নতুন অ্যাকশন প্ল্যান শুরু করুন'),
        ('update_action', 'Update an existing action', 'বিদ্যমান একটি পদক্ষেপ হালনাগাদ করুন'),
    ]:
        ch.append(_ch('ap_mode', k, en, bn))
    for pct in (0, 25, 50, 75, 100):
        ch.append(_ch('ap_completion', str(pct), '%d%%' % pct, '%d%%' % pct))
    return ch


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  FORM — CIPRB Fistula Campaign (Daily CHW Activity report)                ║
# ║  Source: "Mass Campaign on: End Obstetric Fistula in Bangladesh" —        ║
# ║  "Reporting on Daily activities of CHW". This is a DAILY activity/reach   ║
# ║  roll-up (households visited, population covered, staff/focal head-counts, ║
# ║  suspected/diagnosed/referral/surgery/rehab counts) — NOT individual      ║
# ║  patient registration. Each submission → one fistula.FistulaCampaign row  ║
# ║  (PENDING, single-stage CIPRB). Routed to handle_ciprb_fistula_campaign.  ║
# ╚══════════════════════════════════════════════════════════════════════════╝
def _fistula_campaign_survey():
    # Verbatim header notes from the paper form.
    rows = [
        _sr('note', '_hdr1',
            'Mass Campaign on: End Obstetric Fistula in Bangladesh',
            'বাংলাদেশে প্রসূতিজনিত ফিস্টুলা নির্মূলে গণ-অভিযান'),
        _sr('note', '_hdr2',
            'Reporting on Daily activities of CHW',
            'সিএইচডব্লিউ-এর দৈনিক কার্যক্রমের প্রতিবেদন'),
        # Hidden CIPRB org tag + required GPS (matches the other CIPRB forms).
        _sr('calculate', 'organisation', '', '', calc="'CIPRB'"),
        _sr('geopoint', 'location',
            'GPS location (required — step outside if no signal)',
            'জিপিএস অবস্থান (প্রয়োজনীয়)', required='yes'),
        # Enumerator (person filling the daily form).
        _sr('text', 'enumerator_name',
            'Your name (person filling this form)',
            'আপনার নাম (কে পূরণ করছেন)', required='yes'),
        _sr('text', 'enumerator_designation', 'Designation', 'পদবী'),
        _sr('text', 'enumerator_mobile', 'Mobile number', 'মোবাইল নম্বর',
            constraint='regex(., "^[0-9+ -]{6,20}$") or .=""',
            cmsg='Enter a valid phone number.'),
    ]

    def _n(name, en, bn):
        """A non-negative integer count field (>= 0 guard)."""
        return _sr('integer', name, en, bn,
                   constraint='. >= 0', cmsg='Must be 0 or more.')

    rows += [
        _sr('begin_group', 'grp_daily',
            'Daily CHW activity', 'দৈনিক সিএইচডব্লিউ কার্যক্রম'),
        # ── Location (verbatim English labels).
        _sr('date', 'collection_date', 'Date', 'তারিখ', required='yes'),
        _sr('text', 'union', 'Union (as per name like abc,bcd,asdf etc)',
            'ইউনিয়ন (নাম অনুযায়ী, যেমন abc,bcd,asdf ইত্যাদি)', required='yes'),
        _sr('text', 'upazila', 'Upazila', 'উপজেলা', required='yes'),
        _sr('select_one district', 'district', 'District', 'জেলা', required='yes'),
        # ── Staff head-counts (cadre abbreviations kept in Bangla too).
        _n('staff_hi_ahi', 'No of HI &AHI', 'HI ও AHI সংখ্যা'),
        _n('staff_ha',     'No of HA',      'HA সংখ্যা'),
        _n('staff_chcp',   'No of CHCP',    'CHCP সংখ্যা'),
        _n('staff_fwv',    'No of FWV',     'FWV সংখ্যা'),
        _n('staff_fpi',    'No of FPI',     'FPI সংখ্যা'),
        _n('staff_fwa',    'No of FWA',     'FWA সংখ্যা'),
        _n('staff_chw',    'No of CHW',     'CHW সংখ্যা'),
        # ── Community focal points.
        _n('focal_community', 'Focal point(Community)', 'ফোকাল পয়েন্ট (কমিউনিটি)'),
        _n('focal_epi',       'Focal point(EPI)',       'ফোকাল পয়েন্ট (EPI)'),
        _n('focal_fwc',       'Focal point(FWC)',       'ফোকাল পয়েন্ট (FWC)'),
        _n('focal_cc',        'Focal point(CC)',        'ফোকাল পয়েন্ট (CC)'),
        # ── Reach.
        _n('households_visited', '# of Households Visited', 'পরিদর্শিত পরিবারের সংখ্যা'),
        _n('population_covered', 'No of population covered', 'অন্তর্ভুক্ত জনসংখ্যা'),
        # ── Outcomes.
        _n('suspected_patients', '# of Suspected Patients', 'সন্দেহজনক রোগীর সংখ্যা'),
        _n('diagnosed_patients', '# of Diagnosed Patients', 'নির্ণীত রোগীর সংখ্যা'),
        _n('referral',           '# of Referral',           'রেফারেলের সংখ্যা'),
        _n('surgeries',          '# of Surgeries',          'অস্ত্রোপচারের সংখ্যা'),
        _n('rehabilitation',     '# of Rehabilitation',     'পুনর্বাসনের সংখ্যা'),
        _sr('end_group', 'grp_daily'),
    ]
    return rows


def _fistula_campaign_choices():
    # Only the district select_one is used on the daily form.
    return list(DISTRICT_CHOICES)


FORMS = [
    dict(file='CIPRB-1_Fistula_Question_Bank.xlsx',
         id='ciprb_fistula_questions_v1',
         title='CIPRB 1 — Fistula Question Bank',
         survey=_fistula_survey, choices=_fistula_choices),
    dict(file='CIPRB-1b_Fistula_Campaign.xlsx',
         id='ciprb_fistula_campaign_v1',
         title='CIPRB — Fistula Campaign (Daily CHW Activity)',
         survey=_fistula_campaign_survey, choices=_fistula_campaign_choices),
    dict(file='CIPRB-2_MPDSR_Form_01_Community_Maternal.xlsx',
         id='ciprb_mpdsr_community_maternal_v1',
         title='CIPRB 2 — MPDSR Form 01 (Community Maternal Death)',
         survey=_community_maternal_survey,
         choices=_community_maternal_choices),
    dict(file='CIPRB-3_MPDSR_Form_02_Community_Neonatal.xlsx',
         id='ciprb_mpdsr_community_neonatal_v1',
         title='CIPRB 3 — MPDSR Form 02 (Community Neonatal Death)',
         survey=_community_neonatal_survey,
         choices=_community_neonatal_choices),
    dict(file='CIPRB-4_MPDSR_Form_04_Facility_Maternal.xlsx',
         id='ciprb_mpdsr_facility_maternal_v1',
         title='CIPRB 4 — MPDSR Form 04 (Facility Maternal Death)',
         survey=_facility_maternal_survey,
         choices=_facility_maternal_choices),
    dict(file='CIPRB-5_MPDSR_Form_05_Facility_Neonatal.xlsx',
         id='ciprb_mpdsr_facility_neonatal_v1',
         title='CIPRB 5 — MPDSR Form 05 (Facility Neonatal Death)',
         survey=_facility_neonatal_survey,
         choices=_facility_neonatal_choices),
    dict(file='CIPRB-6_Social_Autopsy.xlsx',
         id='ciprb_social_autopsy_v1',
         title='CIPRB 6 — Social Autopsy (Maternal Death)',
         survey=_social_autopsy_survey,
         choices=_social_autopsy_choices),
    dict(file='CIPRB-7_Notification_Slip_01.xlsx',
         id='ciprb_notification_slip_01_v1',
         title='CIPRB 7 — Death Notification Slip 01',
         survey=lambda: _notification_slip_survey(1),
         choices=_notification_slip_choices),
    dict(file='CIPRB-8_Notification_Slip_02.xlsx',
         id='ciprb_notification_slip_02_v1',
         title='CIPRB 8 — Death Notification Slip 02',
         survey=lambda: _notification_slip_survey(2),
         choices=_notification_slip_choices),
    dict(file='CIPRB-9_Maternal_Near_Miss.xlsx',
         id='ciprb_near_miss_v1',
         title='CIPRB 9 — Maternal Near Miss audit',
         survey=_near_miss_survey, choices=_near_miss_choices),
    dict(file='CIPRB-10_MPDSR_Response_Plan.xlsx',
         id='ciprb_mpdsr_response_plan_v1',
         title='CIPRB 10 — MPDSR Response Plan',
         survey=_response_plan_survey, choices=_response_plan_choices),
]


# ─── Kobo upload helper (mirrors export_phd_clients) ────────────────────────

def _kobo_token():
    return (getattr(settings, 'KOBO_API_TOKEN', '')
            or os.environ.get('KOBO_TOKEN', '')).strip()


def _import_xlsform(xlsx_path: str, asset_uid: str, token: str, stdout):
    """Replace the survey XML in an existing Kobo asset by importing the xlsx
    into it.  Returns True on success."""
    headers = {'Authorization': f'Token {token}'}
    api = f'{KOBO_BASE}/api/v2'
    with open(xlsx_path, 'rb') as fh:
        files = {'file': (os.path.basename(xlsx_path), fh,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'destination': f'{api}/assets/{asset_uid}/', 'library': 'false'}
        r = requests.post(f'{api}/imports/', headers=headers,
                          files=files, data=data, timeout=120)
    if r.status_code not in (200, 201):
        stdout.write(f'    import FAILED ({r.status_code}): {r.text[:200]}')
        return False
    stdout.write('    imported')
    return True


def _create_asset(form_id: str, form_title: str, token: str, stdout):
    """Create a fresh KoboToolbox asset for a CIPRB form and return its uid."""
    headers = {'Authorization': f'Token {token}'}
    api = f'{KOBO_BASE}/api/v2'
    body = {
        'name':       form_title,
        'asset_type': 'survey',
        'settings':   {'description': f'CIPRB form {form_id}'},
    }
    r = requests.post(f'{api}/assets/', headers=headers, json=body, timeout=60)
    if r.status_code not in (200, 201):
        stdout.write(f'    create FAILED ({r.status_code}): {r.text[:200]}')
        return None
    return r.json().get('uid')


def _deploy(asset_uid: str, token: str, stdout):
    headers = {'Authorization': f'Token {token}'}
    api = f'{KOBO_BASE}/api/v2'
    # latest version_id
    v = requests.get(f'{api}/assets/{asset_uid}/versions/?limit=1',
                     headers=headers, timeout=30)
    try:
        vhash = v.json()['results'][0]['uid']
    except Exception:
        stdout.write('    no version yet — skipping deploy')
        return False
    r = requests.patch(
        f'{api}/assets/{asset_uid}/deployment/',
        headers=headers,
        json={'version_id': vhash, 'active': True},
        timeout=60,
    )
    if r.status_code in (200, 201):
        stdout.write('    deployed')
        return True
    # First-time deploy may need POST.
    r2 = requests.post(
        f'{api}/assets/{asset_uid}/deployment/',
        headers=headers,
        json={'version_id': vhash, 'active': True},
        timeout=60,
    )
    if r2.status_code in (200, 201):
        stdout.write('    deployed (POST)')
        return True
    stdout.write(f'    deploy FAILED ({r.status_code}/{r2.status_code}): {r2.text[:160]}')
    return False


class Command(BaseCommand):
    help = 'Build the 9 CIPRB XLSForms (Fistula Question Bank, MPDSR 1/2/4/5, Social Autopsy, 2 notification slips, Maternal Near Miss). Pass --upload to push to Kobo.'

    def add_arguments(self, parser):
        parser.add_argument('--output-dir', default=OUTDIR)
        parser.add_argument('--upload', action='store_true',
            help='Upload to Kobo, create assets when missing, deploy.')
        parser.add_argument('--only', default='',
            help='Build/deploy ONLY the form with this id (e.g. '
                 'ciprb_mpdsr_response_plan_v1). Avoids touching the others.')

    def handle(self, *args, **opts):
        out = opts['output_dir']
        os.makedirs(out, exist_ok=True)
        token = _kobo_token() if opts['upload'] else ''

        if opts['upload'] and not token:
            self.stdout.write(self.style.ERROR(
                'KOBO_TOKEN not set — cannot --upload.'))
            return

        only = opts.get('only', '')
        for f in FORMS:
            if only and f['id'] != only:
                continue
            survey  = f['survey']()
            choices = f['choices']()
            wb = _wb(f['id'], f['title'], survey, choices)
            path = os.path.join(out, f['file'])
            wb.save(path)
            self.stdout.write(self.style.SUCCESS(
                f"  OK  {f['file']:55s}  {len(survey):3d} rows  id: {f['id']}"))

            if opts['upload']:
                # Look up existing asset by form_id (id_string), create if absent.
                self.stdout.write('     uploading…')
                api = f'{KOBO_BASE}/api/v2'
                headers = {'Authorization': f'Token {token}'}
                q = requests.get(
                    f'{api}/assets/?q=settings__id_string:{f["id"]}',
                    headers=headers, timeout=30).json()
                asset_uid = None
                for a in q.get('results', []):
                    if a.get('settings', {}).get('id_string') == f['id']:
                        asset_uid = a.get('uid')
                        break
                if not asset_uid:
                    asset_uid = _create_asset(f['id'], f['title'], token, self.stdout)
                if not asset_uid:
                    continue
                ok = _import_xlsform(path, asset_uid, token, self.stdout)
                if ok:
                    _deploy(asset_uid, token, self.stdout)
                self.stdout.write(f'     uid: {asset_uid}')

        self.stdout.write(f'\nWritten to {os.path.abspath(out)}/')
