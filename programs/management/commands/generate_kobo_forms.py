"""
Management command: generate_kobo_forms

Generates all 17 KoboToolbox XLSForm (.xlsx) files for the Spondon IDMS
programmes data-collection forms.

Usage:
    python manage.py generate_kobo_forms [--output-dir koboforms/]

Output: one .xlsx file per form in XLSForm format, ready to upload to
https://kf.kobotoolbox.org

Field name conventions (must match programs/webhook.py exactly):
  - partner_org: "PHD" or "Bandhu" — skip logic uses this
  - center_code: center.code value from ServiceCenter model
  - GPS: `_geolocation` is auto-added by KoboToolbox from the `location` geopoint
  - client_id: links to Client.client_id in the database

All forms produced here target webhook endpoint:
  POST https://web-production-091fa.up.railway.app/webhook/programs/
  Authorization: Token REDACTED
"""
import os
from datetime import date

from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from programs.models import ServiceCenter


# ─── XLSForm builder helpers ───────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="003F72")   # UNFPA dark blue
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_ROW_FILL_A = PatternFill("solid", fgColor="EAF3FB")    # light blue zebra
_ROW_FILL_B = PatternFill("solid", fgColor="FFFFFF")


def _wb_create():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    return wb


def _add_sheet(wb, name, headers, rows, col_widths=None):
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    # Style header row
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        fill = _ROW_FILL_A if row_idx % 2 == 0 else _ROW_FILL_B
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Column widths
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        if col_widths and col_idx <= len(col_widths):
            ws.column_dimensions[letter].width = col_widths[col_idx - 1]
        else:
            ws.column_dimensions[letter].width = 24
    ws.freeze_panes = 'A2'
    return ws


SURVEY_HEADERS = [
    'type', 'name', 'label::English', 'label::Bangla',
    'hint', 'required', 'relevant', 'constraint', 'constraint_message',
    'default', 'appearance', 'calculation',
]
CHOICES_HEADERS = ['list_name', 'name', 'label::English', 'label::Bangla']
SETTINGS_HEADERS = ['form_title', 'form_id', 'version', 'default_language']


def _survey_row(
    qtype, name, label_en, label_bn='', hint='', required='',
    relevant='', constraint='', constraint_msg='', default='',
    appearance='', calculation='',
):
    return [
        qtype, name, label_en, label_bn,
        hint, required, relevant, constraint, constraint_msg,
        default, appearance, calculation,
    ]


def _choice(list_name, name, label_en, label_bn=''):
    return [list_name, name, label_en, label_bn]


def _build_wb(form_id: str, form_title: str, survey_rows, choices_rows):
    wb = _wb_create()
    _add_sheet(wb, 'survey', SURVEY_HEADERS, survey_rows,
               col_widths=[20, 28, 35, 35, 30, 8, 35, 30, 30, 12, 14, 20])
    _add_sheet(wb, 'choices', CHOICES_HEADERS, choices_rows,
               col_widths=[20, 20, 30, 30])
    _add_sheet(wb, 'settings', SETTINGS_HEADERS, [
        [form_title, form_id, date.today().strftime('%Y%m%d'), 'English']
    ], col_widths=[40, 35, 12, 14])
    return wb


# ─── Shared choice lists ───────────────────────────────────────────────────────

def _common_choices():
    rows = []

    # yes_no
    rows += [
        _choice('yes_no', 'yes', 'Yes', 'হ্যাঁ'),
        _choice('yes_no', 'no', 'No', 'না'),
    ]
    # partner_org
    rows += [
        _choice('partner_org', 'PHD', 'PHD', 'পিএইচডি'),
        _choice('partner_org', 'Bandhu', 'Bandhu', 'বন্ধু'),
    ]
    # visit_type
    rows += [
        _choice('visit_type', 'new', 'New', 'নতুন'),
        _choice('visit_type', 'follow_up', 'Follow-Up', 'ফলো-আপ'),
        _choice('visit_type', 'recurrent', 'Recurrent', 'পুনরাবর্তী'),
    ]
    # hiv_result
    for v, en, bn in [
        ('positive', 'Positive', 'পজিটিভ'),
        ('negative', 'Negative', 'নেগেটিভ'),
        ('indeterminate', 'Indeterminate', 'অনির্ধারিত'),
        ('not_done', 'Not Done', 'করা হয়নি'),
    ]:
        rows.append(_choice('hiv_result', v, en, bn))
    # session_type (HTC)
    rows += [
        _choice('session_type', 'pre', 'Pre-Test Counselling', 'পরীক্ষা-পূর্ব পরামর্শ'),
        _choice('session_type', 'post', 'Post-Test Counselling', 'পরীক্ষা-পরবর্তী পরামর্শ'),
        _choice('session_type', 'ongoing', 'Ongoing Counselling', 'চলমান পরামর্শ'),
    ]
    # screening_type (MH)
    rows += [
        _choice('screening_type', 'depression', 'Depression (Zahiruddin Scale)', 'বিষণ্নতা (জহিরউদ্দিন স্কেল)'),
        _choice('screening_type', 'ptsd', 'PTSD', 'পিটিএসডি'),
    ]
    # severity
    for v, en, bn in [
        ('none', 'No / Minimal', 'নেই / সামান্য'),
        ('mild', 'Mild', 'মৃদু'),
        ('moderate', 'Moderate', 'মাঝারি'),
        ('severe', 'Severe', 'তীব্র'),
        ('extreme', 'Extreme', 'অত্যন্ত তীব্র'),
        ('profound', 'Profound', 'গভীর'),
    ]:
        rows.append(_choice('severity', v, en, bn))
    # log_type (autoclave)
    rows += [
        _choice('log_type', 'autoclave', 'Autoclave', 'অটোক্লেভ'),
        _choice('log_type', 'incinerator', 'Incinerator', 'ইনসিনারেটর'),
    ]
    # referral_type
    for v, en in [
        ('hiv', 'HIV Testing'), ('art', 'ART Enrollment'),
        ('sti_kp', 'STI (KP)'), ('sti_partner', 'STI (Partner)'),
        ('tb', 'TB'), ('gbv', 'GBV Services'),
        ('mental_health', 'Mental Health'), ('srhr', 'SRHR'),
        ('fp', 'Family Planning'), ('general_health', 'General Health'),
        ('hep_c', 'Hepatitis C'), ('legal', 'Legal Services'),
        ('shelter', 'Shelter'), ('child', 'Child Health'),
        ('maternal', 'Maternal Health'), ('diabetic', 'Diabetic'),
        ('other', 'Other'),
    ]:
        rows.append(_choice('referral_type', v, en, ''))
    # event_type (training)
    rows += [
        _choice('event_type', 'orientation', 'Orientation', 'ওরিয়েন্টেশন'),
        _choice('event_type', 'training', 'Training', 'প্রশিক্ষণ'),
        _choice('event_type', 'workshop', 'Workshop', 'কর্মশালা'),
    ]
    # participant_type (training)
    for v, en in [
        ('HM', 'Health Managers (UHC/DGHS/DGFP)'),
        ('MW', 'Midwives / Medical Assistants'),
        ('PE', 'Peer Educators'),
        ('CL', 'Community Leaders'),
        ('GOB', 'District / Upazila GOB Staff'),
        ('MIXED', 'Mixed'),
    ]:
        rows.append(_choice('participant_type', v, en, ''))
    # meeting_type
    rows += [
        _choice('meeting_type', 'GOB', 'GOB / Health Staff', 'সরকারি / স্বাস্থ্যকর্মী'),
        _choice('meeting_type', 'CBO', 'CBO / Community Network', 'সিবিও / সম্প্রদায় নেটওয়ার্ক'),
        _choice('meeting_type', 'internal', 'Internal', 'অভ্যন্তরীণ'),
        _choice('meeting_type', 'multi', 'Multi-Stakeholder', 'বহু-অংশীদার'),
        _choice('meeting_type', 'day_observance', 'Day Observance / Awareness Event', 'দিবস পালন / সচেতনতা অনুষ্ঠান'),
    ]
    # fistula_type — V.V.F / R.V.F / both / other (FistulaCornerCase)
    rows += [
        _choice('fistula_type', 'VVF',   'V.V.F (Vesico-Vaginal)',          'ভি.ভি.এফ'),
        _choice('fistula_type', 'RVF',   'R.V.F (Recto-Vaginal)',           'আর.ভি.এফ'),
        _choice('fistula_type', 'BOTH',  'V.V.F + R.V.F (Combined)',         'ভি.ভি.এফ + আর.ভি.এফ'),
        _choice('fistula_type', 'OTHER', 'Other',                            'অন্যান্য'),
    ]
    # surgery_performed tri-state
    rows += [
        _choice('surgery_performed', 'yes',     'Yes',     'হ্যাঁ'),
        _choice('surgery_performed', 'no',      'No',      'না'),
        _choice('surgery_performed', 'pending', 'Pending', 'অপেক্ষমাণ'),
    ]
    # delivery_mode (campaign visit)
    rows += [
        _choice('delivery_mode', 'home',     'Home',     'বাড়িতে'),
        _choice('delivery_mode', 'facility', 'Facility', 'প্রতিষ্ঠানে'),
        _choice('delivery_mode', 'other',    'Other',    'অন্যান্য'),
    ]
    # delivery_outcome (campaign visit)
    rows += [
        _choice('delivery_outcome', 'LB',  'Live Birth',  'জীবিত জন্ম'),
        _choice('delivery_outcome', 'SB',  'Still Birth', 'মৃত জন্ম'),
        _choice('delivery_outcome', 'UNK', 'Unknown',     'অজানা'),
    ]
    # (yes_no choices already defined at the top of this function — fistula
    # forms reuse them; no duplicate entry needed.)
    # gender
    rows += [
        _choice('gender', '02', 'Female', 'মহিলা'),
        _choice('gender', '01', 'Male', 'পুরুষ'),
        _choice('gender', '03', 'Transgender', 'তৃতীয় লিঙ্গ'),
        _choice('gender', '04', 'Other', 'অন্যান্য'),
    ]
    # target_group
    rows += [
        _choice('target_group', '05', 'FSW', 'এফএসডব্লিউ'),
        _choice('target_group', '06', 'PWID', 'পিডব্লিউআইডি'),
        _choice('target_group', '01', 'MSM', 'এমএসএম'),
        _choice('target_group', '02', 'MSW / Kothi', 'এমএসডব্লিউ / কথি'),
        _choice('target_group', '03', 'Transgender', 'তৃতীয় লিঙ্গ'),
        _choice('target_group', '04', 'Others', 'অন্যান্য'),
    ]
    # survivor gender
    rows += [
        _choice('survivor_gender', 'female', 'Female', 'মহিলা'),
        _choice('survivor_gender', 'male', 'Male', 'পুরুষ'),
        _choice('survivor_gender', 'transgender', 'Transgender', 'তৃতীয় লিঙ্গ'),
        _choice('survivor_gender', 'other', 'Other', 'অন্যান্য'),
    ]
    return rows


def _center_choices():
    """Generate choices for all active ServiceCenters.

    Emits a placeholder row when the DB has no active centers — XLSForm
    rejects forms whose select_one list is empty ("List name not in
    choices sheet"). The placeholder is replaced by real center_code
    values once ServiceCenter rows are seeded for the partner.
    """
    rows = []
    for center in ServiceCenter.objects.filter(is_active=True).order_by('organisation', 'name'):
        rows.append(_choice('center_code', center.code, center.name, center.name_bangla or center.name))
    if not rows:
        rows.append(_choice(
            'center_code', 'PLACEHOLDER',
            'Placeholder — seed ServiceCenters before deployment',
            'প্লেসহোল্ডার — প্রকৃত কেন্দ্র যোগ করুন',
        ))
    return rows


def _common_metadata_rows():
    """
    Start-of-form rows shared by every form:
    - GPS location (required)
    - Partner org (required, choose PHD / Bandhu)
    - Service centre (required, filtered by partner_org)
    """
    return [
        _survey_row('begin_group', 'grp_meta', 'Submission Metadata', 'তথ্য প্রেরণের বিবরণ'),
        _survey_row(
            'geopoint', 'location',
            'GPS Location (required — step outside if no signal)',
            'জিপিএস অবস্থান (প্রয়োজনীয়)',
            required='yes',
        ),
        _survey_row(
            'select_one partner_org', 'partner_org',
            'Organisation', 'সংগঠন',
            required='yes',
        ),
        _survey_row(
            'select_one center_code', 'center_code',
            'Service Centre', 'সেবাকেন্দ্র',
            hint='Select your service centre',
            required='yes',
        ),
        _survey_row('end_group', 'grp_meta', '', ''),
    ]


# ─── Individual form definitions ────────────────────────────────────────────────

def _form_client_registration():
    """KF-01: Mother List / Client Registration."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('begin_group', 'grp_client', 'Client Information', 'ক্লায়েন্টের তথ্য'),
        _survey_row('text', 'client_id', 'Client ID / Register No.', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম', required='yes'),
        _survey_row('text', 'mother_name', 'Mother\'s Name', 'মায়ের নাম'),
        _survey_row('text', 'father_name', 'Father\'s Name / Husband\'s Name', 'বাবার নাম / স্বামীর নাম'),
        _survey_row('integer', 'birth_year', 'Birth Year', 'জন্ম সাল',
                    constraint='. >= 1940 and . <= 2010',
                    constraint_msg='Enter a valid birth year (1940–2010)'),
        _survey_row('select_one gender', 'gender', 'Gender', 'লিঙ্গ', required='yes'),
        _survey_row('select_one target_group', 'target_group_code', 'Target Group (KP)', 'লক্ষ্য গোষ্ঠী', required='yes'),
        _survey_row('text', 'current_address', 'Current Address', 'বর্তমান ঠিকানা', appearance='multiline'),
        _survey_row('text', 'spot_name', 'Spot / Location Name', 'স্পট / এলাকার নাম'),
        _survey_row('select_one yes_no', 'uses_fp_method', 'Uses Family Planning Method?', 'পরিবার পরিকল্পনা পদ্ধতি ব্যবহার করেন?'),
        _survey_row('select_one yes_no', 'has_nid', 'Has NID?', 'এনআইডি আছে?'),
        _survey_row('date', 'enrolled_date', 'Enrollment Date', 'তালিকাভুক্তির তারিখ', required='yes'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
        _survey_row('end_group', 'grp_client', '', ''),
    ]
    return rows


def _form_clinic_visit():
    """KF-02: Patient Record Register / Clinic Visit."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('begin_group', 'grp_patient', 'Patient Information', 'রোগীর তথ্য'),
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name (for reference)', 'ক্লায়েন্টের নাম'),
        _survey_row('date', 'visit_date', 'Visit Date', 'পরিদর্শনের তারিখ', required='yes'),
        _survey_row('select_one visit_type', 'visit_type', 'Visit Type', 'পরিদর্শনের ধরন', required='yes'),
        _survey_row('text', 'monthly_serial', 'Monthly Serial No.', 'মাসিক ক্রমিক নং'),
        _survey_row('end_group', 'grp_patient', '', ''),

        _survey_row('begin_group', 'grp_screening', 'Screenings Done', 'স্ক্রিনিং সম্পন্ন'),
        _survey_row('select_one yes_no', 'sti_screening_done', 'STI Screening Done?', 'এসটিআই স্ক্রিনিং?'),
        _survey_row('select_one yes_no', 'hiv_screening_done', 'HIV Screening Done?', 'এইচআইভি স্ক্রিনিং?'),
        _survey_row('select_one yes_no', 'tb_screening_done', 'TB Screening Done?', 'যক্ষ্মা স্ক্রিনিং?'),
        _survey_row('select_one yes_no', 'diabetic_screening_done', 'Diabetic Screening Done?', 'ডায়াবেটিস স্ক্রিনিং?'),
        _survey_row('select_one yes_no', 'hep_b_screening_done', 'Hepatitis B Screening Done?', 'হেপাটাইটিস বি স্ক্রিনিং?',
                    relevant="${partner_org} = 'PHD'"),
        _survey_row('select_one yes_no', 'hep_c_screening_done', 'Hepatitis C Screening Done?', 'হেপাটাইটিস সি স্ক্রিনিং?',
                    relevant="${partner_org} = 'PHD'"),
        _survey_row('end_group', 'grp_screening', '', ''),

        _survey_row('begin_group', 'grp_diag', 'STI Diagnosis', 'এসটিআই নির্ণয়'),
        _survey_row('select_one yes_no', 'diag_uds', 'UDS (Urethral Discharge Syndrome)', 'ইউডিএস'),
        _survey_row('select_one yes_no', 'diag_vds', 'VDS (Vaginal Discharge Syndrome)', 'ভিডিএস'),
        _survey_row('select_one yes_no', 'diag_gu', 'GU (Genital Ulcer)', 'জিইউ'),
        _survey_row('select_one yes_no', 'diag_pid', 'PID (Pelvic Inflammatory Disease)', 'পিআইডি'),
        _survey_row('select_one yes_no', 'diag_ss', 'SS (Scrotal Swelling)', 'এসএস'),
        _survey_row('select_one yes_no', 'diag_ib', 'IB (Inguinal Bubo)', 'আইবি'),
        _survey_row('select_one yes_no', 'diag_anal_sti', 'Anal STI', 'অ্যানাল এসটিআই'),
        _survey_row('select_one yes_no', 'diag_gh', 'General Health Condition', 'সাধারণ স্বাস্থ্য সমস্যা'),
        _survey_row('select_one yes_no', 'diag_psd', 'Psychosexual Disorder', 'মনোযৌন ব্যাধি'),
        _survey_row('select_one yes_no', 'diag_mental_health', 'Mental Health Condition', 'মানসিক স্বাস্থ্য সমস্যা'),
        _survey_row('select_one yes_no', 'diag_other', 'Other Diagnosis', 'অন্যান্য নির্ণয়'),
        _survey_row('text', 'diag_other_specify', 'Specify Other Diagnosis', 'অন্যান্য নির্দিষ্ট করুন',
                    relevant="${diag_other} = 'yes'"),
        _survey_row('end_group', 'grp_diag', '', ''),

        _survey_row('begin_group', 'grp_treatment', 'Treatment', 'চিকিৎসা'),
        _survey_row('text', 'treatment_provided', 'Treatment Provided', 'প্রদত্ত চিকিৎসা', appearance='multiline'),
        _survey_row('integer', 'condom_demo_sessions', 'Condom Demo Sessions', 'কনডম প্রদর্শনী'),
        _survey_row('integer', 'condoms_distributed', 'Condoms Distributed', 'বিতরণকৃত কনডম'),
        _survey_row('select_one yes_no', 'sti_counselling_provided', 'STI Counselling Provided?', 'এসটিআই পরামর্শ দেওয়া হয়েছে?'),
        _survey_row('text', 'partner_management', 'Partner Management', 'সঙ্গী ব্যবস্থাপনা'),
        _survey_row('end_group', 'grp_treatment', '', ''),

        _survey_row('begin_group', 'grp_referrals', 'Referrals Made', 'প্রেরণ করা হয়েছে'),
        _survey_row('select_one yes_no', 'referral_tb', 'TB Referral?', 'যক্ষ্মার জন্য রেফার?'),
        _survey_row('select_one yes_no', 'referral_sti_kp', 'STI Referral (KP)?', 'এসটিআই রেফার (KP)?'),
        _survey_row('select_one yes_no', 'referral_sti_partner', 'STI Referral (Partner)?', 'এসটিআই রেফার (সঙ্গী)?'),
        _survey_row('select_one yes_no', 'referral_general_health', 'General Health Referral?', 'সাধারণ স্বাস্থ্য রেফার?'),
        _survey_row('select_one yes_no', 'referral_hiv_testing', 'HIV Testing Referral?', 'এইচআইভি পরীক্ষার রেফার?'),
        _survey_row('select_one yes_no', 'referral_mental_health', 'Mental Health Referral?', 'মানসিক স্বাস্থ্য রেফার?'),
        _survey_row('select_one yes_no', 'referral_diabetic', 'Diabetic Referral?', 'ডায়াবেটিস রেফার?'),
        _survey_row('select_one yes_no', 'referral_fp', 'Family Planning Referral?', 'পরিবার পরিকল্পনা রেফার?'),
        _survey_row('end_group', 'grp_referrals', '', ''),

        _survey_row('begin_group', 'grp_followup', 'Follow-Up & Other', 'ফলো-আপ ও অন্যান্য'),
        _survey_row('date', 'follow_up_due_date', 'Follow-Up Due Date', 'ফলো-আপের নির্ধারিত তারিখ'),
        _survey_row('select_one yes_no', 'adr_monitoring', 'ADR Monitoring Required?', 'এডিআর মনিটরিং প্রয়োজন?'),
        _survey_row('text', 'pregnancy_status', 'Pregnancy Status', 'গর্ভাবস্থার অবস্থা',
                    relevant="${partner_org} = 'PHD'"),
        _survey_row('text', 'anc_status', 'ANC Status', 'এএনসি অবস্থা',
                    relevant="${partner_org} = 'PHD'"),
        _survey_row('text', 'prepared_by', 'Prepared By (Name & Designation)', 'প্রস্তুতকারীর নাম ও পদবি'),
        _survey_row('end_group', 'grp_followup', '', ''),
    ]
    return rows


def _form_hiv_sti_test():
    """KF-03: HIV/STI Test Result."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('date', 'testing_date', 'Testing Date', 'পরীক্ষার তারিখ', required='yes'),
        _survey_row('text', 'lab_id', 'Lab Sample ID', 'ল্যাব স্যাম্পল আইডি'),
        _survey_row('select_one hiv_result', 'hiv_result', 'HIV Result', 'এইচআইভি ফলাফল', required='yes'),
        _survey_row('select_one hiv_result', 'syphilis_result', 'Syphilis Result', 'সিফিলিস ফলাফল', required='yes'),
        _survey_row('select_one hiv_result', 'hep_b_result', 'Hepatitis B Result', 'হেপাটাইটিস বি ফলাফল'),
        _survey_row('select_one hiv_result', 'hep_c_result', 'Hepatitis C Result', 'হেপাটাইটিস সি ফলাফল'),
        _survey_row('select_one yes_no', 'in_window_period', 'Client in Window Period?', 'ক্লায়েন্ট উইন্ডো পিরিয়ডে?'),
        _survey_row('date', 'retest_date', 'Retest Date (if applicable)', 'পুনরায় পরীক্ষার তারিখ',
                    relevant="${in_window_period} = 'yes'"),
        _survey_row('text', 'art_linkage_status', 'ART Linkage Status', 'এআরটি লিংকেজ অবস্থা',
                    relevant="${hiv_result} = 'positive'"),
        _survey_row('text', 'counsellor_name', 'Counsellor Name', 'পরামর্শদাতার নাম'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_adr_record():
    """KF-13: Adverse Drug Reaction."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('date', 'report_date', 'Report Date', 'রিপোর্টের তারিখ', required='yes'),
        _survey_row('text', 'drugs_given', 'Drug(s) Given', 'প্রদত্ত ওষুধ', required='yes', appearance='multiline'),
        _survey_row('select_one yes_no', 'adverse_effect_present', 'Adverse Effect Present?', 'পার্শ্বপ্রতিক্রিয়া আছে?', required='yes'),
        _survey_row('text', 'adverse_effect_description', 'Describe Adverse Effect', 'পার্শ্বপ্রতিক্রিয়ার বিবরণ',
                    relevant="${adverse_effect_present} = 'yes'", appearance='multiline'),
        _survey_row('date', 'followup_date', 'Follow-Up Date', 'ফলো-আপের তারিখ'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_autoclave_log():
    """KF-16: Autoclave / Incinerator Log."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('date', 'log_date', 'Log Date', 'লগের তারিখ', required='yes'),
        _survey_row('select_one log_type', 'log_type', 'Type', 'ধরন', required='yes'),
        # Autoclave fields
        _survey_row('begin_group', 'grp_autoclave', 'Autoclave Details', 'অটোক্লেভের বিবরণ',
                    relevant="${log_type} = 'autoclave'"),
        _survey_row('text', 'items_autoclaved', 'Items Autoclaved', 'অটোক্লেভ করা সামগ্রী', appearance='multiline'),
        _survey_row('select_one yes_no', 'temp_121_achieved', '121°C Temperature Achieved?', '১২১°সে তাপমাত্রা অর্জিত?'),
        _survey_row('select_one yes_no', 'tape_test_passed', 'Tape Test Passed?', 'টেপ পরীক্ষায় উত্তীর্ণ?'),
        _survey_row('text', 'done_by', 'Done By', 'সম্পাদনকারী'),
        _survey_row('end_group', 'grp_autoclave', '', ''),
        # Incinerator fields
        _survey_row('begin_group', 'grp_incinerator', 'Incinerator Details', 'ইনসিনারেটরের বিবরণ',
                    relevant="${log_type} = 'incinerator'"),
        _survey_row('text', 'material_type', 'Material Type', 'উপকরণের ধরন'),
        _survey_row('text', 'quantity', 'Quantity', 'পরিমাণ'),
        _survey_row('text', 'supervised_by', 'Supervised By', 'তত্ত্বাবধায়ক'),
        _survey_row('end_group', 'grp_incinerator', '', ''),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_antenatal_card():
    """PHD only: ANC Visit Record."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('date', 'visit_date', 'Visit Date', 'পরিদর্শনের তারিখ', required='yes'),
        _survey_row('integer', 'anc_visit_number', 'ANC Visit Number (1–4)', 'এএনসি পরিদর্শন নং',
                    required='yes',
                    constraint='. >= 1 and . <= 8',
                    constraint_msg='Enter visit number 1–8'),
        _survey_row('text', 'trimester', 'Trimester (1st/2nd/3rd)', 'ত্রৈমাসিক'),
        _survey_row('date', 'lmp_date', 'LMP Date', 'শেষ মাসিকের তারিখ'),
        _survey_row('date', 'edd', 'Expected Delivery Date', 'প্রত্যাশিত প্রসবের তারিখ'),
        _survey_row('text', 'blood_pressure', 'Blood Pressure (e.g. 120/80)', 'রক্তচাপ'),
        _survey_row('decimal', 'weight_kg', 'Weight (kg)', 'ওজন (কেজি)',
                    constraint='. > 20 and . < 200',
                    constraint_msg='Enter weight in kg (20–200)'),
        _survey_row('select_one yes_no', 'referred', 'Referred?', 'রেফার করা হয়েছে?'),
        _survey_row('text', 'referred_to', 'Referred To', 'কোথায় রেফার',
                    relevant="${referred} = 'yes'"),
        _survey_row('text', 'prepared_by', 'Prepared By', 'প্রস্তুতকারী'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_htc_counselling():
    """KF-04: HIV Testing & Counselling."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('select_one session_type', 'session_type', 'Session Type', 'সেশনের ধরন', required='yes'),
        _survey_row('date', 'session_date', 'Session Date', 'সেশনের তারিখ', required='yes'),
        _survey_row('integer', 'age_at_session', 'Client Age', 'ক্লায়েন্টের বয়স'),
        _survey_row('begin_group', 'grp_risk', 'Risk Assessment', 'ঝুঁকি মূল্যায়ন'),
        _survey_row('text', 'partner_count', 'No. of Sexual Partners', 'যৌন সঙ্গীর সংখ্যা'),
        _survey_row('text', 'condom_use', 'Condom Use Frequency', 'কনডম ব্যবহারের ফ্রিকোয়েন্সি'),
        _survey_row('select_one yes_no', 'needle_sharing', 'Needle Sharing?', 'সূঁচ শেয়ার করেন?'),
        _survey_row('select_one yes_no', 'blood_transfusion', 'Blood Transfusion History?', 'রক্ত সংক্রমণের ইতিহাস?'),
        _survey_row('text', 'partner_hiv_positive', 'Partner HIV Positive?', 'সঙ্গী এইচআইভি পজিটিভ?'),
        _survey_row('select_one yes_no', 'client_pregnant', 'Client Pregnant?', 'ক্লায়েন্ট গর্ভবতী?'),
        _survey_row('text', 'pregnancy_trimester', 'Pregnancy Trimester', 'গর্ভাবস্থার ত্রৈমাসিক',
                    relevant="${client_pregnant} = 'yes'"),
        _survey_row('end_group', 'grp_risk', '', ''),
        _survey_row('begin_group', 'grp_checklist', 'Counsellor Checklist', 'পরামর্শদাতার চেকলিস্ট'),
        _survey_row('select_one yes_no', 'covered_hiv_sti_prevention', 'HIV/STI Prevention Covered?', 'এইচআইভি/এসটিআই প্রতিরোধ আলোচনা হয়েছে?'),
        _survey_row('select_one yes_no', 'covered_risk_assessment', 'Risk Assessment Covered?', 'ঝুঁকি মূল্যায়ন আলোচনা হয়েছে?'),
        _survey_row('select_one yes_no', 'covered_behavior_change', 'Behaviour Change Covered?', 'আচরণ পরিবর্তন আলোচনা হয়েছে?'),
        _survey_row('select_one yes_no', 'covered_support_systems', 'Support Systems Covered?', 'সহায়তা ব্যবস্থা আলোচনা হয়েছে?'),
        _survey_row('select_one yes_no', 'client_consented', 'Client Consented?', 'ক্লায়েন্ট সম্মতি দিয়েছেন?', required='yes'),
        _survey_row('end_group', 'grp_checklist', '', ''),
        _survey_row('text', 'counsellor_name', 'Counsellor Name', 'পরামর্শদাতার নাম'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_individual_counselling():
    """KF-09: Individual Counselling Session."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('date', 'session_date', 'Session Date', 'সেশনের তারিখ', required='yes'),
        _survey_row('text', 'counsellor_name', 'Counsellor Name', 'পরামর্শদাতার নাম'),
        _survey_row('begin_group', 'grp_issues', 'Issues Discussed', 'আলোচিত বিষয়'),
        _survey_row('select_one yes_no', 'issue_sti', 'STI Issue?', 'এসটিআই সমস্যা?'),
        _survey_row('select_one yes_no', 'issue_general_health', 'General Health Issue?', 'সাধারণ স্বাস্থ্য সমস্যা?'),
        _survey_row('select_one yes_no', 'issue_fp', 'Family Planning Issue?', 'পরিবার পরিকল্পনা সমস্যা?'),
        _survey_row('select_one yes_no', 'issue_drug_use', 'Drug Use Issue?', 'মাদক ব্যবহার সমস্যা?'),
        _survey_row('select_one yes_no', 'issue_psychosocial', 'Psychosocial Issue?', 'মনোসামাজিক সমস্যা?'),
        _survey_row('select_one yes_no', 'issue_gbv', 'GBV Issue?', 'জিবিভি সমস্যা?'),
        _survey_row('select_one yes_no', 'issue_other', 'Other Issue?', 'অন্যান্য সমস্যা?'),
        _survey_row('end_group', 'grp_issues', '', ''),
        _survey_row('integer', 'condom_distributed', 'Condoms Distributed', 'বিতরণকৃত কনডম'),
        _survey_row('integer', 'iec_materials', 'IEC Materials Distributed', 'আইইসি উপকরণ বিতরণ'),
        _survey_row('begin_group', 'grp_referrals', 'Referrals Made', 'রেফার করা হয়েছে'),
        _survey_row('select_one yes_no', 'referral_mental_health', 'Mental Health Referral?', 'মানসিক স্বাস্থ্য রেফার?'),
        _survey_row('select_one yes_no', 'referral_legal', 'Legal Services Referral?', 'আইনি সেবা রেফার?'),
        _survey_row('select_one yes_no', 'referral_htc', 'HTC Referral?', 'এইচটিসি রেফার?'),
        _survey_row('select_one yes_no', 'referral_gbv', 'GBV Services Referral?', 'জিবিভি সেবা রেফার?'),
        _survey_row('end_group', 'grp_referrals', '', ''),
        _survey_row('select_one yes_no', 'drug_habit_noted', 'Drug Habit Noted?', 'মাদকাসক্তি লক্ষ্য করা হয়েছে?'),
        _survey_row('text', 'drug_names', 'Drug Name(s)', 'মাদকের নাম',
                    relevant="${drug_habit_noted} = 'yes'"),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_mh_screening():
    """KF-05/06: Mental Health Screening."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('select_one screening_type', 'screening_type', 'Screening Type', 'স্ক্রিনিংয়ের ধরন', required='yes'),
        _survey_row('date', 'screening_date', 'Screening Date', 'স্ক্রিনিংয়ের তারিখ', required='yes'),
        _survey_row('text', 'psycho_number', 'Psycho Registration Number', 'মানসিক স্বাস্থ্য নিবন্ধন নং'),
        _survey_row('text', 'counsellor_name', 'Counsellor / Screener Name', 'পরামর্শদাতার নাম'),
    ]
    # 20 Zahiruddin depression scale items (keeping brief for mobile)
    rows.append(_survey_row('begin_group', 'grp_items', 'Scale Items (Depression/PTSD)',
                            'স্কেল আইটেম',
                            relevant="${screening_type} = 'depression'"))
    for i in range(1, 21):
        rows.append(_survey_row(
            'integer', f'mh_q{i}', f'Item {i} (score 1–5)', f'আইটেম {i} (১–৫)',
            constraint='. >= 1 and . <= 5',
            constraint_msg='Enter score 1 to 5',
        ))
    rows.append(_survey_row('end_group', 'grp_items', '', ''))
    rows += [
        _survey_row('decimal', 'total_score', 'Total Score', 'মোট স্কোর'),
        _survey_row('select_one severity', 'severity_category', 'Severity Category', 'তীব্রতার মাত্রা'),
        _survey_row('select_one yes_no', 'referred_for_counselling', 'Referred for Counselling?', 'পরামর্শের জন্য রেফার?'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_gbv_case():
    """GBV Case Report."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('begin_group', 'grp_dates', 'Case Dates', 'মামলার তারিখ'),
        _survey_row('date', 'interview_date', 'Interview Date', 'সাক্ষাৎকারের তারিখ', required='yes'),
        _survey_row('date', 'incident_date', 'Incident Date', 'ঘটনার তারিখ', required='yes'),
        _survey_row('end_group', 'grp_dates', '', ''),
        # Sensitive PII (encrypted at rest)
        _survey_row('begin_group', 'grp_survivor', 'Survivor Information', 'বেঁচে থাকা ব্যক্তির তথ্য'),
        _survey_row('text', 'survivor_name', 'Survivor Name (ENCRYPTED)', 'বেঁচে থাকা ব্যক্তির নাম (এনক্রিপ্টেড)',
                    hint='This information is encrypted and only accessible to GBV officers'),
        _survey_row('text', 'survivor_contact', 'Survivor Contact (ENCRYPTED)', 'যোগাযোগ নম্বর (এনক্রিপ্টেড)'),
        _survey_row('text', 'survivor_address', 'Survivor Address (ENCRYPTED)', 'ঠিকানা (এনক্রিপ্টেড)'),
        _survey_row('integer', 'survivor_age', 'Survivor Age', 'বয়স'),
        _survey_row('select_one survivor_gender', 'survivor_gender_identity', 'Gender Identity', 'লিঙ্গ পরিচয়'),
        _survey_row('select_one yes_no', 'survivor_disability', 'Disability?', 'প্রতিবন্ধিতা?'),
        _survey_row('end_group', 'grp_survivor', '', ''),
        # Violence type
        _survey_row('begin_group', 'grp_violence', 'Type of Violence', 'সহিংসতার ধরন'),
        _survey_row('select_one yes_no', 'gbv_sexual', 'Sexual Violence?', 'যৌন সহিংসতা?'),
        _survey_row('select_one yes_no', 'gbv_physical', 'Physical Violence?', 'শারীরিক সহিংসতা?'),
        _survey_row('select_one yes_no', 'gbv_economic', 'Economic Violence?', 'অর্থনৈতিক সহিংসতা?'),
        _survey_row('select_one yes_no', 'gbv_psychological', 'Psychological Violence?', 'মানসিক সহিংসতা?'),
        _survey_row('end_group', 'grp_violence', '', ''),
        # Perpetrator
        _survey_row('begin_group', 'grp_perp', 'Perpetrator Information', 'অপরাধীর তথ্য'),
        _survey_row('text', 'perpetrator_name', 'Perpetrator Name (ENCRYPTED)', 'অপরাধীর নাম (এনক্রিপ্টেড)'),
        _survey_row('text', 'perpetrator_address', 'Perpetrator Address (ENCRYPTED)', 'অপরাধীর ঠিকানা (এনক্রিপ্টেড)'),
        _survey_row('integer', 'perpetrator_count', 'No. of Perpetrators', 'অপরাধীর সংখ্যা', default='1'),
        _survey_row('text', 'perpetrator_gender', 'Perpetrator Gender', 'অপরাধীর লিঙ্গ'),
        _survey_row('text', 'perpetrator_relationship', 'Relationship to Survivor', 'বেঁচে থাকা ব্যক্তির সাথে সম্পর্ক'),
        _survey_row('end_group', 'grp_perp', '', ''),
        # History & Services
        _survey_row('select_one yes_no', 'prior_reporting', 'Prior Reporting?', 'পূর্বে রিপোর্ট করা হয়েছে?'),
        _survey_row('select_one yes_no', 'prior_gbv_history', 'Prior GBV History?', 'পূর্বের জিবিভি ইতিহাস?'),
        _survey_row('begin_group', 'grp_services', 'Services Needed', 'প্রয়োজনীয় সেবা'),
        _survey_row('select_one yes_no', 'needs_medical', 'Medical Services?', 'চিকিৎসা সেবা?'),
        _survey_row('select_one yes_no', 'needs_legal', 'Legal Services?', 'আইনি সেবা?'),
        _survey_row('select_one yes_no', 'needs_shelter', 'Shelter Services?', 'আশ্রয় সেবা?'),
        _survey_row('select_one yes_no', 'needs_psychosocial', 'Psychosocial Support?', 'মনোসামাজিক সহায়তা?'),
        _survey_row('end_group', 'grp_services', '', ''),
        _survey_row('text', 'local_action_taken', 'Local Action Taken', 'স্থানীয় ব্যবস্থা', appearance='multiline'),
        _survey_row('text', 'case_officer_name', 'Case Officer Name', 'কেস অফিসারের নাম'),
        _survey_row('text', 'supervisor_name', 'Supervisor Name', 'সুপারভাইজারের নাম'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_outreach_session():
    """KF-08: Daily Outreach Session."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('date', 'session_date', 'Session Date', 'সেশনের তারিখ', required='yes'),
        _survey_row('text', 'peer_educator_name', 'Peer Educator Name', 'সহকর্মী শিক্ষাবিদের নাম', required='yes'),
        _survey_row('text', 'spot_name', 'Spot Name', 'স্পটের নাম'),
        _survey_row('begin_group', 'grp_contacts', 'Contacts & Distribution', 'যোগাযোগ ও বিতরণ'),
        _survey_row('integer', 'individual_contacts', 'Individual Contacts', 'ব্যক্তিগত যোগাযোগ', required='yes'),
        _survey_row('integer', 'individual_health_edu_count', 'Individual Health Education', 'ব্যক্তিগত স্বাস্থ্য শিক্ষা'),
        _survey_row('integer', 'group_health_edu_count', 'Group Health Education Sessions', 'গ্রুপ স্বাস্থ্য শিক্ষা সেশন'),
        _survey_row('integer', 'condoms_distributed_free', 'Condoms Distributed (Free)', 'বিতরণকৃত কনডম (বিনামূল্যে)'),
        _survey_row('integer', 'lubricants_distributed_free', 'Lubricants Distributed (Free)', 'বিতরণকৃত লুব্রিকেন্ট'),
        _survey_row('integer', 'iec_bcc_materials_distributed', 'IEC/BCC Materials Distributed', 'আইইসি/বিসিসি উপকরণ'),
        _survey_row('end_group', 'grp_contacts', '', ''),
        _survey_row('begin_group', 'grp_sessions', 'Sessions Conducted', 'পরিচালিত সেশন'),
        _survey_row('integer', 'hiv_aids_sti_knowledge_sessions', 'HIV/AIDS/STI Knowledge Sessions', 'এইচআইভি/এইডস/এসটিআই জ্ঞান সেশন'),
        _survey_row('integer', 'gbv_sessions', 'GBV Sessions', 'জিবিভি সেশন'),
        _survey_row('end_group', 'grp_sessions', '', ''),
        _survey_row('begin_group', 'grp_referrals', 'Referrals Made', 'রেফার করা হয়েছে'),
        _survey_row('integer', 'referral_mental_health', 'Mental Health Referrals', 'মানসিক স্বাস্থ্য রেফার'),
        _survey_row('integer', 'referral_legal_services', 'Legal Services Referrals', 'আইনি সেবা রেফার'),
        _survey_row('integer', 'referral_htc_hts', 'HTC/HTS Referrals', 'এইচটিসি/এইচটিএস রেফার'),
        _survey_row('integer', 'referral_gbv', 'GBV Referrals', 'জিবিভি রেফার'),
        _survey_row('integer', 'referral_other', 'Other Referrals', 'অন্যান্য রেফার'),
        _survey_row('end_group', 'grp_referrals', '', ''),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_group_education():
    """KF-10: Group Education Session."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('date', 'session_date', 'Session Date', 'সেশনের তারিখ', required='yes'),
        _survey_row('text', 'spot_name', 'Spot / Venue Name', 'স্পট / ভেন্যু'),
        _survey_row('text', 'facilitator_name', 'Facilitator Name', 'সহায়তাকারীর নাম'),
        _survey_row('text', 'topic', 'Topic / Session Theme', 'বিষয় / সেশনের থিম', required='yes'),
        _survey_row('integer', 'participant_count', 'Total Participants', 'মোট অংশগ্রহণকারী', required='yes'),
        _survey_row('integer', 'male_count', 'Male Participants', 'পুরুষ অংশগ্রহণকারী'),
        _survey_row('integer', 'female_count', 'Female Participants', 'মহিলা অংশগ্রহণকারী'),
        _survey_row('integer', 'tg_count', 'Transgender Participants', 'তৃতীয় লিঙ্গ অংশগ্রহণকারী'),
        _survey_row('integer', 'duration_minutes', 'Duration (minutes)', 'সময়কাল (মিনিট)'),
        _survey_row('integer', 'materials_distributed', 'Materials Distributed', 'বিতরণকৃত উপকরণ'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_referral():
    """Referral Form."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID', 'ক্লায়েন্ট আইডি', required='yes'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('date', 'referral_date', 'Referral Date', 'রেফারেলের তারিখ', required='yes'),
        _survey_row('select_one referral_type', 'referral_type', 'Referral Type', 'রেফারেলের ধরন', required='yes'),
        _survey_row('text', 'referral_reason', 'Reason for Referral', 'রেফারেলের কারণ', appearance='multiline'),
        _survey_row('text', 'referred_to', 'Referred To (Facility/Person)', 'কোথায়/কার কাছে রেফার', required='yes'),
        _survey_row('text', 'referred_by_name', 'Referred By (Name)', 'রেফারকারীর নাম'),
        _survey_row('text', 'referred_by_designation', 'Referred By (Designation)', 'রেফারকারীর পদবি'),
        _survey_row('date', 'follow_up_date', 'Follow-Up Date', 'ফলো-আপের তারিখ'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_hygiene_kit():
    """KF-12: Safety & Hygiene Kit Distribution (Bandhu)."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('text', 'client_id', 'Client ID (if individual distribution)', 'ক্লায়েন্ট আইডি (ব্যক্তিগত বিতরণের ক্ষেত্রে)'),
        _survey_row('text', 'client_name', 'Client Name', 'ক্লায়েন্টের নাম'),
        _survey_row('date', 'distribution_date', 'Distribution Date', 'বিতরণের তারিখ', required='yes'),
        _survey_row('integer', 'condom_count', 'Condoms Distributed', 'বিতরণকৃত কনডম', required='yes'),
        _survey_row('select_one yes_no', 'condom_demo', 'Condom Demonstration Done?', 'কনডম প্রদর্শনী হয়েছে?'),
        _survey_row('select_one yes_no', 'awareness_session', 'Awareness Session Conducted?', 'সচেতনতা সেশন হয়েছে?'),
        _survey_row('integer', 'iec_distributed', 'IEC Materials Distributed', 'আইইসি উপকরণ বিতরণ'),
        _survey_row('select_one yes_no', 'clinical_service_provided', 'Clinical Service Provided?', 'ক্লিনিকাল সেবা দেওয়া হয়েছে?'),
        _survey_row('select_one yes_no', 'counselling_provided', 'Counselling Provided?', 'পরামর্শ দেওয়া হয়েছে?'),
        _survey_row('select_one yes_no', 'referral_done', 'Referral Done?', 'রেফার করা হয়েছে?'),
        _survey_row('select_one yes_no', 'group_session', 'Group Session?', 'গ্রুপ সেশন?'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_training_event():
    """KF-20: Training / Orientation / Workshop."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('date', 'event_date', 'Event Start Date', 'অনুষ্ঠানের শুরুর তারিখ', required='yes'),
        _survey_row('date', 'event_end_date', 'Event End Date (if multi-day)', 'অনুষ্ঠানের শেষ তারিখ (বহু-দিনের ক্ষেত্রে)'),
        _survey_row('select_one event_type', 'event_type', 'Event Type', 'অনুষ্ঠানের ধরন', required='yes'),
        _survey_row('select_one participant_type', 'participant_type', 'Primary Participant Group', 'প্রধান অংশগ্রহণকারী গোষ্ঠী', required='yes'),
        _survey_row('text', 'topic', 'Topic / Module Title', 'বিষয় / মডিউল', required='yes'),
        _survey_row('text', 'location_text', 'Venue (Name & Address)', 'ভেন্যু (নাম ও ঠিকানা)'),
        _survey_row('text', 'district', 'District', 'জেলা'),
        _survey_row('integer', 'total_participants', 'Total Participants', 'মোট অংশগ্রহণকারী', required='yes'),
        _survey_row('integer', 'male_participants', 'Male Participants', 'পুরুষ অংশগ্রহণকারী'),
        _survey_row('integer', 'female_participants', 'Female Participants', 'মহিলা অংশগ্রহণকারী'),
        _survey_row('integer', 'tg_participants', 'Transgender Participants', 'তৃতীয় লিঙ্গ অংশগ্রহণকারী'),
        _survey_row('text', 'facilitator', 'Facilitator(s) Name', 'সহায়তাকারীর নাম'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_coord_meeting():
    """KF-19: Coordination Meeting."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('date', 'meeting_date', 'Meeting Date', 'সভার তারিখ', required='yes'),
        _survey_row('select_one meeting_type', 'meeting_type', 'Meeting Type', 'সভার ধরন', required='yes'),
        _survey_row('text', 'location_text', 'Venue', 'ভেন্যু'),
        _survey_row('text', 'district', 'District', 'জেলা'),
        _survey_row('integer', 'participant_count', 'No. of Participants', 'অংশগ্রহণকারীর সংখ্যা', required='yes'),
        _survey_row('text', 'agenda', 'Agenda', 'এজেন্ডা', appearance='multiline'),
        _survey_row('text', 'key_decisions', 'Key Decisions / Action Points', 'মূল সিদ্ধান্ত / কর্মপরিকল্পনা', appearance='multiline'),
        _survey_row('text', 'prepared_by', 'Prepared By', 'প্রস্তুতকারী'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_fistula_corner():
    """CIPRB Fistula Corner — District Hospital diagnostic register.

    Mirrors the Bengali paper register handed over by Rafi (প্রসবজনিত
    ফিস্টুলা রেজিস্ট্রার). PII fields are plaintext on submission; the
    backend's EncryptedCharField swaps them for Fernet ciphertext at rest.
    Field staff at District Hospital fill this for every diagnosed case.
    """
    rows = _common_metadata_rows()
    rows += [
        # Patient PII
        _survey_row('text',    'patient_name',  'Patient Name',        'রোগীর নাম', required='yes'),
        _survey_row('text',    'husband_name',  "Husband's Name",      'স্বামীর নাম'),
        _survey_row('text',    'mobile_number', 'Mobile Number',       'মোবাইল নম্বর'),
        _survey_row('integer', 'age_years',     'Age (years)',         'বয়স (বছর)'),
        # Address
        _survey_row('text', 'village',  'Village',  'গ্রাম'),
        _survey_row('text', 'union',    'Union',    'ইউনিয়ন'),
        _survey_row('text', 'upazila',  'Upazila',  'উপজেলা'),
        _survey_row('text', 'district', 'District', 'জেলা', required='yes'),
        # Dates
        _survey_row('date', 'suspected_date',      'Suspected Date',      'সাসপেক্টেড তারিখ'),
        _survey_row('date', 'identification_date', 'Identification Date', 'সনাক্তকরণ তারিখ'),
        _survey_row('date', 'diagnosis_date',      'Diagnosis Date',      'ডায়াগনোসিস তারিখ', required='yes'),
        # Informant
        _survey_row('text', 'informant_name',        'Informant Name',        'তথ্যদাতার নাম'),
        _survey_row('text', 'informant_designation', 'Informant Designation', 'তথ্যদাতার পদবী'),
        # Clinical
        _survey_row('text', 'suffering_duration', 'Duration of Suffering',
                    'ভোগান্তির সময়কাল',
                    hint='Free-text — e.g. "3 years", "8 months"'),
        _survey_row('text', 'fistula_cause', 'Cause of Fistula',
                    'ফিস্টুলার কারণ', appearance='multiline',
                    hint='e.g. "prolonged labour" / "দীর্ঘ সময়ের প্রসব"'),
        _survey_row('select_one fistula_type', 'fistula_type',
                    'Fistula Type', 'ফিস্টুলার ধরন', required='yes'),
        # Service provider
        _survey_row('text', 'service_provider_name',        'Service Provider Name',        'সেবা প্রদানকারীর নাম'),
        _survey_row('text', 'service_provider_designation', 'Service Provider Designation', 'সেবা প্রদানকারীর পদবী'),
        # Referral chain
        _survey_row('date', 'referral_date',  'Referral Date',  'রেফারেল তারিখ'),
        _survey_row('text', 'referral_place', 'Referral Place', 'রেফারেল স্থান'),
        _survey_row('select_one surgery_performed', 'surgery_performed',
                    'Surgery Performed?', 'অপারেশন হয়েছে?'),
        _survey_row('text', 'referral_outcome', 'Referral Outcome / Result',
                    'রেফারেল ফলাফল', appearance='multiline'),
        # Remarks
        _survey_row('text', 'remarks', 'Remarks', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_fistula_campaign_visit():
    """CIPRB Fistula Campaign — house-to-house screening register.

    Mirrors the Sunamganj campaign xlsx individual sheet. One row per
    suspected fistula case identified during the campaign sweep. PII
    encrypted at rest via the same EncryptedCharField pattern.
    """
    rows = _common_metadata_rows()
    rows += [
        _survey_row('date', 'visit_date', 'Visit Date', 'ভিজিটের তারিখ', required='yes'),
        # Patient PII
        _survey_row('text',    'patient_name',   'Patient Name',     'রোগীর নাম', required='yes'),
        _survey_row('text',    'husband_name',   "Husband's Name",   'স্বামীর নাম'),
        _survey_row('text',    'contact_number', 'Contact Number',   'যোগাযোগ নম্বর'),
        _survey_row('integer', 'age_years',      'Age (years)',      'বয়স (বছর)'),
        # Patient demographics
        _survey_row('text', 'education',          'Education',           'শিক্ষাগত যোগ্যতা'),
        _survey_row('text', 'profession',         'Profession',          'পেশা'),
        _survey_row('text', 'husband_profession', "Husband's Profession", 'স্বামীর পেশা'),
        # Address
        _survey_row('text', 'village',  'Village',  'গ্রাম'),
        _survey_row('text', 'union',    'Union',    'ইউনিয়ন'),
        _survey_row('text', 'upazila',  'Upazila',  'উপজেলা'),
        _survey_row('text', 'district', 'District', 'জেলা', required='yes'),
        _survey_row('select_one yes_no', 'from_haor',
                    'From Haor (wetland)?', 'হাওর এলাকা থেকে?'),
        # Obstetric history
        _survey_row('select_one delivery_mode', 'delivery_mode',
                    'Mode of Last Delivery', 'শেষ প্রসবের মাধ্যম'),
        _survey_row('select_one delivery_outcome', 'delivery_outcome',
                    'Delivery Outcome', 'প্রসবের ফলাফল'),
        _survey_row('text', 'suffering_duration', 'Duration of Suffering',
                    'ভোগান্তির সময়কাল',
                    hint='Free-text — e.g. "30 years", "8 months"'),
        _survey_row('text', 'info_source', 'Source of Information',
                    'তথ্যের উৎস',
                    hint='DRC / FWA / Midwife / Self'),
        # Remarks
        _survey_row('text', 'remarks', 'Remarks', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


def _form_mobile_camp():
    """KF-18: Mobile Health Camp (PHD only)."""
    rows = _common_metadata_rows()
    rows += [
        _survey_row('date', 'camp_date', 'Camp Date', 'ক্যাম্পের তারিখ', required='yes'),
        _survey_row('text', 'brothel_name', 'Brothel / Location Name', 'পতিতাপল্লি / অবস্থানের নাম', required='yes'),
        _survey_row('text', 'location_text', 'Full Address / Directions', 'সম্পূর্ণ ঠিকানা'),
        _survey_row('begin_group', 'grp_services', 'Services Delivered', 'প্রদত্ত সেবা'),
        _survey_row('integer', 'clients_served', 'Total Clients Served', 'মোট সেবাপ্রাপ্ত ক্লায়েন্ট', required='yes'),
        _survey_row('integer', 'hiv_tests_done', 'HIV Tests Done', 'এইচআইভি পরীক্ষা'),
        _survey_row('integer', 'sti_screenings_done', 'STI Screenings Done', 'এসটিআই স্ক্রিনিং'),
        _survey_row('integer', 'counselling_sessions', 'Counselling Sessions', 'পরামর্শ সেশন'),
        _survey_row('integer', 'referrals_made', 'Referrals Made', 'রেফার'),
        _survey_row('integer', 'condoms_distributed', 'Condoms Distributed', 'বিতরণকৃত কনডম'),
        _survey_row('end_group', 'grp_services', '', ''),
        _survey_row('text', 'services_description', 'Additional Services Description', 'অতিরিক্ত সেবার বিবরণ', appearance='multiline'),
        _survey_row('text', 'team_members', 'Team Members', 'দলের সদস্য', appearance='multiline'),
        _survey_row('text', 'notes', 'Notes', 'মন্তব্য', appearance='multiline'),
    ]
    return rows


# ─── Form manifest ─────────────────────────────────────────────────────────────

FORMS = [
    {
        'filename': 'KF-01_Client_Registration.xlsx',
        'id': 'spondon_client_reg_v1',
        'title': 'Spondon KF-01 — Client Registration (Mother List)',
        'survey_fn': _form_client_registration,
    },
    {
        'filename': 'KF-02_Clinic_Visit.xlsx',
        'id': 'spondon_clinic_visit_v1',
        'title': 'Spondon KF-02 — Clinic Visit (Patient Record Register)',
        'survey_fn': _form_clinic_visit,
    },
    {
        'filename': 'KF-03_HIV_STI_Test.xlsx',
        'id': 'spondon_hiv_sti_test_v1',
        'title': 'Spondon KF-03 — HIV/STI Test Result',
        'survey_fn': _form_hiv_sti_test,
    },
    {
        'filename': 'KF-04_HTC_Counselling.xlsx',
        'id': 'spondon_htc_counsel_v1',
        'title': 'Spondon KF-04 — HTC Counselling',
        'survey_fn': _form_htc_counselling,
    },
    {
        'filename': 'KF-05_MH_Screening.xlsx',
        'id': 'spondon_mh_screening_v1',
        'title': 'Spondon KF-05/06 — Mental Health Screening',
        'survey_fn': _form_mh_screening,
    },
    {
        'filename': 'KF-08_Outreach_Session.xlsx',
        'id': 'spondon_outreach_v1',
        'title': 'Spondon KF-08 — Outreach Session',
        'survey_fn': _form_outreach_session,
    },
    {
        'filename': 'KF-09_Individual_Counselling.xlsx',
        'id': 'spondon_counselling_v1',
        'title': 'Spondon KF-09 — Individual Counselling Session',
        'survey_fn': _form_individual_counselling,
    },
    {
        'filename': 'KF-10_Group_Education.xlsx',
        'id': 'spondon_group_edu_v1',
        'title': 'Spondon KF-10 — Group Education Session',
        'survey_fn': _form_group_education,
    },
    {
        'filename': 'KF-12_Hygiene_Kit.xlsx',
        'id': 'spondon_hygiene_kit_v1',
        'title': 'Spondon KF-12 — Safety & Hygiene Kit Distribution',
        'survey_fn': _form_hygiene_kit,
    },
    {
        'filename': 'KF-13_ADR_Record.xlsx',
        'id': 'spondon_adr_record_v1',
        'title': 'Spondon KF-13 — Adverse Drug Reaction Record',
        'survey_fn': _form_adr_record,
    },
    {
        'filename': 'KF-16_Autoclave_Log.xlsx',
        'id': 'spondon_autoclave_log_v1',
        'title': 'Spondon KF-16 — Autoclave / Incinerator Log',
        'survey_fn': _form_autoclave_log,
    },
    {
        'filename': 'KF-18_Mobile_Health_Camp.xlsx',
        'id': 'spondon_mobile_camp_v1',
        'title': 'Spondon KF-18 — Mobile Health Camp (PHD)',
        'survey_fn': _form_mobile_camp,
    },
    {
        'filename': 'KF-19_Coordination_Meeting.xlsx',
        'id': 'spondon_coord_meeting_v1',
        'title': 'Spondon KF-19 — Coordination Meeting',
        'survey_fn': _form_coord_meeting,
    },
    {
        'filename': 'KF-20_Training_Event.xlsx',
        'id': 'spondon_training_event_v1',
        'title': 'Spondon KF-20 — Training / Orientation / Workshop',
        'survey_fn': _form_training_event,
    },
    {
        'filename': 'KF-Referral.xlsx',
        'id': 'spondon_referral_v1',
        'title': 'Spondon — Referral Form',
        'survey_fn': _form_referral,
    },
    {
        'filename': 'KF-GBV_Case.xlsx',
        'id': 'spondon_gbv_case_v1',
        'title': 'Spondon — GBV Case Report (CONFIDENTIAL)',
        'survey_fn': _form_gbv_case,
    },
    {
        'filename': 'KF-ANC_Antenatal_Card.xlsx',
        'id': 'spondon_antenatal_card_v1',
        'title': 'Spondon — Antenatal Card (PHD)',
        'survey_fn': _form_antenatal_card,
    },
    # CIPRB fistula forms — backend models + handlers landed in
    # fistula.0003 + programs.webhook.FORM_HANDLERS. Schema confirmed
    # from the Bengali Fistula Corner register photo + Sunamganj
    # campaign xlsx Rafi handed over.
    {
        'filename': 'KF-Fistula_Corner.xlsx',
        'id': 'spondon_fistula_corner_v1',
        'title': 'Spondon — Fistula Corner (CIPRB District Hospital)',
        'survey_fn': _form_fistula_corner,
    },
    {
        'filename': 'KF-Fistula_Campaign_Visit.xlsx',
        'id': 'spondon_fistula_campaign_v1',
        'title': 'Spondon — Fistula Campaign Visit (CIPRB House Screening)',
        'survey_fn': _form_fistula_campaign_visit,
    },
]


# ─── Management command ────────────────────────────────────────────────────────

class Command(BaseCommand):
    help = (
        'Generate all 17 KoboToolbox XLSForm (.xlsx) files for Spondon IDMS. '
        'Upload each file to https://kf.kobotoolbox.org → New Project → Upload XLS.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            default='koboforms',
            help='Directory to write XLS files to (default: koboforms/)',
        )

    def handle(self, *args, **options):
        out_dir = options['output_dir']
        os.makedirs(out_dir, exist_ok=True)

        common_choices = _common_choices()
        center_choices = _center_choices()

        self.stdout.write(f'\n  Generating {len(FORMS)} KoboToolbox XLS forms -> {out_dir}/\n')

        for form in FORMS:
            survey_rows = form['survey_fn']()
            choices_rows = common_choices + center_choices
            wb = _build_wb(form['id'], form['title'], survey_rows, choices_rows)
            path = os.path.join(out_dir, form['filename'])
            wb.save(path)
            self.stdout.write(self.style.SUCCESS(f'  OK  {form["filename"]:50s}  id_string: {form["id"]}'))

        self.stdout.write(f'\n  Done - {len(FORMS)} files written to {os.path.abspath(out_dir)}/\n')
        self.stdout.write(self.style.WARNING(
            '\n  NEXT STEPS:\n'
            '  1. Upload each .xlsx to https://kf.kobotoolbox.org -> New Project -> Upload XLS\n'
            '  2. Add REST Service webhook: URL = https://web-production-091fa.up.railway.app/webhook/programs/\n'
            '     HTTP headers: Authorization = Token REDACTED\n'
            '  3. Deploy form and share URL with field staff via Telegram\n'
            '  4. Verify center_code choices match your seeded ServiceCenter records\n'
        ))
