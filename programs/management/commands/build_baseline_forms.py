# -*- coding: utf-8 -*-
"""
Build the two D5 Baseline/Endline KoboToolbox XLSForms (bilingual EN/BN).

  1. Hijra / Gender-diverse Baseline   id: ciprb_baseline_hijra_v1
       Source: "Hijra questionnaire after validation_English_June 22.docx"
               + "..._bengali__June22.docx"   (184 questions, 8 districts)
  2. Female Sex Workers Baseline        id: ciprb_baseline_fsw_v1
       Source: "FSW_questionnaire_street_brothel_updated_June22_english.docx"
               + "..._bengali.docx"           (194 questions, 9 districts × 12 sites)

The baseline STUDY is conducted by CIPRB (D5). Bandhu (Hijra) and PHD (FSW) are
the service-delivery partners whose populations are assessed; they do not own or
enter this data. Both forms therefore carry organisation='CIPRB' and are
distinguished only by a hidden `population` field (hijra / fsw).

Conventions mirror build_ciprb_forms.py:
  - Bilingual: label::English + label::Bangla, digitised WORD-FOR-WORD from the
    validated questionnaires (incl. the ‡ sensitivity marker and the standard
    98 = Don't know / 99 = Decline-to-answer codes, kept literal).
  - Single-page (style='theme-grid') so skip logic works on phones.
  - GPS captured at CLUSTER/SITE level only (one geopoint), never residence.
  - A unique questionnaire serial + a pulldata() duplicate-block (the fistula
    pattern), so the same respondent cannot be entered twice.
  - PHQ-9 only: both instruments name "PHQ-9 and GAD-7" in the Module 8
    instructions but contain NO GAD-7 items — digitised verbatim (PHQ-9 only);
    flagged to CIPRB as a content question.

Run:
    python manage.py build_baseline_forms                 # writes xlsx
    python manage.py build_baseline_forms --only ciprb_baseline_hijra_v1
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.core.management.base import BaseCommand

HERE   = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.normpath(os.path.join(HERE, '..', '..', '..', '..', 'koboforms_baseline'))

_HFILL = PatternFill("solid", fgColor="003F72")
_HFONT = Font(color="FFFFFF", bold=True, size=10)

SURVEY_HDR = [
    'type', 'name', 'label::English', 'label::Bangla',
    'hint', 'required', 'relevant', 'constraint', 'constraint_message',
    'default', 'appearance', 'calculation',
]
CHOICES_HDR  = ['list_name', 'name', 'label::English', 'label::Bangla']
SETTINGS_HDR = ['form_title', 'form_id', 'version', 'default_language', 'style']


def _sr(qtype, name, en='', bn='', hint='', required='',
        relevant='', constraint='', cmsg='', default='', app='', calc=''):
    return [qtype, name, en, bn, hint, required, relevant,
            constraint, cmsg, default, app, calc]


def _ch(lst, name, en, bn=''):
    return [lst, name, en, bn]


def _wb(form_id, form_title, survey, choices):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in [
        ('survey',   SURVEY_HDR,  survey),
        ('choices',  CHOICES_HDR, choices),
        ('settings', SETTINGS_HDR,
         [[form_title, form_id, '20260625', 'English', 'theme-grid']]),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for ci in range(1, len(headers) + 1):
            c = ws.cell(1, ci)
            c.font = _HFONT
            c.fill = _HFILL
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 30
        ws.freeze_panes = 'A2'
    return wb


# ─── Standard codes used across every module (kept literal: 98 / 99) ──────────
DK   = ('dont_know', "Don't know", 'জানি না', '98')
DEC  = ('decline',   'Decline to answer', 'উত্তর দিতে অনিচ্ছুক', '99')


def _dk_dec(list_name, include_dk=True, include_dec=True):
    """Append the standard 98/99 choices to a list, with their literal codes."""
    out = []
    if include_dk:
        out.append([list_name, DK[3], DK[1], DK[2]])
    if include_dec:
        out.append([list_name, DEC[3], DEC[1], DEC[2]])
    return out


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  HIJRA / GENDER-DIVERSE BASELINE  —  id: ciprb_baseline_hijra_v1          ║
# ║  Owner population: Bandhu service area · 8 districts · 184 questions       ║
# ╚══════════════════════════════════════════════════════════════════════════╝

# A201 district codes (Hijra instrument): Sunamganj 1 … Bandarban 8.
HIJRA_DISTRICTS = [
    ('sunamganj',   'Sunamganj',   'সুনামগঞ্জ',   '1'),
    ('habiganj',    'Habiganj',    'হবিগঞ্জ',     '2'),
    ('manikganj',   'Manikganj',   'মানিকগঞ্জ',   '3'),
    ('narayanganj', 'Narayanganj', 'নারায়ণগঞ্জ', '4'),
    ('chandpur',    'Chandpur',    'চাঁদপুর',     '5'),
    ('noakhali',    'Noakhali',    'নোয়াখালী',   '6'),
    ('chittagong',  'Chittagong',  'চট্টগ্রাম',   '7'),
    ('bandarban',   'Bandarban',   'বান্দরবান',   '8'),
]

# XPath-1.0 trim+upper of the typed serial (Enketo has no upper-case()).
_NORM_SERIAL = ("translate(normalize-space(${questionnaire_serial}),"
                "'abcdefghijklmnopqrstuvwxyz',"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ')")


def _hijra_meta():
    """Interview-identification block (cover sheet + A1) — CIPRB-owned, GPS at
    cluster level, with the unique questionnaire serial + duplicate-block."""
    return [
        _sr('begin_group', 'grp_admin', 'Interview identification',
            'সাক্ষাৎকার সনাক্তকরণ'),
        # Hidden routing fields — partner stays CIPRB; population splits the two.
        _sr('calculate', 'organisation', '', '', calc="'CIPRB'"),
        _sr('calculate', 'population',   '', '', calc="'hijra'"),
        _sr('calculate', 'survey_round', '', '', calc="'baseline'"),

        _sr('text', 'questionnaire_serial',
            'Questionnaire Serial No.', 'প্রশ্নপত্রের সিরিয়াল নং',
            required='yes',
            constraint=("pulldata('respondents_hijra','serial','serial',"
                        + _NORM_SERIAL + ")=''"),
            cmsg='⚠ This serial is already recorded. Use a new, unique serial. / '
                 '⚠ এই সিরিয়ালটি ইতিমধ্যে রেকর্ড করা হয়েছে। নতুন, অনন্য সিরিয়াল ব্যবহার করুন।',
            hint='Must be unique per questionnaire.'),
        _sr('calculate', '_dup_serial',
            calc=("pulldata('respondents_hijra','serial','serial',"
                  + _NORM_SERIAL + ")")),
        _sr('note', '_dup_warn',
            '⚠ This serial is already recorded — do not enter the same respondent twice.',
            '⚠ এই সিরিয়ালটি ইতিমধ্যে রেকর্ড করা হয়েছে — একই উত্তরদাতাকে দুবার লিখবেন না।',
            relevant="${questionnaire_serial}!='' and ${_dup_serial}!=''"),

        _sr('text', 'cluster_site_code', 'Cluster/Site Code', 'ক্লাস্টার/সাইট কোড',
            required='yes'),
        _sr('select_one district', 'district', 'District', 'জেলা', required='yes'),
        _sr('calculate', 'district_code', calc=_hijra_dist_code_calc()),

        _sr('text', 'interviewer_name_code', 'Name & Code of Interviewer',
            'সাক্ষাৎকারগ্রহণকারীর নাম ও কোড', required='yes'),
        _sr('text', 'supervisor_name_code', 'Name & Code of Supervisor',
            'সুপারভাইজারের নাম ও কোড'),
        _sr('date', 'interview_date', 'Interview Date (Day/Month/Year)',
            'সাক্ষাৎকারের তারিখ (দিন/মাস/বছর)', required='yes', default='today()'),
        _sr('time', 'start_time', 'Start Time', 'শুরুর সময়'),
        _sr('integer', 'interview_attempts', 'Number of Interview Attempts',
            'সাক্ষাৎকারের প্রচেষ্টার সংখ্যা'),
        _sr('select_one interview_language', 'interview_language',
            'Interview Language', 'সাক্ষাৎকারের ভাষা'),
        _sr('text', 'interview_language_other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="${interview_language}='2'"),
        _sr('select_one interview_method', 'interview_method',
            'Method', 'পদ্ধতি'),

        # GPS — cluster/site level only (A105/A106), never the respondent's home.
        _sr('geopoint', 'location',
            'Geo-reference — Cluster/Site GPS (required; cluster level only)',
            'জিও-রেফারেন্স — ক্লাস্টার/সাইট জিপিএস (প্রয়োজনীয়; শুধু ক্লাস্টার পর্যায়ে)',
            required='yes'),
        _sr('end_group', 'grp_admin'),
    ]


def _hijra_dist_code_calc():
    expr = "''"
    for slug, _en, _bn, code in reversed(HIJRA_DISTRICTS):
        expr = f"if(${{district}}='{slug}','{code}',{expr})"
    return expr


def _hijra_consent():
    """Informed-consent script (read aloud verbatim) + recorded consent.
    Everything after screening is gated on consent + eligibility via _proceed."""
    return [
        _sr('note', 'consent_instr',
            '[Read aloud verbatim in Bengali. Hand the respondent an information '
            'sheet. Allow time for questions. Do not begin the interview until '
            'consent is obtained.]',
            '[বাংলায় হুবহু পড়ে শোনাতে হবে। অংশগ্রহণকারীকে তথ্যপত্র দিন। প্রশ্ন করার '
            'সুযোগ দিন। সম্মতি না পাওয়া পর্যন্ত সাক্ষাৎকার শুরু করবেন না।]'),
        _sr('note', 'consent_script',
            '"Good morning/good afternoon. My name is ____________________. I am '
            'working on behalf of [name of data collection organization]. We are '
            'conducting this interview as part of the baseline research for the '
            "project 'Strengthening SRHR Needs of Key Populations in Selected "
            "Districts of Bangladesh', implemented in collaboration with Bandhu "
            'Social Welfare Society and UNFPA Bangladesh. The purpose of this '
            'interview is to learn about the current health situation, needs, and '
            'experiences of the hijra/gender-diverse community in this district, so '
            'that services including new wellness centres can be planned according '
            "to the community's needs. Your participation is entirely voluntary. "
            'Your name or identity will not be disclosed in any report or public '
            'document. Your answers will be stored under a code number. The link '
            'between contact information and code numbers will be stored separately '
            'in a secure location and will be destroyed once data collection and '
            'verification are complete. Your information will be kept completely '
            'confidential. It will not be shared with anyone outside the research '
            'team — including family members, community leaders, employers, police, '
            'or government offices. Whether or not you participate will have no '
            'effect on any services you receive now or in the future, including '
            'services from Bandhu. There are no right or wrong answers here. Some '
            'questions may be personal — such as about health, relationships, or '
            'difficult experiences. You may choose not to answer any question, take '
            'a break, or stop the interview at any time. No reason is required and '
            'there will be no negative consequences. The interview takes '
            'approximately 50–60 minutes. No payment will be made for participation '
            'beyond [designated hospitality/stipend per protocol]. While there may '
            'be no direct benefit to you personally, the findings of this research '
            'will help plan better services for your community. If you have '
            'questions or complaints about the research, you may contact: Principal '
            'Investigator: ___________ Mobile: _____________ Or, Ethics Review '
            'Committee Representative: _______________ Mobile:____________ Do you '
            'have any questions now? May I begin the interview?"',
            '"সুপ্রভাত/শুভ অপরাহ্ন। আমার নাম ____________________। আমি [তথ্য সংগ্রহকারী '
            'প্রতিষ্ঠানের নাম]-এর পক্ষ থেকে কাজ করছি। বন্ধু সোশ্যাল ওয়েলফেয়ার সোসাইটি এবং '
            'UNFPA বাংলাদেশ-এর সহযোগিতায় পরিচালিত ‘বাংলাদেশের নির্বাচিত জেলায় মূল '
            'জনগোষ্ঠীর SRHR চাহিদা শক্তিশালীকরণ’ প্রকল্পের বেসলাইন গবেষণার অংশ হিসেবে আমরা '
            'এই সাক্ষাৎকার নিচ্ছি। এই সাক্ষাৎকারের উদ্দেশ্য হলো এই জেলার হিজড়া/প্রান্তিক '
            'জনগোষ্ঠীর বর্তমান স্বাস্থ্য পরিস্থিতি, চাহিদা এবং অভিজ্ঞতা সম্পর্কে জানা, যাতে '
            'নতুন ওয়েলনেস সেন্টারসহ বিভিন্ন সেবা কমিউনিটির প্রয়োজন অনুযায়ী পরিকল্পনা করা '
            'যায়। আপনার অংশগ্রহণ সম্পূর্ণ স্বেচ্ছাসেবী। কোনো রিপোর্ট বা প্রকাশ্য নথিতে আপনার '
            'নাম বা পরিচয় প্রকাশ করা হবে না। আপনার উত্তর একটি কোড নম্বরের মাধ্যমে সংরক্ষণ '
            'করা হবে। যোগাযোগের তথ্য ও কোড নম্বরের সংযোগকারী তালিকা আলাদাভাবে নিরাপদে '
            'সংরক্ষণ করা হবে এবং তথ্য সংগ্রহ ও যাচাই শেষ হলে ধ্বংস করে ফেলা হবে। আপনার তথ্য '
            'সম্পূর্ণ গোপন রাখা হবে। গবেষণা দলের বাইরে কারো সাথে—যেমন: পরিবারের সদস্য, '
            'কমিউনিটি নেতা, নিয়োগকর্তা, পুলিশ, সরকারি অফিস কাউকেই আপনার তথ্য শেয়ার করা হবে '
            'না। আপনি অংশগ্রহণ করুন বা না করুন, বর্তমানে বা ভবিষ্যতে আপনি যে কোনো সেবা পান '
            'তাতে কোনো প্রভাব পড়বে না, বন্ধুর সেবার ক্ষেত্রেও নয়। এখানে সঠিক বা ভুল কোনো '
            'উত্তর নেই। কিছু প্রশ্ন ব্যক্তিগত হতে পারে—যেমন স্বাস্থ্য, সম্পর্ক বা কঠিন অভিজ্ঞতা '
            'সম্পর্কে। আপনি চাইলে যেকোনো প্রশ্নের উত্তর না দিতে পারেন, বিরতি নিতে পারেন, '
            'যেকোনো সময় সাক্ষাৎকার বন্ধ করতে পারেন। এজন্য কোনো কারণ দেখাতে হবে না এবং কোনো '
            'নেতিবাচক প্রভাবও পড়বে না। সাক্ষাৎকারে প্রায় ৫০–৬০ মিনিট সময় লাগবে। [প্রোটোকল '
            'অনুযায়ী নির্ধারিত আপ্যায়ন/ভাতা] ছাড়া অংশগ্রহণের জন্য কোনো অর্থ প্রদান করা হবে '
            'না। যদিও আপনার সরাসরি কোনো লাভ নাও হতে পারে, তবে এই গবেষণার ফলাফল আপনার '
            'কমিউনিটির জন্য উন্নত সেবা পরিকল্পনায় সহায়তা করবে। গবেষণা সম্পর্কে কোনো প্রশ্ন বা '
            'অভিযোগ থাকলে যোগাযোগ করতে পারেন: প্রধান গবেষক: ___________ মোবাইল: '
            '_____________ অথবা, নৈতিক পর্যালোচনা কমিটির প্রতিনিধি: _______________ '
            'মোবাইল:____________ এখন কি আপনার কোনো প্রশ্ন আছে? আমি কি সাক্ষাৎকার শুরু '
            'করতে পারি?"'),
        _sr('select_one consent_yn', 'consent',
            'Consent given', 'সম্মতি প্রদান করেছেন', required='yes',
            hint='If No, thank the respondent and end the interview.'),
        _sr('note', 'consent_no_end',
            'No consent — thank the respondent and end the interview.',
            'সম্মতি নেই — উত্তরদাতাকে ধন্যবাদ জানিয়ে সাক্ষাৎকার শেষ করুন।',
            relevant="${consent}='2'"),
    ]


def _hijra_screening():
    """Eligibility screening S1–S4. Sets _eligible; the rest of the form is
    gated on consent='1' AND _eligible='1' via the _proceed flag."""
    return [
        _sr('begin_group', 'grp_screen', 'Eligibility Screening', 'যোগ্যতা যাচাই',
            relevant="${consent}='1'"),
        _sr('select_one selection_method', 's1_selection',
            'Interviewer: Record how the respondent was selected. (Do not read aloud)',
            'সাক্ষাৎকারগ্রহণকারী: উত্তরদাতাকে কীভাবে নির্বাচন করা হয়েছে তা লিখুন। (উচ্চস্বরে পড়বেন না)',
            required='yes'),
        _sr('text', 's1_selection_other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="${s1_selection}='4'"),
        _sr('integer', 's2_age',
            'What is your current age (in completed years)?',
            'আপনার বর্তমান বয়স কত (সম্পূর্ণ বছর)?',
            required='yes', constraint='. >= 0 and . <= 120'),
        _sr('note', 's2_ineligible',
            'Under 18 — end interview and mark as ineligible.',
            '১৮ বছরের কম — সাক্ষাৎকার শেষ করুন এবং অযোগ্য হিসেবে চিহ্নিত করুন।',
            relevant="${s2_age}!='' and ${s2_age}<18"),
        _sr('select_one yn12', 's3_member',
            'Do you consider yourself a member of the hijra/gender-diverse community? '
            '(e.g., hijra, transgender, gay, lesbian, bisexual, intersex, queer/non-binary)',
            'আপনি কি নিজেকে হিজড়া/প্রান্তিক জনগোষ্ঠীর সদস্য মনে করেন? (যেমন: হিজড়া, '
            'ট্রান্সজেন্ডার, গে, লেসবিয়ান, বাইসেক্সুয়াল, ইন্টারসেক্স, কুইয়ার/নন-বাইনারি)',
            required='yes', hint='‡'),
        _sr('select_one residence_len', 's4_residence',
            'How long have you been living in this district?',
            'আপনি কতদিন ধরে এই জেলায় বসবাস করছেন?', required='yes'),
        _sr('note', 's_ineligible_end',
            'Respondent is not eligible — end the interview.',
            'উত্তরদাতা যোগ্য নন — সাক্ষাৎকার শেষ করুন।',
            relevant="(${s2_age}!='' and ${s2_age}<18) or ${s3_member}='2' or ${s4_residence}='1'"),
        _sr('end_group', 'grp_screen'),
        # Gate flag — every module below is relevant only when this is '1'.
        _sr('calculate', '_proceed',
            calc="if(${consent}='1' and ${s2_age}>=18 and ${s3_member}='1' "
                 "and ${s4_residence}='2','1','0')"),
    ]


_B107_SOURCES = [
    ('b107_employment',  'Employment/salary',            'চাকরি/বেতন'),
    ('b107_business',    'Business',                     'ব্যবসা'),
    ('b107_remittance',  'Remittance',                   'প্রবাসী আয় (রেমিট্যান্স)'),
    ('b107_agriculture', 'Agriculture (annual income ÷ 12)', 'কৃষি (বার্ষিক আয় ÷ 12)'),
    ('b107_rent',        'House/land rent',              'বাড়ি/জমি ভাড়া'),
    ('b107_badhai',      'Badhai/dancing or community collection',
                         'বাধাই/নাচ বা কমিউনিটি সংগ্রহ'),
    ('b107_sexwork',     'Sex work',                     'যৌনকর্ম'),
    ('b107_other',       'Other (specify)',              'অন্যান্য (উল্লেখ করুন)'),
]


def _hijra_module1():
    """Module 1 — Background and Socio-Demographic Information (A1, A2, A3, B).
    Each section is its own group, gated on _proceed='1'."""
    P = "${_proceed}='1'"
    rows = [
        _sr('note', 'm1_head',
            'Module 1. Background and Socio-Demographic Information',
            'মডিউল ১. পটভূমি ও সামাজিক-জনসংখ্যাতাত্ত্বিক তথ্য', relevant=P),

        # ── A1. Respondent and Interview Identification ───────────────────────
        _sr('begin_group', 'grp_a1',
            'A1. Respondent and Interview Identification Information',
            'A1. উত্তরদাতা ও সাক্ষাৎকার সনাক্তকরণ তথ্য', relevant=P),
        _sr('text', 'a101_respondent_id', 'A101. Respondent ID Number',
            'A101. উত্তরদাতার আইডি নম্বর'),
        _sr('select_one respondent_type', 'a102_respondent_type',
            'A102. Respondent Type Code ‡', 'A102. উত্তরদাতার ধরন কোড ‡',
            hint='Classification used only for sampling/quota monitoring. GPS is '
                 'taken at cluster level only, not at the residence.'),
        _sr('end_group', 'grp_a1'),

        # ── A2. Respondent's Basic Information and Profile ────────────────────
        _sr('begin_group', 'grp_a2',
            "A2. Respondent's Basic Information and Profile",
            'A2. উত্তরদাতার মৌলিক তথ্য ও পরিচিতি', relevant=P),
        _sr('text', 'a202_upazila', 'A202. Upazila/Thana/City Corporation',
            'A202. উপজেলা/থানা/সিটি কর্পোরেশন'),
        _sr('text', 'a203_union', 'A203. Union/Ward', 'A203. ইউনিয়ন/ওয়ার্ড'),
        _sr('select_one area', 'a204_area', 'A204. Area', 'A204. এলাকা'),
        _sr('integer', 'a205_age', 'A205. Completed Age (in years)',
            'A205. সম্পূর্ণ বয়স (বছরে)', constraint='. >= 0 and . <= 120'),
        _sr('select_one religion', 'a206_religion', 'A206. Religion (code)',
            'A206. ধর্ম (কোড)'),
        _sr('text', 'a206_religion_other', 'Other religion (specify)',
            'অন্যান্য ধর্ম (উল্লেখ করুন)', relevant="${a206_religion}='5'"),
        _sr('select_one ethnicity', 'a207_ethnicity', 'A207. Ethnic identity',
            'A207. জাতিগত পরিচয়'),
        _sr('text', 'a207_ethnicity_other', 'Ethnic minority (specify)',
            'ক্ষুদ্র নৃগোষ্ঠী (উল্লেখ করুন)', relevant="${a207_ethnicity}='2'"),
        _sr('select_one marital', 'a208_marital',
            'A208. Current marital/relationship status (code)',
            'A208. বর্তমান বৈবাহিক/সম্পর্কের অবস্থা (কোড)'),
        _sr('select_one education', 'a209_education',
            'A209. Highest educational qualification (code)',
            'A209. সর্বোচ্চ শিক্ষাগত যোগ্যতা (কোড)'),
        _sr('select_one yn12', 'a210_student', 'A210. Currently a student?',
            'A210. বর্তমানে কি শিক্ষার্থী?'),
        _sr('select_one mobile_phone', 'a211_mobile',
            'A211. Do you have your own mobile phone?',
            'A211. আপনার নিজের মোবাইল ফোন আছে কি?'),
        _sr('select_one yn12', 'a212_nid',
            'A212. Do you have a National Identity Card (NID)?',
            'A212. আপনার কি জাতীয় পরিচয়পত্র (NID) আছে?'),
        _sr('select_one nid_match', 'a213_nid_match',
            'A213. ‡ Does the sex listed on your NID match your current identity?',
            'A213. ‡ আপনার NID-তে উল্লেখিত লিঙ্গ কি আপনার বর্তমান পরিচয়ের সাথে মিলে?'),
        _sr('text', 'a214_ancestral_district', 'A214. District of ancestral home',
            'A214. পৈতৃক বাড়ির জেলা'),
        _sr('integer', 'a215_years_since_left',
            'A215. How many years has it been since you left your ancestral home? '
            '(If still living there or with family, write 0)',
            'A215. পৈতৃক বাড়ি ছেড়ে আসার পর কত বছর হয়েছে? (যদি এখনও সেখানে থাকেন বা '
            'পরিবারের সাথেই থাকেন, তাহলে 0 লিখুন)'),
        _sr('end_group', 'grp_a2'),

        # ── A3. Gender Identity and Sexual Orientation ‡ ──────────────────────
        _sr('begin_group', 'grp_a3',
            'A3. Gender Identity and Sexual Orientation ‡',
            'A3. জেন্ডার পরিচয় ও যৌন অভিমুখিতা ‡', relevant=P),
        _sr('note', 'a3_instr',
            '[Two-step method. Read each question and answer options using '
            'community-recognised Bengali terminology. Assure the respondent that '
            'there are no right or wrong answers and that they may choose not to '
            'answer. Sensitivity training for data collectors is mandatory for this section.]',
            '[দুই-ধাপ পদ্ধতি। প্রতিটি প্রশ্ন এবং উত্তর বিকল্প কমিউনিটির স্বীকৃত বাংলা পরিভাষা '
            'ব্যবহার করে পড়ে শোনান। উত্তরদাতাকে আশ্বস্ত করুন যে এখানে সঠিক বা ভুল কোনো উত্তর '
            'নেই এবং চাইলে তিনি উত্তর না-ও দিতে পারেন। এই অংশের জন্য তথ্যসংগ্রাহকদের '
            'সংবেদনশীলতা বিষয়ক প্রশিক্ষণ বাধ্যতামূলক।]'),
        _sr('select_one a301_sex', 'a301',
            'A301 ‡ What sex was registered for you at birth?',
            'A301 ‡ জন্মের সময় আপনার জন্য কোন লিঙ্গ নিবন্ধিত করা হয়েছিল?'),
        _sr('select_one a302_gender', 'a302_gender',
            'A302 ‡ How do you currently describe your gender identity? [One answer. '
            "Accept the respondent's own terminology and code to the nearest "
            "category; if needed, write exact term next to 'Other'.]",
            'A302 ‡ বর্তমানে আপনি আপনার জেন্ডার পরিচয় কীভাবে বর্ণনা করেন? [একটি উত্তর। '
            'উত্তরদাতার নিজস্ব পরিভাষা গ্রহণ করুন এবং নিকটতম শ্রেণিতে কোড দিন; প্রয়োজনে '
            '‘অন্যান্য’-এর পাশে হুবহু লিখুন।]'),
        _sr('text', 'a302_other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="${a302_gender}='7'"),
        _sr('end_group', 'grp_a3'),

        # ── B. Household/Dera Composition and Economic Status ─────────────────
        _sr('note', 'mB_head', 'B. Household/Dera Composition and Economic Status',
            'B. পরিবার/ডেরা গঠন ও অর্থনৈতিক অবস্থা', relevant=P),
        _sr('begin_group', 'grp_b', 'B. Household / Dera and Economic Status',
            'B. পরিবার/ডেরা ও অর্থনৈতিক অবস্থা', relevant=P),
        _sr('select_one live_with', 'b101_live_with',
            'B101 ‡ Who do you currently live with?',
            'B101 ‡ আপনি বর্তমানে কার সাথে থাকেন?'),
        _sr('text', 'b101_other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="${b101_live_with}='6'"),
        _sr('integer', 'b102_hh_members',
            'B102 How many household members, including yourself, usually live and '
            'eat together?',
            'B102 আপনিসহ আপনার পরিবারের কতজন সদস্য সাধারণত একসাথে থাকেন এবং খাবার খান?',
            relevant="${b101_live_with}!='2'"),
        _sr('integer', 'b103_dera_members',
            'B103 Ask only if living in a dera: How many people in total, including '
            'the guruma, live in the dera?',
            'B103 শুধুমাত্র ডেরায় থাকলে জিজ্ঞাসা করুন: গুরুমাসহ ডেরায় মোট কতজন থাকেন?',
            relevant="${b101_live_with}='2'"),
        _sr('integer', 'b104_share',
            'B104 Last month, how much money (taka) did you receive as your share?',
            'B104 গত মাসে আপনার অংশ হিসেবে কত টাকা পেয়েছেন?',
            relevant="${b101_live_with}='2'"),
        _sr('integer', 'b105_shared_exp',
            'B105 Last month, what was the total shared expenditure of all members '
            "(food, rent, essentials, utilities, guruma's share, etc.)?",
            'B105 গত মাসে সব সদস্যের মোট যৌথ ব্যয় (খাদ্য, বাড়িভাড়া, নিত্যপ্রয়োজনীয় '
            'জিনিস, ইউটিলিটি বিল, গুরুমার অংশ ইত্যাদি) কত ছিল?',
            relevant="${b101_live_with}='2'"),
        _sr('calculate', 'b106_personal_income',
            calc="if(${b103_dera_members}!='' and number(${b103_dera_members})>0,"
                 " number(${b104_share}) - (number(${b105_shared_exp}) div "
                 "number(${b103_dera_members})), '')"),
        _sr('note', 'b106_note',
            'B106 Total personal income (taka/month) = B104 − (B105 ÷ B103): '
            '${b106_personal_income}',
            'B106 মোট ব্যক্তিগত আয় (টাকা/মাস) = B104 − (B105 ÷ B103): '
            '${b106_personal_income}', relevant="${b101_live_with}='2'"),
        _sr('note', 'b107_head',
            'B107 ‡ If not living in a dera: Over the past one month, how much money '
            'did your household earn from various sources? (Write 0 if none.) [Read '
            'each source neutrally; do not comment even if income comes from '
            'traditional occupation or sex work.]',
            'B107 ‡ যদি ডেরায় না থাকেন: গত এক মাসে আপনার পরিবারের বিভিন্ন উৎস থেকে কত '
            'টাকা আয় হয়েছে? (না থাকলে 0 লিখুন)। [প্রতিটি উৎস নিরপেক্ষভাবে পড়ুন; ঐতিহ্যগত '
            'পেশা বা যৌনকর্ম থেকে আয় হলেও কোনো মন্তব্য করবেন না।]',
            relevant="${b101_live_with}!='2'"),
    ]
    rows += [
        _sr('integer', name, '• ' + en, '• ' + bn, relevant="${b101_live_with}!='2'")
        for name, en, bn in _B107_SOURCES
    ]
    _b107_total = " + ".join(
        f"(if(${{{n}}}!='',number(${{{n}}}),0))" for n, _e, _b in _B107_SOURCES)
    rows += [
        _sr('calculate', 'b107_total', calc=_b107_total),
        _sr('note', 'b107_total_note', 'B107 Total: ${b107_total}',
            'B107 মোট: ${b107_total}', relevant="${b101_live_with}!='2'"),
        _sr('select_one yn12', 'b108_worked',
            'B108 In the past 7 days, did you work at least 1 hour to earn a living?',
            'B108 গত ৭ দিনে জীবিকা অর্জনের জন্য অন্তত ১ ঘণ্টা কাজ করেছেন কি?'),
        _sr('select_one yn12', 'b109_ready',
            'B109 In the past 7 days, were you ready/available to work?',
            'B109 গত ৭ দিনে কাজ করার জন্য আপনি প্রস্তুত/উপলব্ধ ছিলেন কি?',
            relevant="${b108_worked}='2'"),
        _sr('select_one yn12', 'b110_looked',
            'B110 In the past 7 days, did you look for work?',
            'B110 গত ৭ দিনে আপনি কি কাজ খুঁজেছেন?', relevant="${b108_worked}='2'"),
        _sr('select_one occupation', 'b111_main_occupation',
            'B111 ‡ What is your main occupation — i.e., the work from which you earn '
            'the most?',
            'B111 ‡ আপনার প্রধান পেশা কী—অর্থাৎ যে কাজ থেকে সবচেয়ে বেশি আয় হয়?'),
        _sr('text', 'b111_other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="${b111_main_occupation}='15'"),
        _sr('select_multiple occupation', 'b112_secondary',
            'B112 Do you have any secondary/additional occupation? (Multiple answers possible)',
            'B112 আপনার কি কোনো দ্বিতীয়/অতিরিক্ত পেশা আছে? (একাধিক উত্তর হতে পারে)'),
        _sr('text', 'b112_other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="selected(${b112_secondary},'15')"),
        _sr('end_group', 'grp_b'),
    ]
    return rows


def _hijra_module1_choices():
    rows = [
        _ch('respondent_type', '1', 'Hijra', 'হিজড়া'),
        _ch('respondent_type', '2', 'Transgender Woman', 'ট্রান্সজেন্ডার নারী'),
        _ch('respondent_type', '3', 'Transgender Man', 'ট্রান্সজেন্ডার পুরুষ'),
        _ch('respondent_type', '4', 'Man who identifies as MSM', 'MSM পরিচয়দানকারী পুরুষ'),
        _ch('respondent_type', '5', 'Lesbian Woman', 'লেসবিয়ান নারী'),
        _ch('respondent_type', '6', 'Bisexual Person', 'উভকামী (বাইসেক্সুয়াল) ব্যক্তি'),
        _ch('respondent_type', '7', 'Intersex Person', 'ইন্টারসেক্স ব্যক্তি'),
        _ch('respondent_type', '8', 'Queer/Non-binary/Other Gender-diverse Person',
            'কুইয়ার/নন-বাইনারি/অন্যান্য জেন্ডার বৈচিত্র্যময় ব্যক্তি'),

        _ch('area', '1', 'Urban', 'শহর'),
        _ch('area', '2', 'Peri-urban', 'শহরতলি'),
        _ch('area', '3', 'Rural', 'গ্রাম'),

        _ch('religion', '1', 'Islam', 'ইসলাম'),
        _ch('religion', '2', 'Hinduism', 'হিন্দু'),
        _ch('religion', '3', 'Christianity', 'খ্রিস্টান'),
        _ch('religion', '4', 'Buddhism', 'বৌদ্ধ'),
        _ch('religion', '5', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
        _ch('religion', '6', 'No religion', 'কোনো ধর্ম নেই'),

        _ch('ethnicity', '1', 'Bengali', 'বাঙালি'),
        _ch('ethnicity', '2', 'Ethnic minority (specify)', 'ক্ষুদ্র নৃগোষ্ঠী (উল্লেখ করুন)'),

        _ch('marital', '1', 'Never married, currently no partner',
            'কখনও বিয়ে হয়নি, বর্তমানে কোনো সঙ্গী নেই'),
        _ch('marital', '2', 'Never married, has a regular partner',
            'কখনও বিয়ে হয়নি, নিয়মিত সঙ্গী আছে'),
        _ch('marital', '3', 'Married (legal/social)', 'বিবাহিত (আইনগত/সামাজিক)'),
        _ch('marital', '4', 'Cohabiting without marriage', 'বিয়ে ছাড়া একসাথে বসবাস করছেন'),
        _ch('marital', '5', 'Separated/abandoned', 'আলাদা থাকেন/পরিত্যক্ত'),
        _ch('marital', '6', 'Divorced', 'তালাকপ্রাপ্ত'),
        _ch('marital', '7', 'Widowed', 'বিধবা/বিপত্নীক'),
        _ch('marital', '8', 'Hijra relationship (Parik/Neheri)', 'হিজড়া সম্পর্ক (পরিক/নেহেরি)'),

        _ch('education', '00', 'No formal education', 'কোনো প্রাতিষ্ঠানিক শিক্ষা নেই'),
        _ch('education', '01', 'Grades 1–4 completed (write grade completed)',
            '১ম–৪র্থ শ্রেণি সম্পন্ন (যে শ্রেণি সম্পন্ন করেছেন তা লিখুন)'),
        _ch('education', '05', 'Grades 5–7 completed', '৫ম–৭ম শ্রেণি সম্পন্ন'),
        _ch('education', '08', 'Grades 8–9 completed', '৮ম–৯ম শ্রেণি সম্পন্ন'),
        _ch('education', '10', 'SSC/Dakhil or equivalent', 'এসএসসি/দাখিল বা সমমান'),
        _ch('education', '12', 'HSC/Alim or equivalent', 'এইচএসসি/আলিম বা সমমান'),
        _ch('education', '14', "Bachelor's degree", 'স্নাতক'),
        _ch('education', '16', "Master's degree or above", 'স্নাতকোত্তর বা তার বেশি'),
        _ch('education', '19', 'Vocational/Trade course', 'কারিগরি/ট্রেড কোর্স'),
        _ch('education', '20', 'Madrasa (ungraded)/Other', 'মাদ্রাসা (গ্রেডবিহীন)/অন্যান্য'),

        _ch('mobile_phone', '1', 'Yes, smartphone', 'হ্যাঁ, স্মার্টফোন'),
        _ch('mobile_phone', '2', 'Yes, basic phone', 'হ্যাঁ, সাধারণ ফোন'),
        _ch('mobile_phone', '3', 'No', 'না'),

        _ch('nid_match', '1', 'Yes', 'হ্যাঁ'),
        _ch('nid_match', '2', 'No', 'না'),
        _ch('nid_match', '3', 'No NID', 'NID নেই'),
        _ch('nid_match', '99', 'Decline to answer', 'উত্তর দিতে অনিচ্ছুক'),

        _ch('a301_sex', '1', 'Male', 'পুরুষ'),
        _ch('a301_sex', '2', 'Female', 'নারী'),
        _ch('a301_sex', '3', 'Intersex/Other', 'ইন্টারসেক্স/অন্যান্য'),
        _ch('a301_sex', '99', 'Decline to answer', 'উত্তর দিতে অনিচ্ছুক'),

        _ch('a302_gender', '1', 'Male', 'পুরুষ'),
        _ch('a302_gender', '2', 'Female', 'নারী'),
        _ch('a302_gender', '3', 'Hijra', 'হিজড়া'),
        _ch('a302_gender', '4', 'Transgender Woman', 'ট্রান্সজেন্ডার নারী'),
        _ch('a302_gender', '5', 'Transgender Man', 'ট্রান্সজেন্ডার পুরুষ'),
        _ch('a302_gender', '6', 'Non-binary/Queer', 'নন-বাইনারি/কুইয়ার'),
        _ch('a302_gender', '7', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
        _ch('a302_gender', '99', 'Decline to answer', 'উত্তর দিতে অনিচ্ছুক'),

        _ch('live_with', '1', 'Own family (parents/siblings/relatives)',
            'নিজের পরিবার (মা-বাবা/ভাইবোন/আত্মীয়)'),
        _ch('live_with', '2', 'In a dera with a guruma', 'গুরুমার সাথে ডেরায়'),
        _ch('live_with', '3', 'With husband/partner (parik)', 'স্বামী/সঙ্গীর (পরিক) সাথে'),
        _ch('live_with', '4', 'With friends/roommate/mess/hostel',
            'বন্ধু/রুমমেট/মেস/হোস্টেলে'),
        _ch('live_with', '5', 'Alone', 'একা'),
        _ch('live_with', '6', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),

        _ch('occupation', '01', 'Government/NGO employment', 'সরকারি/বেসরকারি চাকরি'),
        _ch('occupation', '02', 'Non-agricultural day labour', 'অ-কৃষি দিনমজুরি'),
        _ch('occupation', '03', 'Agricultural worker', 'কৃষি শ্রমিক'),
        _ch('occupation', '04', 'Own business', 'নিজস্ব ব্যবসা'),
        _ch('occupation', '05', 'Badhai/blessing collection', 'বাধাই/আশীর্বাদ দিয়ে অর্থ সংগ্রহ'),
        _ch('occupation', '06', 'Dance performance collection', 'নাচের মাধ্যমে অর্থ সংগ্রহ'),
        _ch('occupation', '07', 'Market tolling', 'বাজার তোলা'),
        _ch('occupation', '08', 'Sex work', 'যৌনকর্ম'),
        _ch('occupation', '09', 'Beautician/parlour work', 'বিউটিশিয়ান/পার্লারের কাজ'),
        _ch('occupation', '10', 'Sewing/handicrafts', 'সেলাই/হস্তশিল্প'),
        _ch('occupation', '11', 'NGO/CBO peer educator or worker',
            'এনজিও/সিবিও পিয়ার এডুকেটর বা কর্মী'),
        _ch('occupation', '12', 'Student', 'শিক্ষার্থী'),
        _ch('occupation', '13', 'Unemployed', 'বেকার'),
        _ch('occupation', '14', 'Unable to work due to illness/disability',
            'অসুস্থতা/প্রতিবন্ধিতার কারণে কাজ করতে অক্ষম'),
        _ch('occupation', '15', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
    ]
    return rows


def _hijra_survey():
    rows = []
    rows += _hijra_meta()
    rows += _hijra_consent()
    rows += _hijra_screening()
    rows += _hijra_module1()
    from ._hijra_modules import hijra_module_survey
    mod = hijra_module_survey()
    # Module 4 skip fix (structural, by name — NOT a string replace). Source:
    # Q4.3 "2 → Q4.12" hides Q4.4–Q4.11 when the respondent had no sex in the
    # past 12 months. The transcriber gated q4_4–q4_6 but missed q4_7–q4_11, so
    # add `${q4_3}!='2'` to exactly those — never to q4_3 itself (self-cycle) or
    # q4_12+ (the skip target onward), which a global replace wrongly did.
    _Q4_GATE = {'q4_4', 'q4_5', 'q4_5_num', 'q4_6', 'q4_7_intro', 'grp_q4_7',
                'q4_7_a', 'q4_7_b', 'q4_7_c', 'q4_7_d', 'q4_8', 'q4_9', 'q4_10', 'q4_11'}
    for r in mod:
        if r[1] in _Q4_GATE and r[6] and '${q4_3}' not in r[6]:
            r[6] = r[6] + " and ${q4_3}!='2'"
    rows += mod
    return rows


def _hijra_choices():
    rows = []
    # District (A201) — value = slug, used by district_code calc.
    rows += [_ch('district', slug, en, bn) for slug, en, bn, _code in HIJRA_DISTRICTS]
    # Admin / screening lists.
    rows += [
        _ch('consent_yn', '1', 'Yes', 'হ্যাঁ'),
        _ch('consent_yn', '2', 'No', 'না'),
        _ch('yn12', '1', 'Yes', 'হ্যাঁ'),
        _ch('yn12', '2', 'No', 'না'),
        _ch('interview_language', '1', 'Bengali', 'বাংলা'),
        _ch('interview_language', '2', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
        _ch('interview_method', '1', 'Face-to-face interview', 'সামনাসামনি সাক্ষাৎকার'),
        _ch('interview_method', '2', 'CAPI Tablet', 'CAPI ট্যাবলেট'),
        _ch('selection_method', '1', 'Peer referral/respondent-driven',
            'Peer referral/respondent-driven'),
        _ch('selection_method', '2', 'CBO/DIC list', 'CBO/DIC তালিকা'),
        _ch('selection_method', '3', 'Hotspot/fixed venue', 'হটস্পট/নির্দিষ্ট স্থান'),
        _ch('selection_method', '4', 'Other', 'অন্যান্য'),
        _ch('residence_len', '1', 'Less than 6 months', '৬ মাসের কম'),
        _ch('residence_len', '2', '6 months or more', '৬ মাস বা তার বেশি'),
    ]
    rows += _hijra_module1_choices()
    from ._hijra_modules import hijra_module_choices
    rows += hijra_module_choices()
    return rows


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  FEMALE SEX WORKERS BASELINE  —  id: ciprb_baseline_fsw_v1                ║
# ║  Owner population: PHD service area · 9 districts × 12 sites · 194 Qs      ║
# ╚══════════════════════════════════════════════════════════════════════════╝

# A102 district stratum codes (FSW instrument): Rajbari 1 … Jamalpur 9.
FSW_DISTRICTS = [
    ('rajbari',    'Rajbari',    'রাজবাড়ী',   '1'),
    ('faridpur',   'Faridpur',   'ফরিদপুর',   '2'),
    ('jashore',    'Jashore',    'যশোর',      '3'),
    ('khulna',     'Khulna',     'খুলনা',     '4'),
    ('bagerhat',   'Bagerhat',   'বাগেরহাট',  '5'),
    ('patuakhali', 'Patuakhali', 'পটুয়াখালী', '6'),
    ('tangail',    'Tangail',    'টাঙ্গাইল',   '7'),
    ('mymensingh', 'Mymensingh', 'ময়মনসিংহ',  '8'),
    ('jamalpur',   'Jamalpur',   'জামালপুর',  '9'),
]

# A102 site / Primary Sampling Unit codes (11 named brothels + Street-based).
FSW_SITES = [
    ('01', 'Doulotdia Brothel (Rajbari)',          'দৌলতদিয়া ব্রোথেল (রাজবাড়ী)'),
    ('02', 'Rathkhola Brothel (Faridpur)',         'রথখোলা ব্রোথেল (ফরিদপুর)'),
    ('03', 'C&B Ghat Brothel (Faridpur)',          'সি অ্যান্ড বি ঘাট ব্রোথেল (ফরিদপুর)'),
    ('04', 'Babubazar Brothel (Jashore)',          'বাবুবাজার ব্রোথেল (যশোর)'),
    ('05', 'Maroari Mondir Brothel (Jashore)',     'মারোয়াড়ি মন্দির ব্রোথেল (যশোর)'),
    ('06', 'Baniasanta Brothel (Khulna)',          'বানিয়াশান্তা ব্রোথেল (খুলনা)'),
    ('07', 'Kachua Patti Brothel (Bagerhat)',      'কচুয়া পট্টি ব্রোথেল (বাগেরহাট)'),
    ('08', 'Puraton Hospital Road Brothel (Patuakhali)', 'পুরাতন হাসপাতাল রোড ব্রোথেল (পটুয়াখালী)'),
    ('09', 'Kandapara Brothel (Tangail)',          'কান্দাপাড়া ব্রোথেল (টাঙ্গাইল)'),
    ('10', 'Ganginarpar Brothel (Mymensingh)',     'গাঙ্গিনারপাড় ব্রোথেল (ময়মনসিংহ)'),
    ('11', 'Ranibazar Brothel (Jamalpur)',         'রানীবাজার ব্রোথেল (জামালপুর)'),
    ('12', 'Street-based',                          'স্ট্রিট ভিত্তিক'),
]


def _fsw_meta():
    return [
        _sr('begin_group', 'grp_admin', 'Interview identification',
            'সাক্ষাৎকার সনাক্তকরণ'),
        _sr('calculate', 'organisation', '', '', calc="'CIPRB'"),
        _sr('calculate', 'population',   '', '', calc="'fsw'"),
        _sr('calculate', 'survey_round', '', '', calc="'baseline'"),
        _sr('text', 'questionnaire_serial',
            'Questionnaire Serial No.', 'প্রশ্নমালার ক্রমিক নং', required='yes',
            constraint=("pulldata('respondents_fsw','serial','serial',"
                        + _NORM_SERIAL + ")=''"),
            cmsg='⚠ This serial is already recorded. Use a new, unique serial. / '
                 '⚠ এই সিরিয়ালটি ইতিমধ্যে রেকর্ড করা হয়েছে। নতুন, অনন্য সিরিয়াল ব্যবহার করুন।',
            hint='Must be unique per questionnaire.'),
        _sr('calculate', '_dup_serial',
            calc=("pulldata('respondents_fsw','serial','serial',"
                  + _NORM_SERIAL + ")")),
        _sr('note', '_dup_warn',
            '⚠ This serial is already recorded — do not enter the same respondent twice.',
            '⚠ এই সিরিয়ালটি ইতিমধ্যে রেকর্ড করা হয়েছে — একই উত্তরদাতাকে দুবার লিখবেন না।',
            relevant="${questionnaire_serial}!='' and ${_dup_serial}!=''"),
        _sr('select_one district', 'district', 'District Code', 'জেলা কোড', required='yes'),
        _sr('select_one fsw_site', 'site_code', 'Brothel Site Code', 'ব্রোথেল সাইট কোড',
            required='yes'),
        _sr('text', 'interviewer_name_code', 'Interviewer Name & Code',
            'সাক্ষাৎকার গ্রহণকারীর নাম ও কোড', required='yes'),
        _sr('text', 'supervisor_name_code', 'Supervisor Name & Code',
            'সুপারভাইজারের নাম ও কোড'),
        _sr('date', 'interview_date', 'Interview Date (DD/MM/YYYY)',
            'সাক্ষাৎকারের তারিখ (দিন/মাস/বছর)', required='yes', default='today()'),
        _sr('time', 'start_time', 'Start Time', 'শুরুর সময়'),
        _sr('integer', 'interview_attempts', 'Interview Attempt Number',
            'সাক্ষাৎকার গ্রহণের প্রচেষ্টার সংখ্যা'),
        _sr('select_one interview_language', 'interview_language',
            'Interview Language', 'সাক্ষাৎকারের ভাষা'),
        _sr('text', 'interview_language_other', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)',
            relevant="${interview_language}='2'"),
        _sr('geopoint', 'location',
            'Geographic Location (Cluster Level) — Latitude/Longitude (required; cluster level only)',
            'ভৌগোলিক অবস্থান (ক্লাস্টার পর্যায়) — অক্ষাংশ/দ্রাঘিমাংশ (প্রয়োজনীয়; শুধু ক্লাস্টার পর্যায়ে)',
            required='yes'),
        _sr('end_group', 'grp_admin'),
    ]


def _fsw_consent():
    return [
        _sr('note', 'consent_instr',
            '[Read aloud verbatim in Bangla. Provide the information sheet to the '
            'respondent and allow her to ask questions. Do not begin the interview '
            'before obtaining consent.]',
            '[বাংলায় হুবহু পড়ে শোনাতে হবে। উত্তরদাতাকে তথ্যপত্র দিতে হবে এবং প্রশ্ন করার '
            'সুযোগ দিতে হবে। সম্মতি গ্রহণের আগে সাক্ষাৎকার শুরু করা যাবে না।]'),
        _sr('note', 'consent_script',
            '"As-salamu alaykum / Good morning / Good afternoon. My name is '
            '__________________. I am working on behalf of __________________ (name of '
            'data collection organization / CIPRB). We are conducting a baseline study '
            "for a project titled 'Strengthening Access to Integrated Sexual and "
            'Reproductive Health and Rights (SRHR) Services for Brothel- and '
            "Street-Based Female Sex Workers in Selected Districts of Bangladesh.' The "
            'project is being implemented by PHD and is funded by UNFPA. The purpose of '
            'this interview is to learn about the health situation, needs, and '
            'experiences of female sex workers in your area so that future services, '
            'including Wellness Centres, can be planned according to their needs. Your '
            'participation is entirely voluntary. No information that could reveal your '
            'name or identity will be used in any report or publication. Your answers '
            'will be stored under a code number and all information will be kept '
            'strictly confidential. The information you provide will not be shared with '
            'your family, sardarni, landlady, other residents, clients, community '
            'leaders, police, or any government institution. You may decline to answer '
            'any question, take a break, or stop the interview at any time without any '
            'harm or disadvantage to you. The interview will take approximately 30–40 '
            'minutes. Do you have any questions? May I begin the interview?"',
            '"আসসালামু আলাইকুম/শুভ সকাল/শুভ অপরাহ্ন। আমার নাম __________________। আমি '
            '__________________ (তথ্য সংগ্রহকারী প্রতিষ্ঠানের নাম/ CIPRB)-এর পক্ষে কাজ করছি। '
            'আমরা ‘বাংলাদেশের নির্বাচিত জেলায়  ব্রোথেল ও স্ট্রিট ভিত্তিক নারী যৌনকর্মীদের জন্য '
            'সমন্বিত যৌন ও প্রজনন স্বাস্থ্য ও অধিকার (SRHR) সেবায় প্রবেশাধিকার শক্তিশালীকরণ’ '
            'শীর্ষক প্রকল্পের একটি বেসলাইন গবেষণা পরিচালনা করছি। প্রকল্পটি PHD বাস্তবায়ন করছে '
            'এবং UNFPA এর অর্থায়নে পরিচালিত হচ্ছে। এই সাক্ষাৎকারের উদ্দেশ্য হলো আপনার এলাকার '
            'নারী যৌনকর্মীদের স্বাস্থ্য পরিস্থিতি, চাহিদা এবং অভিজ্ঞতা সম্পর্কে জানা, যাতে '
            'ভবিষ্যতে ওয়েলনেস সেন্টারসহ বিভিন্ন সেবা তাদের প্রয়োজন অনুযায়ী পরিকল্পনা করা যায়। '
            'আপনার অংশগ্রহণ সম্পূর্ণ স্বেচ্ছাসেবী। আপনার নাম বা পরিচয় প্রকাশ করতে পারে এমন কোনো '
            'তথ্য কোনো প্রতিবেদন বা প্রকাশনায় ব্যবহার করা হবে না। আপনার উত্তর একটি কোড নম্বরের '
            'মাধ্যমে সংরক্ষণ করা হবে এবং সব তথ্য কঠোরভাবে গোপন রাখা হবে। আপনার দেওয়া তথ্য আপনার '
            'পরিবার, সর্দারনি, বাড়িওয়ালি, অন্য বাসিন্দা, খদ্দের, কমিউনিটি নেতা, পুলিশ বা কোনো '
            'সরকারি প্রতিষ্ঠানের সঙ্গে শেয়ার করা হবে না। আপনি চাইলে যেকোনো প্রশ্নের উত্তর না '
            'দিতে পারেন, বিরতি নিতে পারেন অথবা যেকোনো সময় সাক্ষাৎকার বন্ধ করতে পারেন। এতে '
            'আপনার কোনো ক্ষতি বা অসুবিধা হবে না। সাক্ষাৎকারটি সম্পন্ন করতে প্রায় ৩০–৪০ মিনিট '
            'সময় লাগবে। আপনার কোনো প্রশ্ন আছে কি? আমি কি সাক্ষাৎকার শুরু করতে পারি?"'),
        _sr('select_one consent_yn', 'consent', 'Consent given', 'সম্মতি প্রদান করেছেন',
            required='yes'),
        _sr('note', 'consent_no_end',
            'No consent — thank the respondent and end the interview.',
            'সম্মতি নেই — উত্তরদাতাকে ধন্যবাদ জানিয়ে সাক্ষাৎকার শেষ করুন।',
            relevant="${consent}='2'"),
    ]


def _fsw_screening():
    return [
        _sr('begin_group', 'grp_screen', 'Eligibility Screening', 'যোগ্যতা যাচাই',
            relevant="${consent}='1'"),
        _sr('integer', 's1_age',
            'What is your current age in years? [If respondent is unsure, help '
            'estimate age using national ID, significant life events, or '
            'calendar/festival-based information.]',
            'আপনার বর্তমান বয়স কত বছর? [উত্তরদাতা নিশ্চিত না হলে জাতীয় পরিচয়পত্র, গুরুত্বপূর্ণ '
            'জীবনের ঘটনা বা উৎসব/পঞ্জিকা ভিত্তিক তথ্য ব্যবহার করে বয়স নির্ধারণে সহায়তা করুন।]',
            required='yes', constraint='. >= 0 and . <= 120'),
        _sr('note', 's1_ineligible',
            'Under 18 — politely conclude the interview; record as ineligible and '
            'follow the referral protocol for minors.',
            '১৮ বছরের কম হলে → বিনয়ের সঙ্গে সাক্ষাৎকার সমাপ্ত করুন; অযোগ্য হিসেবে রেকর্ড করুন '
            'এবং অপ্রাপ্তবয়স্কদের জন্য নির্ধারিত রেফারেল প্রোটোকল অনুসরণ করুন।',
            relevant="${s1_age}!='' and ${s1_age}<18"),
        _sr('select_one yn12', 's2_sexwork',
            'In the past 6 months, have you worked as a sex worker at least once? '
            '[Use community-acceptable, neutral language as trained. Never use the '
            'word "prostitute" or any derogatory term.]',
            'গত ৬ মাসের মধ্যে, আপনি কি অন্তত একবার যৌনকর্মী হিসেবে কাজ করছেন? [প্রশিক্ষণে '
            'ব্যবহৃত সম্প্রদায়-গ্রহণযোগ্য ও নিরপেক্ষ ভাষা ব্যবহার করুন। কখনোই ‘পতিতা’ বা '
            'অবমাননাকর কোনো শব্দ ব্যবহার করবেন না।]',
            required='yes', hint='‡'),
        _sr('select_one s3_residence', 's3_residence',
            'How long have you been staying or working in this area (brothel or street)?',
            'আপনি কতদিন ধরে এই এলাকায় (ব্রোথেল ও স্ট্রিট) থাকছেন অথবা এখান থেকে কাজ করছেন?',
            required='yes'),
        _sr('select_one s4_list', 's4_sampling',
            'For the interviewer: Confirm that the respondent was selected from the '
            'random sampling list prepared for this site (brothel or street). '
            '[Do not read aloud.]',
            'সাক্ষাৎকার গ্রহণকারীর জন্য: নিশ্চিত করুন যে উত্তরদাতা এই সাইটে (ব্রোথেল ও স্ট্রিট)র '
            'জন্য প্রস্তুতকৃত দৈবচয়ন (Random) তালিকা থেকে নির্বাচিত হয়েছেন। [প্রশ্নটি উত্তরদাতাকে '
            'পড়ে শোনানো হবে না।]'),
        _sr('text', 's4_replacement_reason', 'State reason (replacement respondent)',
            'কারণ উল্লেখ করুন (প্রতিস্থাপিত উত্তরদাতা)', relevant="${s4_sampling}='2'"),
        _sr('note', 's_ineligible_end',
            'Respondent is not eligible — politely conclude the interview.',
            'উত্তরদাতা যোগ্য নন — বিনয়ের সঙ্গে সাক্ষাৎকার সমাপ্ত করুন।',
            relevant="(${s1_age}!='' and ${s1_age}<18) or ${s2_sexwork}='2' or ${s3_residence}='1'"),
        _sr('end_group', 'grp_screen'),
        _sr('calculate', '_proceed',
            calc="if(${consent}='1' and ${s1_age}>=18 and ${s2_sexwork}='1' "
                 "and ${s3_residence}='2','1','0')"),
    ]


def _fsw_survey():
    rows = []
    rows += _fsw_meta()
    rows += _fsw_consent()
    rows += _fsw_screening()
    from ._fsw_modules import fsw_module_survey
    rows += fsw_module_survey()
    return rows


def _fsw_choices():
    rows = []
    rows += [_ch('district', slug, en, bn) for slug, en, bn, _c in FSW_DISTRICTS]
    rows += [_ch('fsw_site', code, en, bn) for code, en, bn in FSW_SITES]
    rows += [
        _ch('consent_yn', '1', 'Yes', 'হ্যাঁ'),
        _ch('consent_yn', '2', 'No', 'না'),
        _ch('yn12', '1', 'Yes', 'হ্যাঁ'),
        _ch('yn12', '2', 'No', 'না'),
        _ch('interview_language', '1', 'Bangla', 'বাংলা'),
        _ch('interview_language', '2', 'Other (specify)', 'অন্যান্য (উল্লেখ করুন)'),
        _ch('s3_residence', '1', 'Less than 6 months', '৬ মাসের কম'),
        _ch('s3_residence', '2', '6 months or more', '৬ মাস বা তার বেশি'),
        _ch('s4_list', '1', 'Yes, she is on the selected list', 'হ্যাঁ, নির্বাচিত তালিকায় আছেন'),
        _ch('s4_list', '2', 'Replacement respondent per protocol (state reason)',
            'প্রোটোকল অনুযায়ী প্রতিস্থাপিত উত্তরদাতা (কারণ উল্লেখ করুন)'),
    ]
    from ._fsw_modules import fsw_module_choices
    rows += fsw_module_choices()
    return rows


# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  FORM REGISTRY + COMMAND                                                  ║
# ╚══════════════════════════════════════════════════════════════════════════╝

FORMS = [
    {
        'id': 'ciprb_baseline_hijra_v1',
        'title': 'Baseline Survey — Hijra / Gender-diverse Population (CIPRB)',
        'file': 'CIPRB_Baseline_Hijra.xlsx',
        'survey':  _hijra_survey,
        'choices': _hijra_choices,
    },
    {
        'id': 'ciprb_baseline_fsw_v1',
        'title': 'Baseline Survey — Female Sex Workers (Brothel & Street) (CIPRB)',
        'file': 'CIPRB_Baseline_FSW.xlsx',
        'survey':  _fsw_survey,
        'choices': _fsw_choices,
    },
]


class Command(BaseCommand):
    help = ('Build the two D5 baseline XLSForms (Hijra + FSW), bilingual EN/BN, '
            'verbatim. Deploy separately via _deploy_one.py into known asset UIDs.')

    def add_arguments(self, parser):
        parser.add_argument('--output-dir', default=OUTDIR)
        parser.add_argument('--only', default='',
            help='Build ONLY the form with this id (e.g. ciprb_baseline_hijra_v1).')

    def handle(self, *args, **opts):
        out = opts['output_dir']
        os.makedirs(out, exist_ok=True)
        only = opts.get('only', '')
        for f in FORMS:
            if only and f['id'] != only:
                continue
            survey  = f['survey']()
            choices = f['choices']()
            wb = _wb(f['id'], f['title'], survey, choices)
            path = os.path.join(out, f['file'])
            wb.save(path)
            self.stdout.write(self.style.SUCCESS(
                f"  OK  {f['file']:34s}  {len(survey):3d} rows  id: {f['id']}"))
        self.stdout.write(f'\nWritten to {os.path.abspath(out)}/')
