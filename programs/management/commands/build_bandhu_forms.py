# -*- coding: utf-8 -*-
"""
Build Bandhu's KoboToolbox XLSForms — corrective rebuild from the FINAL tools
(MIS Tools_100626.xlsx) + Bandhu's marked corrections (M&E Tools correction).

Three forms now (was two):

  bandhu_mother_list_v1     — F-1.1 Mother List (beneficiary registration →
                              creates the Client; source for pulldata autofill)
  bandhu_service_log_v1     — per-client registers (F-01, F-05, F-06, F-02,
                              F-03, Counseling, Referral, F-08)
  bandhu_activity_ops_v1    — aggregate/event/ops (F-04, F-10, F-11, F-12,
                              F-13, F-07 KP Clinic Info, F-09 Wellness Center
                              Info, F-14 e-billboard)

Corrections applied:
  - TG codes UNIFIED everywhere: MSM=01, MSW=02, FSW=03, EVA=04, TG/Hijra=05,
    Others=06 (the per-form schemes are gone).
  - "DIC" → "Wellness Center" (centre dropdown shows the name only).
  - F-01: drop Host/Rohingya, add Lubricant, Referral is multi-select.
  - F-04: add Lubricant distribution + Mental Health & GBV referrals;
    "Peer Educator".
  - F-08: "Wellness Center"; Linked with = AAS/MAB/CAAP/Bandhu/Govt. ART
    Center/Others.
  - F-11: "Name" (not Legal Name); Gender adds TG/Hijra.
  - F-12: participants by gender = Man / Woman / TG/Hijra / Others.
  - F-07 & F-09 added as info forms (update the centre roster).
"""
import os
import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.core.management.base import BaseCommand

HERE   = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.normpath(os.path.join(HERE, '..', '..', '..', '..', 'koboforms'))
KOBO_BASE = 'https://kf.kobotoolbox.org'

_HFILL = PatternFill("solid", fgColor="6A1B9A")
_HFONT = Font(color="FFFFFF", bold=True, size=10)

SURVEY_HDR = [
    'type', 'name', 'label::English', 'label::Bangla',
    'hint', 'required', 'relevant', 'constraint', 'constraint_message',
    'default', 'appearance', 'calculation',
]
CHOICES_HDR = ['list_name', 'name', 'label::English', 'label::Bangla']
SETTINGS_HDR = ['form_title', 'form_id', 'version', 'default_language', 'style']


def _sr(qtype, name, en='', bn='', hint='', required='',
        relevant='', constraint='', cmsg='', default='', app='', calc=''):
    return [qtype, name, en, bn, hint, required, relevant,
            constraint, cmsg, default, app, calc]


def _ch(lst, name, en, bn=''):
    # Fall back to the English+code label when no Bangla is given, so a coded
    # option never renders BLANK in Bangla mode (KoboCollect language toggle).
    return [lst, name, en, bn or en]


def _id_lookup(idfield):
    """Autofill rows for a service-form client ID — read the registered name
    from bandhu_clients.csv (the Mother List, synced via export_bandhu_clients)
    and show it read-only, or warn the ID is not registered. Mirrors the PHD
    Service Log lookup. Spread into a survey list right after the ID field:
    `_sr('text','pr_client_id',...), *_id_lookup('pr_client_id'),`."""
    nm = '_%s_name' % idfield
    norm = ("translate(normalize-space(${%s}),"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')" % idfield)
    return [
        _sr('calculate', nm, calc="pulldata('bandhu_clients','name','id_no',%s)" % norm),
        _sr('note', '_%s_ok' % idfield,
            '👤 Client: ${%s}' % nm, '👤 ক্লায়েন্ট: ${%s}' % nm,
            relevant="${%s}!='' and ${%s}!=''" % (idfield, nm)),
        _sr('note', '_%s_warn' % idfield,
            '⚠ This ID is not in the Mother List — register the client first, or check the ID.',
            '⚠ এই আইডি মাদার লিস্টে নেই — আগে ক্লায়েন্ট নিবন্ধন করুন বা আইডি যাচাই করুন।',
            relevant="${%s}!='' and ${%s}=''" % (idfield, nm)),
    ]


def _age_from_ml(idfield, agefield, label_en='Age (as per Mother List)',
                 label_bn='বয়স (মাদার লিস্ট অনুযায়ী)'):
    """Auto-fill a client's age from bandhu_clients.csv (the Mother List) rather
    than typing it (Ashis review pt 4). Returns a hidden calculate that pulls the
    registered age by ID + a read-only note showing it. Use in place of a manual
    age field: `*_age_from_ml('pr_client_id', 'pr_age'),`. The backend still reads
    the same field name — the value now comes from the lookup, not manual entry."""
    norm = ("translate(normalize-space(${%s}),"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ')" % idfield)
    return [
        _sr('calculate', agefield,
            calc="pulldata('bandhu_clients','age','id_no',%s)" % norm),
        _sr('note', '_%s_show' % agefield,
            '%s: ${%s}' % (label_en, agefield), '%s: ${%s}' % (label_bn, agefield),
            relevant="${%s}!=''" % agefield),
    ]


def _wb(form_id, form_title, survey, choices):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in [
        ('survey',   SURVEY_HDR,   survey),
        ('choices',  CHOICES_HDR,  choices),
        ('settings', SETTINGS_HDR, [[form_title, form_id, '20260620', 'English', 'theme-grid']]),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for ci in range(1, len(headers) + 1):
            c = ws.cell(1, ci)
            c.font = _HFONT
            c.fill = _HFILL
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(ri, ci, value=val).alignment = Alignment(wrap_text=True, vertical='top')
        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 28
        ws.freeze_panes = 'A2'
    return wb


# ─── Shared submission header ─────────────────────────────────────────────────

def _meta(filler_label_en='Your name (person filling this form)',
          filler_label_bn='আপনার নাম'):
    return [
        _sr('begin_group', 'grp_meta', 'Submission info', 'তথ্য প্রেরণ'),
        _sr('calculate', 'organisation', '', '', calc="'Bandhu'"),
        _sr('geopoint', 'location',
            'GPS location (required — step outside if no signal)',
            'জিপিএস অবস্থান (প্রয়োজনীয়)', required='yes'),
        _sr('date', 'collection_date', 'Date', 'তারিখ', required='yes'),
        _sr('select_one bandhu_centre', 'centre_id',
            'Wellness Centre', 'ওয়েলনেস সেন্টার', required='yes',
            hint='Select your Wellness Centre / KP clinic.'),
        _sr('text', 'enumerator_name', filler_label_en, filler_label_bn, required='yes'),
        _sr('end_group', 'grp_meta'),
    ]


def _centre_choices():
    """Bandhu's 8 Wellness Centres + Dhaka KP clinic. Label shows the centre
    name + its 2-digit district code — the code that prefixes every beneficiary
    ID at that centre — so the worker can see/verify it. value = centre code."""
    from .seed_centers import BONDHU_DICS
    out = []
    for c in BONDHU_DICS:
        dcode = BANDHU_DISTRICT_CODE.get(c.get('district', ''), '00')
        en = '%s (%s)' % (c['name'], dcode)
        bn = '%s (%s)' % (c.get('name_bangla', c['name']), dcode)
        out.append(_ch('bandhu_centre', c['code'], en, bn))
    return out


# ─── Beneficiary-ID district codes (Bandhu handwritten note, 2026-06-20) ───────
# Each Wellness Centre sits in one district. The beneficiary ID is composed as
# {2-digit district code}-{4-digit serial} so every ID has one fixed shape and
# the service forms' pulldata lookup always resolves. Codes are Bandhu's own,
# keyed by the centre's district.
BANDHU_DISTRICT_CODE = {
    'Bandarban': '01', 'Chittagong': '02', 'Chattogram': '02',
    'Chandpur': '03', 'Noakhali': '04', 'Sunamganj': '05',
    'Habiganj': '06', 'Manikganj': '07', 'Narayanganj': '08',
    'Dhaka': '09',   # Dhaka KP clinic — confirmed as the 9th Bandhu code (2026-06-20).
}


def _centre_to_district_code():
    """centre_id (BND-DIC-xx / BND-KPC-xx) → 2-digit Bandhu district code."""
    from .seed_centers import BONDHU_DICS
    return {c['code']: BANDHU_DISTRICT_CODE.get(c.get('district', ''), '00')
            for c in BONDHU_DICS}


def _bandhu_dist_code_calc():
    """Nested if() mapping the selected centre to its 2-digit district code.
    Enketo is XPath 1.0 (no lookup tables), so an explicit if() chain."""
    expr = "'00'"
    for code, dcode in reversed(list(_centre_to_district_code().items())):
        expr = "if(${centre_id}='%s','%s',%s)" % (code, dcode, expr)
    return expr


# ─── Reusable choice lists (corrected, unified) ───────────────────────────────

def _shared_choices():
    rows = []
    rows += [_ch('yes_no', 'yes', 'Yes', 'হ্যাঁ'), _ch('yes_no', 'no', 'No', 'না')]

    # UNIFIED TG code — every tool now uses this exact list.
    for v, en in [('01', 'MSM'), ('02', 'MSW'), ('03', 'FSW'),
                  ('04', 'EVA'), ('05', 'TG/Hijra'), ('06', 'Others')]:
        rows.append(_ch('tg_code', v, f'{en} ({v})'))

    for v, en in [('01', 'General'), ('02', 'Diverse')]:
        rows.append(_ch('general_diverse', v, f'{en} ({v})'))

    # F-01 referral codes (now multi-select)
    for v, en in [('01', 'STI'), ('02', 'GH'), ('03', 'Counseling'),
                  ('04', 'Mental Health'), ('05', 'FP'), ('06', 'Legal'),
                  ('07', 'Lab Test'), ('08', 'Other')]:
        rows.append(_ch('f01_referral', v, f'{en} ({v})'))

    for v, en, bn in [('negative', 'HIV-', 'এইচআইভি নেগেটিভ'),
                      ('positive', 'HIV+', 'এইচআইভি পজিটিভ'),
                      ('indeterminate', 'Indeterminate', 'অনির্ণেয়')]:
        rows.append(_ch('hiv_result', v, en, bn))

    for v, en in [('new', 'New'), ('follow_up', 'Follow up'),
                  ('recurrent', 'Recurrent (within last 6 months)')]:
        rows.append(_ch('sti_case', v, en))

    for v, en in [('uds', 'UDS'), ('vds', 'VDS'), ('gu', 'GU'), ('pid', 'PID'),
                  ('ss', 'SS'), ('ib', 'IB'), ('anal_sti', 'Anal STIs'),
                  ('sti_other', 'Other STI'), ('gh', 'GH'), ('psd', 'PSD'),
                  ('mental_health', 'Mental health')]:
        rows.append(_ch('diagnosis', v, en))

    for v, en in [('within_7', 'Within 7 days'), ('more_7', 'More than 7 days')]:
        rows.append(_ch('seek_timing', v, en))

    for v, en in [('treat_center', 'Treatment at centre'),
                  ('medicine_gdp_pe', 'Medicine through GDP/PE'),
                  ('prescription_only', 'Only prescription provided')]:
        rows.append(_ch('partner_mgmt', v, en))

    for v, en in [('tb', 'TB (suspected, for diagnosis/management)'),
                  ('sti_kp', 'STI-KP (non-responsive, complicated)'),
                  ('sti_partner', 'STI-Partner (non-responsive, complicated)'),
                  ('general_health', 'General Health'), ('htc_hts', 'HTC/HTS'),
                  ('art', 'ART'), ('maternal', 'Maternal Health'),
                  ('fp', 'Family Planning'), ('gbv', 'GBV'),
                  ('legal', 'Legal Support'), ('other', 'Others')]:
        rows.append(_ch('referred_for', v, en))

    # F-05 Patient Record — its OWN Referral Cases column set (tool F-05 r7);
    # not the Referral Register list. Exactly: TB | STI-KP | STI-Partner |
    # General Health | HIV Testing | Mental Health | HTC | FP methods.
    for v, en in [('tb', 'TB (suspected, for diagnosis/management)'),
                  ('sti_kp', 'STI-KP (non-responsive, complicated)'),
                  ('sti_partner', 'STI-Partner (non-responsive, complicated)'),
                  ('general_health', 'General Health'),
                  ('hiv_testing', 'HIV Testing'),
                  ('mental_health', 'Mental Health'),
                  ('htc', 'HTC'),
                  ('fp', 'FP methods')]:
        rows.append(_ch('f05_referral', v, en))

    # F-10 Mobile Health Camp — its OWN Referral Cases column set (tool F-10 r7):
    # TB | STI | General Health | ART Linkage | Mental Health | GBV | Legal | FP.
    for v, en in [('tb', 'TB (suspected, for diagnosis/management)'),
                  ('sti', 'STI (non-responsive, complicated)'),
                  ('general_health', 'General Health'),
                  ('art_linkage', 'ART Linkage (if HIV+)'),
                  ('mental_health', 'Mental Health'),
                  ('gbv', 'GBV'),
                  ('legal', 'Legal Support'),
                  ('fp', 'FP methods')]:
        rows.append(_ch('f10_referral', v, en))

    for v, en in [('man', 'Man'), ('women', 'Women'), ('hijras', 'Hijras'), ('others', 'Others')]:
        rows.append(_ch('sex_with', v, en))
    for v, en in [('anal', 'Anal'), ('oral', 'Oral'), ('peno_vaginal', 'Peno-Vaginal')]:
        rows.append(_ch('sex_activity', v, en))
    for v, en in [('some_times', 'Some times'), ('all_times', 'All Times'),
                  ('use_last_sex', 'Used last sex'),
                  ('not_use_last_sex', 'Not used last sex'), ('never', 'Never')]:
        rows.append(_ch('condom_use', v, en))
    for v, en in [('mental_health', 'Mental Health'), ('gbv', 'GBV')]:
        rows.append(_ch('mh_counsel_type', v, en))

    for v, en in [('yes', 'Yes'), ('no', 'No'), ('drop_out', 'Drop out')]:
        rows.append(_ch('receiving_art', v, en))
    # F-08 Linked with — corrected list
    for v, en in [('aas', 'AAS'), ('mab', 'MAB'), ('caap', 'CAAP'),
                  ('bandhu', 'Bandhu'), ('govt_art', 'Govt. ART Center'),
                  ('others', 'Others')]:
        rows.append(_ch('linked_with', v, en))

    for v, en in [('sti', 'STI'), ('general_health', 'General Health'),
                  ('fp', 'FP (Family Planning)'), ('mental_health', 'Mental Health'),
                  ('harmful_drug', 'Harmful use drugs as infestation'),
                  ('psychosocial', 'Phycosocial Counseling'),
                  ('gbv', 'GBV'), ('other', 'Others')]:
        rows.append(_ch('counsel_issue', v, en))
    for v, en in [('demonstration', 'Demonstration'), ('distribution', 'Distribution')]:
        rows.append(_ch('counsel_condom', v, en))
    for v, en in [('mental_health', 'Mental Health'), ('legal', 'Legal Services'),
                  ('htc_hts', 'HTC/HTS'), ('gbv', 'GBV'),
                  ('complicated_sti', 'Complicated STI'), ('other', 'Others (specify)')]:
        rows.append(_ch('counsel_referral', v, en))

    # F-11 attendance gender (TG/Hijra added) + age band
    for v, en in [('man', 'Man'), ('woman', 'Woman'),
                  ('tg_hijra', 'TG/Hijra'), ('others', 'Others')]:
        rows.append(_ch('att_gender', v, en))
    for v, en in [('18_24', '18-24'), ('25_30', '25-30'),
                  ('31_35', '31-35'), ('gt35', '>35')]:
        rows.append(_ch('age_band', v, en))

    for v, en in [('in_person', 'In person'), ('online', 'Online'), ('hybrid', 'Hybrid')]:
        rows.append(_ch('event_modality', v, en))
    # event_kind: NOT a column on the F-12 paper tool, but the backend routes
    # F-12 events to the right indicator (Training vs Coordination vs Observance)
    # by this value (bandhu_handlers._bnd_event). Kept deliberately for that
    # reason — flagged to Rafi as a fidelity exception, not an accidental field.
    for v, en in [
        ('orientation_managers', 'Orientation — health managers/supervisors'),
        ('training_midwives',    'Training — midwives/providers'),
        ('training_peers',       'Training — community leaders/peer educators'),
        ('coord_gob',            'Coordination meeting — GOB/NGO'),
        ('coord_cbo',            'Coordination meeting — CBO/network'),
        ('observance',           'Observance event — World AIDS Day etc.'),
        ('other',                'Other'),   # Ashis review pt 6 — routed to INTERNAL (no indicator)
    ]:
        rows.append(_ch('event_kind', v, en))

    # Mother List parameters
    for v, en in [('1', 'Illiterate'), ('2', 'Primary'), ('3', 'Secondary'),
                  ('4', 'Higher Secondary'), ('5', 'Graduate / Masters')]:
        rows.append(_ch('education', v, f'{en} ({v})'))
    for v, en in [('1', 'Single (never married)'), ('2', 'Married'),
                  ('3', 'Widowed'), ('4', 'Separated'), ('5', 'Divorced')]:
        rows.append(_ch('marital', v, f'{en} ({v})'))
    # Mother List col 12 — income source / occupation (tool F-1.1 coded list).
    for v, en, bn in [('1', 'Service / job holder', 'চাকুরিজীবী'),
                      ('2', 'Businessman', 'ব্যবসায়ী'), ('3', 'Student', 'ছাত্র'),
                      ('4', 'Sex work', 'যৌনপেশা'), ('5', 'Unemployed', 'বেকার'),
                      ('6', 'Others', 'অন্যান্য')]:
        rows.append(_ch('occupation', v, f'{en} ({v})', f'{bn} ({v})'))

    # Mother List col 17 — current status (tool F-1.1 r5).
    for v, en in [('1', 'Not found'), ('2', 'In jail'), ('3', 'Left the place'),
                  ('4', 'Others'), ('5', 'Dead')]:
        rows.append(_ch('ml_status', v, f'{en} ({v})'))
    # Yes-1 / No-0 coded (Mother List cols 14-16).
    for v, en in [('1', 'Yes (1)'), ('0', 'No (0)')]:
        rows.append(_ch('yn_code', v, en))

    return rows


# ─── FORM 0 — Mother List (registration) ──────────────────────────────────────

def _mother_list_survey():
    rows = _meta('Information collector (Peer Educator)', 'তথ্য সংগ্রাহক (পিয়ার এডুকেটর)')
    rows += [
        _sr('begin_group', 'grp_ml', 'Mother List — Beneficiary Registration',
            'মাদার লিস্ট — সুবিধাভোগী নিবন্ধন'),
        # Beneficiary ID — auto-composed as {district code}-{4-digit serial}.
        # District code comes from the chosen centre (worker can't mistype it);
        # the worker enters only the serial. One fixed format (DD-NNNN) so the
        # service forms' pulldata lookup always finds the registered mother —
        # fixes the old free-text IDs (02001 / 020007 / 020010) that never
        # matched on lookup and broke "registration".
        _sr('calculate', 'centre_district_code', calc=_bandhu_dist_code_calc()),
        _sr('text', 'ml_serial', 'Beneficiary serial (4 digits)',
            'সুবিধাভোগী ক্রমিক (৪ অঙ্ক)', required='yes',
            constraint="regex(., '^[0-9]{4}$')",
            cmsg='Enter exactly 4 digits, e.g. 0001. / ঠিক ৪ অঙ্ক দিন, যেমন 0001।',
            hint='Next free number at your centre (0001, 0002, …). The district '
                 'code is added automatically. / আপনার কেন্দ্রের পরবর্তী নম্বর; '
                 'জেলা কোড স্বয়ংক্রিয়ভাবে যুক্ত হবে।'),
        _sr('calculate', 'ml_id_no',
            calc="concat(${centre_district_code}, '-', ${ml_serial})"),
        _sr('note', '_ml_id_show', '🆔 Beneficiary ID: ${ml_id_no}',
            '🆔 সুবিধাভোগী আইডি: ${ml_id_no}', relevant="${ml_serial}!=''"),
        # Duplicate-ID warning from the bandhu_clients.csv attachment.
        _sr('calculate', '_dup_name',
            calc=("pulldata('bandhu_clients','name','id_no',"
                  "translate(normalize-space(${ml_id_no}),"
                  "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'))")),
        _sr('note', '_dup_warn',
            '⚠ This ID is already registered for ${_dup_name}. Use the existing ID in service forms.',
            '⚠ এই আইডি ইতিমধ্যে ${_dup_name} এর জন্য নিবন্ধিত।',
            relevant="${ml_id_no}!='' and ${_dup_name}!=''"),
        _sr('text', 'ml_name', 'Name', 'নাম', required='yes'),
        _sr('text', 'ml_parent_name', "Father's / Mother's name", 'পিতা/মাতার নাম'),
        _sr('integer', 'ml_birth_year', 'Birth year', 'জন্ম সাল',
            constraint='. >= 1940 and . <= 2012', cmsg='1940–2012'),
        _sr('select_one tg_code', 'ml_gender', 'Gender (TG code)', 'লিঙ্গ (টিজি কোড)'),
        _sr('text', 'ml_address', 'Current address (area name)', 'বর্তমান ঠিকানা (এলাকার নাম)'),
        _sr('text', 'ml_spot', 'Spot name', 'স্পটের নাম'),
        _sr('select_one education', 'ml_education', 'Education', 'শিক্ষা'),
        _sr('select_one marital', 'ml_marital', 'Marital status', 'বৈবাহিক অবস্থা'),
        # Skip children count for never-married (marital '1') — Ashis review pt 2.
        _sr('integer', 'ml_children_u18', 'Number of children under 18', '১৮ বছরের নিচে সন্তান সংখ্যা',
            relevant="${ml_marital}!='1'"),
        _sr('select_one occupation', 'ml_occupation', 'Income source (occupation)', 'আয়ের উৎস (পেশা)'),
        _sr('note', '_ml_avg', 'Average sex-work contacts (fill the applicable period):', ''),
        _sr('integer', 'ml_avg_day', 'Per day', 'দৈনিক'),
        _sr('integer', 'ml_avg_week', 'Per week', 'সপ্তাহে'),
        _sr('integer', 'ml_avg_month', 'Per month', 'মাসে'),
        _sr('integer', 'ml_avg_year', 'Per year', 'বছরে'),
        # Tool F-1.1 cols 14–17.
        _sr('select_one yn_code', 'ml_needle_drug',
            'Takes any drug/intoxicant via needle-syringe? (Yes-1/No-0)',
            'সুঁই-সিরিঞ্জের মাধ্যমে কোনো নেশা গ্রহণ করেন কি না? (হ্যাঁ-১/না-০)'),
        _sr('select_one yn_code', 'ml_has_nid',
            'Has a National ID card? (Yes-1/No-0)',
            'জাতীয় পরিচয়পত্র আছে কি না? (হ্যাঁ-১/না-০)'),
        # FP not asked for never-married (marital '1') — Ashis review pt 2.
        _sr('select_one yn_code', 'ml_fp_method',
            'Uses any family-planning method? (Yes-1/No-0)',
            'পরিবার পরিকল্পনার কোনো পদ্ধতি ব্যবহার করেন কি না? (হ্যাঁ-১/না-০)',
            relevant="${ml_marital}!='1'"),
        _sr('select_one ml_status', 'ml_current_status',
            'Current status', 'বর্তমান অবস্থা'),
        _sr('text', 'ml_remarks', 'Remarks', 'মন্তব্য'),
        _sr('end_group', 'grp_ml'),
    ]
    return rows


def _mother_list_choices():
    return list(_centre_choices()) + _shared_choices()


# ─── FORM 1 — Service Log ──────────────────────────────────────────────────────

def _service_log_survey():
    R = lambda b: f"${{record_type}}='{b}'"
    rows = _meta()
    rows += [_sr('select_one sl_record_type', 'record_type',
                 'Which register are you recording?', 'কোন রেজিস্টার পূরণ করছেন?', required='yes',
                 hint='Pick the register/form you are filling now. / এখন যে রেজিস্টার/ফর্ম পূরণ করছেন তা নির্বাচন করুন।')]

    # F-01 Wellness Centre Service Logbook — CONSOLIDATED register-or-log.
    # The worker types only the 4-digit serial; the DD- district prefix is added
    # automatically so the ID always matches the Mother List. A NEW client (ID
    # not found) is registered inline via the ml_* block; a returning client
    # skips straight to services. One sheet: registration + services in F-01.
    rows += [
        _sr('begin_group', 'grp_logbook', 'F-01 · Wellness Centre Service Logbook',
            'F-01 · ওয়েলনেস সেন্টার সার্ভিস লগবুক', relevant=R('wellness_logbook')),
        _sr('date', 'log_date', 'Date', 'তারিখ', required='yes'),
        _sr('calculate', 'logbook_dist_code', calc=_bandhu_dist_code_calc()),
        _sr('text', 'log_serial', 'Beneficiary serial (4 digits)',
            'সুবিধাভোগী ক্রমিক (৪ অঙ্ক)', required='yes',
            constraint="regex(., '^[0-9]{4}$')",
            cmsg='Enter exactly 4 digits, e.g. 0001. / ঠিক ৪ অঙ্ক দিন, যেমন 0001।',
            hint='The district code is added automatically. / জেলা কোড স্বয়ংক্রিয়ভাবে যুক্ত হবে।'),
        _sr('calculate', 'log_client_id',
            calc="concat(${logbook_dist_code}, '-', ${log_serial})"),
        _sr('note', '_log_id_show', '🆔 Client ID: ${log_client_id}',
            '🆔 ক্লায়েন্ট আইডি: ${log_client_id}', relevant="${log_serial}!=''"),
        *_id_lookup('log_client_id'),

        # ── New client → register inline (only when the ID is not in the Mother List) ──
        _sr('begin_group', 'grp_log_reg', 'New client — register',
            'নতুন ক্লায়েন্ট — নিবন্ধন',
            relevant="${log_client_id}!='' and ${_log_client_id_name}=''"),
        _sr('note', '_reg_hint', 'This ID is new — please register the client.',
            'এই আইডি নতুন — অনুগ্রহ করে ক্লায়েন্ট নিবন্ধন করুন।'),
        _sr('text', 'ml_name', 'Name', 'নাম', required='yes'),
        _sr('text', 'ml_parent_name', "Father's / Mother's name", 'পিতা/মাতার নাম'),
        _sr('integer', 'ml_birth_year', 'Birth year', 'জন্ম সাল',
            constraint='. >= 1940 and . <= 2012', cmsg='1940–2012'),
        _sr('select_one tg_code', 'ml_gender', 'Gender (TG code)', 'লিঙ্গ (টিজি কোড)'),
        _sr('text', 'ml_address', 'Current address (area name)', 'বর্তমান ঠিকানা (এলাকার নাম)'),
        _sr('text', 'ml_spot', 'Spot name', 'স্পটের নাম'),
        _sr('select_one education', 'ml_education', 'Education', 'শিক্ষা'),
        _sr('select_one marital', 'ml_marital', 'Marital status', 'বৈবাহিক অবস্থা'),
        _sr('integer', 'ml_children_u18', 'Number of children under 18',
            '১৮ বছরের নিচে সন্তান সংখ্যা', relevant="${ml_marital}!='1'"),
        _sr('select_one occupation', 'ml_occupation', 'Income source (occupation)',
            'আয়ের উৎস (পেশা)'),
        _sr('integer', 'ml_avg_day', 'Average sex-work contacts per day', 'দৈনিক গড় যোগাযোগ'),
        _sr('select_one yn_code', 'ml_needle_drug',
            'Takes any drug via needle-syringe? (Yes-1/No-0)',
            'সুঁই-সিরিঞ্জে নেশা গ্রহণ করেন? (হ্যাঁ-১/না-০)'),
        _sr('select_one yn_code', 'ml_has_nid',
            'Has a National ID card? (Yes-1/No-0)', 'জাতীয় পরিচয়পত্র আছে? (হ্যাঁ-১/না-০)'),
        _sr('select_one yn_code', 'ml_fp_method',
            'Uses any family-planning method? (Yes-1/No-0)',
            'পরিবার পরিকল্পনার পদ্ধতি ব্যবহার করেন? (হ্যাঁ-১/না-০)',
            relevant="${ml_marital}!='1'"),
        _sr('select_one ml_status', 'ml_current_status', 'Current status', 'বর্তমান অবস্থা'),
        _sr('end_group', 'grp_log_reg'),

        # ── Services provided this visit ──
        _sr('select_one tg_code', 'log_tg', 'TG (Code)', 'টিজি কোড'),
        _sr('note', '_log_services', 'Services provided (enter count where applicable):', ''),
        _sr('integer', 'log_condom', 'Condom', 'কনডম'),
        _sr('integer', 'log_condom_demo', 'Condom demo', 'কনডম ডেমো'),
        _sr('integer', 'log_lubricant', 'Lubricant', 'লুব্রিকেন্ট'),
        _sr('integer', 'log_awareness', 'Awareness session', 'সচেতনতা সেশন'),
        _sr('integer', 'log_iec', 'IEC distribution', 'আইইসি বিতরণ'),
        _sr('select_one yes_no', 'log_sti_screening', 'STI screening done', 'এসটিআই স্ক্রিনিং'),
        _sr('select_one yes_no', 'log_clinical', 'Clinical services', 'ক্লিনিক্যাল সেবা'),
        _sr('select_one yes_no', 'log_htc', 'HTC (HIV test)', 'এইচটিসি (এইচআইভি পরীক্ষা)'),
        _sr('select_one yes_no', 'log_mental_health', 'Mental health', 'মানসিক স্বাস্থ্য'),
        _sr('select_one yes_no', 'log_gbv', 'GBV', 'জিবিভি'),
        _sr('select_one yes_no', 'log_legal', 'Legal support', 'আইনি সহায়তা'),
        _sr('select_one yes_no', 'log_counseling', 'Counseling', 'কাউন্সেলিং'),
        _sr('select_one yes_no', 'log_recreation', 'Recreation', 'বিনোদন'),
        _sr('select_one yes_no', 'log_group_edu', 'Attended group education session', 'দলগত শিক্ষা সেশন'),
        _sr('select_multiple f01_referral', 'log_referral', 'Referral', 'রেফারেল'),
        _sr('text', 'log_remarks', 'Remarks', 'মন্তব্য', app='multiline'),
        _sr('end_group', 'grp_logbook'),
    ]

    # F-05 Patient Record Register
    rows += [
        _sr('begin_group', 'grp_patient', 'F-05 · Patient Record Register',
            'F-05 · রোগীর রেকর্ড রেজিস্টার', relevant=R('patient_record')),
        _sr('date', 'pr_date', 'Date', 'তারিখ', required='yes'),
        _sr('text', 'pr_client_id', 'ID No.', 'আইডি নম্বর', required='yes'),
        *_id_lookup('pr_client_id'),
        _sr('select_one tg_code', 'pr_tg', 'TG Code', 'টিজি কোড'),
        _sr('select_one general_diverse', 'pr_general_diverse', 'General / Diverse', 'সাধারণ / বৈচিত্র্যময়'),
        *_age_from_ml('pr_client_id', 'pr_age'),
        _sr('select_one yes_no', 'pr_screening_sti_hiv', 'Screening (STI & HIV) Health', 'এসটিআই ও এইচআইভি স্ক্রিনিং'),
        _sr('select_one yes_no', 'pr_tb_screening', 'TB Screening', 'টিবি স্ক্রিনিং'),
        _sr('select_one yes_no', 'pr_gh_screening', 'GH Screening', 'জিএইচ স্ক্রিনিং'),
        _sr('text', 'pr_chief_complaints', 'Chief Complaints', 'প্রধান অভিযোগ', app='multiline'),
        _sr('select_one sti_case', 'pr_sti_case', 'STI Case', 'এসটিআই কেস'),
        _sr('select_multiple diagnosis', 'pr_diagnosis', 'Diagnosis', 'রোগ নির্ণয়'),
        _sr('text', 'pr_treatment', 'Treatment provided (medicine name & quantity)', 'চিকিৎসা'),
        _sr('select_one seek_timing', 'pr_seek_timing', 'Seeking treatment after onset of STI symptoms', 'উপসর্গের পর চিকিৎসা গ্রহণ'),
        _sr('integer', 'pr_condom_demo', '# of condoms used for demonstration', 'ডেমোর কনডম সংখ্যা'),
        _sr('select_one yes_no', 'pr_sti_counseling', 'STI counseling provided', 'এসটিআই কাউন্সেলিং'),
        _sr('select_one partner_mgmt', 'pr_partner_mgmt', 'Partner management', 'পার্টনার ম্যানেজমেন্ট'),
        _sr('date', 'pr_followup_due', 'Follow-up due date', 'ফলোআপ নির্ধারিত তারিখ'),
        _sr('date', 'pr_followup_done', 'Follow-up done date', 'ফলোআপ সম্পন্ন তারিখ'),
        _sr('select_one yes_no', 'pr_adr', 'Adverse Drug Reaction monitoring', 'ওষুধের বিরূপ প্রতিক্রিয়া'),
        _sr('select_multiple f05_referral', 'pr_referral', 'Referral cases', 'রেফারেল'),
        _sr('end_group', 'grp_patient'),
    ]

    # F-06 HTC Service Register
    rows += [
        _sr('begin_group', 'grp_htc', 'F-06 · HTC Service Register',
            'F-06 · এইচটিসি সার্ভিস রেজিস্টার', relevant=R('htc')),
        _sr('text', 'htc_client_id', "Client's ID", 'ক্লায়েন্ট আইডি', required='yes'),
        *_id_lookup('htc_client_id'),
        *_age_from_ml('htc_client_id', 'htc_age'),
        _sr('integer', 'htc_age_manual', 'Age in year (enter if not in the Mother List)',
            'বয়স (বছর) — মাদার লিস্টে না থাকলে লিখুন',
            relevant="${htc_client_id}!='' and ${htc_age}=''",
            constraint='. >= 0 and . <= 120', cmsg='0–120'),
        _sr('select_one tg_code', 'htc_tg', 'TG Code', 'টিজি কোড'),
        _sr('select_one yes_no', 'htc_partner_testing', 'Partner testing', 'পার্টনার টেস্টিং'),
        _sr('select_one yes_no', 'htc_pretest', 'Pretest counseling', 'প্রি-টেস্ট কাউন্সেলিং'),
        _sr('date', 'htc_date_tested', 'Date tested', 'পরীক্ষার তারিখ'),
        _sr('select_one hiv_result', 'htc_result', 'HIV test result', 'এইচআইভি ফলাফল'),
        _sr('select_one yes_no', 'htc_posttest', 'Post-test counseling', 'পোস্ট-টেস্ট কাউন্সেলিং'),
        _sr('select_one yes_no', 'htc_referred_art', 'Referred / linkage to ART', 'এআরটি রেফারেল'),
        _sr('select_one yes_no', 'htc_received_result', 'Client received result', 'ক্লায়েন্ট ফলাফল পেয়েছেন'),
        _sr('end_group', 'grp_htc'),
    ]

    # F-02 GBV Register
    rows += [
        _sr('begin_group', 'grp_gbv', 'F-02 · GBV Register',
            'F-02 · জিবিভি রেজিস্টার', relevant=R('gbv')),
        _sr('text', 'gbv_client_id', "Client's ID", 'ক্লায়েন্ট আইডি', required='yes'),
        *_id_lookup('gbv_client_id'),
        *_age_from_ml('gbv_client_id', 'gbv_age'),
        _sr('integer', 'gbv_age_manual', 'Survivor age (enter if not in the Mother List)', 'ভুক্তভোগীর বয়স (মাদার লিস্টে না থাকলে লিখুন)', required='yes', relevant="${gbv_client_id}!='' and ${gbv_age}=''", constraint='. >= 0 and . <= 120', cmsg='0–120'),
        _sr('select_one tg_code', 'gbv_tg', 'TG Code', 'টিজি কোড'),
        _sr('text', 'gbv_complaint', 'Complaint on GBV', 'জিবিভি অভিযোগ', app='multiline'),
        _sr('text', 'gbv_primary_service', 'Primary service provided on GBV', 'প্রাথমিক সেবা'),
        _sr('select_one yes_no', 'gbv_ref_other_center', 'Referred to GBV service (other centre)', 'অন্য কেন্দ্রে রেফার'),
        _sr('select_one yes_no', 'gbv_ref_mental_health', 'Referred to mental health service', 'মানসিক স্বাস্থ্য রেফার'),
        _sr('select_one yes_no', 'gbv_ref_legal', 'Referred to legal support', 'আইনি সহায়তা রেফার'),
        _sr('end_group', 'grp_gbv'),
    ]

    # F-03 Mental Health Counseling Register
    rows += [
        _sr('begin_group', 'grp_mh', 'F-03 · Mental Health Counseling Register',
            'F-03 · মানসিক স্বাস্থ্য কাউন্সেলিং', relevant=R('mh_counseling')),
        _sr('date', 'mh_date', 'Date', 'তারিখ', required='yes'),
        _sr('text', 'mh_client_id', 'Client ID #', 'ক্লায়েন্ট আইডি', required='yes'),
        *_id_lookup('mh_client_id'),
        _sr('select_one tg_code', 'mh_tg', 'TG Code', 'টিজি কোড'),
        *_age_from_ml('mh_client_id', 'mh_age'),
        _sr('text', 'mh_cruising_spot', 'Cruising spot', 'ক্রুজিং স্পট'),
        _sr('select_multiple sex_with', 'mh_sex_with', 'Recently had sex with', 'সম্প্রতি যৌন সম্পর্ক'),
        _sr('text', 'mh_sex_since', 'Practicing sexual activities since', 'কবে থেকে যৌন কার্যকলাপ'),
        _sr('select_multiple sex_activity', 'mh_sex_activity', 'Type of sexual activity', 'যৌন কার্যকলাপের ধরন'),
        _sr('select_one condom_use', 'mh_condom_use', 'Practice of condom use', 'কনডম ব্যবহার'),
        _sr('text', 'mh_condom_since', 'Continuously using condom since', 'কবে থেকে নিয়মিত কনডম'),
        _sr('select_one yes_no', 'mh_drug_history', 'History of drug use', 'মাদক ব্যবহারের ইতিহাস'),
        # Only ask which drugs when history of drug use = Yes — Ashis review pt 5.
        _sr('text', 'mh_drug_names', 'What drugs/substances used', 'কোন মাদক',
            relevant="${mh_drug_history}='yes'"),
        _sr('select_multiple mh_counsel_type', 'mh_counsel_type', 'Counseling', 'কাউন্সেলিং'),
        _sr('text', 'mh_issue', 'Complaint / Query / Issue(s)', 'অভিযোগ / সমস্যা', app='multiline'),
        _sr('text', 'mh_description', 'Description (in detail)', 'বিস্তারিত বিবরণ', app='multiline'),
        _sr('text', 'mh_counsel_details', 'Counselling details', 'কাউন্সেলিং বিবরণ', app='multiline'),
        _sr('text', 'mh_referral', 'Referral', 'রেফারেল'),
        _sr('text', 'mh_remarks', 'Remarks', 'মন্তব্য'),
        _sr('end_group', 'grp_mh'),
    ]

    # Daily Counseling form
    rows += [
        _sr('begin_group', 'grp_counsel', 'Daily Counseling Form', 'দৈনিক কাউন্সেলিং ফর্ম',
            relevant=R('counseling_daily')),
        _sr('date', 'cn_date', 'Date', 'তারিখ', required='yes'),
        _sr('text', 'cn_client_id', 'ID # of client', 'ক্লায়েন্ট আইডি', required='yes'),
        *_id_lookup('cn_client_id'),
        _sr('select_one tg_code', 'cn_tg', 'Target Group (TG) Code', 'টিজি কোড'),
        _sr('select_multiple counsel_issue', 'cn_issues', 'Issues of counseling', 'কাউন্সেলিং বিষয়'),
        _sr('select_multiple counsel_condom', 'cn_condom', 'Condom', 'কনডম'),
        _sr('select_multiple counsel_referral', 'cn_referral', 'Referral', 'রেফারেল'),
        _sr('end_group', 'grp_counsel'),
    ]

    # Referral Register
    rows += [
        _sr('begin_group', 'grp_referral', 'Referral Register', 'রেফারেল রেজিস্টার',
            relevant=R('referral')),
        _sr('date', 'rf_date', 'Date', 'তারিখ', required='yes'),
        _sr('text', 'rf_client_id', 'ID No.', 'আইডি নম্বর', required='yes'),
        *_id_lookup('rf_client_id'),
        _sr('text', 'rf_reason', 'Reasons for referral (problem)', 'রেফারেলের কারণ', app='multiline'),
        _sr('select_multiple referred_for', 'rf_referred_for', 'Referred for', 'যে কারণে রেফার'),
        _sr('text', 'rf_where', 'Where referred?', 'কোথায় রেফার'),
        _sr('date', 'rf_receiving_date', 'Date of receiving referral service', 'সেবা গ্রহণের তারিখ'),
        _sr('date', 'rf_followup_date', 'Date of follow up', 'ফলোআপ তারিখ'),
        _sr('text', 'rf_remarks', 'Remarks', 'মন্তব্য'),
        _sr('end_group', 'grp_referral'),
    ]

    # F-08 Detailed data of HIV identified
    rows += [
        _sr('begin_group', 'grp_hiv', 'F-08 · Detailed Data of HIV Identified',
            'F-08 · এইচআইভি শনাক্তের বিস্তারিত', relevant=R('hiv_identified')),
        _sr('text', 'hv_period', 'Period', 'সময়কাল'),
        _sr('date', 'hv_date_testing', 'Date of testing', 'পরীক্ষার তারিখ'),
        _sr('text', 'hv_name', 'Name', 'নাম'),
        _sr('text', 'hv_client_uid', 'Client UID', 'ক্লায়েন্ট ইউআইডি'),
        _sr('text', 'hv_service_package', 'Service package', 'সেবা প্যাকেজ'),
        _sr('text', 'hv_gender', 'Gender', 'লিঙ্গ'),
        _sr('integer', 'hv_age', 'Age', 'বয়স'),
        _sr('text', 'hv_marital', 'Marital status', 'বৈবাহিক অবস্থা'),
        _sr('text', 'hv_education', 'Education', 'শিক্ষা'),
        _sr('text', 'hv_occupation', 'Current occupation', 'বর্তমান পেশা'),
        _sr('text', 'hv_prev_occupation', 'Previous occupation', 'পূর্ববর্তী পেশা'),
        _sr('text', 'hv_migration', 'Migration status (last 3 months)', 'অভিবাসন অবস্থা'),
        _sr('select_one yes_no', 'hv_linked_care', 'Linked to care, Rx & support', 'যত্নে সংযুক্ত'),
        _sr('select_one linked_with', 'hv_linked_with', 'Linked with', 'যার সাথে সংযুক্ত'),
        _sr('select_one receiving_art', 'hv_receiving_art', 'Receiving ART', 'এআরটি গ্রহণ'),
        _sr('date', 'hv_art_start', 'ART starting date', 'এআরটি শুরুর তারিখ'),
        _sr('text', 'hv_spouse_occupation', 'Occupation of spouse', 'স্বামী/স্ত্রীর পেশা'),
        _sr('select_one yes_no', 'hv_spouse_tested', 'HIV testing of spouse', 'স্বামী/স্ত্রীর পরীক্ষা'),
        _sr('date', 'hv_spouse_test_date', 'Date of testing (spouse)', 'পরীক্ষার তারিখ'),
        _sr('select_one hiv_result', 'hv_spouse_result', 'Test result of spouse', 'ফলাফল'),
        _sr('select_one yes_no', 'hv_child_tested', 'HIV testing of child', 'সন্তানের পরীক্ষা'),
        _sr('date', 'hv_child_test_date', 'Date of testing (child)', 'পরীক্ষার তারিখ'),
        _sr('select_one hiv_result', 'hv_child_result', 'Test result of child', 'ফলাফল'),
        _sr('text', 'hv_remarks', 'Remarks', 'মন্তব্য'),
        _sr('end_group', 'grp_hiv'),
    ]
    return rows


def _service_log_choices():
    rows = list(_centre_choices()) + _shared_choices()
    for v, en, bn in [
        ('wellness_logbook', 'Wellness Centre Service Logbook (F-01)', 'F-01 লগবুক'),
        ('patient_record',   'Patient Record Register (F-05)', 'F-05 রোগীর রেকর্ড'),
        ('htc',              'HTC Service Register (F-06)', 'F-06 এইচটিসি'),
        ('gbv',              'GBV Register (F-02)', 'F-02 জিবিভি'),
        ('mh_counseling',    'Mental Health Counseling (F-03)', 'F-03 মানসিক স্বাস্থ্য'),
        ('counseling_daily', 'Daily Counseling Form', 'দৈনিক কাউন্সেলিং'),
        ('referral',         'Referral Register', 'রেফারেল রেজিস্টার'),
        ('hiv_identified',   'F-08 HIV Identified (detailed)', 'F-08 এইচআইভি শনাক্ত'),
    ]:
        rows.append(_ch('sl_record_type', v, en, bn))
    return rows


# ─── FORM 2 — Activity & Operations ───────────────────────────────────────────

def _activity_ops_survey():
    R = lambda b: f"${{record_type}}='{b}'"
    rows = _meta()
    rows += [_sr('select_one ao_record_type', 'record_type',
                 'What are you recording?', 'কী নথিভুক্ত করছেন?', required='yes')]

    # F-04 Daily Outreach Monitoring
    rows += [
        _sr('begin_group', 'grp_outreach', 'F-04 · Daily Outreach Monitoring',
            'F-04 · দৈনিক আউটরীচ মনিটরিং', relevant=R('outreach')),
        _sr('date', 'or_date', 'Date', 'তারিখ', required='yes'),
        _sr('text', 'or_id', 'ID No. (write name if new)', 'আইডি নম্বর (নতুন হলে নাম লিখুন)'),
        _sr('text', 'or_peer_educator', 'Peer Educator name', 'পিয়ার এডুকেটরের নাম'),
        _sr('text', 'or_spot', 'Spot name', 'স্পটের নাম'),
        _sr('integer', 'or_condom', 'Condom distributed', 'কনডম বিতরণ'),
        _sr('integer', 'or_lubricant', 'Lubricant distributed', 'লুব্রিকেন্ট বিতরণ'),
        _sr('integer', 'or_awareness', 'Awareness education sessions', 'সচেতনতা মূলক শিক্ষা'),
        _sr('integer', 'or_iec', 'IEC/BCC distribution (number)', 'আইইসি/বিসিসি বিতরণ'),
        _sr('note', '_or_ref', 'Referrals (from outreach to service centre):', 'রেফারেল:'),
        _sr('integer', 'or_ref_single_education', 'Referral — Single (unit) education', 'রেফারেল — একক শিক্ষা'),
        _sr('integer', 'or_ref_sti', 'Referral — STI', 'রেফারেল — যৌনরোগ'),
        _sr('integer', 'or_ref_gh', 'Referral — General Health', 'রেফারেল — সাধারণ স্বাস্থ্য'),
        _sr('integer', 'or_ref_mental_health', 'Referral — Mental Health', 'রেফারেল — মানসিক স্বাস্থ্য'),
        _sr('integer', 'or_ref_gbv', 'Referral — GBV', 'রেফারেল — জিবিভি'),
        _sr('integer', 'or_ref_counseling', 'Referral — Counseling', 'রেফারেল — কাউন্সেলিং'),
        _sr('integer', 'or_ref_recreation', 'Referral — Recreation', 'রেফারেল — বিনোদন'),
        _sr('text', 'or_remarks', 'Remarks', 'মন্তব্য'),
        _sr('end_group', 'grp_outreach'),
    ]

    # F-10 Mobile Health Camp — Patient Record
    rows += [
        _sr('begin_group', 'grp_camp', 'F-10 · Mobile Health Camp — Patient Record',
            'F-10 · মোবাইল হেলথ ক্যাম্প', relevant=R('mobile_camp')),
        _sr('date', 'mc_date', 'Date', 'তারিখ', required='yes'),
        _sr('text', 'mc_client_id', 'ID No.', 'আইডি নম্বর', required='yes'),
        *_id_lookup('mc_client_id'),
        _sr('select_one tg_code', 'mc_tg', 'TG Code', 'টিজি কোড'),
        _sr('select_one general_diverse', 'mc_general_diverse', 'General / Diverse', 'সাধারণ / বৈচিত্র্যময়'),
        *_age_from_ml('mc_client_id', 'mc_age'),
        _sr('select_one yes_no', 'mc_screening_sti_hiv', 'Screening (STI & HIV) Health', 'স্ক্রিনিং'),
        _sr('select_one yes_no', 'mc_tb_screening', 'TB Screening', 'টিবি স্ক্রিনিং'),
        _sr('select_one yes_no', 'mc_gh_screening', 'GH Screening', 'জিএইচ স্ক্রিনিং'),
        _sr('select_one yes_no', 'mc_sti_service', 'Service provided — STI', 'এসটিআই সেবা'),
        _sr('select_one yes_no', 'mc_gh_service', 'Service provided — GH', 'সেবা — সাধারণ স্বাস্থ্য'),
        _sr('select_one yes_no', 'mc_psd_service', 'Service provided — PSD', 'সেবা — পিএসডি'),
        _sr('select_one yes_no', 'mc_mental_health_service', 'Service provided — Mental health', 'সেবা — মানসিক স্বাস্থ্য'),
        _sr('select_one hiv_result', 'mc_hiv_result', 'HIV testing result', 'এইচআইভি ফলাফল'),
        _sr('text', 'mc_treatment', 'Treatment provided (medicine name & quantity)', 'চিকিৎসা'),
        _sr('select_one seek_timing', 'mc_seek_timing', 'Seeking treatment after onset of STI symptoms', 'চিকিৎসা গ্রহণ'),
        _sr('integer', 'mc_condom_demo', '# of condoms used for demonstration', 'ডেমোর কনডম'),
        _sr('select_one yes_no', 'mc_sti_counseling', 'STI counseling provided', 'এসটিআই কাউন্সেলিং'),
        _sr('select_one partner_mgmt', 'mc_partner_mgmt', 'Partner management', 'পার্টনার ম্যানেজমেন্ট'),
        _sr('date', 'mc_followup_due', 'Follow-up due date', 'ফলোআপ তারিখ'),
        _sr('date', 'mc_followup_done', 'Follow-up done date', 'ফলোআপ সম্পন্ন'),
        _sr('select_one yes_no', 'mc_art_linkage', 'ART linkage (if HIV positive)', 'এআরটি লিংকেজ'),
        _sr('select_multiple f10_referral', 'mc_referral', 'Referral cases', 'রেফারেল'),
        _sr('end_group', 'grp_camp'),
    ]

    # F-11 Attendance Sheet
    rows += [
        _sr('begin_group', 'grp_attendance', 'F-11 · Attendance Sheet', 'F-11 · উপস্থিতি শিট',
            relevant=R('attendance')),
        _sr('note','_at_kobo_only','Note: F-11 attendance entries are kept in KoboToolbox only and are NOT counted in SIMPLE.','দ্রষ্টব্য: F-11 উপস্থিতি শুধু KoboToolbox-এ থাকে, SIMPLE-এ গণনা হয় না।'),
        _sr('date', 'at_date', 'Date', 'তারিখ'),
        _sr('text', 'at_name', 'Name', 'নাম'),
        _sr('text', 'at_designation', 'Designation and organization', 'পদবি ও সংস্থা'),
        _sr('text', 'at_contact', 'Email ID and contact number', 'ইমেইল ও যোগাযোগ'),
        _sr('select_one att_gender', 'at_gender', 'Gender', 'লিঙ্গ'),
        _sr('select_one age_band', 'at_age_band', 'Age group', 'বয়স গ্রুপ'),
        _sr('select_one yes_no', 'at_photo_consent', 'Photo consent', 'ছবির সম্মতি'),
        _sr('end_group', 'grp_attendance'),
    ]

    # F-12 Event Report
    rows += [
        _sr('begin_group', 'grp_event', 'F-12 · Event Report', 'F-12 · ইভেন্ট রিপোর্ট',
            relevant=R('event_report')),
        _sr('text', 'ev_activity', 'Activity name', 'কার্যক্রমের নাম'),
        _sr('text', 'ev_objective', 'Objectives', 'উদ্দেশ্য', app='multiline'),
        _sr('text', 'ev_ir', 'Short-term result (IR)', 'স্বল্পমেয়াদী ফলাফল (IR)', app='multiline'),
        _sr('select_one event_kind', 'ev_kind', 'Event type', 'ইভেন্টের ধরন', required='yes',
            hint='Required so the report routes to the correct indicator.'),
        # If "Other", capture what kind of event it was — Ashis review pt 6.
        _sr('text', 'ev_kind_other', 'If Other, specify the event type', 'অন্যান্য হলে ধরন উল্লেখ করুন',
            relevant="${ev_kind}='other'", required='yes'),
        _sr('text', 'ev_place', 'Place', 'স্থান'),
        _sr('date', 'ev_date', 'Date', 'তারিখ'),
        # Participants — age-band × gender cross-tab (tool F-12). The gender
        # totals below are auto-calculated from the grid so the dashboard still
        # receives ev_man/ev_woman/ev_tg_hijra/ev_other.
        _sr('note', '_ev_part',
            'Participants by age band and gender (enter the count in each cell):',
            'বয়স ও জেন্ডার অনুযায়ী অংশগ্রহণকারী (প্রতিটি ঘরে সংখ্যা লিখুন):'),
        _sr('note', '_ev_a1', '18-24 years', '১৮-২৪ বছর'),
        _sr('integer', 'ev_a1824_woman', 'Woman (18-24)', 'নারী (১৮-২৪)'),
        _sr('integer', 'ev_a1824_man', 'Man (18-24)', 'পুরুষ (১৮-২৪)'),
        _sr('integer', 'ev_a1824_tg', 'TG/Hijra (18-24)', 'টিজি/হিজড়া (১৮-২৪)'),
        _sr('integer', 'ev_a1824_other', 'Others (18-24)', 'অন্যান্য (১৮-২৪)'),
        _sr('note', '_ev_a2', '25-30 years', '২৫-৩০ বছর'),
        _sr('integer', 'ev_a2530_woman', 'Woman (25-30)', 'নারী (২৫-৩০)'),
        _sr('integer', 'ev_a2530_man', 'Man (25-30)', 'পুরুষ (২৫-৩০)'),
        _sr('integer', 'ev_a2530_tg', 'TG/Hijra (25-30)', 'টিজি/হিজড়া (২৫-৩০)'),
        _sr('integer', 'ev_a2530_other', 'Others (25-30)', 'অন্যান্য (২৫-৩০)'),
        _sr('note', '_ev_a3', '31-35 years', '৩১-৩৫ বছর'),
        _sr('integer', 'ev_a3135_woman', 'Woman (31-35)', 'নারী (৩১-৩৫)'),
        _sr('integer', 'ev_a3135_man', 'Man (31-35)', 'পুরুষ (৩১-৩৫)'),
        _sr('integer', 'ev_a3135_tg', 'TG/Hijra (31-35)', 'টিজি/হিজড়া (৩১-৩৫)'),
        _sr('integer', 'ev_a3135_other', 'Others (31-35)', 'অন্যান্য (৩১-৩৫)'),
        _sr('note', '_ev_a4', 'Above 35 years', '৩৫ বছরের উপরে'),
        _sr('integer', 'ev_agt35_woman', 'Woman (>35)', 'নারী (৩৫+)'),
        _sr('integer', 'ev_agt35_man', 'Man (>35)', 'পুরুষ (৩৫+)'),
        _sr('integer', 'ev_agt35_tg', 'TG/Hijra (>35)', 'টিজি/হিজড়া (৩৫+)'),
        _sr('integer', 'ev_agt35_other', 'Others (>35)', 'অন্যান্য (৩৫+)'),
        _sr('note', '_ev_a5', 'Age unknown / refused', 'বয়স জানাতে অনিচ্ছুক'),
        _sr('integer', 'ev_aunk_woman', 'Woman (age unknown)', 'নারী (বয়স অজানা)'),
        _sr('integer', 'ev_aunk_man', 'Man (age unknown)', 'পুরুষ (বয়স অজানা)'),
        _sr('integer', 'ev_aunk_tg', 'TG/Hijra (age unknown)', 'টিজি/হিজড়া (বয়স অজানা)'),
        _sr('integer', 'ev_aunk_other', 'Others (age unknown)', 'অন্যান্য (বয়স অজানা)'),
        _sr('calculate', 'ev_woman',
            calc="coalesce(${ev_a1824_woman},0)+coalesce(${ev_a2530_woman},0)+coalesce(${ev_a3135_woman},0)+coalesce(${ev_agt35_woman},0)+coalesce(${ev_aunk_woman},0)"),
        _sr('calculate', 'ev_man',
            calc="coalesce(${ev_a1824_man},0)+coalesce(${ev_a2530_man},0)+coalesce(${ev_a3135_man},0)+coalesce(${ev_agt35_man},0)+coalesce(${ev_aunk_man},0)"),
        _sr('calculate', 'ev_tg_hijra',
            calc="coalesce(${ev_a1824_tg},0)+coalesce(${ev_a2530_tg},0)+coalesce(${ev_a3135_tg},0)+coalesce(${ev_agt35_tg},0)+coalesce(${ev_aunk_tg},0)"),
        _sr('calculate', 'ev_other',
            calc="coalesce(${ev_a1824_other},0)+coalesce(${ev_a2530_other},0)+coalesce(${ev_a3135_other},0)+coalesce(${ev_agt35_other},0)+coalesce(${ev_aunk_other},0)"),
        _sr('calculate', 'ev_total',
            calc="${ev_woman}+${ev_man}+${ev_tg_hijra}+${ev_other}"),
        _sr('note', '_ev_total_show', 'Total participants: ${ev_total}',
            'মোট অংশগ্রহণকারী: ${ev_total}'),
        _sr('select_one event_modality', 'ev_modality', 'Event modality', 'ইভেন্টের ধরন'),
        _sr('integer', 'ev_attend_ge80', 'Attendance ≥ 80% (count)', '৮০%+ উপস্থিতি'),
        _sr('integer', 'ev_attend_lt80', 'Attendance < 80% (count)', '৮০%-এর কম উপস্থিতি'),
        _sr('integer', 'ev_iec', 'IEC/BCC distribution (number)', 'আইইসি বিতরণ'),
        _sr('select_one yes_no', 'ev_as_per_plan', 'Activity done as per plan?', 'পরিকল্পনা অনুযায়ী হয়েছে?'),
        _sr('text', 'ev_status_explain', 'If not as per plan, please explain', 'না হলে ব্যাখ্যা করুন',
            app='multiline', relevant="${ev_as_per_plan}='no'"),
        _sr('text', 'ev_chief_guest', 'Chief guest (name, organization, designation)', 'প্রধান অতিথি'),
        _sr('text', 'ev_chair', 'Chair / facilitator', 'সভাপতি'),
        _sr('text', 'ev_discussion', 'Discussion and decision', 'আলোচনা ও সিদ্ধান্ত', app='multiline'),
        _sr('text', 'ev_output', 'Output / outcome', 'ফলাফল', app='multiline'),
        _sr('text', 'ev_limitation', 'Limitation / feedback', 'সীমাবদ্ধতা'),
        _sr('text', 'ev_learning', 'Learning', 'শিক্ষণীয় বিষয়'),
        _sr('text', 'ev_comments', 'Comments', 'মন্তব্য'),
        _sr('image', 'ev_attachment', 'Attachment (photo)', 'সংযুক্তি (ছবি)'),
        _sr('end_group', 'grp_event'),
    ]

    # F-13 Stock Register
    rows += [
        _sr('begin_group', 'grp_stock', 'F-13 · Stock Register', 'F-13 · স্টক রেজিস্টার',
            relevant=R('stock')),
        _sr('note','_st_kobo_only','Note: F-13 stock entries are kept in KoboToolbox only and are NOT counted in SIMPLE.','দ্রষ্টব্য: F-13 স্টক শুধু KoboToolbox-এ থাকে, SIMPLE-এ গণনা হয় না।'),
        _sr('text', 'st_item', 'Item description', 'পণ্যের বিবরণ'),
        _sr('date', 'st_date', 'Date', 'তারিখ'),
        _sr('text', 'st_from_to', 'Received from / Issued to', 'গ্রহণ/বিতরণ'),
        _sr('integer', 'st_opening', 'Opening balance', 'প্রারম্ভিক মজুদ'),
        _sr('integer', 'st_received', 'Quantity received', 'গৃহীত পরিমাণ'),
        _sr('integer', 'st_distributed', 'Quantity distribution', 'বিতরণকৃত পরিমাণ'),
        _sr('integer', 'st_balance', 'Stock balance', 'মজুদ স্থিতি'),
        _sr('text', 'st_comments', 'Comments', 'মন্তব্য'),
        _sr('end_group', 'grp_stock'),
    ]

    # F-07 KP Clinic Information
    rows += [
        _sr('begin_group', 'grp_kpc', 'F-07 · KP Clinic Information', 'F-07 · কেপি ক্লিনিক তথ্য',
            relevant=R('kp_clinic_info')),
        _sr('text', 'kc_name', 'Name of KP Clinic', 'কেপি ক্লিনিকের নাম'),
        _sr('text', 'kc_address', 'Address', 'ঠিকানা', app='multiline'),
        _sr('text', 'kc_incharge', 'Name of Wellness Center Incharge & Contact #', 'ওয়েলনেস সেন্টার ইনচার্জের নাম ও যোগাযোগ'),
        _sr('integer', 'kc_num_staff', 'Number of staff', 'কর্মী সংখ্যা'),
        _sr('select_one yes_no', 'kc_equipped', 'Well equipped with all logistics', 'সব লজিস্টিকসসহ সুসজ্জিত'),
        _sr('select_one yes_no', 'kc_functional', 'Functional', 'কার্যকর'),
        _sr('text', 'kc_remarks', 'Remarks', 'মন্তব্য'),
        _sr('end_group', 'grp_kpc'),
    ]

    # F-09 Wellness Center Information
    rows += [
        _sr('begin_group', 'grp_wci', 'F-09 · Wellness Center Information', 'F-09 · ওয়েলনেস সেন্টার তথ্য',
            relevant=R('wellness_center_info')),
        _sr('text', 'wc_name', 'Name of Wellness Center', 'ওয়েলনেস সেন্টারের নাম'),
        _sr('text', 'wc_address', 'Address', 'ঠিকানা', app='multiline'),
        _sr('text', 'wc_incharge', 'Name of Wellness Center Incharge & Contact #', 'ওয়েলনেস সেন্টার ইনচার্জের নাম ও যোগাযোগ'),
        _sr('integer', 'wc_num_staff', 'Number of staff', 'কর্মী সংখ্যা'),
        _sr('text', 'wc_cruising_spot', 'Name of cruising spot(s)', 'ক্রুজিং স্পটের নাম'),
        _sr('select_one yes_no', 'wc_functional', 'Functional', 'কার্যকর'),
        _sr('text', 'wc_remarks', 'Remarks', 'মন্তব্য'),
        _sr('end_group', 'grp_wci'),
    ]

    # F-14 e-billboard screenshot
    rows += [
        _sr('begin_group', 'grp_ebillboard', 'F-14 · e-Billboard Screenshot', 'F-14 · ই-বিলবোর্ড স্ক্রিনশট',
            relevant=R('ebillboard')),
        _sr('note', '_eb_note',
            'F-14 has no data fields — attach the screenshot of the displayed e-billboard message.',
            'F-14 তে কোনো তথ্য ক্ষেত্র নেই — প্রদর্শিত ই-বিলবোর্ডের স্ক্রিনশট সংযুক্ত করুন।'),
        _sr('date', 'eb_date', 'Date displayed', 'প্রদর্শনের তারিখ'),
        _sr('text', 'eb_location', 'Location (Wellness Centre / area)', 'অবস্থান (ওয়েলনেস সেন্টার / এলাকা)'),
        _sr('image', 'eb_screenshot', 'Screenshot of message', 'বার্তার স্ক্রিনশট', required='yes'),
        _sr('end_group', 'grp_ebillboard'),
    ]
    return rows


def _activity_ops_choices():
    rows = list(_centre_choices()) + _shared_choices()
    for v, en, bn in [
        ('outreach',            'F-04 Daily Outreach Monitoring', 'F-04 আউটরীচ'),
        ('mobile_camp',         'F-10 Mobile Health Camp', 'F-10 মোবাইল ক্যাম্প'),
        ('attendance',          'F-11 Attendance Sheet', 'F-11 উপস্থিতি'),
        ('event_report',        'F-12 Event Report', 'F-12 ইভেন্ট রিপোর্ট'),
        ('stock',               'F-13 Stock Register', 'F-13 স্টক'),
        ('kp_clinic_info',      'F-07 KP Clinic Information', 'F-07 কেপি ক্লিনিক তথ্য'),
        ('wellness_center_info','F-09 Wellness Center Information', 'F-09 ওয়েলনেস সেন্টার তথ্য'),
        ('ebillboard',          'F-14 e-Billboard Screenshot', 'F-14 ই-বিলবোর্ড'),
    ]:
        rows.append(_ch('ao_record_type', v, en, bn))
    return rows


# ─── Kobo upload helpers ──────────────────────────────────────────────────────

def _kobo_token():
    return (getattr(settings, 'KOBO_API_TOKEN', '')
            or os.environ.get('KOBO_TOKEN', '')).strip()


def _import_xlsform(xlsx_path, asset_uid, token, stdout):
    headers = {'Authorization': f'Token {token}'}
    api = f'{KOBO_BASE}/api/v2'
    with open(xlsx_path, 'rb') as fh:
        files = {'file': (os.path.basename(xlsx_path), fh,
                          'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'destination': f'{api}/assets/{asset_uid}/', 'library': 'false'}
        r = requests.post(f'{api}/imports/', headers=headers, files=files, data=data, timeout=120)
    if r.status_code not in (200, 201):
        stdout.write(f'    import FAILED ({r.status_code}): {r.text[:200]}')
        return False
    stdout.write('    imported')
    return True


def _deploy(asset_uid, token, stdout):
    headers = {'Authorization': f'Token {token}'}
    api = f'{KOBO_BASE}/api/v2'
    v = requests.get(f'{api}/assets/{asset_uid}/versions/?limit=1', headers=headers, timeout=30)
    try:
        vhash = v.json()['results'][0]['uid']
    except Exception:
        stdout.write('    no version yet — skipping deploy')
        return False
    r = requests.patch(f'{api}/assets/{asset_uid}/deployment/',
                       headers=headers, json={'version_id': vhash, 'active': True}, timeout=60)
    if r.status_code in (200, 201):
        stdout.write('    redeployed')
        return True
    r2 = requests.post(f'{api}/assets/{asset_uid}/deployment/',
                       headers=headers, json={'version_id': vhash, 'active': True}, timeout=60)
    if r2.status_code in (200, 201):
        stdout.write('    deployed (POST)')
        return True
    stdout.write(f'    deploy FAILED ({r.status_code}/{r2.status_code}): {r2.text[:160]}')
    return False


# ─── Command ──────────────────────────────────────────────────────────────────

FORMS = [
    {'file': 'Bandhu-0_Mother_List.xlsx', 'id': 'bandhu_mother_list_v1',
     'title': 'Bandhu 0 — Mother List', 'survey': _mother_list_survey, 'choices': _mother_list_choices},
    {'file': 'Bandhu-1_Service_Log.xlsx', 'id': 'bandhu_service_log_v1',
     'title': 'Bandhu 1 — Service Log', 'survey': _service_log_survey, 'choices': _service_log_choices},
    {'file': 'Bandhu-2_Activity_Operations.xlsx', 'id': 'bandhu_activity_ops_v1',
     'title': 'Bandhu 2 — Activity & Operations', 'survey': _activity_ops_survey, 'choices': _activity_ops_choices},
]


class Command(BaseCommand):
    help = 'Build the 3 Bandhu XLSForms (Mother List + Service Log + Activity & Operations).'

    def add_arguments(self, parser):
        parser.add_argument('--output-dir', default=OUTDIR)
        parser.add_argument('--upload', action='store_true',
            help='Import each xlsx into its existing Kobo asset and redeploy.')

    def handle(self, *args, **options):
        out = options['output_dir']
        os.makedirs(out, exist_ok=True)
        token = _kobo_token() if options['upload'] else ''
        if options['upload'] and not token:
            self.stdout.write(self.style.ERROR('KOBO_TOKEN not set — cannot --upload.'))
            return

        for f in FORMS:
            survey  = f['survey']()
            choices = f['choices']()
            wb = _wb(f['id'], f['title'], survey, choices)
            path = os.path.join(out, f['file'])
            wb.save(path)
            self.stdout.write(self.style.SUCCESS(
                f"  OK  {f['file']:38s}  {len(survey)} survey rows  id: {f['id']}"))

            if options['upload']:
                self.stdout.write('     uploading…')
                api = f'{KOBO_BASE}/api/v2'
                headers = {'Authorization': f'Token {token}'}
                q = requests.get(f'{api}/assets/?q=settings__id_string:{f["id"]}',
                                 headers=headers, timeout=30).json()
                asset_uid = None
                for a in q.get('results', []):
                    if a.get('settings', {}).get('id_string') == f['id']:
                        asset_uid = a.get('uid')
                        break
                if not asset_uid:
                    allq = requests.get(f'{api}/assets/?limit=300', headers=headers, timeout=30).json()
                    for a in allq.get('results', []):
                        if (a.get('name') or '') == f['title']:
                            asset_uid = a.get('uid')
                            break
                if not asset_uid:
                    self.stdout.write(self.style.ERROR(
                        f'    no Kobo asset for {f["id"]} — create a blank asset named "{f["title"]}" first.'))
                    continue
                if _import_xlsform(path, asset_uid, token, self.stdout):
                    _deploy(asset_uid, token, self.stdout)

        self.stdout.write(f'\nWritten to {os.path.abspath(out)}/')
