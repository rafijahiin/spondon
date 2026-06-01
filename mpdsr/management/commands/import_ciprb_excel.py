"""
Import Sayeed's historical-baseline Excel files into the platform.

Usage:
    python manage.py import_ciprb_excel --all          # ingest everything
    python manage.py import_ciprb_excel --denominators # MPDSR district denominators
    python manage.py import_ciprb_excel --facility     # MPDSR FDN/FDR per facility
    python manage.py import_ciprb_excel --action-plan  # MPDSR action plan summaries
    python manage.py import_ciprb_excel --va-md        # Verbal autopsy maternal deaths
    python manage.py import_ciprb_excel --fistula      # Fistula Identified records

Paths default to C:/Users/HP/Downloads/{fielddatareports_mpdsr,fistuladatareport}.
Override with --mpdsr-root / --fistula-root.

Every import is idempotent — re-running skips existing rows by deterministic key.
"""
from __future__ import annotations

import datetime
import os
import re

from django.core.management.base import BaseCommand
from django.utils import timezone

from mpdsr.models import (
    MPDSRCase, MPDSRDistrictDenominator, MPDSRFacilityCount,
    MPDSRActionPlanSummary, DeathType, ReviewStatus, PlaceOfDeath,
)
from fistula.models import FistulaCornerCase, FistulaCampaign
from submissions.models import FormType

MPDSR_ROOT_DEFAULT   = r'C:/Users/HP/Downloads/fielddatareports_mpdsr'
FISTULA_ROOT_DEFAULT = r'C:/Users/HP/Downloads/fistuladatareport'


def _int(v, default=0):
    """Coerce any cell value to a non-negative int. Empty/None/'NaN' -> default."""
    if v is None or v == '':
        return default
    try:
        return max(0, int(float(v)))
    except (TypeError, ValueError):
        return default


def _float(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _str(v):
    if v is None:
        return ''
    return str(v).strip()


def _parse_date(v):
    """Try hard to parse a date out of messy Excel cells.

    Accepts: datetime, 'DD/MM/YYYY', 'DD.MM.YY', 'DD.MM.YYYY', etc.
    Returns None on failure.
    """
    if v is None or v == '':
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    # Try common formats
    for fmt in ('%d/%m/%Y', '%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Strip time portion
    s2 = s.split(' ')[0]
    for fmt in ('%d/%m/%Y', '%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(s2, fmt).date()
        except ValueError:
            continue
    return None


# ─── 1. MPDSR District Denominators ──────────────────────────────────────────

def import_denominators(mpdsr_root: str) -> tuple[int, int]:
    """District-level Project Deaths 2026 estimates.

    Source: MPDSR Report_2026.xlsx :: District Wise
    Structure: r0-r2 are header rows, r3+ are district rows.
    Columns: 0=Coordinator, 1=District Name, 2=MD, 3=ND, 4=SB (Project Deaths 2026)
    """
    import openpyxl
    path = os.path.join(mpdsr_root, 'MPDSR Report_ 2026.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['District Wise']

    created = 0
    updated = 0
    for row in ws.iter_rows(min_row=4, max_row=1000, values_only=True):
        district = _str(row[1])
        if not district or district.lower() in ('grand total', 'total', 'district name'):
            continue
        md = _float(row[2])
        nd = _float(row[3])
        sb = _float(row[4])
        if md is None and nd is None and sb is None:
            continue

        obj, was_created = MPDSRDistrictDenominator.objects.update_or_create(
            district=district,
            defaults={
                'project_deaths_md': md,
                'project_deaths_nd': nd,
                'project_deaths_sb': sb,
                'source': 'excel_2026',
            },
        )
        if was_created: created += 1
        else: updated += 1

    return created, updated


# ─── 2. MPDSR FDN & FDR per facility ─────────────────────────────────────────

def import_facility_counts(mpdsr_root: str) -> tuple[int, int]:
    """Per-facility FDN (notifications) + FDR (reviews), MD/ND/SB.

    Source: MPDSR Report_2026.xlsx :: FDN & FDR
    Row 0/1/2 are merged header rows. r3+ are data rows.
    Columns: 0=District (only on first row of each district group),
             1=Facility, 2-4=FDN MD/ND/SB, 5-7=FDR MD/ND/SB
    """
    import openpyxl
    path = os.path.join(mpdsr_root, 'MPDSR Report_ 2026.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['FDN & FDR']

    created = 0
    updated = 0
    current_district = ''
    for row in ws.iter_rows(min_row=4, max_row=1000, values_only=True):
        d = _str(row[0])
        if d:
            current_district = d
        district = current_district
        facility = _str(row[1])
        if not facility or facility.lower() in ('grand total', 'total'):
            continue
        if not district:
            continue

        obj, was_created = MPDSRFacilityCount.objects.update_or_create(
            district=district,
            facility_name=facility,
            period='2026',
            defaults={
                'fdn_md': _int(row[2]),
                'fdn_nd': _int(row[3]),
                'fdn_sb': _int(row[4]),
                'fdr_md': _int(row[5]),
                'fdr_nd': _int(row[6]),
                'fdr_sb': _int(row[7]),
                'source': 'excel_2026',
            },
        )
        if was_created: created += 1
        else: updated += 1

    return created, updated


# ─── 3. MPDSR Action Plan summaries ──────────────────────────────────────────

def import_action_plan(mpdsr_root: str) -> tuple[int, int]:
    """Per-district response-plan implementation summary.

    Source: MPDSR Action Plan_ Progress.xlsx — 8 sheets, one per district x DM/UM.
    Sheet structure has complex merged headers, but each sheet contains:
      - Place of meeting (r1, c1)
      - Meeting date (r2, c1)
      - Number of participants (r3, c1)
      - Rows of activities planned vs implemented in subsequent rows
    """
    import openpyxl
    path = os.path.join(mpdsr_root, 'MPDSR Action Plan_ Progress.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    created = 0
    updated = 0
    for sheet_name in wb.sheetnames:
        # Sheet name pattern: {District}_{DM|UM}
        m = re.match(r'(.+?)_(DM|UM)\s*$', sheet_name)
        if not m:
            continue
        district = m.group(1).strip()
        level = m.group(2)

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
        if len(rows) < 5:
            continue

        place_of_meeting = _str(rows[1][1]) if len(rows[1]) > 1 else ''
        meeting_date = _str(rows[2][1]) if len(rows[2]) > 1 else ''
        participants_raw = rows[3][1] if len(rows[3]) > 1 else None
        participants = _int(participants_raw, default=0) or None

        # Count planned vs implemented across all meeting blocks. The sheet
        # repeats blocks of 4 columns: [MPDSR System Strengthening, Actions to
        # reduce, Common modifiable, Activities implemented]. Each row in
        # those columns is an item. Empty = no entry.
        meetings_planned = 0
        activities_planned = 0
        activities_implemented = 0
        # Header row for activity blocks is rows[4]. Find columns where header
        # contains 'Activities implemented' vs 'Actions to reduce'.
        header_row = rows[4] if len(rows) > 4 else []
        implemented_cols = []
        action_cols = []
        for ci, cell in enumerate(header_row):
            label = _str(cell).lower()
            if 'activities implement' in label:
                implemented_cols.append(ci)
            elif 'actions to reduce' in label or 'mpdsr system' in label:
                action_cols.append(ci)
        # Each implemented column block also indicates a meeting/committee instance
        meetings_planned = len(implemented_cols)

        for r in rows[5:]:
            for ci in action_cols:
                if ci < len(r) and _str(r[ci]):
                    activities_planned += 1
            for ci in implemented_cols:
                if ci < len(r) and _str(r[ci]):
                    activities_implemented += 1

        obj, was_created = MPDSRActionPlanSummary.objects.update_or_create(
            district=district,
            level=level,
            defaults={
                'place_of_meeting': place_of_meeting,
                'meeting_date': meeting_date,
                'participants': participants,
                'meetings_planned': meetings_planned,
                'activities_planned': activities_planned,
                'activities_implemented': activities_implemented,
                'source': 'excel_action_plan_2026',
            },
        )
        if was_created: created += 1
        else: updated += 1

    return created, updated


# ─── 4. Verbal Autopsy MD → MPDSRCase rows ───────────────────────────────────

def import_va_md(mpdsr_root: str) -> tuple[int, int]:
    """Maternal-death verbal-autopsy records become MPDSRCase rows with
    source='excel_va_md_2026'. Idempotent via case_hash.

    Source: MD_Verbel Atopsy_20265 Data (2).xlsx — 16 sheets per district.
    Header rows 0-1 (duplicate), data starts row 2.
    """
    import openpyxl
    path = os.path.join(mpdsr_root, 'MD_Verbel Atopsy_20265 Data (2).xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    created = 0
    skipped = 0
    for sheet_name in wb.sheetnames:
        # Extract district from sheet name like 'VA_MD_Sunamganj_Shakhayat'.
        m = re.match(r'VA_MD_\s*([\w\' ]+?)\s*_', sheet_name)
        if not m:
            continue
        district = m.group(1).strip()
        ws = wb[sheet_name]

        for ri, row in enumerate(ws.iter_rows(min_row=3, max_row=400, values_only=True)):
            sl = _str(row[0])
            if not sl or sl.lower() in ('sl', 'total'):
                continue
            form_received_date = _parse_date(row[1] if len(row) > 1 else None)
            interview_date     = _parse_date(row[4] if len(row) > 4 else None)
            # Pick the most reliable date for date_of_death (interview_date as proxy)
            dod = interview_date or form_received_date or datetime.date(2026, 1, 1)

            # Hash key for idempotency
            case_hash = f'VA-MD-{district}-{sheet_name[:20]}-{sl}'

            # Skip if already imported under this hash
            if MPDSRCase.objects.filter(case_hash=case_hash).exists():
                skipped += 1
                continue

            MPDSRCase.objects.create(
                case_hash=case_hash,
                partner='CIPRB',
                district=district,
                sub_form_type='va_md',
                date_of_death=dod,
                death_type=DeathType.MATERNAL,
                cause_of_death='',  # Detailed cause is in many extra columns; left blank for v1
                place_of_death=PlaceOfDeath.FACILITY,
                facility_name='',
                age_years=None,
                status=ReviewStatus.REPORTED,
                audit_trail=[{
                    'timestamp': timezone.now().isoformat(),
                    'user': 'excel_importer',
                    'action': 'Imported from Verbal Autopsy 2026 Excel',
                    'notes': f'Sheet: {sheet_name}, SL: {sl}',
                }],
                source='excel_va_md_2026',
            )
            created += 1

    return created, skipped


# ─── 5. Fistula Identified → FistulaCornerCase ───────────────────────────────

def import_fistula_identified(fistula_root: str) -> tuple[int, int]:
    """List of patients identified at Fistula Corners. Maps onto the existing
    FistulaCornerCase model so the Patient Funnel + Pie chart populate.

    Source: Fistula_Indicator Wise 2026 Data.xlsx :: Identified sheet.
    Row 0/1 are headers; data starts row 2.
    Columns: 0=SN, 1=Name, 2=Age, 3=Husband, 4=Phone, 5=Village, 6=Union,
             7=Upazila, 8=District, 9=Duration, 10=Date of identification,
             11=Place of identification, 12=Type of Fistula, 13=Remark
    """
    import openpyxl
    path = os.path.join(fistula_root, 'Fistula_Indicator Wise 2026 Data.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Identified']

    created = 0
    skipped = 0
    for ri, row in enumerate(ws.iter_rows(min_row=3, max_row=2000, values_only=True)):
        sn = _str(row[0])
        if not sn or sn.lower() in ('sn', 'total'):
            continue
        name     = _str(row[1])
        age      = _int(row[2], default=0) or None
        husband  = _str(row[3])
        phone    = _str(row[4])
        village  = _str(row[5])
        union    = _str(row[6])
        upazila  = _str(row[7])
        district = _str(row[8])
        duration = _str(row[9])
        ident_date = _parse_date(row[10])
        place    = _str(row[11])
        ftype    = _str(row[12]).upper()
        remark   = _str(row[13])

        if not name or not district:
            continue

        # Map raw fistula_type string to FISTULA_TYPE_CHOICES
        if 'VVF' in ftype:
            mapped_type = 'VVF'
        elif 'RVF' in ftype:
            mapped_type = 'RVF'
        elif 'BOTH' in ftype:
            mapped_type = 'BOTH'
        elif ftype:
            mapped_type = 'OTHER'
        else:
            mapped_type = ''

        # Idempotency key — synthetic hash from sheet position
        case_hash = f'FIS-IDENT-{district[:8]}-{sn}'.replace(' ', '')[:30]

        if FistulaCornerCase.objects.filter(case_hash=case_hash).exists():
            skipped += 1
            continue

        FistulaCornerCase.objects.create(
            case_hash=case_hash,
            patient_name=name,
            husband_name=husband,
            mobile_number=phone[:20] if phone else '',
            age_years=age,
            village=village,
            union=union,
            upazila=upazila,
            district=district,
            identification_date=ident_date,
            diagnosis_date=ident_date,
            suffering_duration=duration,
            fistula_type=mapped_type,
            referral_place=place,
            remarks=remark,
            source='excel_fistula_2026',
        )
        created += 1

    return created, skipped


def _parse_date(v):
    """Cells in 'Sunamganj-Daily Data Sheet' come as datetime objects or
    DD.MM.YY / DD/MM/YYYY strings. Return a date or None."""
    if v is None or v == '':
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = _str(v)
    for fmt in ('%d.%m.%y', '%d.%m.%Y', '%d/%m/%y', '%d/%m/%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def import_mass_campaign(path: str) -> tuple[int, int]:
    """Ingest 'Mass Campaign on_ End Obstetric Fistula in Bangladesh.xlsx'
    daily roll-up sheet into FistulaCampaign rows. Maps the form columns
    Animesh asked for — households visited + population covered — into
    structured fields so the CIPRB dashboard tiles render real numbers."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    created = skipped = 0
    for sheet_name in wb.sheetnames:
        if 'Daily' not in sheet_name:
            continue
        ws = wb[sheet_name]
        # Header row is row 3 in this workbook; data starts row 4.
        for row in ws.iter_rows(min_row=4, values_only=True):
            campaign_date = _parse_date(row[0])
            if not campaign_date:
                continue
            union    = _str(row[1])
            upazila  = _str(row[2])
            district = _str(row[3])
            households = _int(row[15])
            population = _int(row[16])
            if households == 0 and population == 0:
                continue   # empty row / total row
            # Compact deterministic key — case_hash is max_length=30, so
            # encode district+upazila+date through a short hash to avoid
            # collisions truncating the date component.
            import hashlib
            digest = hashlib.md5(
                f'{district}|{upazila}|{campaign_date.isoformat()}'.encode()
            ).hexdigest()[:18]
            key = f'MC-{digest}'   # 21 chars, well under the 30-char ceiling
            obj, was_created = FistulaCampaign.objects.update_or_create(
                case_hash=key,
                defaults=dict(
                    partner='CIPRB',
                    campaign_date=campaign_date,
                    district=district,
                    upazila=upazila,
                    union=union,
                    households_visited=households,
                    population_covered=population,
                    suspected_fistula_cases=_int(row[17]),
                    confirmed_fistula_cases=_int(row[18]),
                    cases_referred=_int(row[19]),
                    cases_surgery_completed=_int(row[20]),
                ),
            )
            if was_created:
                created += 1
            else:
                skipped += 1
    return created, skipped


class Command(BaseCommand):
    help = "Import Sayeed's CIPRB Excel files as historical baseline."

    def add_arguments(self, parser):
        parser.add_argument('--all', action='store_true', help='Run every importer')
        parser.add_argument('--denominators', action='store_true')
        parser.add_argument('--facility', action='store_true')
        parser.add_argument('--action-plan', action='store_true')
        parser.add_argument('--va-md', action='store_true')
        parser.add_argument('--fistula', action='store_true')
        parser.add_argument('--mass-campaign', action='store_true',
            help='Import Mass Campaign daily roll-up (households + population)')
        parser.add_argument('--mass-campaign-file', default=r'C:/Users/HP/Downloads/Mass Campaign on_ End Obstetric Fistula in Bangladesh.xlsx')
        parser.add_argument('--mpdsr-root', default=MPDSR_ROOT_DEFAULT)
        parser.add_argument('--fistula-root', default=FISTULA_ROOT_DEFAULT)

    def handle(self, *args, **opts):
        do_all = opts['all']
        mpdsr_root = opts['mpdsr_root']
        fistula_root = opts['fistula_root']

        if do_all or opts['denominators']:
            self.stdout.write('Importing district denominators...')
            c, u = import_denominators(mpdsr_root)
            self.stdout.write(self.style.SUCCESS(f'  OK created={c} updated={u}'))

        if do_all or opts['facility']:
            self.stdout.write('Importing FDN & FDR facility counts...')
            c, u = import_facility_counts(mpdsr_root)
            self.stdout.write(self.style.SUCCESS(f'  OK created={c} updated={u}'))

        if do_all or opts['action_plan']:
            self.stdout.write('Importing action plan summaries...')
            c, u = import_action_plan(mpdsr_root)
            self.stdout.write(self.style.SUCCESS(f'  OK created={c} updated={u}'))

        if do_all or opts['va_md']:
            self.stdout.write('Importing verbal autopsy maternal-death records...')
            c, s = import_va_md(mpdsr_root)
            self.stdout.write(self.style.SUCCESS(f'  OK created={c} skipped={s}'))

        if do_all or opts['fistula']:
            self.stdout.write('Importing Fistula Identified records...')
            c, s = import_fistula_identified(fistula_root)
            self.stdout.write(self.style.SUCCESS(f'  OK created={c} skipped={s}'))

        if do_all or opts['mass_campaign']:
            self.stdout.write('Importing Mass Campaign daily roll-ups...')
            c, s = import_mass_campaign(opts['mass_campaign_file'])
            self.stdout.write(self.style.SUCCESS(f'  OK created={c} updated={s}'))

        if not (do_all or any(opts[k] for k in
                ('denominators', 'facility', 'action_plan', 'va_md', 'fistula', 'mass_campaign'))):
            self.stdout.write('Pass --all or a specific flag. See --help.')
