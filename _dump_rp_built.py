import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
F = r'_ciprb_build/CIPRB-10_MPDSR_Response_Plan.xlsx'
wb = openpyxl.load_workbook(F)
out = []
ws = wb['survey']
hdr = [c.value for c in ws[1]]
def col(n):
    return hdr.index(n) if n in hdr else None
ti, ni = col('type'), col('name')
le, lb = col('label::English'), col('label::Bangla')
out.append('===== SURVEY =====')
for r in range(2, ws.max_row + 1):
    t = ws.cell(r, ti + 1).value
    if not t:
        continue
    nm = ws.cell(r, ni + 1).value or ''
    en = ws.cell(r, le + 1).value if le is not None else ''
    bn = ws.cell(r, lb + 1).value if lb is not None else ''
    out.append('%-22s | %-22s | EN=%s' % (t, nm, en))
    if bn:
        out.append('%-47s | BN=%s' % ('', bn))
cs = wb['choices']
chdr = [c.value for c in cs[1]]
li = chdr.index('list_name'); cn = chdr.index('name')
ce = chdr.index('label::English'); cb = chdr.index('label::Bangla')
out.append('\n===== CHOICES (rp_subcat) =====')
for r in range(2, cs.max_row + 1):
    if cs.cell(r, li + 1).value == 'rp_subcat':
        out.append('  %s | EN=%s | BN=%s' % (cs.cell(r, cn + 1).value,
                   cs.cell(r, ce + 1).value, cs.cell(r, cb + 1).value))
text = '\n'.join(out)
open(r'_rp_built_dump.md', 'w', encoding='utf-8').write(text)
print(text)
