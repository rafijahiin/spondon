# -*- coding: utf-8 -*-
"""Generate the PHD FSW Registry (Mother List) XLSForm for KoboToolbox.
Bilingual (English / Bangla). Run: python koboforms/_build_fsw_registry.py"""
import openpyxl, os

wb = openpyxl.Workbook()

# ── survey sheet ──────────────────────────────────────────────────────────────
sv = wb.active
sv.title = 'survey'
COLS = ['type', 'name', 'label::English (en)', 'label::Bangla (bn)',
        'hint::English (en)', 'required', 'relevant', 'constraint',
        'constraint_message::English (en)', 'appearance', 'default']
sv.append(COLS)


def row(**k):
    sv.append([k.get(c, '') for c in COLS])


L = 'label::English (en)'
B = 'label::Bangla (bn)'
H = 'hint::English (en)'
C = 'constraint'
CM = 'constraint_message::English (en)'

# metadata
for t, n in [('start', 'start'), ('end', 'end'), ('today', 'today'),
             ('username', 'username'), ('deviceid', 'deviceid')]:
    row(type=t, name=n)

# consent gate (sensitive population — informed consent first)
row(type='note', name='intro',
    **{L: 'PHD — FSW Registry (Mother List). Register each woman ONCE. The ID No. created here is used by every service form to auto-fill her details.',
       B: 'পিএইচডি — যৌনকর্মী নিবন্ধন (মাদারলিস্ট)। প্রত্যেক নারীকে একবার নিবন্ধন করুন। এখানে তৈরি আইডি নম্বরটি প্রতিটি সেবা ফর্মে তার তথ্য স্বয়ংক্রিয়ভাবে আনতে ব্যবহৃত হবে।'})
row(type='select_one yes_no', name='consent', required='yes',
    **{L: 'Informed consent obtained from the participant?',
       B: 'অংশগ্রহণকারীর কাছ থেকে অবহিত সম্মতি নেওয়া হয়েছে?'})
row(type='note', name='no_consent', relevant="${consent}='0'",
    **{L: 'Without consent the registration cannot continue. Please stop here.',
       B: 'সম্মতি ছাড়া নিবন্ধন চালিয়ে যাওয়া যাবে না। অনুগ্রহ করে এখানে থামুন।'})

REL = "${consent}='1'"

# ── identity ──────────────────────────────────────────────────────────────────
row(type='begin_group', name='identity', relevant=REL, **{L: 'Identity', B: 'পরিচয়'})
row(type='geopoint', name='gps', required='yes',
    **{L: 'Location (GPS)', B: 'অবস্থান (জিপিএস)',
       H: 'Enable phone GPS — location tagging is mandatory.'})
row(type='select_one wellness_center', name='wellness_center', required='yes',
    **{L: 'Wellness Center', B: 'ওয়েলনেস সেন্টার',
       H: 'Finalise the 9 centre names / IDs at the workshop.'})
row(type='text', name='id_no', required='yes',
    **{C: "regex(., '^[A-Za-z0-9-]+$')",
       L: 'ID No. (unique)', B: 'আইডি নম্বর (অনন্য)',
       H: 'Permanent unique ID for this woman. Letters, numbers and dashes only.',
       CM: 'Use only letters, numbers and dashes.'})
row(type='text', name='name', required='yes', **{L: 'Name', B: 'নাম'})
row(type='text', name='mother_name', **{L: "Mother's name", B: 'মাতার নাম'})
row(type='integer', name='birth_year', **{C: '. > 1940 and . <= 2010',
    L: 'Year of birth', B: 'জন্ম সাল', CM: 'Enter a valid birth year.'})
row(type='select_one gender', name='gender', required='yes', default='2',
    **{L: 'Sex', B: 'লিঙ্গ'})
row(type='text', name='permanent_address', appearance='multiline',
    **{L: 'Permanent address', B: 'স্থায়ী ঠিকানা'})
row(type='end_group', name='identity_end')

# ── socio-economic ────────────────────────────────────────────────────────────
row(type='begin_group', name='socio', relevant=REL, **{L: 'Socio-economic', B: 'আর্থ-সামাজিক'})
row(type='select_one education', name='education', **{L: 'Education', B: 'শিক্ষা'})
row(type='select_one marital_status', name='marital_status', **{L: 'Marital status', B: 'বৈবাহিক অবস্থা'})
row(type='select_one occupation', name='occupation', **{L: 'Occupation', B: 'পেশা'})
row(type='integer', name='age_started_profession', **{C: '. >= 0 and . <= 80',
    L: 'Age started this profession', B: 'কত বছর বয়স থেকে এই পেশা শুরু'})
row(type='integer', name='clients_freq_value', **{C: '. >= 0',
    L: 'Average number of sex-work contacts', B: 'সাধারণত গড়ে কতবার যৌনকাজ করেন'})
row(type='select_one freq_period', name='clients_freq_period', **{L: '...per', B: '...প্রতি'})
row(type='integer', name='children_under_18', **{C: '. >= 0 and . <= 20',
    L: 'Children under 18', B: '১৮ বছরের নিচে সন্তান সংখ্যা'})
row(type='end_group', name='socio_end')

# ── health / FP / status ──────────────────────────────────────────────────────
row(type='begin_group', name='health', relevant=REL, **{L: 'Health & status', B: 'স্বাস্থ্য ও অবস্থা'})
row(type='select_one yes_no', name='substance_use',
    **{L: 'Substance use (cigarette / drugs)?', B: 'নেশা গ্রহণ করেন কি না'})
row(type='text', name='substance_type', relevant="${substance_use}='1'",
    **{L: 'Which substance(s)?', B: 'কোন নেশা?'})
row(type='select_one yes_no', name='has_nid', **{L: 'Has National ID?', B: 'জাতীয় পরিচয়পত্র আছে কি না'})
row(type='select_one yes_no', name='fp_use',
    **{L: 'Using any family-planning method?', B: 'পরিবার পরিকল্পনার কোনো পদ্ধতি ব্যবহার করেন কি না'})
row(type='select_one fp_method', name='fp_method', relevant="${fp_use}='1'",
    **{L: 'Which FP method?', B: 'কোন পদ্ধতি?'})
row(type='select_one setting', name='setting', **{L: 'Sex-work setting', B: 'যৌনকাজের স্থান'})
row(type='select_one living_place', name='living_place', **{L: 'Living place', B: 'বসবাসের স্থান'})
row(type='select_one present_status', name='present_status', default='6',
    **{L: 'Present status', B: 'বর্তমান অবস্থা'})
row(type='text', name='phone', **{L: 'Phone (optional, for follow-up)', B: 'ফোন (ঐচ্ছিক)'})
row(type='text', name='remarks', appearance='multiline', **{L: 'Remarks', B: 'মন্তব্য'})
row(type='end_group', name='health_end')

# ── choices sheet ─────────────────────────────────────────────────────────────
ch = wb.create_sheet('choices')
ch.append(['list_name', 'name', 'label::English (en)', 'label::Bangla (bn)'])


def add(lst, items):
    for n, en, bn in items:
        ch.append([lst, n, en, bn])


add('yes_no', [('1', 'Yes', 'হ্যাঁ'), ('0', 'No', 'না')])
add('gender', [('1', 'Male', 'পুরুষ'), ('2', 'Female', 'নারী')])
add('marital_status', [('1', 'Single — never married', 'অবিবাহিত'), ('2', 'Married', 'বিবাহিত'),
    ('3', 'Widowed', 'বিধবা'), ('4', 'Separated', 'আলাদা'), ('5', 'Divorced', 'তালাকপ্রাপ্ত')])
add('education', [('1', 'Illiterate', 'নিরক্ষর'), ('2', 'Primary', 'প্রাথমিক'), ('3', 'Secondary', 'মাধ্যমিক'),
    ('4', 'Higher Secondary', 'উচ্চ মাধ্যমিক'), ('5', 'Graduate / Masters', 'স্নাতক/স্নাতকোত্তর'), ('6', 'Others', 'অন্যান্য')])
add('occupation', [('1', 'Service', 'চাকরি'), ('2', 'Business', 'ব্যবসা'), ('3', 'Student', 'ছাত্রী'),
    ('4', 'Housewife', 'গৃহিণী'), ('5', 'Unemployed', 'বেকার'), ('6', 'Others', 'অন্যান্য'),
    ('7', 'Labor / day labor', 'শ্রমিক/দিনমজুর'), ('8', 'Health worker', 'স্বাস্থ্যকর্মী')])
add('setting', [('1', 'Street', 'রাস্তা'), ('2', 'Hotel', 'হোটেল'), ('3', 'House', 'বাড়ি'), ('4', 'Both', 'উভয়')])
add('living_place', [('1', 'Street', 'রাস্তা'), ('2', 'House', 'বাড়ি'), ('3', 'Both', 'উভয়')])
add('present_status', [('6', 'Active', 'সক্রিয়'), ('1', 'Lost to follow up', 'ফলোআপ থেকে হারিয়ে গেছে'),
    ('2', 'In jail', 'জেলে'), ('3', 'Migrated', 'স্থানান্তরিত'), ('5', 'Died', 'মৃত'), ('4', 'Others', 'অন্যান্য')])
add('freq_period', [('1', 'Per day', 'প্রতিদিন'), ('2', 'Per week', 'সপ্তাহে'), ('3', 'Per month', 'মাসে'), ('4', 'Per year', 'বছরে')])
add('fp_method', [('1', 'Pill', 'বড়ি'), ('2', 'Condom', 'কনডম'), ('3', 'Injectable', 'ইনজেকশন'),
    ('4', 'IUD', 'আইইউডি'), ('5', 'Implant', 'ইমপ্ল্যান্ট'), ('6', 'None', 'কোনোটি নয়'), ('7', 'Other', 'অন্যান্য')])
add('wellness_center', [('wc_0%d' % i, 'Wellness Centre 0%d' % i, 'ওয়েলনেস সেন্টার ০%d' % i) for i in range(1, 10)])

# ── settings sheet ────────────────────────────────────────────────────────────
st = wb.create_sheet('settings')
st.append(['form_title', 'form_id', 'default_language', 'version', 'style'])
st.append(['PHD — FSW Registry (Mother List)', 'phd_fsw_registry_v1', 'English (en)', '2026060501', 'pages theme-grid'])

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'KF-FSW_Registry.xlsx')
wb.save(out)
print('WROTE', out)
print('survey rows:', sv.max_row - 1, '| choices rows:', ch.max_row - 1)
