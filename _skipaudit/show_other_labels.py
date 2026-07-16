import openpyxl, sys, os, re
sys.stdout.reconfigure(encoding='utf-8')
BD = r'C:\Users\HP\Documents\koboforms_baseline'
def show(path, label):
    wb = openpyxl.load_workbook(os.path.join(BD, path))
    sv, ch = wb['survey'], wb['choices']
    hdr = [c.value for c in sv[1]]; ti, ni = hdr.index('type'), hdr.index('name')
    rows = [[c.value for c in r] for r in sv.iter_rows(min_row=2)]
    names = {r[ni] for r in rows if r[ni]}
    chh = [c.value for c in ch[1]]; li, ci, lai = chh.index('list_name'), chh.index('name'), chh.index('label::English')
    lists = {}
    for r in ch.iter_rows(min_row=2):
        v = [c.value for c in r]
        if v[li]: lists.setdefault(v[li], []).append((str(v[ci]), str(v[lai] or '')))
    print(f'\n===== {label} =====')
    for r in rows:
        m = re.match(r'(select_one|select_multiple)\s+(\S+)', str(r[ti] or ''))
        if not m: continue
        field, ln = r[ni], m.group(2)
        oc = next((c for c,l in lists.get(ln,[]) if re.search(r'\bother\b', l, re.I)), None)
        if oc and f'{field}_other' not in names:
            olab = next(l for c,l in lists[ln] if c==oc)
            print(f'  {field:22s} code {oc:>3}  ->  "{olab}"')
show('CIPRB_Baseline_Hijra.xlsx','HIJRA')
show('CIPRB_Baseline_FSW.xlsx','FSW')
