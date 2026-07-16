"""Audit BOTH baseline forms for the systematic gap NK keeps flagging:
a choice list has an 'Other' option but the survey has no <field>_other text box.
Reads the generated xlsx (final truth)."""
import openpyxl, sys, os, re
sys.stdout.reconfigure(encoding='utf-8')
BD = r'C:\Users\HP\Documents\koboforms_baseline'

def audit(path, label):
    wb = openpyxl.load_workbook(os.path.join(BD, path))
    sv = wb['survey']; ch = wb['choices']
    hdr = [c.value for c in sv[1]]
    ti, ni, ri = hdr.index('type'), hdr.index('name'), hdr.index('relevant')
    rows = [[c.value for c in r] for r in sv.iter_rows(min_row=2)]
    names = {r[ni] for r in rows if r[ni]}
    # choices: list_name -> [(code,label)]
    chh = [c.value for c in ch[1]]
    li, ci, lai = chh.index('list_name'), chh.index('name'), chh.index('label::English')
    lists = {}
    for r in ch.iter_rows(min_row=2):
        v = [c.value for c in r]
        if v[li]:
            lists.setdefault(v[li], []).append((str(v[ci]), str(v[lai] or '')))
    def other_code(listname):
        for code, lab in lists.get(listname, []):
            if re.search(r'\bother\b', lab, re.I):
                return code
        return None
    gaps = []
    for r in rows:
        t = str(r[ti] or '')
        m = re.match(r'(select_one|select_multiple)\s+(\S+)', t)
        if not m:
            continue
        field, listname = r[ni], m.group(2)
        oc = other_code(listname)
        if oc is None:
            continue
        if f'{field}_other' not in names:
            gaps.append((field, listname, oc, r[ni]))
    print(f'\n===== {label}: {len(gaps)} Other-choice fields MISSING a text box =====')
    for field, ln, oc, _ in gaps:
        print(f'  {field:22s} list={ln:18s} Other-code={oc}')
    return gaps

audit('CIPRB_Baseline_Hijra.xlsx', 'HIJRA')
audit('CIPRB_Baseline_FSW.xlsx', 'FSW')
