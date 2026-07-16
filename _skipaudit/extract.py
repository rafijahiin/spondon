"""Extract EVERY skip-logic item from both baseline forms with full context for
exhaustive verification. Produces per-form item lists + a missing-skip candidate
list + mechanical pre-checks."""
import os, re, sys, json
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.join(os.path.expanduser('~'), 'Documents', 'koboforms_baseline')
OUT = os.path.dirname(os.path.abspath(__file__))
FORMS = [('hijra', 'CIPRB_Baseline_Hijra.xlsx'), ('fsw', 'CIPRB_Baseline_FSW.xlsx')]

REF = re.compile(r"\$\{(\w+)\}")
# comparisons like ${f}='v' or ${f}!='v'
CMP = re.compile(r"\$\{(\w+)\}\s*(!?=)\s*'([^']*)'")
# label keywords implying conditional display / routing
SKIP_KW = re.compile(
    r"(ask only if|ask if|only if|if q\d|if a\d|if b\d|if s\d|if c\d|if yes|if no|"
    r"skip to|go to|proceed to|→|=\s*2\b|=\s*1\b|shows any|if respondent|"
    r"if she|if he|if the respondent|if selected|if reported)",
    re.I)

summary = {}

for key, fname in FORMS:
    wb = openpyxl.load_workbook(os.path.join(BASE, fname))
    sv = wb['survey']; ch = wb['choices']
    h = [c.value for c in sv[1]]
    I = {name: h.index(name) for name in h}
    ti, ni, le, lb = I['type'], I['name'], I['label::English'], I['label::Bangla']
    ri, ai, ci = I['relevant'], I['appearance'], I['calculation']
    coni = I['constraint']

    rows = [list(r) for r in sv.iter_rows(min_row=2, values_only=True)]

    # choices: list_name -> {code: label}
    chh = [c.value for c in ch[1]]
    cl, cn, clab = chh.index('list_name'), chh.index('name'), chh.index('label::English')
    choices = {}
    for r in ch.iter_rows(min_row=2, values_only=True):
        if r[cl] is not None:
            choices.setdefault(r[cl], {})[str(r[cn])] = r[clab]

    # field dictionaries
    ftype, flist, flabel, fcalc, fpos = {}, {}, {}, {}, {}
    for i, r in enumerate(rows):
        nm = r[ni]
        if not nm:
            continue
        t = (r[ti] or '')
        base = t.split()[0] if t else ''
        ftype[nm] = base
        flabel[nm] = r[le]
        fpos[nm] = i
        if base in ('select_one', 'select_multiple'):
            parts = t.split()
            if len(parts) > 1:
                flist[nm] = parts[1]
        if base == 'calculate':
            fcalc[nm] = r[ci]

    def field_dict(names):
        d = {}
        for f in sorted(set(names)):
            info = {'type': ftype.get(f, 'UNKNOWN')}
            if f in flist:
                info['choices'] = choices.get(flist[f], {})
            if f in fcalc:
                info['calculation'] = fcalc[f]
            if flabel.get(f):
                info['label'] = str(flabel[f])[:80]
            d[f] = info
        return d

    # walk with group stack to compute ancestor gates
    items = []
    missing = []
    stack = []  # list of (group_name, group_relevant)
    for i, r in enumerate(rows):
        t = (r[ti] or ''); base = t.split()[0] if t else ''
        nm = r[ni]
        own_rel = r[ri]
        if base == 'end_group' or base == 'end_repeat':
            if stack:
                stack.pop()
            continue

        ancestor_gates = [{'group': g, 'relevant': gr} for g, gr in stack if gr]
        eff = [g['relevant'] for g in ancestor_gates] + ([own_rel] if own_rel else [])

        if base in ('begin_group', 'begin_repeat'):
            stack.append((nm, own_rel))

        label_en = r[le] or ''
        label_implies = bool(SKIP_KW.search(str(label_en)))

        # Build an item if it carries its own relevant, OR label implies a skip
        if own_rel or (label_implies and base not in ('end_group',)):
            refs = set()
            for expr in ([own_rel] if own_rel else []) + [g['relevant'] for g in ancestor_gates]:
                refs |= set(REF.findall(expr or ''))

            # mechanical checks on OWN relevant
            mech = {'value_issues': [], 'multi_eq': [], 'forward_ref': [],
                    'self_ref': False, 'unknown_ref': []}
            if own_rel:
                for f, op, v in CMP.findall(own_rel):
                    if f in flist:
                        if v not in choices.get(flist[f], {}):
                            mech['value_issues'].append(f"${{{f}}}{op}'{v}' : '{v}' not a code of {flist[f]}")
                        if ftype.get(f) == 'select_multiple':
                            mech['multi_eq'].append(f"${{{f}}}{op}'{v}' : {f} is select_multiple (use selected())")
                for f in REF.findall(own_rel):
                    if f not in ftype:
                        mech['unknown_ref'].append(f)
                    elif fpos.get(f, 10**9) > i:
                        mech['forward_ref'].append(f)
                    if f == nm:
                        mech['self_ref'] = True

            items.append({
                'idx': i + 2,  # xlsx row number
                'name': nm,
                'type': t,
                'base': base,
                'appearance': r[ai],
                'label_en': str(label_en),
                'label_bn': str(r[lb] or ''),
                'own_relevant': own_rel,
                'ancestor_gates': ancestor_gates,
                'effective_gates': eff,
                'constraint': r[coni],
                'calculation': r[ci],
                'label_implies_skip': label_implies,
                'referenced_fields': field_dict(refs),
                'mechanical': mech,
            })

            # missing-skip candidate: label implies a conditional but NOTHING gates it
            if label_implies and not eff and base in (
                    'select_one', 'select_multiple', 'integer', 'decimal', 'text', 'note', 'date'):
                missing.append({
                    'idx': i + 2, 'name': nm, 'type': t,
                    'label_en': str(label_en)[:120],
                    'why': 'label implies a conditional but no own relevant and no ancestor gate',
                })

    with open(os.path.join(OUT, f'{key}_items.json'), 'w', encoding='utf-8') as fh:
        json.dump(items, fh, ensure_ascii=False, indent=1)

    # global field dictionary so agents can resolve ANY field a label mentions
    # (needed to detect conditions the relevant OMITTED)
    fdict = {}
    for nm in ftype:
        info = {'type': ftype[nm], 'label': str(flabel.get(nm) or '')[:100]}
        if nm in flist:
            info['choices'] = choices.get(flist[nm], {})
        if nm in fcalc:
            info['calculation'] = fcalc[nm]
        fdict[nm] = info
    with open(os.path.join(OUT, f'{key}_fielddict.json'), 'w', encoding='utf-8') as fh:
        json.dump(fdict, fh, ensure_ascii=False, indent=1)

    # split items into chunks of CHUNK for the verification workflow
    CHUNK = 14
    manifest = []
    for start in range(0, len(items), CHUNK):
        cid = f'{key}_chunk_{start//CHUNK:02d}'
        sl = items[start:start + CHUNK]
        with open(os.path.join(OUT, cid + '.json'), 'w', encoding='utf-8') as fh:
            json.dump(sl, fh, ensure_ascii=False, indent=1)
        manifest.append({'chunk': cid, 'form': key, 'count': len(sl),
                         'names': [it['name'] for it in sl]})
    with open(os.path.join(OUT, f'{key}_manifest.json'), 'w', encoding='utf-8') as fh:
        json.dump(manifest, fh, ensure_ascii=False, indent=1)
    summary[key] = {
        'total_items': len(items),
        'with_own_relevant': sum(1 for it in items if it['own_relevant']),
        'label_implies_only': sum(1 for it in items if it['label_implies_skip'] and not it['own_relevant']),
        'mech_value_issues': sum(len(it['mechanical']['value_issues']) for it in items),
        'mech_multi_eq': sum(len(it['mechanical']['multi_eq']) for it in items),
        'mech_forward_ref': sum(len(it['mechanical']['forward_ref']) for it in items),
        'mech_unknown_ref': sum(len(it['mechanical']['unknown_ref']) for it in items),
        'missing_candidates': len(missing),
    }
    with open(os.path.join(OUT, f'{key}_missing.json'), 'w', encoding='utf-8') as fh:
        json.dump(missing, fh, ensure_ascii=False, indent=1)

with open(os.path.join(OUT, 'summary.json'), 'w', encoding='utf-8') as fh:
    json.dump(summary, fh, ensure_ascii=False, indent=1)
print(json.dumps(summary, ensure_ascii=False, indent=2))
