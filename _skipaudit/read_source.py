import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
p = r'C:\Users\HP\Downloads\Data_Collection_Plans_FSW_and_TGD-1.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
print('SHEETS:', wb.sheetnames)
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f'\n===== {sh}  ({ws.max_row} rows x {ws.max_column} cols) =====')
    # header row
    hdr = [c.value for c in ws[1]]
    print('HDR:', [str(h)[:22] for h in hdr])
    # print rows mentioning the NK-flagged items or skip/specify keywords
    KEY = ('4.18','4.19','4.7','4.8','b202','b203','b204','7.19','7.20','7.21','skip','specify','other','condom')
    hits = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        joined = ' | '.join(str(c) for c in r if c is not None)
        low = joined.lower()
        if any(k in low for k in KEY):
            print('  •', joined[:300])
            hits += 1
            if hits > 40:
                print('  ...(more)'); break
