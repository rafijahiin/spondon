import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
F = r'C:\Users\HP\Documents\koboforms\Bandhu-1_Service_Log.xlsx'
wb = openpyxl.load_workbook(F)
ws = wb['choices']
hdr = [c.value for c in ws[1]]
li = hdr.index('list_name'); ni = hdr.index('name')
ei = hdr.index('label::English'); bi = hdr.index('label::Bangla')
WANT = ['bandhu_centre', 'tg_code', 'general_diverse', 'education', 'marital',
        'ml_status', 'occupation', 'yn_code', 'f01_referral', 'sti_case',
        'diagnosis', 'hiv_result', 'counsel_issue', 'referred_for']
seen = {}
for r in range(2, ws.max_row + 1):
    ln = ws.cell(r, li + 1).value
    if ln in WANT:
        en = ws.cell(r, ei + 1).value
        bn = ws.cell(r, bi + 1).value
        seen.setdefault(ln, []).append((ws.cell(r, ni + 1).value, en, bn))
for ln in WANT:
    if ln in seen:
        print(f'\n=== {ln} ===')
        for name, en, bn in seen[ln][:6]:
            print(f'   value={name!r:10} EN={en!r:34} BN={bn!r}')
# also check settings default language
sset = wb['settings']
print('\nsettings:', [c.value for c in sset[1]], '->', [c.value for c in sset[2]])
