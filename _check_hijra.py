# -*- coding: utf-8 -*-
"""Structural QA on the built Hijra XLSForm: choice-list integrity (every
select references a defined list) + confirm the m4 skip fix landed."""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
P = r'C:\Users\HP\Documents\koboforms_baseline\CIPRB_Baseline_Hijra.xlsx'
wb = openpyxl.load_workbook(P)
sv, ch = wb['survey'], wb['choices']

refs, names = set(), []
for row in sv.iter_rows(min_row=2, values_only=True):
    t = row[0] or ''
    if isinstance(t, str) and (t.startswith('select_one ') or t.startswith('select_multiple ')):
        refs.add(t.split(' ', 1)[1].strip())
    if row[1]:
        names.append(row[1])
defined = {row[0] for row in ch.iter_rows(min_row=2, values_only=True) if row[0]}

orphans = sorted(refs - defined)
dup_names = sorted({n for n in names if names.count(n) > 1})
print('survey rows:', sv.max_row - 1, '| choice rows:', ch.max_row - 1)
print('select lists referenced:', len(refs), '| defined:', len(defined))
print('ORPHAN lists (referenced but no choices):', orphans or 'NONE ✓')
print('DUPLICATE question names:', dup_names or 'NONE ✓')

print('--- m4 skip fix ---')
for row in sv.iter_rows(min_row=2, values_only=True):
    if row[1] in ('q4_8', 'q4_9', 'q4_10', 'q4_11'):
        print(' ', row[1], 'guards q4_3:', 'q4_3' in (row[6] or ''))

ok = not orphans and not dup_names
print('\nSTRUCTURAL QA:', 'PASS ✓' if ok else 'ISSUES ABOVE ✗')
