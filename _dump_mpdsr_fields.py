import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
FORMS = [
    ('F01 Community Maternal', 'CIPRB-2_MPDSR_Form_01_Community_Maternal.xlsx'),
    ('F02 Community Neonatal', 'CIPRB-3_MPDSR_Form_02_Community_Neonatal.xlsx'),
    ('F04 Facility Maternal',  'CIPRB-4_MPDSR_Form_04_Facility_Maternal.xlsx'),
    ('F05 Facility Neonatal',  'CIPRB-5_MPDSR_Form_05_Facility_Neonatal.xlsx'),
]
out = []
for title, fn in FORMS:
    wb = openpyxl.load_workbook('_ciprb_build/' + fn)
    ws = wb['survey']
    hdr = [c.value for c in ws[1]]
    ti, ni = hdr.index('type'), hdr.index('name')
    le = hdr.index('label::English') if 'label::English' in hdr else None
    out.append('\n========== %s ==========' % title)
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, ti + 1).value
        if not t or t in ('begin_group', 'end_group', 'begin_repeat', 'end_repeat', 'note', 'calculate'):
            continue
        nm = ws.cell(r, ni + 1).value or ''
        en = (ws.cell(r, le + 1).value if le is not None else '') or ''
        out.append('%-26s %-22s %s' % (str(t)[:26], str(nm)[:22], str(en)[:60]))
txt = '\n'.join(out)
open('_mpdsr_fields_dump.md', 'w', encoding='utf-8').write(txt)
print('forms dumped; total lines', len(out))
